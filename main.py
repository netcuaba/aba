from fastapi import FastAPI, Request, Form, Depends, UploadFile, File, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, DateTime, ForeignKey, and_, or_, extract, func, UniqueConstraint
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from datetime import datetime, date, timedelta
import os
import io
import re
import unicodedata
import calendar
from typing import Optional, Tuple
from urllib.parse import quote
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Tạo database
SQLALCHEMY_DATABASE_URL = "sqlite:///./transport.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Tạo templates với custom filters
templates = Jinja2Templates(directory="templates")

# Thêm custom filter để parse JSON
def from_json(value):
    import json
    try:
        return json.loads(value) if value else []
    except:
        return []

# Thêm custom filter để escape JSON cho JavaScript (an toàn hơn |safe)
def tojson(value):
    import json
    import markupsafe
    try:
        # Xử lý None
        if value is None:
            return markupsafe.Markup('null')
        
        # Nếu đã là string
        if isinstance(value, str):
            # Nếu là string rỗng, trả về empty array
            if not value.strip():
                return markupsafe.Markup('[]')
            try:
                # Thử parse để kiểm tra xem có phải JSON string không
                parsed = json.loads(value)
                # Stringify lại để đảm bảo format đúng và escape đúng cách
                return markupsafe.Markup(json.dumps(parsed, ensure_ascii=False))
            except (json.JSONDecodeError, ValueError):
                # Nếu không phải JSON string hợp lệ, escape như string thông thường
                return markupsafe.Markup(json.dumps(value, ensure_ascii=False))
        
        # Nếu là object/list/dict, stringify trực tiếp
        return markupsafe.Markup(json.dumps(value, ensure_ascii=False))
    except Exception as e:
        # Nếu có lỗi, trả về empty array
        return markupsafe.Markup('[]')

# Custom filter để lấy attribute an toàn
def safe_getattr(value, attr_name=''):
    """Lấy attribute an toàn, trả về empty string nếu không có
    Usage trong template: {{ record|safe_getattr('route_type') }}
    """
    if value is None or not attr_name:
        return ''
    try:
        if hasattr(value, attr_name):
            result = getattr(value, attr_name, '')
            # Trả về empty string nếu None hoặc empty
            if result is None:
                return ''
            return str(result) if result else ''
        return ''
    except (AttributeError, TypeError, Exception):
        return ''

# Custom filter để convert UTC time sang UTC+7 (Bangkok, Hanoi, Jakarta)
def to_local_time(utc_datetime):
    """Convert UTC datetime sang UTC+7 và format thành dd/mm/yyyy HH:MM
    Usage trong template: {{ account.last_login|to_local_time }}
    """
    if utc_datetime is None:
        return 'Chưa đăng nhập'
    
    try:
        from datetime import timedelta
        # Thêm 7 giờ vào UTC time
        local_time = utc_datetime + timedelta(hours=7)
        # Format thành dd/mm/yyyy HH:MM
        return local_time.strftime('%d/%m/%Y %H:%M')
    except (AttributeError, TypeError, Exception) as e:
        # Fallback về format mặc định nếu có lỗi
        try:
            return utc_datetime.strftime('%d/%m/%Y %H:%M')
        except:
            return str(utc_datetime)

# Mapping từ page_path sang permission_code
PAGE_PERMISSION_MAP = {
    "/": "home.view",
    "/operations": "operations.view",
    "/employees": "employee.view",
    "/vehicles": "vehicle.view",
    "/routes": "route.view",
    "/timekeeping-v1": "timekeeping.view",
    "/maintenance": "maintenance.view",
    "/theo-doi-dau-v2": "fuel.view",
    "/salary-calculation-v2": "salary.view",
    "/finance-report": "finance.report.view",
    "/financial-statistics": "finance.statistics.view",
    "/administrative": "administrative.view",
    "/accounts": "account.view",  # Account list (admin only)
    "/revenue": "revenue.view",
    "/daily-new": "daily.view",
}

def get_permission_code_for_page(page_path: str) -> Optional[str]:
    """Lấy permission code tương ứng với page_path"""
    return PAGE_PERMISSION_MAP.get(page_path)

# Helper function để kiểm tra quyền trong template
def has_page_access(role: str, page_path: str, user_id: Optional[int] = None, db: Optional[Session] = None) -> bool:
    """Open access: luôn cho phép hiển thị menu/route."""
    return True

# Helper function để lấy ngày hiện tại cho templates
def get_today():
    """Trả về ngày hiện tại để sử dụng trong templates"""
    return date.today()

# Đăng ký filters và global functions
templates.env.filters["from_json"] = from_json
templates.env.filters["tojson"] = tojson
templates.env.filters["safe_getattr"] = safe_getattr
templates.env.filters["to_local_time"] = to_local_time
templates.env.globals["has_page_access"] = has_page_access
templates.env.globals["today"] = get_today

# Models
class Employee(Base):
    __tablename__ = "employees"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    birth_date = Column(Date)  # Ngày tháng năm sinh
    phone = Column(String)
    cccd = Column(String)  # Số CCCD
    cccd_issue_date = Column(Date)  # Ngày cấp CCCD
    cccd_expiry = Column(Date)  # Ngày hết hạn CCCD
    driving_license = Column(String)  # Số bằng lái xe
    license_expiry = Column(Date)  # Ngày hết hạn bằng lái
    documents = Column(String)  # Đường dẫn file upload giấy tờ (JSON array)
    status = Column(Integer, default=1)  # 1: Active, 0: Inactive
    employee_status = Column(String, default="Đang làm việc")  # Trạng thái: Đang làm việc, Đã nghỉ việc, Nghỉ phép dài hạn
    position = Column(String)  # Chức vụ: Giám đốc, Phó Giám đốc, Lái xe, Nhân viên văn phòng
    social_insurance_salary = Column(Integer)  # Mức lương tham gia BHXH (số nguyên)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships removed - no longer linked to routes

class Vehicle(Base):
    __tablename__ = "vehicles"
    
    id = Column(Integer, primary_key=True, index=True)
    license_plate = Column(String, unique=True, nullable=False)
    vehicle_type = Column(String, default="Xe Nhà")  # Loại xe: "Xe Nhà" hoặc "Xe Đối tác"
    capacity = Column(Float)  # Trọng tải
    fuel_consumption = Column(Float)  # Tiêu hao nhiên liệu
    inspection_expiry = Column(Date)  # Ngày hết hạn đăng kiểm
    inspection_documents = Column(String)  # Đường dẫn file upload sổ đăng kiểm (JSON array)
    phu_hieu_expired_date = Column(Date)  # Ngày hết hạn phù hiệu vận tải
    phu_hieu_files = Column(String)  # Đường dẫn file upload phù hiệu vận tải (JSON array)
    status = Column(Integer, default=1)  # 1: Active, 0: Inactive
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    routes = relationship("Route", back_populates="vehicle")
    maintenances = relationship("VehicleMaintenance", back_populates="vehicle")
    assignments = relationship("VehicleAssignment", back_populates="vehicle")

class VehicleAssignment(Base):
    """Bảng quản lý khoán xe cho tài xế"""
    __tablename__ = "vehicle_assignments"
    
    id = Column(Integer, primary_key=True, index=True)
    vehicle_id = Column(Integer, ForeignKey("vehicles.id"), nullable=False)
    employee_id = Column(Integer, ForeignKey("employees.id"), nullable=False)
    assignment_date = Column(Date, nullable=False)  # Ngày nhận xe
    end_date = Column(Date, nullable=True)  # Ngày kết thúc khoán (null nếu đang khoán)
    transfer_reason = Column(String, nullable=True)  # Lý do thu hồi/chuyển xe
    internal_note = Column(String, nullable=True)  # Ghi chú nội bộ
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    vehicle = relationship("Vehicle", back_populates="assignments")
    employee = relationship("Employee")

class VehicleMaintenance(Base):
    """Bảng quản lý bảo dưỡng xe"""
    __tablename__ = "vehicle_maintenances"
    
    id = Column(Integer, primary_key=True, index=True)
    vehicle_id = Column(Integer, ForeignKey("vehicles.id"), nullable=False)
    maintenance_date = Column(Date, nullable=False)  # Ngày bảo dưỡng
    maintenance_km = Column(Float, nullable=False)  # Số km bảo dưỡng
    vat_rate = Column(Float, default=0)  # VAT (%): 0, 5, 8, 10
    total_amount = Column(Float, default=0)  # Tổng cộng (chưa VAT)
    total_with_vat = Column(Float, default=0)  # Tổng cộng (bao gồm VAT)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    vehicle = relationship("Vehicle", back_populates="maintenances")
    items = relationship("VehicleMaintenanceItem", back_populates="maintenance", cascade="all, delete-orphan")

class VehicleMaintenanceItem(Base):
    """Bảng chi tiết hạng mục bảo dưỡng"""
    __tablename__ = "vehicle_maintenance_items"
    
    id = Column(Integer, primary_key=True, index=True)
    maintenance_id = Column(Integer, ForeignKey("vehicle_maintenances.id"), nullable=False)
    content = Column(String, nullable=False)  # Nội dung bảo dưỡng
    unit = Column(String)  # Đơn vị tính (ĐVT)
    quantity = Column(Float, default=0)  # Số lượng (SL)
    unit_price = Column(Float, default=0)  # Đơn giá
    discount_percent = Column(Float, default=0)  # Giảm giá (%)
    total_price = Column(Float, default=0)  # Thành tiền = SL × Đơn giá × (1 − Giảm giá / 100)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    maintenance = relationship("VehicleMaintenance", back_populates="items")

class Route(Base):
    __tablename__ = "routes"
    
    id = Column(Integer, primary_key=True, index=True)
    route_code = Column(String, nullable=False)  # NA_002, NA_004, etc.
    route_name = Column(String, nullable=False)
    distance = Column(Float)  # KM/Chuyến
    unit_price = Column(Float)  # Đơn giá (VNĐ)
    route_type = Column(String, nullable=False, default="Nội Tỉnh")  # Loại tuyến: Nội thành, Nội Tỉnh, Liên Tỉnh
    bridge_fee = Column(Float, default=0)  # Phí cầu đường (VNĐ) - chỉ cho Nội Tỉnh/Liên Tỉnh
    loading_fee = Column(Float, default=0)  # Phí chờ tải (VNĐ) - chỉ cho Nội Tỉnh/Liên Tỉnh
    monthly_salary = Column(Float)  # Lương tuyến/tháng
    vehicle_id = Column(Integer, ForeignKey("vehicles.id"), nullable=True)
    is_active = Column(Integer, default=1)
    status = Column(Integer, default=1)  # 1: Active, 0: Inactive
    route_status = Column(String, default="ONL")  # Trạng thái tuyến: ONL hoặc OFF
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    vehicle = relationship("Vehicle", back_populates="routes")
    daily_routes = relationship("DailyRoute", back_populates="route")

class DailyRoute(Base):
    __tablename__ = "daily_routes"
    
    id = Column(Integer, primary_key=True, index=True)
    route_id = Column(Integer, ForeignKey("routes.id"))
    date = Column(Date, nullable=False)
    distance_km = Column(Float, default=0)  # Số km
    cargo_weight = Column(Float, default=0)  # Tải trọng
    driver_name = Column(String)  # Tên lái xe
    license_plate = Column(String)  # Biển số xe
    employee_name = Column(String)  # Tên nhân viên
    status = Column(String, default="Online")  # Trạng thái: Online hoặc OFF
    notes = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    route = relationship("Route", back_populates="daily_routes")

class FuelRecord(Base):
    __tablename__ = "fuel_records"
    
    id = Column(Integer, primary_key=True, index=True)
    date = Column(Date, nullable=False)  # Ngày đổ dầu
    fuel_type = Column(String, default="Dầu DO 0,05S-II")  # Loại dầu
    license_plate = Column(String, nullable=False)  # Biển số xe
    fuel_price_per_liter = Column(Float, default=0)  # Giá xăng dầu hôm nay (đồng/lít)
    liters_pumped = Column(Float, default=0)  # Số lít dầu đã đổ
    cost_pumped = Column(Float, default=0)  # Số tiền dầu đã đổ (tự động tính)
    notes = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    vehicle = relationship("Vehicle", foreign_keys=[license_plate], primaryjoin="FuelRecord.license_plate == Vehicle.license_plate")

class DieselPriceHistory(Base):
    """Bảng lưu lịch sử giá dầu Diesel 0.05S theo từng thời điểm"""
    __tablename__ = "diesel_price_history"
    
    id = Column(Integer, primary_key=True, index=True)
    application_date = Column(Date, nullable=False, unique=True)  # Ngày áp dụng giá (unique để tránh trùng)
    unit_price = Column(Integer, nullable=False)  # Đơn giá dầu Diesel 0.05S (VNĐ) - số nguyên
    created_at = Column(DateTime, default=datetime.utcnow)  # Ngày tạo bản ghi
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)  # Ngày cập nhật

class FinanceRecord(Base):
    __tablename__ = "finance_records"
    
    id = Column(Integer, primary_key=True, index=True)
    date = Column(Date, nullable=False)  # Ngày giao dịch
    category = Column(String, nullable=False)  # Danh mục (Thu/Chi)
    description = Column(String, nullable=False)  # Diễn giải/Tên khách hàng
    route_code = Column(String)  # Mã tuyến
    amount_before_vat = Column(Float, default=0)  # Số tiền (chưa VAT)
    vat_rate = Column(Float, default=0)  # VAT (%)
    discount1_rate = Column(Float, default=0)  # Chiết khấu 1 (%)
    discount2_rate = Column(Float, default=0)  # Chiết khấu 2 (%)
    final_amount = Column(Float, default=0)  # Thành tiền (tự động tính)
    income = Column(Float, default=0)  # Số tiền thu (để tương thích)
    expense = Column(Float, default=0)  # Số tiền chi (để tương thích)
    balance = Column(Float, default=0)  # Thành tiền (tự động tính)
    notes = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)

class FinanceTransaction(Base):
    """Bảng riêng biệt chuyên quản lý dữ liệu thu chi độc lập"""
    __tablename__ = "finance_transactions"
    
    id = Column(Integer, primary_key=True, index=True)
    transaction_type = Column(String, nullable=False)  # Thu/Chi
    category = Column(String, nullable=False)  # Danh mục
    date = Column(Date, nullable=False)  # Ngày thu/chi
    description = Column(String, nullable=False)  # Diễn giải
    route_code = Column(String)  # Mã tuyến (nếu có)
    amount = Column(Float, default=0)  # Số tiền chưa VAT
    vat = Column(Float, default=0)  # VAT (%)
    discount1 = Column(Float, default=0)  # Chiết khấu 1 (%)
    discount2 = Column(Float, default=0)  # Chiết khấu 2 (%)
    total = Column(Float, default=0)  # Thành tiền
    note = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class RevenueRecord(Base):
    """Bảng quản lý doanh thu hàng ngày theo tuyến"""
    __tablename__ = "revenue_records"
    
    id = Column(Integer, primary_key=True, index=True)
    date = Column(Date, nullable=False)  # Ngày ghi nhận doanh thu
    route_id = Column(Integer, ForeignKey("routes.id"), nullable=False)  # ID tuyến
    route_type = Column(String, default="Tăng cường Nội Tỉnh")  # Loại tuyến: Nội thành, Tăng cường Nội Tỉnh, Tăng cường Liên Tỉnh
    distance_km = Column(Float, default=0)  # Khoảng cách (có thể chỉnh sửa từ routes)
    unit_price = Column(Integer, default=0)  # Đơn giá (VNĐ/km) - số nguyên
    bridge_fee = Column(Integer, default=0)  # Phí cầu đường - số nguyên
    loading_fee = Column(Integer, default=0)  # Phí dừng tải - số nguyên
    late_penalty = Column(Integer, default=0)  # Trễ Ontime - số nguyên
    status = Column(String, default="Online")  # Trạng thái: Online/Offline
    total_amount = Column(Integer, default=0)  # Thành tiền = (Khoảng cách x Đơn giá) + Phí cầu đường + Phí dừng tải – Trễ Ontime
    manual_total = Column(Integer, default=0)  # Thành tiền nhập thủ công (dùng khi Offline hoặc muốn ghi đè)
    route_name = Column(String)  # Lộ trình (cho tuyến tăng cường)
    license_plate = Column(String)  # Biển số xe
    driver_name = Column(String)  # Tên tài xế
    notes = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    route = relationship("Route")

class Account(Base):
    """Bảng quản lý tài khoản người dùng"""
    __tablename__ = "accounts"
    
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, nullable=False)  # Tên đăng nhập
    password_hash = Column(String, nullable=True)  # Mật khẩu đã hash (nullable để backward compatibility)
    password = Column(String, nullable=True)  # Mật khẩu plain text (legacy, kept for backward compatibility)
    full_name = Column(String)  # Họ tên
    email = Column(String)  # Email
    phone = Column(String)  # Số điện thoại
    role = Column(String, default="User")  # Phân quyền: Admin, Manager, User (legacy field, kept for backward compatibility)
    status = Column(String, default="Active")  # Trạng thái: Active, Inactive
    is_active = Column(Integer, default=1)  # 1: Active, 0: Inactive (chuẩn RBAC)
    is_locked = Column(Integer, default=0)  # 0: Mở, 1: Khoá
    locked_at = Column(DateTime, nullable=True)  # Thời điểm khoá
    locked_by = Column(Integer, ForeignKey("accounts.id"), nullable=True)  # Ai khoá
    last_login = Column(DateTime, nullable=True)  # Lần đăng nhập cuối
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    locker = relationship("Account", remote_side=[id], foreign_keys=[locked_by])
    user_roles = relationship("UserRole", primaryjoin="Account.id == UserRole.user_id", back_populates="user", cascade="all, delete-orphan")
    user_permissions = relationship("UserPermission", back_populates="user", cascade="all, delete-orphan")

class Role(Base):
    """Bảng quản lý vai trò"""
    __tablename__ = "roles"
    
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, nullable=True)  # Mã vai trò: ADMIN, MANAGER, USER (nullable để backward compatibility)
    name = Column(String, unique=True, nullable=False)  # Tên vai trò: Super Admin, Admin Operations, etc.
    description = Column(String)  # Mô tả vai trò
    is_system_role = Column(Integer, default=0)  # 1: System role (cannot delete), 0: Custom role
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    user_roles = relationship("UserRole", back_populates="role", cascade="all, delete-orphan")
    role_permissions = relationship("RolePermission", back_populates="role", cascade="all, delete-orphan")

class UserRole(Base):
    """Bảng mapping vai trò cho user (many-to-many)"""
    __tablename__ = "user_roles"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("accounts.id", ondelete="CASCADE"), nullable=False)
    role_id = Column(Integer, ForeignKey("roles.id", ondelete="CASCADE"), nullable=False)
    assigned_by = Column(Integer, ForeignKey("accounts.id"), nullable=True)  # Who assigned this role
    assigned_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    user = relationship("Account", foreign_keys=[user_id], back_populates="user_roles")
    role = relationship("Role", back_populates="user_roles")
    assigner = relationship("Account", foreign_keys=[assigned_by])

class Permission(Base):
    """Bảng quản lý quyền truy cập"""
    __tablename__ = "permissions"
    
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, nullable=True)  # Mã quyền: user.view, user.edit, role.manage (primary identifier) - nullable để backward compatibility
    name = Column(String, unique=True, nullable=False)  # Tên quyền (VD: employees.view, employees.create) - kept for backward compatibility
    description = Column(String)  # Mô tả quyền
    page_path = Column(String)  # Đường dẫn page (VD: /employees) - kept for backward compatibility
    action = Column(String)  # Hành động: view, create, update, delete - kept for backward compatibility
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    role_permissions = relationship("RolePermission", back_populates="permission", cascade="all, delete-orphan")
    user_permissions = relationship("UserPermission", back_populates="permission", cascade="all, delete-orphan")

class RolePermission(Base):
    """Bảng mapping quyền cho vai trò"""
    __tablename__ = "role_permissions"
    
    id = Column(Integer, primary_key=True, index=True)
    role_id = Column(Integer, ForeignKey("roles.id", ondelete="CASCADE"), nullable=False)  # Changed from role String to role_id FK
    permission_id = Column(Integer, ForeignKey("permissions.id", ondelete="CASCADE"), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    role = relationship("Role", back_populates="role_permissions")
    permission = relationship("Permission", back_populates="role_permissions")

class UserPermission(Base):
    """Bảng mapping quyền cho từng user cụ thể (override role permissions)"""
    __tablename__ = "user_permissions"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("accounts.id", ondelete="CASCADE"), nullable=False)
    permission_id = Column(Integer, ForeignKey("permissions.id", ondelete="CASCADE"), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    user = relationship("Account", foreign_keys=[user_id], back_populates="user_permissions")
    permission = relationship("Permission", back_populates="user_permissions")

class Document(Base):
    """Bảng quản lý tài liệu hành chính (Legal, Administrative/HR, Tax)"""
    __tablename__ = "documents"
    
    id = Column(Integer, primary_key=True, index=True)
    
    # Category: legal, administrative, tax
    category = Column(String, nullable=False)  # legal, administrative, tax
    
    # Document type
    document_type = Column(String, nullable=False)  # e.g., 'contract', 'license', 'tax_return'
    
    # Related entity (polymorphic)
    related_entity_type = Column(String, nullable=True)  # e.g., 'vehicle', 'employee', 'company'
    related_entity_id = Column(Integer, nullable=True)  # ID of related entity
    
    # Document details
    title = Column(String, nullable=False)
    file_path = Column(String, nullable=False)  # Relative path to file
    
    # Dates
    issued_date = Column(Date, nullable=True)
    expiry_date = Column(Date, nullable=True)  # Nullable
    
    # Status
    status = Column(String, default="active")  # active, expired, archived
    
    # Metadata
    description = Column(String, nullable=True)
    notes = Column(String, nullable=True)
    
    # Audit fields
    created_at = Column(DateTime, default=datetime.utcnow)
    created_by = Column(Integer, ForeignKey("accounts.id"), nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by = Column(Integer, ForeignKey("accounts.id"), nullable=True)
    
    # Relationships
    creator = relationship("Account", foreign_keys=[created_by])
    updater = relationship("Account", foreign_keys=[updated_by])

class AuditLog(Base):
    """Bảng nhật ký hệ thống - ghi lại mọi thay đổi"""
    __tablename__ = "audit_logs"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("accounts.id"), nullable=False)  # Người thực hiện
    action = Column(String, nullable=False)  # Hành động: create, update, delete, lock, unlock, reset_password
    entity_type = Column(String, nullable=False)  # Loại entity: account, permission, etc.
    entity_id = Column(Integer, nullable=True)  # ID của entity bị thay đổi
    old_values = Column(String)  # Giá trị cũ (JSON)
    new_values = Column(String)  # Giá trị mới (JSON)
    description = Column(String)  # Mô tả chi tiết
    ip_address = Column(String)  # IP address
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    user = relationship("Account", foreign_keys=[user_id])

class TimekeepingTable(Base):
    """Bảng quản lý bảng chấm công V1"""
    __tablename__ = "timekeeping_tables"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)  # Tên bảng chấm công
    from_date = Column(Date, nullable=False)  # Từ ngày
    to_date = Column(Date, nullable=False)  # Đến ngày
    created_at = Column(DateTime, default=datetime.utcnow)  # Ngày tạo


class TimekeepingDetail(Base):
    """Dữ liệu chi tiết cho từng sheet/tuyến trong bảng chấm công V1"""
    __tablename__ = "timekeeping_details"

    id = Column(Integer, primary_key=True, index=True)
    table_id = Column(Integer, ForeignKey("timekeeping_tables.id"), nullable=False)
    sheet_name = Column(String, nullable=False)  # Tên sheet (route_code hoặc route_name)
    route_code = Column(String)
    route_name = Column(String)
    route_type = Column(String)
    itinerary = Column(String)  # Lộ trình
    date = Column(Date, nullable=False)
    license_plate = Column(String)
    driver_name = Column(String)
    trip_code = Column(String)  # Mã chuyến
    notes = Column(String)  # Ghi chú
    status = Column(String, default="Onl")  # Status: Onl hoặc OFF
    distance_km = Column(Float, default=0)
    unit_price = Column(Float, default=0)
    bridge_fee = Column(Float, default=0)
    loading_fee = Column(Float, default=0)
    total_amount = Column(Float, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class RoutePrice(Base):
    """Bảng quản lý giá tuyến theo ngày áp dụng"""
    __tablename__ = "route_prices"
    
    id = Column(Integer, primary_key=True, index=True)
    route_id = Column(Integer, ForeignKey("routes.id"), nullable=False)  # ID tuyến
    unit_price = Column(Integer, nullable=False)  # Đơn giá (VNĐ) - số nguyên
    fuel_price = Column(Integer, nullable=False)  # Áp dụng giá dầu (VNĐ) - số nguyên
    application_date = Column(Date, nullable=False)  # Ngày áp giá
    update_name = Column(String)  # Tên bản cập nhật giá tuyến
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    route = relationship("Route")

class SalaryMonthly(Base):
    """Bảng lưu snapshot lương tháng cho từng lái xe"""
    __tablename__ = "salary_monthly"
    
    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, ForeignKey("employees.id"), nullable=False)  # ID nhân viên/lái xe
    month = Column(Integer, nullable=False)  # Tháng (1-12)
    year = Column(Integer, nullable=False)  # Năm
    # Các trường manual
    bao_hiem_xh = Column(Integer, default=0)  # Bảo hiểm XH (VNĐ)
    rua_xe = Column(Integer, default=0)  # Rửa xe (VNĐ)
    tien_trach_nhiem = Column(Integer, default=0)  # Tiền trách nhiệm (VNĐ)
    ung_luong = Column(Integer, default=0)  # Ứng lương (VNĐ)
    sua_xe = Column(Integer, default=0)  # Sửa xe (VNĐ)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    employee = relationship("Employee")
    
    # Unique constraint: một nhân viên chỉ có một bản ghi cho mỗi tháng/năm
    __table_args__ = (
        UniqueConstraint('employee_id', 'month', 'year', name='uq_salary_monthly_employee_month_year'),
    )


# Tạo bảng
Base.metadata.create_all(bind=engine)

# Migration: Thêm các cột mới vào bảng accounts nếu chưa có
def migrate_accounts():
    """Thêm các cột mới vào bảng accounts nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'accounts' not in inspector.get_table_names():
            print("Table accounts does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('accounts')]
        
        with engine.connect() as conn:
            # Thêm các cột mới nếu chưa có
            if 'full_name' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN full_name VARCHAR"))
                print("Added column full_name to accounts table")
            
            if 'email' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN email VARCHAR"))
                print("Added column email to accounts table")
            
            if 'phone' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN phone VARCHAR"))
                print("Added column phone to accounts table")
            
            if 'status' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN status VARCHAR DEFAULT 'Active'"))
                # Cập nhật các record cũ thành Active
                conn.execute(text("UPDATE accounts SET status = 'Active' WHERE status IS NULL"))
                conn.commit()
                print("Added column status to accounts table")
            
            if 'is_locked' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN is_locked INTEGER DEFAULT 0"))
                print("Added column is_locked to accounts table")
            
            if 'locked_at' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN locked_at DATETIME"))
                print("Added column locked_at to accounts table")
            
            if 'locked_by' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN locked_by INTEGER"))
                print("Added column locked_by to accounts table")
            
            if 'last_login' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN last_login DATETIME"))
                print("Added column last_login to accounts table")
            
            # RBAC refactor: Thêm password_hash và is_active
            if 'password_hash' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN password_hash VARCHAR"))
                # Copy password sang password_hash cho các accounts hiện có
                conn.execute(text("UPDATE accounts SET password_hash = password WHERE password_hash IS NULL"))
                print("Added column password_hash to accounts table")
            
            if 'is_active' not in existing_columns:
                conn.execute(text("ALTER TABLE accounts ADD COLUMN is_active INTEGER DEFAULT 1"))
                # Set is_active = 1 cho các accounts hiện có
                conn.execute(text("UPDATE accounts SET is_active = 1 WHERE is_active IS NULL"))
                print("Added column is_active to accounts table")
            
            conn.commit()
            
    except Exception as e:
        print(f"Migration error for accounts: {e}")

# Chạy migration
migrate_accounts()

# Migration: Thêm các cột mới vào bảng revenue_records nếu chưa có
def migrate_revenue_records():
    """Thêm các cột mới vào bảng revenue_records nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'revenue_records' not in inspector.get_table_names():
            print("Table revenue_records does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('revenue_records')]
        
        new_columns = {
            'route_type': 'VARCHAR',
            'route_name': 'VARCHAR',
            'license_plate': 'VARCHAR',
            'driver_name': 'VARCHAR'
        }
        
        with engine.connect() as conn:
            for col_name, col_type in new_columns.items():
                if col_name not in existing_columns:
                    try:
                        conn.execute(text(f"ALTER TABLE revenue_records ADD COLUMN {col_name} {col_type}"))
                        conn.commit()
                        print(f"Added column {col_name} to revenue_records")
                    except Exception as e:
                        print(f"Error adding column {col_name}: {e}")
                        conn.rollback()
    except Exception as e:
        print(f"Migration error: {e}")

# Migration: Thêm các cột mới vào bảng timekeeping_details nếu chưa có
def migrate_timekeeping_details():
    """Thêm các cột mới vào bảng timekeeping_details nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'timekeeping_details' not in inspector.get_table_names():
            print("Table timekeeping_details does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('timekeeping_details')]
        
        new_columns = {
            'trip_code': 'VARCHAR',
            'notes': 'VARCHAR',
            'status': 'VARCHAR'
        }
        
        with engine.connect() as conn:
            for col_name, col_type in new_columns.items():
                if col_name not in existing_columns:
                    try:
                        conn.execute(text(f"ALTER TABLE timekeeping_details ADD COLUMN {col_name} {col_type}"))
                        conn.commit()
                        print(f"Added column {col_name} to timekeeping_details")
                        
                        # Nếu là cột status, set giá trị mặc định 'Onl' cho các row hiện có
                        if col_name == 'status':
                            conn.execute(text("UPDATE timekeeping_details SET status = 'Onl' WHERE status IS NULL"))
                            conn.commit()
                            print(f"Set default value 'Onl' for existing rows in status column")
                    except Exception as e:
                        print(f"Error adding column {col_name}: {e}")
                        conn.rollback()
    except Exception as e:
        print(f"Migration error: {e}")

# Migration: Thêm cột update_name vào bảng route_prices nếu chưa có
def migrate_route_prices():
    """Thêm cột update_name vào bảng route_prices nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'route_prices' not in inspector.get_table_names():
            print("Table route_prices does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('route_prices')]
        
        if 'update_name' not in existing_columns:
            with engine.connect() as conn:
                try:
                    conn.execute(text("ALTER TABLE route_prices ADD COLUMN update_name VARCHAR"))
                    conn.commit()
                    print("Added column update_name to route_prices")
                except Exception as e:
                    print(f"Error adding column update_name: {e}")
                    conn.rollback()
    except Exception as e:
        print(f"Migration error: {e}")

# Helper function để lấy giá tuyến theo ngày
def get_route_price_by_date(db: Session, route_id: int, target_date: date) -> Optional[RoutePrice]:
    """
    Lấy giá tuyến áp dụng cho một ngày cụ thể.
    Trả về giá tuyến có application_date <= target_date và gần nhất với target_date.
    Nếu không tìm thấy, trả về None.
    """
    route_price = db.query(RoutePrice).filter(
        RoutePrice.route_id == route_id,
        RoutePrice.application_date <= target_date
    ).order_by(RoutePrice.application_date.desc()).first()
    
    return route_price

# Helper function để lấy giá dầu theo ngày
def get_fuel_price_by_date(db: Session, target_date: date) -> Optional[DieselPriceHistory]:
    """
    Lấy giá dầu Diesel 0.05S áp dụng cho một ngày cụ thể.
    Trả về giá dầu có application_date <= target_date và gần nhất với target_date.
    Nếu không tìm thấy, trả về None.
    """
    fuel_price = db.query(DieselPriceHistory).filter(
        DieselPriceHistory.application_date <= target_date
    ).order_by(DieselPriceHistory.application_date.desc()).first()
    
    return fuel_price

# Helper function để lấy định mức nhiên liệu của xe
def is_route_off_on_date(db: Session, route_code: str, date: date, license_plate: str) -> bool:
    """
    Kiểm tra xem tuyến có bị OFF trong ngày đó không.
    
    Logic:
    - Tìm DailyRoute với cùng date, license_plate, và route_code
    - Nếu tất cả DailyRoute của route đó trong ngày đều có status = OFF → return True
    - Nếu có ít nhất 1 DailyRoute có status = ONLINE/ON → return False
    - Nếu không tìm thấy DailyRoute → return False (không có dữ liệu chấm công hàng ngày)
    
    Args:
        db: Database session
        route_code: Mã tuyến (route_code)
        date: Ngày cần kiểm tra
        license_plate: Biển số xe
    
    Returns:
        True nếu route bị OFF trong ngày đó, False nếu không
    """
    try:
        # Tìm Route theo route_code
        route = db.query(Route).filter(
            Route.route_code == route_code.strip(),
            Route.status == 1,
            Route.is_active == 1
        ).first()
        
        if not route:
            # Không tìm thấy route → không có dữ liệu → không OFF
            return False
        
        # Tìm DailyRoute với cùng route_id, date, và license_plate
        daily_routes = db.query(DailyRoute).filter(
            DailyRoute.route_id == route.id,
            DailyRoute.date == date,
            DailyRoute.license_plate == license_plate.strip()
        ).all()
        
        if not daily_routes:
            # Không có DailyRoute → không có dữ liệu chấm công hàng ngày → không OFF
            return False
        
        # Kiểm tra: nếu TẤT CẢ DailyRoute đều OFF → route bị OFF
        all_off = True
        for dr in daily_routes:
            dr_status = (dr.status or "").strip().upper()
            if dr_status in ["ONLINE", "ON"]:
                all_off = False
                break
        
        return all_off
    
    except Exception as e:
        print(f"Error checking route status for {route_code} on {date}: {e}")
        # Nếu có lỗi, mặc định là không OFF để tránh bỏ sót dữ liệu
        return False

def get_vehicle_fuel_consumption(db: Session, license_plate: str) -> Optional[float]:
    """
    Lấy định mức nhiên liệu (lít/100km) của xe theo biển số.
    Trả về giá trị định mức hoặc None nếu không tìm thấy.
    """
    if not license_plate or not license_plate.strip():
        return None
    
    vehicle = db.query(Vehicle).filter(
        Vehicle.license_plate == license_plate.strip(),
        Vehicle.status == 1
    ).first()
    
    if vehicle and vehicle.fuel_consumption is not None:
        return vehicle.fuel_consumption
    
    return None

def check_vehicle_assignment_for_trip(db: Session, license_plate: str, driver_name: str, trip_date: date) -> Tuple[bool, Optional[str]]:
    """
    Kiểm tra xem xe có đang được khoán cho đúng lái xe tại thời điểm chạy chuyến không.
    
    Điều kiện để tính tiền dầu:
    1. Có biển số xe
    2. Có lái xe
    3. Xe đang ở trạng thái Khoán xe = Active
    4. Lái xe của chuyến = Lái xe đang khoán xe
    5. Ngày chuyến nằm trong khoảng khoán (assignment_date <= trip_date < end_date hoặc end_date is null)
    
    Args:
        db: Database session
        license_plate: Biển số xe
        driver_name: Tên lái xe
        trip_date: Ngày chạy chuyến
    
    Returns:
        Tuple[bool, Optional[str]]: 
        - (True, None) nếu đúng khoán
        - (False, reason) nếu không đúng khoán (reason là lý do)
    """
    # Kiểm tra điều kiện cơ bản
    if not license_plate or not license_plate.strip():
        return (False, "Không có biển số xe")
    
    if not driver_name or not driver_name.strip():
        return (False, "Không có lái xe")
    
    if not trip_date:
        return (False, "Không có ngày chạy chuyến")
    
    # Lấy thông tin xe
    vehicle = db.query(Vehicle).filter(
        Vehicle.license_plate == license_plate.strip(),
        Vehicle.status == 1
    ).first()
    
    if not vehicle:
        return (False, "Xe không tồn tại hoặc đã bị vô hiệu hóa")
    
    # Xe đối tác không tính tiền dầu
    if vehicle.vehicle_type == "Xe Đối tác":
        return (False, "Xe đối tác")
    
    # Lấy thông tin lái xe
    employee = db.query(Employee).filter(
        Employee.name == driver_name.strip(),
        Employee.status == 1
    ).first()
    
    if not employee:
        return (False, "Lái xe không tồn tại trong hệ thống")
    
    # Kiểm tra khoán xe tại ngày chạy chuyến
    # Tìm assignment hợp lệ: assignment_date <= trip_date < end_date (hoặc end_date is null)
    assignment = db.query(VehicleAssignment).join(Vehicle).filter(
        Vehicle.license_plate == license_plate.strip(),
        VehicleAssignment.employee_id == employee.id,
        VehicleAssignment.assignment_date <= trip_date,
        or_(
            VehicleAssignment.end_date.is_(None),
            VehicleAssignment.end_date > trip_date
        )
    ).first()
    
    if not assignment:
        # Kiểm tra xem có assignment nào cho xe này không (để biết lý do)
        any_assignment = db.query(VehicleAssignment).join(Vehicle).filter(
            Vehicle.license_plate == license_plate.strip()
        ).first()
        
        if not any_assignment:
            return (False, "Xe chưa được khoán cho ai")
        else:
            # Xe đã được khoán nhưng không phải cho lái xe này hoặc không đúng thời điểm
            return (False, "Xe không khoán cho lái xe này tại thời điểm chạy chuyến")
    
    # Tất cả điều kiện đều thỏa mãn
    return (True, None)

# Helper function để tính dầu khoán (DK) và tiền dầu
def calculate_fuel_quota(result: TimekeepingDetail, db: Session) -> dict:
    """
    Tính số lít dầu khoán (DK) và tiền dầu cho một chuyến.
    
    QUY ĐỊNH: Tiền dầu CHỈ ĐƯỢC TÍNH khi:
    - Có biển số xe
    - Có lái xe
    - Xe đang ở trạng thái Khoán xe = Active
    - Lái xe của chuyến = Lái xe đang khoán xe
    - Ngày chuyến nằm trong khoảng khoán
    
    Nếu không đúng → Tiền dầu = 0
    
    Trả về dictionary với các key:
    - dk_liters: Số lít dầu khoán (float, tối đa 2 chữ số thập phân)
    - fuel_cost: Tiền dầu (int, số nguyên)
    - fuel_price: Đơn giá dầu (int, None nếu không có)
    - fuel_consumption: Định mức nhiên liệu (float, None nếu không có)
    - warning: Thông báo cảnh báo (string, None nếu không có)
    - assignment_status: Trạng thái khoán xe ("valid", "invalid", "no_assignment", "partner_vehicle")
    - assignment_reason: Lý do không tính tiền dầu (string, None nếu tính được)
    """
    # Khởi tạo kết quả
    result_dict = {
        "dk_liters": 0.0,
        "fuel_cost": 0,
        "fuel_price": None,
        "fuel_consumption": None,
        "warning": None,
        "assignment_status": None,
        "assignment_reason": None
    }
    
    # Kiểm tra nếu status là OFF, không tính
    if result.status and (result.status.strip().upper() == "OFF"):
        return result_dict
    
    # Lấy thông tin cơ bản
    trip_date = result.date
    license_plate = result.license_plate
    driver_name = result.driver_name
    distance_km = result.distance_km or 0
    
    if not trip_date or not license_plate or distance_km <= 0:
        return result_dict
    
    # 🔍 KIỂM TRA ROUTE STATUS: Nếu route bị OFF trong ngày đó → KHÔNG tính dầu
    route_code_to_check = result.route_code or result.route_name or ""
    if route_code_to_check:
        if is_route_off_on_date(db, route_code_to_check, trip_date, license_plate.strip()):
            result_dict["warning"] = "Tuyến bị OFF trong ngày này"
            return result_dict
    
    # 🔍 BƯỚC 1: Kiểm tra điều kiện khoán xe (BẮT BUỘC)
    is_valid_assignment, assignment_reason = check_vehicle_assignment_for_trip(
        db, license_plate, driver_name, trip_date
    )
    
    if not is_valid_assignment:
        # Không đúng khoán → Tiền dầu = 0
        result_dict["assignment_status"] = "invalid" if assignment_reason else "no_assignment"
        result_dict["assignment_reason"] = assignment_reason
        # Vẫn tính DK và các thông tin khác để hiển thị, nhưng fuel_cost = 0
        # (Có thể bỏ qua phần tính toán nếu muốn tối ưu)
        return result_dict
    
    # Đánh dấu là khoán hợp lệ
    result_dict["assignment_status"] = "valid"
    
    # 1. Lấy định mức nhiên liệu của xe
    fuel_consumption = get_vehicle_fuel_consumption(db, license_plate)
    result_dict["fuel_consumption"] = fuel_consumption
    
    if fuel_consumption is None or fuel_consumption <= 0:
        result_dict["warning"] = "Xe chưa có định mức nhiên liệu"
        return result_dict
    
    # 2. Lấy giá dầu theo ngày chuyến
    fuel_price_record = get_fuel_price_by_date(db, trip_date)
    
    if fuel_price_record is None or fuel_price_record.unit_price is None:
        result_dict["warning"] = "Chưa có đơn giá dầu cho ngày này"
        return result_dict
    
    fuel_price = fuel_price_record.unit_price
    result_dict["fuel_price"] = fuel_price
    
    # 3. Tính số lít dầu khoán (DK)
    # DK = Km chuyến × Định mức nhiên liệu / 100
    dk_liters = (distance_km * fuel_consumption) / 100.0
    # Làm tròn đến 2 chữ số thập phân
    dk_liters = round(dk_liters, 2)
    result_dict["dk_liters"] = dk_liters
    
    # 4. Tính tiền dầu (CHỈ TÍNH KHI ĐÚNG KHOÁN)
    # Tiền dầu = DK × Đơn giá dầu
    fuel_cost = dk_liters * fuel_price
    # Làm tròn theo quy tắc toán học (số nguyên)
    fuel_cost = round(fuel_cost)
    result_dict["fuel_cost"] = int(fuel_cost)
    
    return result_dict

# Chạy migration
try:
    migrate_revenue_records()
except Exception as e:
    print(f"Migration error (may be expected if table doesn't exist yet): {e}")

try:
    migrate_timekeeping_details()
except Exception as e:
    print(f"Migration error for timekeeping_details (may be expected if table doesn't exist yet): {e}")

try:
    migrate_route_prices()
except Exception as e:
    print(f"Migration error for route_prices (may be expected if table doesn't exist yet): {e}")

# Migration: Thêm cột discount_percent vào bảng vehicle_maintenance_items nếu chưa có
def migrate_maintenance_items():
    """Thêm cột discount_percent vào bảng vehicle_maintenance_items nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'vehicle_maintenance_items' not in inspector.get_table_names():
            print("Table vehicle_maintenance_items does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('vehicle_maintenance_items')]
        
        if 'discount_percent' not in existing_columns:
            with engine.connect() as conn:
                try:
                    conn.execute(text("ALTER TABLE vehicle_maintenance_items ADD COLUMN discount_percent FLOAT DEFAULT 0"))
                    conn.commit()
                    print("Added column discount_percent to vehicle_maintenance_items")
                except Exception as e:
                    print(f"Error adding column discount_percent: {e}")
                    conn.rollback()
    except Exception as e:
        print(f"Migration error for vehicle_maintenance_items: {e}")

try:
    migrate_maintenance_items()
except Exception as e:
    print(f"Migration error for vehicle_maintenance_items (may be expected if table doesn't exist yet): {e}")

# Migration: Thêm code field vào roles và permissions (RBAC refactor)
# Trả về True nếu migration thành công, False nếu thất bại
def migrate_rbac_code_fields():
    """Thêm code field vào roles và permissions table"""
    from sqlalchemy import inspect, text
    
    migration_success = True
    
    try:
        inspector = inspect(engine)
        
        # Migrate roles table
        if 'roles' in inspector.get_table_names():
            existing_columns = [col['name'] for col in inspector.get_columns('roles')]
            
            if 'code' not in existing_columns:
                with engine.connect() as conn:
                    trans = conn.begin()
                    try:
                        # Step 1: Add column WITHOUT UNIQUE constraint (SQLite không hỗ trợ)
                        conn.execute(text("ALTER TABLE roles ADD COLUMN code VARCHAR"))
                        
                        # Step 2: Update existing roles với code
                        conn.execute(text("UPDATE roles SET code = 'ADMIN' WHERE name = 'Super Admin' OR name = 'Admin'"))
                        conn.execute(text("UPDATE roles SET code = 'MANAGER' WHERE name = 'Admin Operations'"))
                        conn.execute(text("UPDATE roles SET code = 'USER' WHERE name = 'Viewer' OR name = 'User'"))
                        
                        trans.commit()
                        print("Added column code to roles table")
                        
                        # Step 3: Create UNIQUE INDEX sau khi đã có dữ liệu (ngoài transaction)
                        # Kiểm tra xem index đã tồn tại chưa
                        indexes = inspector.get_indexes('roles')
                        index_names = [idx['name'] for idx in indexes]
                        if 'idx_roles_code_unique' not in index_names:
                            with engine.connect() as conn2:
                                conn2.execute(text("CREATE UNIQUE INDEX idx_roles_code_unique ON roles(code)"))
                                conn2.commit()
                                print("Created UNIQUE INDEX on roles.code")
                    except Exception as e:
                        trans.rollback()
                        print(f"Error adding code to roles: {e}")
                        migration_success = False
        
        # Migrate permissions table
        if 'permissions' in inspector.get_table_names():
            existing_columns = [col['name'] for col in inspector.get_columns('permissions')]
            
            if 'code' not in existing_columns:
                with engine.connect() as conn:
                    trans = conn.begin()
                    try:
                        # Step 1: Add column WITHOUT UNIQUE constraint (SQLite không hỗ trợ)
                        conn.execute(text("ALTER TABLE permissions ADD COLUMN code VARCHAR"))
                        
                        # Step 2: Update existing permissions với code (map từ page_path + action)
                        conn.execute(text("UPDATE permissions SET code = 'user.view' WHERE page_path = '/user-management' AND action = 'view' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'user.create' WHERE page_path = '/user-management' AND action = 'create' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'user.edit' WHERE page_path = '/user-management' AND action = 'update' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'user.delete' WHERE page_path = '/user-management' AND action = 'delete' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'role.view' WHERE page_path = '/role-management' AND action = 'view' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'role.create' WHERE page_path = '/role-management' AND action = 'create' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'role.edit' WHERE page_path = '/role-management' AND action = 'update' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'role.delete' WHERE page_path = '/role-management' AND action = 'delete' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'account.view' WHERE page_path = '/accounts' AND action = 'view' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'account.edit' WHERE page_path = '/accounts' AND action = 'update' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'administrative.view' WHERE page_path = '/administrative' AND action = 'view' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'administrative.create' WHERE page_path = '/administrative' AND action = 'create' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'administrative.update' WHERE page_path = '/administrative' AND action = 'update' AND code IS NULL"))
                        conn.execute(text("UPDATE permissions SET code = 'administrative.delete' WHERE page_path = '/administrative' AND action = 'delete' AND code IS NULL"))
                        
                        trans.commit()
                        print("Added column code to permissions table")
                        
                        # Step 3: Create UNIQUE INDEX sau khi đã có dữ liệu (ngoài transaction)
                        # Kiểm tra xem index đã tồn tại chưa
                        indexes = inspector.get_indexes('permissions')
                        index_names = [idx['name'] for idx in indexes]
                        if 'idx_permissions_code_unique' not in index_names:
                            with engine.connect() as conn2:
                                conn2.execute(text("CREATE UNIQUE INDEX idx_permissions_code_unique ON permissions(code)"))
                                conn2.commit()
                                print("Created UNIQUE INDEX on permissions.code")
                    except Exception as e:
                        trans.rollback()
                        print(f"Error adding code to permissions: {e}")
                        migration_success = False
        
    except Exception as e:
        print(f"Migration error for RBAC code fields: {e}")
        migration_success = False
    
    return migration_success

# Chạy migration và lưu kết quả
rbac_migration_success = False
try:
    rbac_migration_success = migrate_rbac_code_fields()
except Exception as e:
    print(f"Migration error for RBAC code fields (may be expected): {e}")
    rbac_migration_success = False

# Migration: Thêm các cột mới vào bảng vehicle_assignments nếu chưa có
def migrate_vehicle_assignments():
    """Thêm các cột transfer_reason và internal_note vào bảng vehicle_assignments nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'vehicle_assignments' not in inspector.get_table_names():
            print("Table vehicle_assignments does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('vehicle_assignments')]
        
        new_columns = {
            'transfer_reason': 'VARCHAR',
            'internal_note': 'VARCHAR'
        }
        
        with engine.connect() as conn:
            for col_name, col_type in new_columns.items():
                if col_name not in existing_columns:
                    try:
                        conn.execute(text(f"ALTER TABLE vehicle_assignments ADD COLUMN {col_name} {col_type}"))
                        conn.commit()
                        print(f"Added column {col_name} to vehicle_assignments")
                    except Exception as e:
                        print(f"Error adding column {col_name}: {e}")
                        conn.rollback()
    except Exception as e:
        print(f"Migration error for vehicle_assignments: {e}")

try:
    migrate_vehicle_assignments()
except Exception as e:
    print(f"Migration error for vehicle_assignments (may be expected if table doesn't exist yet): {e}")

# Migration: Thêm cột social_insurance_salary vào bảng employees nếu chưa có
def migrate_employee_social_insurance_salary():
    """Thêm cột social_insurance_salary vào bảng employees nếu chưa có"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'employees' not in inspector.get_table_names():
            print("Table employees does not exist yet, will be created by create_all")
            return
        
        existing_columns = [col['name'] for col in inspector.get_columns('employees')]
        
        if 'social_insurance_salary' not in existing_columns:
            with engine.connect() as conn:
                try:
                    conn.execute(text("ALTER TABLE employees ADD COLUMN social_insurance_salary INTEGER"))
                    conn.commit()
                    print("Added column social_insurance_salary to employees")
                except Exception as e:
                    print(f"Error adding column social_insurance_salary: {e}")
                    conn.rollback()
    except Exception as e:
        print(f"Migration error for employees.social_insurance_salary: {e}")

try:
    migrate_employee_social_insurance_salary()
except Exception as e:
    print(f"Migration error for employees.social_insurance_salary (may be expected if table doesn't exist yet): {e}")

# Migration: Thêm cột route_status vào bảng routes nếu chưa có và set mặc định ONL
def migrate_route_status():
    """Thêm cột route_status vào bảng routes nếu chưa có và set mặc định ONL cho các tuyến cũ"""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(engine)
        # Kiểm tra xem bảng có tồn tại không
        if 'routes' not in inspector.get_table_names():
            print("Table routes does not exist yet, will be created by create_all")
            return True
        
        existing_columns = [col['name'] for col in inspector.get_columns('routes')]
        
        with engine.connect() as conn:
            # Thêm cột route_status nếu chưa có
            if 'route_status' not in existing_columns:
                conn.execute(text("ALTER TABLE routes ADD COLUMN route_status VARCHAR DEFAULT 'ONL'"))
                # Set mặc định ONL cho tất cả các tuyến cũ
                conn.execute(text("UPDATE routes SET route_status = 'ONL' WHERE route_status IS NULL"))
                conn.commit()
                print("Added column route_status to routes table and set default ONL for existing routes")
            else:
                # Nếu cột đã tồn tại nhưng có giá trị NULL, set mặc định ONL
                conn.execute(text("UPDATE routes SET route_status = 'ONL' WHERE route_status IS NULL"))
                conn.commit()
                print("Updated NULL route_status values to ONL")
        
        return True
    except Exception as e:
        print(f"Migration error for routes.route_status: {e}")
        return False

# Chạy migration route_status
try:
    migrate_route_status()
except Exception as e:
    print(f"Migration error for routes.route_status (may be expected if table doesn't exist yet): {e}")

# Dependency để lấy database session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_current_user(request: Request):
    """
    Dependency to get current logged-in user from session.
    Returns user info if logged in, None otherwise.
    Note: AuthMiddleware handles redirects, this just returns user info.
    """
    user_id = request.session.get("user_id")
    username = request.session.get("username")
    role = request.session.get("role")
    
    if not user_id or not username:
        return None
    
    return {
        "id": user_id,
        "username": username,
        "role": role or "User"
    }

# Compatibility stub: legacy endpoints still reference require_auth().
# This function is deprecated - use get_current_user dependency instead.
def require_auth():
    # Return None as this is deprecated - routes should use get_current_user dependency
    return None

# Helper function để check user có quyền truy cập trang không
def check_page_access(role: str, page_path: str, user_id: Optional[int] = None, db: Optional[Session] = None) -> bool:
    """Open access: luôn cho phép truy cập page."""
    return True

# Helper function để kiểm tra column có tồn tại không
def column_exists(table_name: str, column_name: str) -> bool:
    """Kiểm tra column có tồn tại trong table không"""
    try:
        from sqlalchemy import inspect
        inspector = inspect(engine)
        if table_name not in inspector.get_table_names():
            return False
        existing_columns = [col['name'] for col in inspector.get_columns(table_name)]
        return column_name in existing_columns
    except Exception:
        return False

def check_permission(db: Session, user_id: Optional[int], permission_code: str, page_path: str = None, action: str = None) -> bool:
    """Open access: luôn cho phép (RBAC disabled)."""
    return True

def has_permission(db: Session, user_id: int, permission_code: str) -> bool:
    """
    Helper function để kiểm tra permission theo code (alias cho check_permission)
    Usage: has_permission(db, user_id, "user.view")
    """
    return check_permission(db, user_id, permission_code)

 # NOTE: require_permission dependency removed (open access).

def get_user_permissions(db: Session, user_id: int) -> dict:
    """
    Lấy tất cả permissions của user (từ roles + user-specific)
    Returns dict với key là permission.code và value là True
    """
    # Get user's roles
    user_roles = db.query(UserRole).filter(UserRole.user_id == user_id).all()
    role_ids = [ur.role_id for ur in user_roles]
    
    # Kiểm tra column code có tồn tại không
    permission_code_exists = column_exists('permissions', 'code')
    role_code_exists = column_exists('roles', 'code')
    
    # Fallback to legacy role field if RBAC tables are empty (backward compatibility)
    if not role_ids:
        account = db.query(Account).filter(Account.id == user_id).first()
        if account and account.role in ["Admin", "Super Admin"]:
            # Return all permissions for legacy Admin/Super Admin
            all_permissions = db.query(Permission).all()
            result = {}
            for p in all_permissions:
                if permission_code_exists and p.code:
                    result[p.code] = True
                else:
                    # Fallback to old format
                    result[f"{p.page_path}:{p.action}"] = True
            return result
        return {}
    
    # Check if Admin (by code or name)
    admin_role = None
    if role_code_exists:
        admin_role = db.query(Role).filter(
            or_(
                Role.code == "ADMIN",
                Role.name == "Admin",
                Role.name == "Super Admin"
            )
        ).first()
    else:
        # Fallback nếu column code chưa tồn tại
        admin_role = db.query(Role).filter(
            or_(
                Role.name == "Admin",
                Role.name == "Super Admin"
            )
        ).first()
    
    if admin_role and admin_role.id in role_ids:
        # Return all permissions for Admin
        all_permissions = db.query(Permission).all()
        result = {}
        for p in all_permissions:
            if permission_code_exists and p.code:
                result[p.code] = True
            else:
                result[f"{p.page_path}:{p.action}"] = True
        return result
    
    # Get permissions from roles
    role_permissions = db.query(RolePermission).filter(
        RolePermission.role_id.in_(role_ids)
    ).all()
    permission_ids = {rp.permission_id for rp in role_permissions}
    
    # Get user-specific permissions (override)
    user_permissions = db.query(UserPermission).filter(
        UserPermission.user_id == user_id
    ).all()
    user_permission_ids = {up.permission_id for up in user_permissions}
    
    # Combine: user permissions override role permissions
    all_permission_ids = permission_ids | user_permission_ids
    
    # Get permission details
    permissions = db.query(Permission).filter(Permission.id.in_(all_permission_ids)).all()
    
    result = {}
    for p in permissions:
        if permission_code_exists and p.code:
            result[p.code] = True
        else:
            result[f"{p.page_path}:{p.action}"] = True
    
    return result

# Helper function để kiểm tra và redirect nếu không có quyền
def check_and_redirect_access(role: str, page_path: str, user_id: Optional[int] = None, db: Optional[Session] = None) -> Optional[RedirectResponse]:
    """Open access: không redirect."""
    return None

# Helper function để lấy IP address từ request
def get_client_ip(request: Request) -> str:
    """Lấy IP address của client"""
    if request.client:
        return request.client.host
    return "unknown"

# Helper function để tạo audit log
def create_audit_log(
    db: Session,
    user_id: int,
    action: str,
    entity_type: str,
    entity_id: Optional[int] = None,
    old_values: Optional[dict] = None,
    new_values: Optional[dict] = None,
    description: Optional[str] = None,
    ip_address: Optional[str] = None
):
    """Tạo audit log entry"""
    import json
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        old_values=json.dumps(old_values, ensure_ascii=False, default=str) if old_values else None,
        new_values=json.dumps(new_values, ensure_ascii=False, default=str) if new_values else None,
        description=description,
        ip_address=ip_address
    )
    db.add(audit_log)
    db.commit()

# Helper function để khởi tạo permissions cho tất cả các pages
def initialize_permissions(db: Session):
    """Khởi tạo permissions cho tất cả các pages trong hệ thống"""
    # Kiểm tra column code có tồn tại không
    permission_code_exists = column_exists('permissions', 'code')
    
    # Danh sách các pages và mô tả
    pages = [
        {"path": "/", "name": "home.view", "code": "home.view", "description": "Trang chủ", "action": "view"},
        {"path": "/operations", "name": "operations.view", "code": "operations.view", "description": "Quản lý vận hành", "action": "view"},
        {"path": "/employees", "name": "employees.view", "code": "employee.view", "description": "Quản lý nhân viên", "action": "view"},
        {"path": "/vehicles", "name": "vehicles.view", "code": "vehicle.view", "description": "Quản lý xe", "action": "view"},
        {"path": "/routes", "name": "routes.view", "code": "route.view", "description": "Quản lý tuyến", "action": "view"},
        {"path": "/maintenance", "name": "maintenance.view", "code": "maintenance.view", "description": "Bảo dưỡng xe", "action": "view"},
        {"path": "/theo-doi-dau-v2", "name": "fuel.view", "code": "fuel.view", "description": "Theo dõi dầu", "action": "view"},
        {"path": "/revenue", "name": "revenue.view", "code": "revenue.view", "description": "Doanh thu", "action": "view"},
        {"path": "/daily-new", "name": "daily.view", "code": "daily.view", "description": "Chấm công hàng ngày", "action": "view"},
        {"path": "/timekeeping-v1", "name": "timekeeping.view", "code": "timekeeping.view", "description": "Bảng chấm công V1", "action": "view"},
        {"path": "/salary-calculation-v2", "name": "salary.view", "code": "salary.view", "description": "Tính lương", "action": "view"},
        {"path": "/salary-summary", "name": "salary.summary.view", "code": "salary.summary.view", "description": "Bảng lương tổng", "action": "view"},
        {"path": "/finance-report", "name": "finance.view", "code": "finance.report.view", "description": "Báo cáo tài chính", "action": "view"},
        {"path": "/financial-statistics", "name": "statistics.view", "code": "finance.statistics.view", "description": "Thống kê tài chính", "action": "view"},
        {"path": "/accounts", "name": "accounts.view", "code": "account.view", "description": "Quản lý tài khoản", "action": "view"},
    ]
    
    for page in pages:
        # Kiểm tra xem permission đã tồn tại chưa (theo page_path hoặc code nếu có)
        if permission_code_exists:
            existing = db.query(Permission).filter(
                or_(
                    Permission.page_path == page["path"],
                    Permission.code == page["code"]
                )
            ).first()
        else:
            existing = db.query(Permission).filter(Permission.page_path == page["path"]).first()
        
        if not existing:
            permission_data = {
                "name": page["name"],
                "description": page["description"],
                "page_path": page["path"],
                "action": page["action"]
            }
            # Chỉ thêm code nếu column tồn tại và code chưa được dùng
            if permission_code_exists:
                # Kiểm tra xem code đã được dùng bởi permission khác chưa
                code_exists = db.query(Permission).filter(Permission.code == page["code"]).first()
                if not code_exists:
                    permission_data["code"] = page["code"]
                else:
                    print(f"Warning: Code '{page['code']}' already exists, skipping for {page['path']}")
            
            permission = Permission(**permission_data)
            db.add(permission)
        elif existing and permission_code_exists and not existing.code:
            # Update existing permission với code nếu chưa có
            # Kiểm tra xem code đã được dùng bởi permission khác chưa
            code_exists = db.query(Permission).filter(
                and_(
                    Permission.code == page["code"],
                    Permission.id != existing.id
                )
            ).first()
            if not code_exists:
                existing.code = page["code"]
            else:
                print(f"Warning: Code '{page['code']}' already exists, cannot update permission {existing.id} ({existing.page_path})")
    
    db.commit()
    print("Permissions initialized successfully")

# Helper function để generate password tự động
def generate_password(length: int = 12) -> str:
    """Tạo mật khẩu tự động"""
    import random
    import string
    # Bao gồm: chữ hoa, chữ thường, số
    uppercase = string.ascii_uppercase
    lowercase = string.ascii_lowercase
    digits = string.digits
    all_chars = uppercase + lowercase + digits
    
    # Đảm bảo có ít nhất 1 ký tự mỗi loại
    password = [
        random.choice(uppercase),
        random.choice(lowercase),
        random.choice(digits)
    ]
    
    # Thêm các ký tự ngẫu nhiên
    password.extend(random.choice(all_chars) for _ in range(length - 3))
    
    # Xáo trộn
    random.shuffle(password)
    return ''.join(password)

# FastAPI app
app = FastAPI(title="Hệ thống quản lý vận chuyển")

# Authentication middleware to protect all routes except /login and static files
class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        # Allow access to login page, logout, and static files
        if request.url.path in ["/login", "/logout"] or request.url.path.startswith("/static/"):
            return await call_next(request)
        
        # Check if user is logged in
        if not request.session.get("user_id"):
            return RedirectResponse(url="/login", status_code=303)
        
        return await call_next(request)

# Add middleware in correct order: SessionMiddleware must be added AFTER AuthMiddleware
# so it runs first (middleware executes in reverse order of addition)
app.add_middleware(AuthMiddleware)
app.add_middleware(SessionMiddleware, secret_key="local-dev-secret-key-12345")

# ==================== FILE UPLOAD HELPER FUNCTIONS ====================
# Cấu trúc thư mục ảnh: Picture/{category}/{subcategory}/
PICTURE_BASE_DIR = "Picture"

def ensure_directory_exists(directory_path: str):
    """Đảm bảo thư mục tồn tại, nếu chưa thì tạo"""
    if not os.path.exists(directory_path):
        os.makedirs(directory_path, exist_ok=True)
    return directory_path

# ==================== DOCUMENTS UPLOAD CONFIGURATION ====================
# Base directory for document uploads
DOCUMENTS_UPLOAD_DIR = "uploads/documents"

# Allowed document file types
ALLOWED_DOCUMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".jpg", ".jpeg", ".png"}

def ensure_document_dirs():
    """Ensure document upload directories exist"""
    # Create main documents directory
    ensure_directory_exists(DOCUMENTS_UPLOAD_DIR)
    
    # Create subdirectories by type/category
    subdirs = ["contracts", "company", "tax", "others"]
    for subdir in subdirs:
        ensure_directory_exists(os.path.join(DOCUMENTS_UPLOAD_DIR, subdir))

def get_document_category_folder(category: str, document_type: str) -> str:
    """Map category/document_type to folder structure"""
    # Map categories to folders
    category_map = {
        "legal": "contracts",
        "administrative": "company", 
        "tax": "tax"
    }
    
    # Default folder based on category
    folder = category_map.get(category, "others")
    
    # Special handling for certain document types
    if document_type and "contract" in document_type.lower():
        folder = "contracts"
    elif document_type and "tax" in document_type.lower():
        folder = "tax"
    
    return folder

def validate_document_file(filename: str) -> Tuple[bool, Optional[str]]:
    """Validate document file type. Returns (is_valid, error_message)"""
    if not filename:
        return False, "No file provided"
    
    file_ext = os.path.splitext(filename)[1].lower()
    if file_ext not in ALLOWED_DOCUMENT_EXTENSIONS:
        allowed_types = ", ".join(sorted(ALLOWED_DOCUMENT_EXTENSIONS))
        return False, f"File type not allowed. Only {allowed_types} files are supported."
    
    return True, None

def save_uploaded_file(
    file: UploadFile,
    category: str,
    subcategory: str,
    entity_id: str,
    entity_type: str = "vehicle"
) -> str:
    """
    Lưu file upload với cấu trúc thư mục mới
    
    Args:
        file: UploadFile object
        category: Loại nghiệp vụ (vehicles, employees, maintenance, fuel, tires, other)
        subcategory: Thư mục con (registration, insurance, vehicle_photos, etc.)
        entity_id: ID của entity (license_plate cho vehicle, id cho employee, etc.)
        entity_type: Loại entity (vehicle, employee, etc.)
    
    Returns:
        str: Relative path từ root project (ví dụ: Picture/vehicles/registration/vehicle_50H14740_20260122103015.jpg)
    """
    # Validate file
    if not file or not file.filename:
        raise ValueError("File không hợp lệ")
    
    # Validate file extension
    allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
    file_extension = os.path.splitext(file.filename)[1].lower()
    if file_extension not in allowed_extensions:
        raise ValueError(f"File extension không được phép: {file_extension}")
    
    # Tạo tên file mới: {entity}_{id}_{yyyyMMddHHmmss}.{ext}
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    # Sanitize entity_id để tránh ký tự đặc biệt trong tên file
    safe_entity_id = re.sub(r'[^\w\-]', '_', str(entity_id))
    new_filename = f"{entity_type}_{safe_entity_id}_{timestamp}{file_extension}"
    
    # Tạo đường dẫn thư mục
    directory_path = os.path.join(PICTURE_BASE_DIR, category, subcategory)
    ensure_directory_exists(directory_path)
    
    # Đường dẫn đầy đủ để lưu file
    full_file_path = os.path.join(directory_path, new_filename)
    
    # Lưu file (sử dụng absolute path để đảm bảo)
    abs_file_path = os.path.abspath(full_file_path)
    abs_dir_path = os.path.dirname(abs_file_path)
    ensure_directory_exists(abs_dir_path)
    
    # Đọc và lưu file
    # Note: file.file.read() sẽ đọc toàn bộ nội dung vào memory
    # Đối với file lớn có thể cần stream, nhưng với ảnh thông thường thì OK
    try:
        with open(abs_file_path, "wb") as buffer:
            content = file.file.read()
            buffer.write(content)
            file.file.seek(0)  # Reset file pointer để có thể đọc lại nếu cần
    except Exception as e:
        raise Exception(f"Lỗi khi lưu file: {str(e)}")
    
    # Trả về relative path từ root project
    return full_file_path.replace("\\", "/")  # Normalize path separators

def get_file_url(file_path: str) -> str:
    """
    Chuyển đổi file path trong DB thành URL để hiển thị
    
    Args:
        file_path: Path từ DB (ví dụ: Picture/vehicles/registration/vehicle_50H14740_20260122103015.jpg)
    
    Returns:
        str: URL để truy cập file (ví dụ: /Picture/vehicles/registration/vehicle_50H14740_20260122103015.jpg)
    """
    if not file_path:
        return ""
    # Normalize path và đảm bảo bắt đầu với /
    normalized_path = file_path.replace("\\", "/")
    if not normalized_path.startswith("/"):
        normalized_path = "/" + normalized_path
    return normalized_path

def delete_file_if_exists(file_path: str):
    """Xóa file vật lý nếu tồn tại"""
    if file_path:
        # Nếu là relative path, chuyển thành absolute
        if not os.path.isabs(file_path):
            abs_path = os.path.abspath(file_path)
        else:
            abs_path = file_path
        
        if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except Exception as e:
                print(f"Lỗi khi xóa file {abs_path}: {e}")

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")
# Mount Picture directory để truy cập ảnh
ensure_directory_exists(PICTURE_BASE_DIR)
app.mount("/Picture", StaticFiles(directory=PICTURE_BASE_DIR), name="picture")

# Mount documents upload directory
ensure_document_dirs()
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

# Ensure administrative documents directory exists (function is defined later in file)

# Templates đã được tạo ở trên với custom filters

# ==================== AUTHENTICATION ROUTES ====================

def hash_password(password: str, *, iterations: int = 210_000) -> str:
    """
    PBKDF2-HMAC-SHA256 password hashing (stdlib only).
    Stored format: pbkdf2_sha256$<iterations>$<salt_b64>$<hash_b64>
    """
    import base64
    import hashlib
    import secrets

    salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return "pbkdf2_sha256$%d$%s$%s" % (
        iterations,
        base64.b64encode(salt).decode("ascii"),
        base64.b64encode(dk).decode("ascii"),
    )

def verify_password(password: str, stored: str) -> bool:
    import base64
    import hashlib
    import hmac

    if not stored or "$" not in stored:
        return False
    try:
        algo, iter_s, salt_b64, hash_b64 = stored.split("$", 3)
        if algo != "pbkdf2_sha256":
            return False
        iterations = int(iter_s)
        salt = base64.b64decode(salt_b64.encode("ascii"))
        expected = base64.b64decode(hash_b64.encode("ascii"))
        dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
        return hmac.compare_digest(dk, expected)
    except Exception:
        return False

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Trang đăng nhập"""
    # If user is already logged in, redirect to home
    if request.session.get("user_id"):
        return RedirectResponse(url="/", status_code=303)
    
    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": None
    })

@app.post("/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Authenticate user and create session"""

    # Find account in database
    account = db.query(Account).filter(Account.username == username).first()
    
    # Check if account exists
    if not account:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Sai tài khoản hoặc mật khẩu"
        })

    # Check password (plain text comparison as requested)
    # Check both password and password_hash fields for backward compatibility
    password_match = False
    if account.password and account.password == password:
        password_match = True
    elif account.password_hash and account.password_hash == password:
        password_match = True
    
    if not password_match:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Sai tài khoản hoặc mật khẩu"
        })
    
    # Check status must be 'Active'
    if account.status != "Active":
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Tài khoản không hoạt động. Vui lòng liên hệ quản trị viên."
        })
    
    # Update last_login
    account.last_login = datetime.utcnow()
    db.commit()
    
    # Store user info in session
    request.session["user_id"] = account.id
    request.session["username"] = account.username
    request.session["role"] = account.role or "User"
    
    # Redirect to home
    return RedirectResponse(url="/", status_code=303)

@app.post("/logout")
async def logout(request: Request):
    """Clear session and redirect to login"""
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)

@app.get("/access-denied", response_class=HTMLResponse)
async def access_denied_page(request: Request, current_user = Depends(get_current_user)):
    """Trang thông báo không có quyền truy cập"""
    return templates.TemplateResponse("access_denied.html", {
        "request": request,
        "current_user": current_user
    })

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # Lấy thống kê tổng quan
    employees_count = db.query(Employee).count()
    vehicles_count = db.query(Vehicle).count()
    routes_count = db.query(Route).filter(Route.is_active == 1).count()
    today = date.today()
    daily_routes_count = db.query(DailyRoute).filter(DailyRoute.date == today).count()
    
    return templates.TemplateResponse("index.html", {
        "request": request,
        "current_user": current_user,
        "employees_count": employees_count,
        "vehicles_count": vehicles_count,
        "routes_count": routes_count,
        "daily_routes_count": daily_routes_count
    })

@app.get("/report", response_class=HTMLResponse)
async def report_page(request: Request):
    """Trang báo cáo tổng hợp - redirect tới trang thống kê"""
    # Redirect tới trang thống kê
    return RedirectResponse(url="/statistics", status_code=303)

@app.get("/employees", response_class=HTMLResponse)
async def employees_page(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    employees = db.query(Employee).filter(Employee.status == 1).all()
    
    # Sắp xếp nhân viên: Ưu tiên 1 (theo trạng thái), Ưu tiên 2 (theo chức vụ)
    def sort_employees(employees):
        # Định nghĩa thứ tự ưu tiên trạng thái
        status_order = {
            "Đang làm việc": 1,
            "Nghỉ phép dài hạn": 2,
            "Đã nghỉ việc": 3
        }
        
        # Định nghĩa thứ tự ưu tiên chức vụ
        position_order = {
            "Giám đốc": 1,
            "Phó Giám đốc": 2,
            "Lái xe": 3,
            "Nhân viên văn phòng": 4
        }
        
        def get_sort_key(emp):
            # Lấy thứ tự trạng thái (mặc định là 1 nếu không có)
            emp_status = emp.employee_status or "Đang làm việc"
            status_priority = status_order.get(emp_status, 4)
            
            # Lấy thứ tự chức vụ (mặc định là 99 nếu không có)
            emp_position = emp.position or ""
            position_priority = position_order.get(emp_position, 99)
            
            # Sắp xếp: trạng thái trước, sau đó chức vụ, cuối cùng là tên
            return (status_priority, position_priority, (emp.name or "").lower())
        
        return sorted(employees, key=get_sort_key)
    
    employees = sort_employees(employees)
    
    return templates.TemplateResponse("employees.html", {
        "request": request,
        "current_user": current_user,
        "employees": employees
    })


@app.get("/employees/documents/{employee_id}")
async def get_employee_documents(employee_id: int, db: Session = Depends(get_db)):
    """API để lấy thông tin giấy tờ của nhân viên"""
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy nhân viên"}
        )
    
    if not employee.documents:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Nhân viên chưa upload giấy tờ"}
        )
    
    try:
        import json
        documents = json.loads(employee.documents)
        
        # Kiểm tra file tồn tại
        existing_documents = []
        for doc in documents:
            # Hỗ trợ cả path cũ (static/uploads/) và path mới (Picture/...)
            if doc.startswith("Picture/"):
                # Path mới - sử dụng trực tiếp
                file_path = doc
                file_url = get_file_url(doc)
            else:
                # Path cũ - giữ nguyên để backward compatibility
                file_path = f"static/uploads/{doc}"
                file_url = f"/static/uploads/{doc}"
            
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                file_extension = os.path.splitext(doc)[1].lower()
                filename = os.path.basename(doc) if "/" in doc else doc
                existing_documents.append({
                    "filename": filename,
                    "url": file_url,
                    "size": file_size,
                    "extension": file_extension,
                    "exists": True
                })
            else:
                filename = os.path.basename(doc) if "/" in doc else doc
                existing_documents.append({
                    "filename": filename,
                    "url": file_url,
                    "exists": False,
                    "error": "File không tồn tại trên server"
                })
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "documents": existing_documents,
                "total": len(existing_documents)
            }
        )
        
    except json.JSONDecodeError:
        # Xử lý dữ liệu cũ (không phải JSON)
        if isinstance(employee.documents, str) and employee.documents.strip():
            file_path = f"static/uploads/{employee.documents}"
            if os.path.exists(file_path):
                return JSONResponse(
                    status_code=200,
                    content={
                        "success": True,
                        "documents": [{
                            "filename": employee.documents,
                            "url": f"/static/uploads/{employee.documents}",
                            "size": os.path.getsize(file_path),
                            "extension": os.path.splitext(employee.documents)[1].lower(),
                            "exists": True
                        }],
                        "total": 1
                    }
                )
        
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Dữ liệu giấy tờ không hợp lệ"}
        )

@app.post("/employees/add")
async def add_employee(
    name: str = Form(...),
    birth_date: str = Form(""),
    phone: str = Form(""),
    cccd: str = Form(""),
    cccd_issue_date: str = Form(""),
    cccd_expiry: str = Form(""),
    driving_license: str = Form(""),
    license_expiry: str = Form(""),
    employee_status: str = Form("Đang làm việc"),
    position: str = Form(""),
    social_insurance_salary: str = Form(""),
    documents: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    # Convert date strings to date objects
    birth_date_obj = None
    cccd_issue_date_obj = None
    cccd_expiry_date = None
    license_expiry_date = None
    
    if birth_date:
        birth_date_obj = datetime.strptime(birth_date, "%Y-%m-%d").date()
    if cccd_issue_date:
        cccd_issue_date_obj = datetime.strptime(cccd_issue_date, "%Y-%m-%d").date()
    if cccd_expiry:
        cccd_expiry_date = datetime.strptime(cccd_expiry, "%Y-%m-%d").date()
    if license_expiry:
        license_expiry_date = datetime.strptime(license_expiry, "%Y-%m-%d").date()
    
    # Parse social_insurance_salary (must be positive integer or None)
    social_insurance_salary_int = None
    if social_insurance_salary and social_insurance_salary.strip():
        try:
            salary_value = int(social_insurance_salary.strip())
            if salary_value > 0:
                social_insurance_salary_int = salary_value
        except ValueError:
            # Invalid input, will be None
            pass
    
    # Tạo employee trước để có ID
    employee = Employee(
        name=name,
        birth_date=birth_date_obj,
        phone=phone, 
        cccd=cccd,
        cccd_issue_date=cccd_issue_date_obj,
        cccd_expiry=cccd_expiry_date,
        driving_license=driving_license,
        license_expiry=license_expiry_date,
        employee_status=employee_status,
        position=position,
        social_insurance_salary=social_insurance_salary_int,
        documents=None  # Tạm thời để None, sẽ cập nhật sau
    )
    db.add(employee)
    db.flush()  # Lấy ID mà không commit
    
    # Handle multiple file uploads - sau khi có employee ID
    documents_paths = []
    if documents:
        for document in documents:
            if document and document.filename:
                try:
                    # Sử dụng helper function để lưu file với cấu trúc mới
                    file_path = save_uploaded_file(
                        file=document,
                        category="employees",
                        subcategory="documents",
                        entity_id=str(employee.id),
                        entity_type="employee"
                    )
                    documents_paths.append(file_path)
                except Exception as e:
                    print(f"Lỗi khi lưu file giấy tờ nhân viên: {e}")
                    continue  # Skip file nếu có lỗi
    
    # Convert documents list to JSON string và cập nhật employee
    documents_json = json.dumps(documents_paths) if documents_paths else None
    employee.documents = documents_json
    
    db.commit()
    return RedirectResponse(url="/employees", status_code=303)

@app.post("/employees/delete/{employee_id}")
async def delete_employee(employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if employee:
        employee.status = 0  # Soft delete
        db.commit()
    return RedirectResponse(url="/employees", status_code=303)

@app.get("/employees/edit/{employee_id}", response_class=HTMLResponse)
async def edit_employee_page(request: Request, employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return RedirectResponse(url="/employees", status_code=303)
    return templates.TemplateResponse("edit_employee.html", {"request": request, "employee": employee})

@app.post("/employees/edit/{employee_id}")
async def edit_employee(
    employee_id: int,
    name: str = Form(...),
    birth_date: str = Form(""),
    phone: str = Form(""),
    cccd: str = Form(""),
    cccd_issue_date: str = Form(""),
    cccd_expiry: str = Form(""),
    driving_license: str = Form(""),
    license_expiry: str = Form(""),
    employee_status: str = Form("Đang làm việc"),
    position: str = Form(""),
    social_insurance_salary: str = Form(""),
    documents: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return RedirectResponse(url="/employees", status_code=303)
    
    # Convert date strings to date objects
    birth_date_obj = None
    cccd_issue_date_obj = None
    cccd_expiry_date = None
    license_expiry_date = None
    
    if birth_date:
        birth_date_obj = datetime.strptime(birth_date, "%Y-%m-%d").date()
    if cccd_issue_date:
        cccd_issue_date_obj = datetime.strptime(cccd_issue_date, "%Y-%m-%d").date()
    if cccd_expiry:
        cccd_expiry_date = datetime.strptime(cccd_expiry, "%Y-%m-%d").date()
    if license_expiry:
        license_expiry_date = datetime.strptime(license_expiry, "%Y-%m-%d").date()
    
    # Handle multiple file uploads
    if documents:
        documents_paths = []
        for document in documents:
            if document and document.filename:
                try:
                    # Sử dụng helper function để lưu file với cấu trúc mới
                    file_path = save_uploaded_file(
                        file=document,
                        category="employees",
                        subcategory="documents",
                        entity_id=str(employee.id),
                        entity_type="employee"
                    )
                    documents_paths.append(file_path)
                except Exception as e:
                    print(f"Lỗi khi lưu file giấy tờ nhân viên: {e}")
                    continue  # Skip file nếu có lỗi
        
        if documents_paths:
            # Get existing documents and append new ones
            existing_documents = []
            if employee.documents:
                try:
                    existing_documents = json.loads(employee.documents)
                except json.JSONDecodeError:
                    existing_documents = []
            
            # Combine existing and new documents
            all_documents = existing_documents + documents_paths
            employee.documents = json.dumps(all_documents)
    
    # Update employee data
    employee.name = name
    employee.birth_date = birth_date_obj
    employee.phone = phone
    employee.cccd = cccd
    employee.cccd_issue_date = cccd_issue_date_obj
    employee.cccd_expiry = cccd_expiry_date
    employee.driving_license = driving_license
    employee.license_expiry = license_expiry_date
    employee.employee_status = employee_status
    employee.position = position
    
    # Update social_insurance_salary only if a new value is provided
    # Không ghi đè dữ liệu cũ khi người dùng không nhập lại giá trị
    if social_insurance_salary and social_insurance_salary.strip():
        try:
            salary_value = int(social_insurance_salary.strip())
            if salary_value > 0:
                employee.social_insurance_salary = salary_value
            else:
                # If 0 or negative, set to None
                employee.social_insurance_salary = None
        except ValueError:
            # Invalid input, keep old value (don't update)
            pass
    # If empty string, keep the existing value (don't update)
    
    db.commit()
    return RedirectResponse(url="/employees", status_code=303)

@app.delete("/employees/documents/{employee_id}")
async def delete_employee_document(
    employee_id: int, 
    filename: str,
    db: Session = Depends(get_db)
):
    """API để xóa giấy tờ của nhân viên"""
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy nhân viên"}
        )
    
    if not employee.documents:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Nhân viên chưa có giấy tờ nào"}
        )
    
    try:
        import json
        documents = json.loads(employee.documents)
        
        # Tìm file trong danh sách (có thể là filename hoặc full path)
        file_to_remove = None
        for doc in documents:
            # So sánh với filename hoặc basename của path
            doc_basename = os.path.basename(doc) if "/" in doc else doc
            if doc == filename or doc_basename == filename:
                file_to_remove = doc
                break
        
        if not file_to_remove:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "File không tồn tại trong danh sách giấy tờ"}
            )
        
        # Xóa file khỏi thư mục lưu trữ
        # Hỗ trợ cả path cũ và mới
        if file_to_remove.startswith("Picture/"):
            file_path = file_to_remove
        else:
            file_path = f"static/uploads/{file_to_remove}"
        
        delete_file_if_exists(file_path)
        
        # Xóa file khỏi danh sách trong DB
        documents.remove(file_to_remove)
        
        if documents:
            # Còn giấy tờ khác, cập nhật danh sách
            employee.documents = json.dumps(documents)
        else:
            # Không còn giấy tờ nào, set null
            employee.documents = None
        
        db.commit()
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "message": "Xóa giấy tờ thành công",
                "remaining_documents": len(documents) if documents else 0
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Lỗi định dạng dữ liệu giấy tờ"}
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

@app.get("/vehicles", response_class=HTMLResponse)
async def vehicles_page(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    today = date.today()
    
    # Lấy danh sách khoán xe với thông tin xe và lái xe
    assignments = db.query(VehicleAssignment).order_by(VehicleAssignment.assignment_date.desc()).all()
    
    # Lấy danh sách lái xe đang làm việc để hiển thị trong dropdown
    drivers = db.query(Employee).filter(
        Employee.position == "Lái xe",
        Employee.employee_status == "Đang làm việc",
        Employee.status == 1
    ).order_by(Employee.name).all()
    
    # Lấy danh sách xe nhà chưa được khoán hoặc đã kết thúc khoán
    xe_nha = [v for v in vehicles if v.vehicle_type == "Xe Nhà"]
    available_vehicles = []
    for vehicle in xe_nha:
        # Kiểm tra xem xe có đang được khoán không
        active_assignment = db.query(VehicleAssignment).filter(
            VehicleAssignment.vehicle_id == vehicle.id,
            VehicleAssignment.end_date.is_(None)
        ).first()
        if not active_assignment:
            available_vehicles.append(vehicle)
    
    return templates.TemplateResponse("vehicles.html", {
        "request": request,
        "current_user": current_user,
        "vehicles": vehicles,
        "today": today,
        "assignments": assignments,
        "drivers": drivers,
        "available_vehicles": available_vehicles
    })

@app.post("/vehicles/add")
async def add_vehicle(
    license_plate: str = Form(...),
    vehicle_type: str = Form("Xe Nhà"),
    capacity: float = Form(0),
    fuel_consumption: float = Form(0),
    inspection_expiry: str = Form(""),
    inspection_documents: list[UploadFile] = File(None),
    phu_hieu_expired_date: str = Form(""),
    phu_hieu_files: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    # Convert date string to date object
    inspection_expiry_date = None
    if inspection_expiry:
        try:
            inspection_expiry_date = datetime.strptime(inspection_expiry, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle multiple file uploads - Sổ đăng kiểm
    documents_paths = []
    if inspection_documents:
        for document in inspection_documents:
            if document and document.filename:
                try:
                    # Sử dụng helper function để lưu file với cấu trúc mới
                    file_path = save_uploaded_file(
                        file=document,
                        category="vehicles",
                        subcategory="registration",
                        entity_id=license_plate,
                        entity_type="vehicle"
                    )
                    documents_paths.append(file_path)
                except Exception as e:
                    print(f"Lỗi khi lưu file sổ đăng kiểm: {e}")
                    continue  # Skip file nếu có lỗi
    
    # Convert documents list to JSON string
    documents_json = json.dumps(documents_paths) if documents_paths else None
    
    # Handle phù hiệu vận tải date
    phu_hieu_expired_date_obj = None
    if phu_hieu_expired_date:
        try:
            phu_hieu_expired_date_obj = datetime.strptime(phu_hieu_expired_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle phù hiệu vận tải file uploads
    phu_hieu_paths = []
    if phu_hieu_files:
        for document in phu_hieu_files:
            if document and document.filename:
                try:
                    # Sử dụng helper function để lưu file với cấu trúc mới
                    file_path = save_uploaded_file(
                        file=document,
                        category="vehicles",
                        subcategory="insurance",
                        entity_id=license_plate,
                        entity_type="vehicle"
                    )
                    phu_hieu_paths.append(file_path)
                except Exception as e:
                    print(f"Lỗi khi lưu file phù hiệu vận tải: {e}")
                    continue  # Skip file nếu có lỗi
    
    # Convert phù hiệu files list to JSON string
    phu_hieu_json = json.dumps(phu_hieu_paths) if phu_hieu_paths else None
    
    vehicle = Vehicle(
        license_plate=license_plate,
        vehicle_type=vehicle_type,
        capacity=capacity,
        fuel_consumption=fuel_consumption,
        inspection_expiry=inspection_expiry_date,
        inspection_documents=documents_json,
        phu_hieu_expired_date=phu_hieu_expired_date_obj,
        phu_hieu_files=phu_hieu_json
    )
    db.add(vehicle)
    db.commit()
    return RedirectResponse(url="/vehicles", status_code=303)

@app.post("/vehicles/delete/{vehicle_id}")
async def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if vehicle:
        vehicle.status = 0  # Soft delete
        db.commit()
    return RedirectResponse(url="/vehicles", status_code=303)

@app.get("/vehicles/edit/{vehicle_id}", response_class=HTMLResponse)
async def edit_vehicle_page(request: Request, vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return RedirectResponse(url="/vehicles", status_code=303)
    return templates.TemplateResponse("edit_vehicle.html", {"request": request, "vehicle": vehicle})

@app.post("/vehicles/edit/{vehicle_id}")
async def edit_vehicle(
    vehicle_id: int,
    license_plate: str = Form(...),
    vehicle_type: str = Form("Xe Nhà"),
    capacity: float = Form(0),
    fuel_consumption: float = Form(0),
    inspection_expiry: str = Form(""),
    inspection_documents: list[UploadFile] = File(None),
    phu_hieu_expired_date: str = Form(""),
    phu_hieu_files: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return RedirectResponse(url="/vehicles", status_code=303)
    
    # Convert date string to date object
    inspection_expiry_date = None
    if inspection_expiry:
        try:
            inspection_expiry_date = datetime.strptime(inspection_expiry, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle multiple file uploads - append to existing documents
    if inspection_documents:
        documents_paths = []
        for document in inspection_documents:
            if document and document.filename:
                try:
                    # Sử dụng helper function để lưu file với cấu trúc mới
                    file_path = save_uploaded_file(
                        file=document,
                        category="vehicles",
                        subcategory="registration",
                        entity_id=vehicle.license_plate,
                        entity_type="vehicle"
                    )
                    documents_paths.append(file_path)
                except Exception as e:
                    print(f"Lỗi khi lưu file sổ đăng kiểm: {e}")
                    continue  # Skip file nếu có lỗi
        
        if documents_paths:
            # Get existing documents and append new ones
            existing_documents = []
            if vehicle.inspection_documents:
                try:
                    existing_documents = json.loads(vehicle.inspection_documents)
                except json.JSONDecodeError:
                    existing_documents = []
            
            # Combine existing and new documents
            all_documents = existing_documents + documents_paths
            vehicle.inspection_documents = json.dumps(all_documents)
    
    # Handle phù hiệu vận tải date
    phu_hieu_expired_date_obj = None
    if phu_hieu_expired_date:
        try:
            phu_hieu_expired_date_obj = datetime.strptime(phu_hieu_expired_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle phù hiệu vận tải file uploads - append to existing files
    if phu_hieu_files:
        phu_hieu_paths = []
        for document in phu_hieu_files:
            if document and document.filename:
                try:
                    # Sử dụng helper function để lưu file với cấu trúc mới
                    file_path = save_uploaded_file(
                        file=document,
                        category="vehicles",
                        subcategory="insurance",
                        entity_id=vehicle.license_plate,
                        entity_type="vehicle"
                    )
                    phu_hieu_paths.append(file_path)
                except Exception as e:
                    print(f"Lỗi khi lưu file phù hiệu vận tải: {e}")
                    continue  # Skip file nếu có lỗi
        
        if phu_hieu_paths:
            # Get existing phù hiệu files and append new ones
            existing_phu_hieu = []
            if vehicle.phu_hieu_files:
                try:
                    existing_phu_hieu = json.loads(vehicle.phu_hieu_files)
                except json.JSONDecodeError:
                    existing_phu_hieu = []
            
            # Combine existing and new phù hiệu files
            all_phu_hieu = existing_phu_hieu + phu_hieu_paths
            vehicle.phu_hieu_files = json.dumps(all_phu_hieu)
    
    # Update vehicle data
    vehicle.license_plate = license_plate
    vehicle.vehicle_type = vehicle_type
    vehicle.capacity = capacity
    vehicle.fuel_consumption = fuel_consumption
    vehicle.inspection_expiry = inspection_expiry_date
    vehicle.phu_hieu_expired_date = phu_hieu_expired_date_obj
    
    db.commit()
    return RedirectResponse(url="/vehicles", status_code=303)

@app.get("/vehicles/documents/{vehicle_id}")
async def get_vehicle_documents(vehicle_id: int, db: Session = Depends(get_db)):
    """API để lấy thông tin sổ đăng kiểm của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    # Log thông tin xe để debug
    print(f"DEBUG: Vehicle ID: {vehicle_id}, License Plate: {vehicle.license_plate}")
    print(f"DEBUG: inspection_documents value: {vehicle.inspection_documents}")
    
    if not vehicle.inspection_documents:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Xe chưa upload sổ đăng kiểm"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.inspection_documents)
        print(f"DEBUG: Parsed documents: {documents}")
        
        # Kiểm tra file tồn tại
        existing_documents = []
        for doc in documents:
            # Hỗ trợ cả path cũ (static/uploads/) và path mới (Picture/...)
            if doc.startswith("Picture/"):
                # Path mới - sử dụng trực tiếp
                file_path = doc
                file_url = get_file_url(doc)
            else:
                # Path cũ - giữ nguyên để backward compatibility
                file_path = f"static/uploads/{doc}"
                file_url = f"/static/uploads/{doc}"
            
            file_exists = os.path.exists(file_path)
            print(f"DEBUG: Checking file: {file_path}, exists: {file_exists}")
            
            if file_exists:
                try:
                    file_size = os.path.getsize(file_path)
                    file_extension = os.path.splitext(doc)[1].lower()
                    # Lấy tên file từ path
                    filename = os.path.basename(doc) if "/" in doc else doc
                    existing_documents.append({
                        "filename": filename,
                        "url": file_url,
                        "size": file_size,
                        "extension": file_extension,
                        "exists": True
                    })
                    print(f"DEBUG: Added document: {doc}, size: {file_size} bytes")
                except Exception as e:
                    print(f"DEBUG: Error getting file size for {doc}: {e}")
                    filename = os.path.basename(doc) if "/" in doc else doc
                    existing_documents.append({
                        "filename": filename,
                        "url": file_url,
                        "exists": False,
                        "error": f"Lỗi khi đọc file: {str(e)}"
                    })
            else:
                filename = os.path.basename(doc) if "/" in doc else doc
                existing_documents.append({
                    "filename": filename,
                    "url": file_url,
                    "exists": False,
                    "error": "File không tồn tại trên server"
                })
                print(f"DEBUG: File not found: {file_path}")
        
        print(f"DEBUG: Returning {len(existing_documents)} documents")
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "documents": existing_documents,
                "total": len(existing_documents)
            }
        )
        
    except json.JSONDecodeError as e:
        print(f"DEBUG: JSON decode error: {e}")
        print(f"DEBUG: Raw inspection_documents: {vehicle.inspection_documents}")
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": f"Dữ liệu sổ đăng kiểm không hợp lệ: {str(e)}"}
        )
    except Exception as e:
        print(f"DEBUG: Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

@app.delete("/vehicles/documents/{vehicle_id}")
async def delete_vehicle_document(
    vehicle_id: int, 
    filename: str,
    db: Session = Depends(get_db)
):
    """API để xóa sổ đăng kiểm của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.inspection_documents:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Xe chưa có sổ đăng kiểm nào"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.inspection_documents)
        
        # Tìm file trong danh sách (có thể là filename hoặc full path)
        file_to_remove = None
        for doc in documents:
            # So sánh với filename hoặc basename của path
            doc_basename = os.path.basename(doc) if "/" in doc else doc
            if doc == filename or doc_basename == filename:
                file_to_remove = doc
                break
        
        if not file_to_remove:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "File không tồn tại trong danh sách sổ đăng kiểm"}
            )
        
        # Xóa file khỏi thư mục lưu trữ
        # Hỗ trợ cả path cũ và mới
        if file_to_remove.startswith("Picture/"):
            file_path = file_to_remove
        else:
            file_path = f"static/uploads/{file_to_remove}"
        
        delete_file_if_exists(file_path)
        
        # Xóa file khỏi danh sách trong DB
        documents.remove(file_to_remove)
        
        if documents:
            # Còn sổ đăng kiểm khác, cập nhật danh sách
            vehicle.inspection_documents = json.dumps(documents)
        else:
            # Không còn sổ đăng kiểm nào, set null
            vehicle.inspection_documents = None
        
        db.commit()
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "message": "Xóa sổ đăng kiểm thành công",
                "remaining_documents": len(documents) if documents else 0
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Lỗi định dạng dữ liệu sổ đăng kiểm"}
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

@app.get("/vehicles/phu-hieu-documents/{vehicle_id}")
async def get_vehicle_phu_hieu_documents(vehicle_id: int, db: Session = Depends(get_db)):
    """API để lấy thông tin phù hiệu vận tải của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.phu_hieu_files:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Xe chưa upload phù hiệu vận tải"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.phu_hieu_files)
        
        # Kiểm tra file tồn tại
        existing_documents = []
        for doc in documents:
            # Hỗ trợ cả path cũ (static/uploads/) và path mới (Picture/...)
            if doc.startswith("Picture/"):
                # Path mới - sử dụng trực tiếp
                file_path = doc
                file_url = get_file_url(doc)
            else:
                # Path cũ - giữ nguyên để backward compatibility
                file_path = f"static/uploads/{doc}"
                file_url = f"/static/uploads/{doc}"
            
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                file_extension = os.path.splitext(doc)[1].lower()
                filename = os.path.basename(doc) if "/" in doc else doc
                existing_documents.append({
                    "filename": filename,
                    "url": file_url,
                    "size": file_size,
                    "extension": file_extension,
                    "exists": True
                })
            else:
                filename = os.path.basename(doc) if "/" in doc else doc
                existing_documents.append({
                    "filename": filename,
                    "url": file_url,
                    "exists": False,
                    "error": "File không tồn tại trên server"
                })
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "documents": existing_documents,
                "total": len(existing_documents)
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Dữ liệu phù hiệu vận tải không hợp lệ"}
        )

@app.delete("/vehicles/phu-hieu-documents/{vehicle_id}")
async def delete_vehicle_phu_hieu_document(
    vehicle_id: int, 
    filename: str,
    db: Session = Depends(get_db)
):
    """API để xóa phù hiệu vận tải của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.phu_hieu_files:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Xe chưa có phù hiệu vận tải nào"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.phu_hieu_files)
        
        # Tìm file trong danh sách (có thể là filename hoặc full path)
        file_to_remove = None
        for doc in documents:
            # So sánh với filename hoặc basename của path
            doc_basename = os.path.basename(doc) if "/" in doc else doc
            if doc == filename or doc_basename == filename:
                file_to_remove = doc
                break
        
        if not file_to_remove:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "File không tồn tại trong danh sách phù hiệu vận tải"}
            )
        
        # Xóa file khỏi thư mục lưu trữ
        # Hỗ trợ cả path cũ và mới
        if file_to_remove.startswith("Picture/"):
            file_path = file_to_remove
        else:
            file_path = f"static/uploads/{file_to_remove}"
        
        delete_file_if_exists(file_path)
        
        # Xóa file khỏi danh sách trong DB
        documents.remove(file_to_remove)
        
        if documents:
            # Còn phù hiệu vận tải khác, cập nhật danh sách
            vehicle.phu_hieu_files = json.dumps(documents)
        else:
            # Không còn phù hiệu vận tải nào, set null
            vehicle.phu_hieu_files = None
        
        db.commit()
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "message": "Xóa phù hiệu vận tải thành công",
                "remaining_documents": len(documents) if documents else 0
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Lỗi định dạng dữ liệu phù hiệu vận tải"}
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

# ===== VEHICLE ASSIGNMENT ROUTES =====

@app.post("/vehicles/assignments/add")
async def add_vehicle_assignment(
    vehicle_id: int = Form(...),
    employee_id: int = Form(...),
    assignment_date: str = Form(...),
    db: Session = Depends(get_db)
):
    """Tạo khoán xe mới cho lái xe"""
    try:
        # Validate vehicle
        vehicle = db.query(Vehicle).filter(
            Vehicle.id == vehicle_id,
            Vehicle.status == 1,
            Vehicle.vehicle_type == "Xe Nhà"
        ).first()
        if not vehicle:
            return JSONResponse({
                "success": False,
                "message": "Xe không tồn tại hoặc không phải xe nhà"
            }, status_code=400)
        
        # Validate employee
        employee = db.query(Employee).filter(
            Employee.id == employee_id,
            Employee.position == "Lái xe",
            Employee.employee_status == "Đang làm việc",
            Employee.status == 1
        ).first()
        if not employee:
            return JSONResponse({
                "success": False,
                "message": "Lái xe không tồn tại hoặc không hợp lệ"
            }, status_code=400)
        
        # Parse date
        try:
            assignment_date_obj = datetime.strptime(assignment_date, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({
                "success": False,
                "message": "Ngày nhận xe không hợp lệ"
            }, status_code=400)
        
        # Kiểm tra xem xe có đang được khoán không
        active_assignment = db.query(VehicleAssignment).filter(
            VehicleAssignment.vehicle_id == vehicle_id,
            VehicleAssignment.end_date.is_(None)
        ).first()
        
        if active_assignment:
            # Kết thúc khoán cũ
            active_assignment.end_date = assignment_date_obj - timedelta(days=1)
            db.add(active_assignment)
        
        # Tạo khoán mới
        new_assignment = VehicleAssignment(
            vehicle_id=vehicle_id,
            employee_id=employee_id,
            assignment_date=assignment_date_obj
        )
        db.add(new_assignment)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Tạo khoán xe thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi tạo khoán xe: {str(e)}"
        }, status_code=500)

@app.get("/api/vehicles/assignments")
async def get_vehicle_assignments(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API lấy danh sách khoán xe"""
    if current_user is None or current_user["role"] != "Admin":
        return JSONResponse({
            "success": False,
            "message": "Không có quyền truy cập"
        }, status_code=403)
    
    assignments = db.query(VehicleAssignment).order_by(
        VehicleAssignment.assignment_date.desc()
    ).all()
    
    result = []
    for assignment in assignments:
        result.append({
            "id": assignment.id,
            "vehicle_id": assignment.vehicle_id,
            "vehicle_license_plate": assignment.vehicle.license_plate if assignment.vehicle else "",
            "employee_id": assignment.employee_id,
            "employee_name": assignment.employee.name if assignment.employee else "",
            "assignment_date": assignment.assignment_date.strftime("%d/%m/%Y") if assignment.assignment_date else "",
            "assignment_date_raw": assignment.assignment_date.strftime("%Y-%m-%d") if assignment.assignment_date else "",
            "end_date": assignment.end_date.strftime("%d/%m/%Y") if assignment.end_date else None,
            "is_active": assignment.end_date is None
        })
    
    return JSONResponse({
        "success": True,
        "data": result
    })

@app.get("/api/vehicles/available")
async def get_available_vehicles(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API lấy danh sách xe nhà chưa được khoán hoặc đã kết thúc khoán"""
    if current_user is None or current_user["role"] != "Admin":
        return JSONResponse({
            "success": False,
            "message": "Không có quyền truy cập"
        }, status_code=403)
    
    vehicles = db.query(Vehicle).filter(
        Vehicle.status == 1,
        Vehicle.vehicle_type == "Xe Nhà"
    ).all()
    
    available = []
    for vehicle in vehicles:
        # Kiểm tra xem xe có đang được khoán không
        active_assignment = db.query(VehicleAssignment).filter(
            VehicleAssignment.vehicle_id == vehicle.id,
            VehicleAssignment.end_date.is_(None)
        ).first()
        if not active_assignment:
            available.append({
                "id": vehicle.id,
                "license_plate": vehicle.license_plate
            })
    
    return JSONResponse({
        "success": True,
        "data": available
    })

@app.get("/api/employees/drivers")
async def get_drivers(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API lấy danh sách lái xe đang làm việc"""
    if current_user is None or current_user["role"] != "Admin":
        return JSONResponse({
            "success": False,
            "message": "Không có quyền truy cập"
        }, status_code=403)
    
    drivers = db.query(Employee).filter(
        Employee.position == "Lái xe",
        Employee.employee_status == "Đang làm việc",
        Employee.status == 1
    ).order_by(Employee.name).all()
    
    result = []
    for driver in drivers:
        result.append({
            "id": driver.id,
            "name": driver.name
        })
    
    return JSONResponse({
        "success": True,
        "data": result
    })

@app.post("/vehicles/assignments/transfer")
async def transfer_vehicle_assignment(
    assignment_id: int = Form(...),
    vehicle_id: int = Form(...),
    old_employee_id: int = Form(...),
    transfer_reason: str = Form(...),
    end_date: str = Form(...),
    new_employee_id: int = Form(...),
    new_assignment_date: str = Form(...),
    internal_note: str = Form(""),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Thu hồi và chuyển xe từ lái xe cũ sang lái xe mới"""
    try:
        # Check permission
        if current_user is None or current_user["role"] != "Admin":
            return JSONResponse({
                "success": False,
                "message": "Không có quyền thực hiện thao tác này"
            }, status_code=403)
        
        # Validate old assignment
        old_assignment = db.query(VehicleAssignment).filter(
            VehicleAssignment.id == assignment_id,
            VehicleAssignment.vehicle_id == vehicle_id,
            VehicleAssignment.employee_id == old_employee_id,
            VehicleAssignment.end_date.is_(None)
        ).first()
        
        if not old_assignment:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy khoán xe đang hoạt động"
            }, status_code=400)
        
        # Validate new employee
        new_employee = db.query(Employee).filter(
            Employee.id == new_employee_id,
            Employee.position == "Lái xe",
            Employee.employee_status == "Đang làm việc",
            Employee.status == 1
        ).first()
        
        if not new_employee:
            return JSONResponse({
                "success": False,
                "message": "Lái xe mới không tồn tại hoặc không hợp lệ"
            }, status_code=400)
        
        # Validate that new employee is not the old employee
        if new_employee_id == old_employee_id:
            return JSONResponse({
                "success": False,
                "message": "Lái xe mới không thể là lái xe cũ"
            }, status_code=400)
        
        # Parse dates
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            new_assignment_date_obj = datetime.strptime(new_assignment_date, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({
                "success": False,
                "message": "Ngày tháng không hợp lệ"
            }, status_code=400)
        
        # Validate dates
        if new_assignment_date_obj <= end_date_obj:
            return JSONResponse({
                "success": False,
                "message": "Ngày nhận xe của lái xe mới phải lớn hơn ngày kết thúc trách nhiệm của lái xe cũ"
            }, status_code=400)
        
        # Check if vehicle is already assigned to someone else on the new assignment date
        conflicting_assignment = db.query(VehicleAssignment).filter(
            VehicleAssignment.vehicle_id == vehicle_id,
            VehicleAssignment.end_date.is_(None)
        ).first()
        
        if conflicting_assignment and conflicting_assignment.id != assignment_id:
            return JSONResponse({
                "success": False,
                "message": "Xe đã được khoán cho lái xe khác"
            }, status_code=400)
        
        # Update old assignment
        old_assignment.end_date = end_date_obj
        old_assignment.transfer_reason = transfer_reason
        old_assignment.internal_note = internal_note if internal_note else None
        db.add(old_assignment)
        
        # Create new assignment
        new_assignment = VehicleAssignment(
            vehicle_id=vehicle_id,
            employee_id=new_employee_id,
            assignment_date=new_assignment_date_obj
        )
        db.add(new_assignment)
        
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Thu hồi và chuyển xe thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi thu hồi và chuyển xe: {str(e)}"
        }, status_code=500)

# ==================== BẢO DƯỠNG XE ====================

@app.get("/maintenance", response_class=HTMLResponse)
async def maintenance_page(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Trang danh sách bảo dưỡng xe"""
    
    # Lấy danh sách xe có loại = "Xe Nhà"
    vehicles = db.query(Vehicle).filter(
        Vehicle.status == 1,
        Vehicle.vehicle_type == "Xe Nhà"
    ).all()
    
    # Tính số km bảo dưỡng gần nhất cho mỗi xe
    # Logic: lấy record có ngày bảo dưỡng ≤ ngày hiện tại và gần nhất
    today = date.today()
    vehicles_with_maintenance = []
    for vehicle in vehicles:
        last_maintenance = db.query(VehicleMaintenance).filter(
            VehicleMaintenance.vehicle_id == vehicle.id,
            VehicleMaintenance.maintenance_date <= today
        ).order_by(VehicleMaintenance.maintenance_date.desc()).first()
        
        last_maintenance_km = None
        last_maintenance_date = None
        if last_maintenance:
            last_maintenance_km = last_maintenance.maintenance_km
            last_maintenance_date = last_maintenance.maintenance_date
        
        vehicles_with_maintenance.append({
            "id": vehicle.id,
            "license_plate": vehicle.license_plate,
            "last_maintenance_km": last_maintenance_km,
            "last_maintenance_date": last_maintenance_date
        })
    
    return templates.TemplateResponse("maintenance.html", {
        "request": request,
        "current_user": current_user,
        "vehicles": vehicles_with_maintenance,
        "today": today
    })

@app.get("/maintenance/detail/{vehicle_id}", response_class=JSONResponse)
async def get_maintenance_detail(vehicle_id: int, db: Session = Depends(get_db)):
    """Lấy danh sách bảo dưỡng của một xe"""
    # Kiểm tra xe có tồn tại và là "Xe Nhà"
    vehicle = db.query(Vehicle).filter(
        Vehicle.id == vehicle_id,
        Vehicle.status == 1,
        Vehicle.vehicle_type == "Xe Nhà"
    ).first()
    
    if not vehicle:
        return JSONResponse({"success": False, "error": "Không tìm thấy xe"}, status_code=404)
    
    # Lấy danh sách bảo dưỡng
    maintenances = db.query(VehicleMaintenance).filter(
        VehicleMaintenance.vehicle_id == vehicle_id
    ).order_by(VehicleMaintenance.maintenance_date.desc()).all()
    
    result = []
    for maintenance in maintenances:
        # Lấy các hạng mục bảo dưỡng
        items = db.query(VehicleMaintenanceItem).filter(
            VehicleMaintenanceItem.maintenance_id == maintenance.id
        ).all()
        
        result.append({
            "id": maintenance.id,
            "maintenance_date": maintenance.maintenance_date.strftime("%Y-%m-%d"),
            "maintenance_km": maintenance.maintenance_km,
            "vat_rate": maintenance.vat_rate,
            "total_amount": maintenance.total_amount,
            "total_with_vat": maintenance.total_with_vat,
            "items": [
                {
                    "id": item.id,
                    "content": item.content,
                    "unit": item.unit or "",
                    "quantity": item.quantity,
                    "unit_price": item.unit_price,
                    "discount_percent": item.discount_percent or 0,
                    "total_price": item.total_price
                }
                for item in items
            ]
        })
    
    return JSONResponse({
        "success": True,
        "vehicle": {
            "id": vehicle.id,
            "license_plate": vehicle.license_plate
        },
        "maintenances": result
    })

@app.post("/maintenance/add")
async def add_maintenance(
    request: Request,
    vehicle_id: int = Form(...),
    maintenance_date: str = Form(...),
    maintenance_km: float = Form(...),
    vat_rate: float = Form(0),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Thêm mới bảo dưỡng xe"""
    try:
        # Kiểm tra xe có tồn tại và là "Xe Nhà"
        vehicle = db.query(Vehicle).filter(
            Vehicle.id == vehicle_id,
            Vehicle.status == 1,
            Vehicle.vehicle_type == "Xe Nhà"
        ).first()
        
        if not vehicle:
            return JSONResponse({"success": False, "error": "Không tìm thấy xe"}, status_code=404)
        
        # Parse ngày bảo dưỡng
        try:
            maintenance_date_obj = datetime.strptime(maintenance_date, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({"success": False, "error": "Ngày bảo dưỡng không hợp lệ"}, status_code=400)
        
        # Lấy dữ liệu items từ form (JSON string)
        form_data = await request.form()
        items_json = form_data.get("items", "[]")
        
        import json
        try:
            items_data = json.loads(items_json)
        except json.JSONDecodeError:
            items_data = []
        
        # Tính tổng tiền
        total_amount = 0
        maintenance_items = []
        
        for item_data in items_data:
            content = item_data.get("content", "").strip()
            if not content:
                continue
            
            unit = item_data.get("unit", "").strip()
            quantity = float(item_data.get("quantity", 0) or 0)
            unit_price = float(item_data.get("unit_price", 0) or 0)
            discount_percent = float(item_data.get("discount_percent", 0) or 0)
            # Thành tiền = SL × Đơn giá × (1 − Giảm giá / 100)
            total_price = quantity * unit_price * (1 - discount_percent / 100)
            
            total_amount += total_price
            
            maintenance_items.append({
                "content": content,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "discount_percent": discount_percent,
                "total_price": total_price
            })
        
        # Tính tổng có VAT
        vat_amount = total_amount * (vat_rate / 100)
        total_with_vat = total_amount + vat_amount
        
        # Tạo bảo dưỡng
        maintenance = VehicleMaintenance(
            vehicle_id=vehicle_id,
            maintenance_date=maintenance_date_obj,
            maintenance_km=maintenance_km,
            vat_rate=vat_rate,
            total_amount=total_amount,
            total_with_vat=total_with_vat
        )
        db.add(maintenance)
        db.flush()  # Để lấy ID
        
        # Tạo các hạng mục bảo dưỡng
        for item_data in maintenance_items:
            item = VehicleMaintenanceItem(
                maintenance_id=maintenance.id,
                content=item_data["content"],
                unit=item_data["unit"],
                quantity=item_data["quantity"],
                unit_price=item_data["unit_price"],
                discount_percent=item_data.get("discount_percent", 0),
                total_price=item_data["total_price"]
            )
            db.add(item)
        
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã thêm bảo dưỡng thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "error": f"Lỗi hệ thống: {str(e)}"
        }, status_code=500)

@app.put("/maintenance/edit/{maintenance_id}")
async def edit_maintenance(
    maintenance_id: int,
    request: Request,
    maintenance_date: str = Form(...),
    maintenance_km: float = Form(...),
    vat_rate: float = Form(0),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Sửa bảo dưỡng xe"""
    if current_user is None:
        return JSONResponse({"success": False, "error": "Chưa đăng nhập"}, status_code=401)
    
    try:
        # Kiểm tra bảo dưỡng có tồn tại
        maintenance = db.query(VehicleMaintenance).filter(
            VehicleMaintenance.id == maintenance_id
        ).first()
        
        if not maintenance:
            return JSONResponse({"success": False, "error": "Không tìm thấy bảo dưỡng"}, status_code=404)
        
        # Kiểm tra xe có tồn tại và là "Xe Nhà"
        vehicle = db.query(Vehicle).filter(
            Vehicle.id == maintenance.vehicle_id,
            Vehicle.status == 1,
            Vehicle.vehicle_type == "Xe Nhà"
        ).first()
        
        if not vehicle:
            return JSONResponse({"success": False, "error": "Không tìm thấy xe"}, status_code=404)
        
        # Parse ngày bảo dưỡng
        try:
            maintenance_date_obj = datetime.strptime(maintenance_date, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({"success": False, "error": "Ngày bảo dưỡng không hợp lệ"}, status_code=400)
        
        # Lấy dữ liệu items từ form (JSON string)
        form_data = await request.form()
        items_json = form_data.get("items", "[]")
        
        import json
        try:
            items_data = json.loads(items_json)
        except json.JSONDecodeError:
            items_data = []
        
        # Xóa các items cũ
        db.query(VehicleMaintenanceItem).filter(
            VehicleMaintenanceItem.maintenance_id == maintenance_id
        ).delete()
        
        # Tính tổng tiền
        total_amount = 0
        maintenance_items = []
        
        for item_data in items_data:
            content = item_data.get("content", "").strip()
            if not content:
                continue
            
            unit = item_data.get("unit", "").strip()
            quantity = float(item_data.get("quantity", 0) or 0)
            unit_price = float(item_data.get("unit_price", 0) or 0)
            discount_percent = float(item_data.get("discount_percent", 0) or 0)
            # Thành tiền = SL × Đơn giá × (1 − Giảm giá / 100)
            total_price = quantity * unit_price * (1 - discount_percent / 100)
            
            total_amount += total_price
            
            maintenance_items.append({
                "content": content,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "discount_percent": discount_percent,
                "total_price": total_price
            })
        
        # Tính tổng có VAT
        vat_amount = total_amount * (vat_rate / 100)
        total_with_vat = total_amount + vat_amount
        
        # Cập nhật bảo dưỡng
        maintenance.maintenance_date = maintenance_date_obj
        maintenance.maintenance_km = maintenance_km
        maintenance.vat_rate = vat_rate
        maintenance.total_amount = total_amount
        maintenance.total_with_vat = total_with_vat
        
        # Tạo các hạng mục bảo dưỡng mới
        for item_data in maintenance_items:
            item = VehicleMaintenanceItem(
                maintenance_id=maintenance.id,
                content=item_data["content"],
                unit=item_data["unit"],
                quantity=item_data["quantity"],
                unit_price=item_data["unit_price"],
                discount_percent=item_data.get("discount_percent", 0),
                total_price=item_data["total_price"]
            )
            db.add(item)
        
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã cập nhật bảo dưỡng thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "error": f"Lỗi hệ thống: {str(e)}"
        }, status_code=500)

@app.delete("/maintenance/delete/{maintenance_id}")
async def delete_maintenance(
    maintenance_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Xóa bảo dưỡng xe"""
    if current_user is None:
        return JSONResponse({"success": False, "error": "Chưa đăng nhập"}, status_code=401)
    
    try:
        # Kiểm tra bảo dưỡng có tồn tại
        maintenance = db.query(VehicleMaintenance).filter(
            VehicleMaintenance.id == maintenance_id
        ).first()
        
        if not maintenance:
            return JSONResponse({"success": False, "error": "Không tìm thấy bảo dưỡng"}, status_code=404)
        
        # Kiểm tra xe có tồn tại và là "Xe Nhà"
        vehicle = db.query(Vehicle).filter(
            Vehicle.id == maintenance.vehicle_id,
            Vehicle.status == 1,
            Vehicle.vehicle_type == "Xe Nhà"
        ).first()
        
        if not vehicle:
            return JSONResponse({"success": False, "error": "Không tìm thấy xe"}, status_code=404)
        
        # Xóa các items (cascade sẽ tự động xóa do relationship)
        db.query(VehicleMaintenanceItem).filter(
            VehicleMaintenanceItem.maintenance_id == maintenance_id
        ).delete()
        
        # Xóa bảo dưỡng
        db.delete(maintenance)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã xóa bảo dưỡng thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "error": f"Lỗi hệ thống: {str(e)}"
        }, status_code=500)

@app.get("/operations", response_class=HTMLResponse)
async def operations_page(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Chỉ Admin mới được truy cập
    if current_user["role"] != "Admin":
        return RedirectResponse(url="/daily-new", status_code=303)
    
    return templates.TemplateResponse("operations.html", {
        "request": request,
        "current_user": current_user
    })

@app.get("/routes", response_class=HTMLResponse)
async def routes_page(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Chỉ Admin mới được truy cập
    if current_user["role"] != "Admin":
        return RedirectResponse(url="/daily-new", status_code=303)
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        # Lọc ra routes không phải "Tăng Cường"
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        
        # Lọc ra routes "Tăng Cường"
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        
        # Sắp xếp routes bình thường theo A-Z
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower() if route.route_code else "")
        
        # Ghép lại: routes bình thường + routes "Tăng Cường"
        return normal_routes_sorted + tang_cuong_routes
    
    routes = sort_routes_with_tang_cuong_at_bottom(routes)
    
    # Lấy danh sách các bản cập nhật giá tuyến (nhóm theo update_name và application_date)
    price_updates = db.query(
        RoutePrice.update_name,
        RoutePrice.application_date,
        func.count(RoutePrice.id).label('route_count'),
        func.min(RoutePrice.created_at).label('created_at')
    ).filter(
        RoutePrice.update_name.isnot(None),
        RoutePrice.update_name != ''
    ).group_by(
        RoutePrice.update_name,
        RoutePrice.application_date
    ).order_by(
        RoutePrice.application_date.desc(),
        RoutePrice.created_at.desc()
    ).all()
    
    return templates.TemplateResponse("routes.html", {
        "request": request, 
        "current_user": current_user, 
        "routes": routes,
        "price_updates": price_updates
    })

@app.post("/routes/add")
async def add_route(
    route_code: str = Form(...),
    route_name: str = Form(...),
    route_type: str = Form(...),
    unit_price: float = Form(...),
    bridge_fee: float = Form(0),
    loading_fee: float = Form(0),
    distance: float = Form(0),
    monthly_salary: float = Form(0),
    route_status: str = Form("ONL"),
    db: Session = Depends(get_db)
):
    route = Route(
        route_code=route_code,
        route_name=route_name,
        route_type=route_type,
        unit_price=unit_price,
        bridge_fee=bridge_fee if bridge_fee else 0,
        loading_fee=loading_fee if loading_fee else 0,
        distance=distance,
        monthly_salary=monthly_salary,
        route_status=route_status if route_status in ["ONL", "OFF"] else "ONL",
        vehicle_id=None  # No vehicle assigned by default
    )
    db.add(route)
    db.commit()
    return RedirectResponse(url="/routes", status_code=303)

@app.post("/routes/delete/{route_id}")
async def delete_route(route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
    if route:
        route.status = 0  # Soft delete
        db.commit()
    return RedirectResponse(url="/routes", status_code=303)

@app.get("/routes/edit/{route_id}", response_class=HTMLResponse)
async def edit_route_page(request: Request, route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
    if not route:
        return RedirectResponse(url="/routes", status_code=303)
    return templates.TemplateResponse("edit_route.html", {
        "request": request, 
        "route": route
    })

@app.post("/routes/edit/{route_id}")
async def edit_route(
    route_id: int,
    route_code: str = Form(...),
    route_name: str = Form(...),
    route_type: str = Form(...),
    unit_price: float = Form(...),
    bridge_fee: float = Form(0),
    loading_fee: float = Form(0),
    distance: float = Form(0),
    monthly_salary: float = Form(0),
    route_status: str = Form("ONL"),
    db: Session = Depends(get_db)
):
    route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
    if not route:
        return RedirectResponse(url="/routes", status_code=303)
    
    route.route_code = route_code
    route.route_name = route_name
    route.route_type = route_type
    route.unit_price = unit_price
    route.bridge_fee = bridge_fee if bridge_fee else 0
    route.loading_fee = loading_fee if loading_fee else 0
    route.distance = distance
    route.monthly_salary = monthly_salary
    route.route_status = route_status if route_status in ["ONL", "OFF"] else "ONL"
    
    db.commit()
    return RedirectResponse(url="/routes", status_code=303)

@app.post("/routes/update-price", include_in_schema=False)
async def update_route_price(
    request: Request,
    db: Session = Depends(get_db)
):
    """Cập nhật giá tuyến theo ngày áp dụng - Bulk update"""
    try:
        # Lấy form data trực tiếp
        form_data = await request.form()
        
        # Debug: In ra tất cả keys để kiểm tra
        print(f"DEBUG: Form data keys: {list(form_data.keys())}")
        print(f"DEBUG: Content-Type: {request.headers.get('content-type', 'N/A')}")
        
        # Lấy dữ liệu từ form
        application_date = form_data.get("application_date")
        update_name = form_data.get("update_name")
        route_ids = form_data.getlist("route_ids")
        unit_prices = form_data.getlist("unit_prices")
        
        print(f"DEBUG: application_date={application_date}, update_name={update_name}")
        print(f"DEBUG: route_ids={route_ids}, unit_prices={unit_prices}")
        
        # Kiểm tra nếu không có dữ liệu
        if not route_ids or not unit_prices:
            return RedirectResponse(url="/routes?error=missing_data", status_code=303)
        
        # Kiểm tra dữ liệu đầu vào
        if not application_date or not update_name:
            return RedirectResponse(url="/routes?error=missing_data", status_code=303)
        
        # Parse ngày áp giá
        try:
            app_date = datetime.strptime(application_date, "%Y-%m-%d").date()
        except ValueError:
            return RedirectResponse(url="/routes?error=invalid_date", status_code=303)
        
        # Kiểm tra số lượng route_ids và unit_prices phải bằng nhau
        if len(route_ids) != len(unit_prices):
            return RedirectResponse(url="/routes?error=invalid_data", status_code=303)
        
        # Xử lý từng tuyến
        success_count = 0
        for route_id_str, unit_price_str in zip(route_ids, unit_prices):
            try:
                route_id = int(route_id_str)
                unit_price = int(unit_price_str)
                
                # Kiểm tra route có tồn tại và không phải "Tăng Cường"
                route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
                if not route:
                    continue
                
                if route.route_code and route.route_code.strip() == "Tăng Cường":
                    continue
                
                # Tạo bản ghi giá tuyến mới
                # Lấy giá dầu mới nhất (có thể từ bảng fuel hoặc mặc định)
                # Tạm thời set fuel_price = 0 vì không có trong form mới
                route_price = RoutePrice(
                    route_id=route_id,
                    unit_price=unit_price,
                    fuel_price=0,  # Có thể cần điều chỉnh sau
                    application_date=app_date,
                    update_name=update_name
                )
                db.add(route_price)
                success_count += 1
            except (ValueError, TypeError) as e:
                print(f"Error processing route {route_id_str}: {e}")
                continue
        
        if success_count == 0:
            return RedirectResponse(url="/routes?error=no_routes_updated", status_code=303)
        
        db.commit()
        return RedirectResponse(url=f"/routes?success=price_updated&count={success_count}", status_code=303)
    except Exception as e:
        print(f"Error updating route prices: {e}")
        db.rollback()
        return RedirectResponse(url="/routes?error=update_failed", status_code=303)

@app.get("/routes/price-update-detail")
async def get_price_update_detail(
    request: Request,
    update_name: str,
    application_date: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Lấy chi tiết một bản cập nhật giá tuyến"""
    if current_user is None or current_user["role"] != "Admin":
        return JSONResponse(status_code=403, content={"error": "Unauthorized"})
    
    try:
        app_date = datetime.strptime(application_date, "%Y-%m-%d").date()
    except ValueError:
        return JSONResponse(status_code=400, content={"error": "Invalid date format"})
    
    # Lấy tất cả các bản ghi giá tuyến của bản cập nhật này
    route_prices = db.query(RoutePrice).filter(
        RoutePrice.update_name == update_name,
        RoutePrice.application_date == app_date
    ).join(Route).order_by(Route.route_code).all()
    
    # Chuyển đổi sang dictionary
    details = []
    for rp in route_prices:
        details.append({
            "id": rp.id,
            "route_id": rp.route_id,
            "route_code": rp.route.route_code if rp.route else "",
            "route_name": rp.route.route_name if rp.route else "",
            "unit_price": rp.unit_price,
            "fuel_price": rp.fuel_price,
            "application_date": rp.application_date.strftime("%Y-%m-%d"),
            "update_name": rp.update_name
        })
    
    return JSONResponse(content={
        "update_name": update_name,
        "application_date": application_date,
        "details": details
    })

@app.post("/routes/price-update-edit")
async def edit_price_update(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Sửa bản cập nhật giá tuyến"""
    if current_user is None or current_user["role"] != "Admin":
        return RedirectResponse(url="/routes?error=unauthorized", status_code=303)
    
    try:
        form_data = await request.form()
        
        # Lấy dữ liệu từ form
        update_name = form_data.get("update_name")
        application_date = form_data.get("application_date")
        route_price_ids = form_data.getlist("route_price_ids")
        unit_prices = form_data.getlist("unit_prices")
        
        if not update_name or not application_date:
            return RedirectResponse(url="/routes?error=missing_data", status_code=303)
        
        try:
            app_date = datetime.strptime(application_date, "%Y-%m-%d").date()
        except ValueError:
            return RedirectResponse(url="/routes?error=invalid_date", status_code=303)
        
        # Lấy tất cả các bản ghi của bản cập nhật này
        route_prices = db.query(RoutePrice).filter(
            RoutePrice.update_name == update_name,
            RoutePrice.application_date == app_date
        ).all()
        
        # Cập nhật tên bản cập nhật nếu có thay đổi
        new_update_name = form_data.get("new_update_name", "").strip()
        if new_update_name and new_update_name != update_name:
            for rp in route_prices:
                rp.update_name = new_update_name
        
        # Cập nhật giá cho từng tuyến
        if route_price_ids and unit_prices and len(route_price_ids) == len(unit_prices):
            for price_id_str, unit_price_str in zip(route_price_ids, unit_prices):
                try:
                    price_id = int(price_id_str)
                    unit_price = int(unit_price_str)
                    
                    route_price = db.query(RoutePrice).filter(RoutePrice.id == price_id).first()
                    if route_price:
                        route_price.unit_price = unit_price
                except (ValueError, TypeError) as e:
                    print(f"Error updating price for route_price_id {price_id_str}: {e}")
                    continue
        
        db.commit()
        return RedirectResponse(url="/routes?success=price_update_edited", status_code=303)
    except Exception as e:
        print(f"Error editing price update: {e}")
        db.rollback()
        return RedirectResponse(url="/routes?error=edit_failed", status_code=303)

# ===== REVENUE MANAGEMENT ROUTES =====

@app.get("/revenue", response_class=HTMLResponse)
async def revenue_page(request: Request, db: Session = Depends(get_db), selected_date: Optional[str] = None, deleted_all: Optional[str] = None, current_user = Depends(get_current_user)):
    """Trang quản lý doanh thu - Tự động tính từ dữ liệu chấm công"""
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Kiểm tra quyền truy cập (User hoặc Admin)
    redirect_response = check_and_redirect_access(current_user["role"], "/revenue", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    today = date.today()
    
    # Xử lý ngày được chọn
    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except ValueError:
            filter_date = today
    else:
        filter_date = today
    
    # Lấy dữ liệu chấm công (DailyRoute) cho ngày được chọn
    daily_routes = db.query(DailyRoute).filter(DailyRoute.date == filter_date).all()
    
    # Tự động tính toán và tạo/cập nhật doanh thu từ dữ liệu chấm công
    revenue_dict = {}
    routes_with_attendance = set()
    
    # Đơn giá mặc định cho tuyến Nội thành
    NOI_THANH_UNIT_PRICE = 227273
    
    # Nhóm DailyRoute theo route_id để xử lý
    daily_routes_by_route = {}
    for daily_route in daily_routes:
        route = daily_route.route
        if not route:
            continue
        
        route_id = route.id
        routes_with_attendance.add(route_id)
        
        # Bỏ qua tuyến Tăng cường - sẽ xử lý riêng (nhập thủ công)
        if route.route_code and route.route_code.strip() == "Tăng Cường":
            continue
        
        if route_id not in daily_routes_by_route:
            daily_routes_by_route[route_id] = []
        daily_routes_by_route[route_id].append(daily_route)
    
    # Xử lý từng route
    for route_id, route_daily_routes in daily_routes_by_route.items():
        # Lấy thông tin route
        route = route_daily_routes[0].route
        if not route:
            continue
        
        # Lọc các chuyến có trạng thái ON (Online)
        # Chỉ tính doanh thu cho các chuyến có status = "Online" hoặc "ON"
        online_daily_routes = [
            dr for dr in route_daily_routes 
            if dr.status and dr.status.upper() in ["ONLINE", "ON"]
        ]
        
        # Kiểm tra xem đã có RevenueRecord chưa
        existing_revenue = db.query(RevenueRecord).filter(
            RevenueRecord.route_id == route_id,
            RevenueRecord.date == filter_date
        ).first()
        
        # Xác định status: Nếu có ít nhất 1 chuyến ON thì status = "Online", ngược lại = "OFF"
        if online_daily_routes:
            status = "Online"
        else:
            # Tất cả chuyến đều OFF
            status = "OFF"
        
        # Lấy license_plate và driver_name từ DailyRoute
        # Ưu tiên lấy từ chuyến có status Online, nếu không có thì lấy từ chuyến đầu tiên
        license_plate = ""
        driver_name = ""
        notes = ""
        if online_daily_routes:
            # Lấy từ chuyến đầu tiên có status Online
            first_online_route = online_daily_routes[0]
            license_plate = first_online_route.license_plate or ""
            driver_name = first_online_route.driver_name or ""
            notes = first_online_route.notes or ""
        elif route_daily_routes:
            # Nếu không có chuyến Online, lấy từ chuyến đầu tiên
            first_route = route_daily_routes[0]
            license_plate = first_route.license_plate or ""
            driver_name = first_route.driver_name or ""
            notes = first_route.notes or ""
        
        # Tính doanh thu tự động dựa trên loại tuyến
        # Chỉ tính doanh thu nếu có ít nhất 1 chuyến ON
        if not online_daily_routes:
            # Tất cả chuyến đều OFF: doanh thu = 0
            total_amount = 0
            distance_km = route.distance or 0 if route.route_type != "Nội thành" else 0
            unit_price = route.unit_price or 0 if route.route_type != "Nội thành" else NOI_THANH_UNIT_PRICE
            bridge_fee = 0
            loading_fee = 0
        elif route.route_type == "Nội thành":
            # Nội thành: Đơn giá cố định 227,273 VNĐ/chuyến
            # Đếm số chuyến ON (mỗi DailyRoute = 1 chuyến)
            trip_count = len(online_daily_routes)
            total_amount = NOI_THANH_UNIT_PRICE * trip_count
            distance_km = 0  # Không dùng km cho Nội thành
            unit_price = NOI_THANH_UNIT_PRICE
            bridge_fee = 0
            loading_fee = 0
        else:
            # Nội Tỉnh hoặc Liên Tỉnh: (Số km × Đơn giá) + Phí cầu đường + Phí chờ tải
            # Lấy km từ khoảng cách đã khai báo tại Page routes (route.distance)
            distance_km = route.distance or 0
            unit_price = route.unit_price or 0
            bridge_fee = route.bridge_fee or 0
            loading_fee = route.loading_fee or 0
            
            base_revenue = distance_km * unit_price
            total_amount = int(base_revenue + bridge_fee + loading_fee)
        
        # Tạo hoặc cập nhật RevenueRecord
        if existing_revenue:
            # Chỉ cập nhật nếu chưa có manual_total (giữ nguyên nếu đã nhập thủ công)
            if existing_revenue.manual_total == 0:
                # Kiểm tra xem distance_km đã được chỉnh sửa chưa (khác route.distance)
                # Nếu đã chỉnh sửa, giữ nguyên số km thực tế
                existing_distance_km = existing_revenue.distance_km or 0
                route_default_distance = route.distance or 0
                
                # Nếu số km hiện tại khác số km mặc định, có nghĩa là đã được chỉnh sửa
                # Trong trường hợp này, giữ nguyên số km thực tế đã chỉnh sửa
                if abs(existing_distance_km - route_default_distance) > 0.01:  # Cho phép sai số nhỏ do float
                    # Đã chỉnh sửa: giữ nguyên distance_km, nhưng tính lại total_amount với số km thực tế
                    distance_km_to_use = existing_distance_km
                else:
                    # Chưa chỉnh sửa: cập nhật bằng số km mặc định
                    distance_km_to_use = distance_km
                
                existing_revenue.distance_km = distance_km_to_use
                existing_revenue.unit_price = unit_price
                existing_revenue.bridge_fee = bridge_fee
                existing_revenue.loading_fee = loading_fee
                existing_revenue.late_penalty = 0
                
                # Tính lại total_amount với số km thực tế (có thể là số km đã chỉnh sửa)
                if route.route_type == "Nội thành":
                    # Nội thành: Đơn giá cố định
                    existing_revenue.total_amount = NOI_THANH_UNIT_PRICE * len(online_daily_routes)
                else:
                    # Nội Tỉnh hoặc Liên Tỉnh: Đơn giá × Số km thực tế
                    base_revenue = distance_km_to_use * unit_price
                    existing_revenue.total_amount = int(base_revenue + bridge_fee + loading_fee)
                
                existing_revenue.status = status
                # Cập nhật license_plate và driver_name nếu chưa có hoặc từ DailyRoute
                if license_plate:
                    existing_revenue.license_plate = license_plate
                if driver_name:
                    existing_revenue.driver_name = driver_name
                if notes:
                    existing_revenue.notes = notes
                existing_revenue.updated_at = datetime.utcnow()
                revenue_record = existing_revenue
        else:
            # Tạo mới
            revenue_record = RevenueRecord(
                date=filter_date,
                route_id=route_id,
                route_type=route.route_type or "Nội Tỉnh",  # Lấy từ route
                distance_km=distance_km,
                unit_price=unit_price,
                bridge_fee=bridge_fee,
                loading_fee=loading_fee,
                late_penalty=0,
                status=status,
                total_amount=total_amount,
                manual_total=0,
                route_name="",
                license_plate=license_plate,
                driver_name=driver_name,
                notes=notes
            )
            db.add(revenue_record)
    
    # Commit các thay đổi tự động
    try:
        db.commit()
        # Tự động cập nhật bản ghi thu nhập trong finance-report sau khi tính doanh thu
        await create_daily_revenue_finance_record(filter_date, db)
    except Exception as e:
        print(f"Error auto-creating revenue records: {e}")
        db.rollback()
    
    # Lấy lại tất cả revenue records sau khi đã tự động tạo/cập nhật
    revenue_records = db.query(RevenueRecord).filter(RevenueRecord.date == filter_date).all()
    
    # Tạo dictionary để dễ tra cứu
    for record in revenue_records:
        route = record.route
        if not route:
            continue
        
        route_id = record.route_id
        # Với "Tăng cường", lưu tất cả records (sẽ được xử lý riêng trong template)
        if route.route_code and route.route_code.strip() == "Tăng Cường":
            if route_id not in revenue_dict:
                revenue_dict[route_id] = []
            revenue_dict[route_id].append(record)
        else:
            # Với các tuyến khác, chỉ lưu record đầu tiên
            if route_id not in revenue_dict:
                revenue_dict[route_id] = record
    
    # Lấy tất cả routes
    all_routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Lọc routes để hiển thị:
    # - Tuyến có chấm công: hiển thị doanh thu đã tự động tính
    # - Tuyến Tăng cường: LUÔN hiển thị trong all_routes (để form có thể lấy được)
    routes_for_display = []
    for route in all_routes:
        if route.route_code and route.route_code.strip() == "Tăng Cường":
            # Tăng cường: luôn hiển thị để nhập thủ công (ngay cả khi chưa có chấm công)
            routes_for_display.append(route)
        elif route.id in routes_with_attendance:
            # Tuyến có chấm công: hiển thị doanh thu đã tự động tính
            routes_for_display.append(route)
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower() if route.route_code else "")
        return normal_routes_sorted + tang_cuong_routes
    
    routes_for_display = sort_routes_with_tang_cuong_at_bottom(routes_for_display)
    all_routes = sort_routes_with_tang_cuong_at_bottom(all_routes)
    
    # Lấy danh sách xe và nhân viên
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).order_by(Vehicle.license_plate).all()
    # Chỉ lấy nhân viên có trạng thái "Đang làm việc"
    employees = db.query(Employee).filter(
        Employee.status == 1,
        Employee.employee_status == "Đang làm việc"
    ).order_by(Employee.name).all()
    
    # Chuyển đổi thành dictionaries để JavaScript có thể sử dụng
    vehicles_list = [{"license_plate": v.license_plate or ""} for v in vehicles]
    employees_list = [{"name": e.name or ""} for e in employees]
    
    return templates.TemplateResponse("revenue.html", {
        "request": request,
        "current_user": current_user,
        "routes": routes_for_display,  # Routes có chấm công hoặc Tăng cường
        "all_routes": all_routes,    # Tất cả routes để hiển thị trong bảng đã ghi nhận
        "revenue_dict": revenue_dict,
        "filter_date": filter_date,
        "today": today,
        "deleted_all": deleted_all,
        "routes_with_attendance": routes_with_attendance,  # Set các route_id có chấm công
        "vehicles": vehicles_list,  # Danh sách xe (dạng dictionary)
        "employees": employees_list  # Danh sách nhân viên (dạng dictionary, chỉ "Đang làm việc")
    })

@app.post("/revenue/add")
async def add_revenue_today(request: Request, db: Session = Depends(get_db)):
    """Thêm doanh thu cho tuyến Tăng cường (chỉ nhập thủ công)"""
    form_data = await request.form()
    
    # Lấy ngày được chọn từ form
    selected_date_str = form_data.get("date")
    if not selected_date_str:
        return RedirectResponse(url="/revenue", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = date.today()
    
    # Lấy route_id từ form
    route_id_str = form_data.get("route_id")
    if not route_id_str:
        return RedirectResponse(url="/revenue", status_code=303)
    
    try:
        route_id = int(route_id_str)
    except (ValueError, TypeError):
        return RedirectResponse(url="/revenue", status_code=303)
    
    # Xử lý từng dòng doanh thu (mỗi dòng là 1 chuyến)
    row_index = 1
    while True:
        # Lấy dữ liệu từ form cho dòng này
        route_name = form_data.get(f"route_name_{row_index}")
        distance_km = form_data.get(f"distance_km_{row_index}")
        unit_price = form_data.get(f"unit_price_{row_index}")
        bridge_fee = form_data.get(f"bridge_fee_{row_index}")
        loading_fee = form_data.get(f"loading_fee_{row_index}")
        total_amount_input = form_data.get(f"total_amount_{row_index}")
        license_plate = form_data.get(f"license_plate_{row_index}")
        driver_name = form_data.get(f"driver_name_{row_index}")
        notes = form_data.get(f"notes_{row_index}")
        
        # Nếu không có route_name thì dừng (hết dòng) - nhưng cần kiểm tra ít nhất 1 trường để tránh dòng trống
        # Kiểm tra nếu không có dữ liệu nào thì dừng
        if not route_name and not distance_km and not unit_price and not license_plate and not driver_name:
            break
        
        # Bỏ qua dòng trống (không có dữ liệu quan trọng)
        if not distance_km and not unit_price:
            row_index += 1
            continue
        
        # Xử lý giá trị
        try:
            distance_km_val = float(distance_km) if distance_km and distance_km.strip() else 0
        except (ValueError, AttributeError):
            distance_km_val = 0
            
        try:
            unit_price_val = int(unit_price) if unit_price and unit_price.strip() else 0
        except (ValueError, AttributeError):
            unit_price_val = 0
            
        try:
            bridge_fee_val = int(bridge_fee) if bridge_fee and bridge_fee.strip() else 0
        except (ValueError, AttributeError):
            bridge_fee_val = 0
            
        try:
            loading_fee_val = int(loading_fee) if loading_fee and loading_fee.strip() else 0
        except (ValueError, AttributeError):
            loading_fee_val = 0
        
        # Tính thành tiền: Đơn giá × Số km + Phí cầu đường + Phí chờ tải
        base_revenue = distance_km_val * unit_price_val
        total_amount = max(0, int(base_revenue + bridge_fee_val + loading_fee_val))
        
        # Tạo record mới cho Tăng cường
        revenue_record = RevenueRecord(
            date=selected_date,
            route_id=route_id,
            route_type="Tăng cường",  # Cố định là "Tăng cường"
            distance_km=distance_km_val,
            unit_price=unit_price_val,
            bridge_fee=bridge_fee_val,
            loading_fee=loading_fee_val,
            late_penalty=0,
            status="Online",
            total_amount=total_amount,
            manual_total=0,  # Không dùng manual_total nữa
            route_name=route_name or "",
            license_plate=license_plate or "",
            driver_name=driver_name or "",
            notes=notes or ""
        )
        db.add(revenue_record)
        
        row_index += 1
    
    try:
        db.commit()
        print(f"Successfully committed Tăng cường revenue records for date {selected_date}")
        
        # Tự động tạo bản ghi thu nhập trong finance-report
        await create_daily_revenue_finance_record(selected_date, db)
        
    except Exception as e:
        print(f"Error committing revenue records: {e}")
        db.rollback()
        return RedirectResponse(url="/revenue", status_code=303)
    
    # Redirect về trang revenue với ngày đã chọn
    return RedirectResponse(url=f"/revenue?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)

@app.get("/revenue/edit/{revenue_id}", response_class=HTMLResponse)
async def edit_revenue_page(request: Request, revenue_id: int, db: Session = Depends(get_db)):
    """Trang sửa doanh thu"""
    try:
        revenue_record = db.query(RevenueRecord).filter(RevenueRecord.id == revenue_id).first()
    except Exception as e:
        print(f"Error querying revenue record for edit: {e}")
        return RedirectResponse(url="/revenue", status_code=303)
    
    if not revenue_record:
        return RedirectResponse(url="/revenue", status_code=303)
    
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    return templates.TemplateResponse("edit_revenue.html", {
        "request": request,
        "revenue_record": revenue_record,
        "routes": routes
    })

@app.post("/revenue/edit/{revenue_id}")
async def edit_revenue(
    revenue_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Cập nhật doanh thu"""
    form_data = await request.form()
    
    try:
        revenue_record = db.query(RevenueRecord).filter(RevenueRecord.id == revenue_id).first()
    except Exception as e:
        print(f"Error querying revenue record for update: {e}")
        return RedirectResponse(url="/revenue", status_code=303)
    
    if not revenue_record:
        return RedirectResponse(url="/revenue", status_code=303)
    
    # Cập nhật thông tin theo cấu trúc mới
    status = form_data.get("status", "Online")
    
    # Xử lý giá trị dựa trên trạng thái
    if status == "Offline":
        # Khi Offline: tất cả trường nhập liệu = 0, chỉ giữ lại status và notes
        distance_km = 0
        unit_price = 0
        bridge_fee = 0
        loading_fee = 0
        late_penalty = 0
        manual_total = 0
    else:
        # Xử lý các giá trị số, đảm bảo không bị lỗi khi chuỗi rỗng
        try:
            distance_km_str = form_data.get("distance_km", "0")
            distance_km = float(distance_km_str) if distance_km_str and distance_km_str.strip() else 0.0
        except (ValueError, AttributeError):
            distance_km = 0.0
        
        try:
            unit_price_str = form_data.get("unit_price", "0")
            unit_price = int(unit_price_str) if unit_price_str and unit_price_str.strip() else 0
        except (ValueError, AttributeError):
            unit_price = 0
        
        try:
            bridge_fee_str = form_data.get("bridge_fee", "0")
            bridge_fee = int(bridge_fee_str) if bridge_fee_str and bridge_fee_str.strip() else 0
        except (ValueError, AttributeError):
            bridge_fee = 0
        
        try:
            loading_fee_str = form_data.get("loading_fee", "0")
            loading_fee = int(loading_fee_str) if loading_fee_str and loading_fee_str.strip() else 0
        except (ValueError, AttributeError):
            loading_fee = 0
        
        try:
            late_penalty_str = form_data.get("late_penalty", "0")
            late_penalty = int(late_penalty_str) if late_penalty_str and late_penalty_str.strip() else 0
        except (ValueError, AttributeError):
            late_penalty = 0
        
        try:
            manual_total_str = form_data.get("manual_total", "0")
            manual_total = int(manual_total_str) if manual_total_str and manual_total_str.strip() else 0
        except (ValueError, AttributeError):
            manual_total = 0
    
    # Lấy route để kiểm tra route_type
    route = db.query(Route).filter(Route.id == revenue_record.route_id).first()
    route_type = route.route_type if route else "Nội Tỉnh"
    
    # Tính thành tiền: ưu tiên manual_total, nếu không có thì dùng công thức
    if manual_total > 0:
        total_amount = manual_total
    elif status == "Offline":
        total_amount = 0  # Offline mà không có manual total thì = 0
    else:
        # Tính doanh thu theo loại tuyến
        if route_type == "Nội thành":
            # Nội thành: Đơn giá theo chuyến, không nhân km
            base_revenue = unit_price
        else:
            # Nội Tỉnh hoặc Liên Tỉnh: Đơn giá × Số km
            base_revenue = distance_km * unit_price
        
        # Công thức: Doanh thu cơ bản + Phí cầu đường + Phí dừng tải – Trễ Ontime
        total_amount = max(0, int(base_revenue + bridge_fee + loading_fee - late_penalty))
    
    revenue_record.distance_km = distance_km
    revenue_record.unit_price = unit_price
    revenue_record.bridge_fee = bridge_fee
    revenue_record.loading_fee = loading_fee
    revenue_record.late_penalty = late_penalty
    revenue_record.status = status
    revenue_record.total_amount = total_amount
    revenue_record.manual_total = manual_total
    revenue_record.notes = form_data.get("notes", "")
    revenue_record.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        
        # Tự động cập nhật bản ghi thu nhập trong finance-report
        await create_daily_revenue_finance_record(revenue_record.date, db)
        
    except Exception as e:
        print(f"Error updating revenue record: {e}")
        db.rollback()
        return RedirectResponse(url="/revenue", status_code=303)
    
    return RedirectResponse(url=f"/revenue?selected_date={revenue_record.date.strftime('%Y-%m-%d')}", status_code=303)

@app.post("/revenue/delete/{revenue_id}")
async def delete_revenue(revenue_id: int, db: Session = Depends(get_db)):
    """Xóa doanh thu"""
    try:
        revenue_record = db.query(RevenueRecord).filter(RevenueRecord.id == revenue_id).first()
    except Exception as e:
        print(f"Error querying revenue record for delete: {e}")
        return RedirectResponse(url="/revenue", status_code=303)
    
    if revenue_record:
        selected_date = revenue_record.date
        try:
            db.delete(revenue_record)
            db.commit()
            
            # Tự động cập nhật bản ghi thu nhập trong finance-report
            await create_daily_revenue_finance_record(selected_date, db)
            
            return RedirectResponse(url=f"/revenue?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)
        except Exception as e:
            print(f"Error deleting revenue record: {e}")
            db.rollback()
            return RedirectResponse(url="/revenue", status_code=303)
    
    return RedirectResponse(url="/revenue", status_code=303)

@app.post("/revenue/delete-all")
async def delete_all_revenue(request: Request, db: Session = Depends(get_db)):
    """Xóa tất cả doanh thu trong ngày"""
    form_data = await request.form()
    selected_date_str = form_data.get("date")
    
    if not selected_date_str:
        return RedirectResponse(url="/revenue", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        return RedirectResponse(url="/revenue", status_code=303)
    
    try:
        # Xóa tất cả revenue records trong ngày
        deleted_count = db.query(RevenueRecord).filter(RevenueRecord.date == selected_date).delete()
        db.commit()
        print(f"Deleted {deleted_count} revenue records for date {selected_date}")
        
        # Tự động cập nhật bản ghi thu nhập trong finance-report
        await create_daily_revenue_finance_record(selected_date, db)
        
        return RedirectResponse(url=f"/revenue?selected_date={selected_date.strftime('%Y-%m-%d')}&deleted_all=true", status_code=303)
    except Exception as e:
        print(f"Error deleting all revenue records: {e}")
        db.rollback()
        return RedirectResponse(url=f"/revenue?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)

@app.get("/daily", response_class=HTMLResponse)
async def daily_page(request: Request, db: Session = Depends(get_db), selected_date: Optional[str] = None):
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    today = date.today()
    
    # Xử lý ngày được chọn
    print(f"DEBUG: selected_date parameter = {selected_date}")
    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            print(f"DEBUG: Parsed filter_date = {filter_date}")
        except ValueError:
            print(f"DEBUG: Invalid date format, using today")
            filter_date = today
    else:
        print(f"DEBUG: No selected_date, using today")
        filter_date = today
    
    # Lọc chuyến đã ghi nhận theo ngày được chọn
    daily_routes = db.query(DailyRoute).filter(DailyRoute.date == filter_date).order_by(DailyRoute.created_at.desc()).all()
    
    # Debug: Print to console
    print(f"DEBUG: Routes count: {len(routes)}")
    print(f"DEBUG: Employees count: {len(employees)}")
    print(f"DEBUG: Vehicles count: {len(vehicles)}")
    print(f"DEBUG: Filter date: {filter_date}")
    print(f"DEBUG: Daily routes count: {len(daily_routes)}")
    if vehicles:
        for v in vehicles:
            print(f"DEBUG: Vehicle: {v.license_plate} (ID: {v.id}, Status: {v.status})")
    else:
        print("DEBUG: No vehicles found!")
        # Check all vehicles regardless of status
        all_vehicles = db.query(Vehicle).all()
        print(f"DEBUG: Total vehicles in DB: {len(all_vehicles)}")
        for v in all_vehicles:
            print(f"DEBUG: All Vehicle: {v.license_plate} (ID: {v.id}, Status: {v.status})")
    
    return templates.TemplateResponse("daily.html", {
        "request": request,
        "routes": routes,
        "employees": employees,
        "vehicles": vehicles,
        "daily_routes": daily_routes,
        "today": today,
        "selected_date": filter_date.strftime('%Y-%m-%d'),
        "filter_date": filter_date
    })

@app.post("/daily/add")
async def add_daily_route(request: Request, db: Session = Depends(get_db)):
    form_data = await request.form()
    
    # Lấy ngày được chọn từ form
    selected_date_str = form_data.get("date")
    if not selected_date_str:
        return RedirectResponse(url="/daily", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = date.today()
    
    # Lấy tất cả routes
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Xử lý từng route
    for route in routes:
        route_id = route.id
        
        # Lấy dữ liệu từ form cho route này
        distance_km = form_data.get(f"distance_km_{route_id}")
        driver_name = form_data.get(f"driver_name_{route_id}")
        license_plate = form_data.get(f"license_plate_{route_id}")
        notes = form_data.get(f"notes_{route_id}")
        
        # Chỉ tạo record nếu có ít nhất một trường được điền
        if distance_km or driver_name or license_plate or notes:
            daily_route = DailyRoute(
                route_id=route_id,
                date=selected_date,
                distance_km=float(distance_km) if distance_km else 0,
                cargo_weight=0,  # Set default value
                driver_name=driver_name or "",
                license_plate=license_plate or "",
                employee_name="",  # Empty since we removed this field
                notes=notes or ""
            )
            db.add(daily_route)
    
    db.commit()
    # Redirect về trang daily với ngày đã chọn
    return RedirectResponse(url=f"/daily?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)

@app.post("/daily/delete/{daily_route_id}")
async def delete_daily_route(daily_route_id: int, request: Request, db: Session = Depends(get_db)):
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if daily_route:
        # Lưu ngày của chuyến bị xóa để redirect về đúng ngày
        deleted_date = daily_route.date
        db.delete(daily_route)
        db.commit()
        return RedirectResponse(url=f"/daily?selected_date={deleted_date.strftime('%Y-%m-%d')}", status_code=303)
    return RedirectResponse(url="/daily", status_code=303)

# New Daily Page with simple date selection
@app.get("/daily-new", response_class=HTMLResponse)
async def daily_new_page(
    request: Request, 
    db: Session = Depends(get_db), 
    selected_date: Optional[str] = None, 
    deleted_all: Optional[str] = None,
    mode: Optional[str] = None,
    selected_month: Optional[str] = None,
    selected_route_id: Optional[int] = None,
    current_user = Depends(get_current_user)
):
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Kiểm tra quyền truy cập (User hoặc Admin)
    redirect_response = check_and_redirect_access(current_user["role"], "/daily-new", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    today = date.today()
    
    # Xử lý mode: by-date hoặc by-route
    if mode == "by-route":
        # Chế độ chấm công theo tuyến
        # Xử lý tháng được chọn
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                filter_month_start = date(year, month, 1)
                # Lấy ngày cuối cùng của tháng
                if month == 12:
                    filter_month_end = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    filter_month_end = date(year, month + 1, 1) - timedelta(days=1)
            except ValueError:
                filter_month_start = date(today.year, today.month, 1)
                if today.month == 12:
                    filter_month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
                else:
                    filter_month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
                selected_month = f"{today.year}-{today.month:02d}"
        else:
            filter_month_start = date(today.year, today.month, 1)
            if today.month == 12:
                filter_month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                filter_month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
            selected_month = f"{today.year}-{today.month:02d}"
        
        # Lấy tất cả tuyến (không lọc)
        all_routes = sorted(routes, key=lambda r: r.route_code.lower() if r.route_code else "")
        
        # Lấy tất cả chuyến trong tháng (lọc theo tuyến nếu có)
        monthly_daily_routes_query = db.query(DailyRoute).filter(
            DailyRoute.date >= filter_month_start,
            DailyRoute.date <= filter_month_end
        )
        
        if selected_route_id:
            monthly_daily_routes_query = monthly_daily_routes_query.filter(DailyRoute.route_id == selected_route_id)
        
        monthly_daily_routes = monthly_daily_routes_query.all()
        
        # Sắp xếp monthly_daily_routes: Mã tuyến A-Z, tuyến "Tăng Cường" luôn ở cuối
        def sort_monthly_daily_routes_by_route_code(monthly_daily_routes):
            # Tách ra các chuyến không phải "Tăng Cường" và các chuyến "Tăng Cường"
            normal_daily_routes = []
            tang_cuong_daily_routes = []
            
            for dr in monthly_daily_routes:
                if dr.route and dr.route.route_code and dr.route.route_code.strip() == "Tăng Cường":
                    tang_cuong_daily_routes.append(dr)
                else:
                    normal_daily_routes.append(dr)
            
            # Sắp xếp các chuyến bình thường: theo mã tuyến A-Z, sau đó theo ngày
            normal_daily_routes_sorted = sorted(
                normal_daily_routes, 
                key=lambda dr: (
                    dr.route.route_code.lower() if dr.route and dr.route.route_code else "",
                    dr.date,
                    dr.created_at
                )
            )
            
            # Sắp xếp các chuyến "Tăng Cường": theo ngày
            tang_cuong_daily_routes_sorted = sorted(
                tang_cuong_daily_routes,
                key=lambda dr: (dr.date, dr.created_at)
            )
            
            # Ghép lại: chuyến bình thường (A-Z) + chuyến "Tăng Cường"
            return normal_daily_routes_sorted + tang_cuong_daily_routes_sorted
        
        monthly_daily_routes = sort_monthly_daily_routes_by_route_code(monthly_daily_routes)
        
        # Format tháng để hiển thị
        selected_month_display = datetime.strptime(selected_month, "%Y-%m").strftime("%m/%Y")
        current_month = f"{today.year}-{today.month:02d}"
        current_month_display = today.strftime("%m/%Y")
        
        # Lấy mã tuyến được chọn (chỉ hiển thị mã tuyến, không hiển thị tên)
        selected_route_name = None
        if selected_route_id:
            selected_route = db.query(Route).filter(Route.id == selected_route_id).first()
            if selected_route:
                selected_route_name = selected_route.route_code or ""
        
        # Chuẩn bị dữ liệu JSON cho JavaScript
        import json
        routes_json = json.dumps([{"id": r.id, "route_code": r.route_code or "", "route_name": r.route_name or ""} for r in all_routes])
        employees_json = json.dumps([{"name": e.name or ""} for e in sorted(employees, key=lambda emp: emp.name.lower() if emp.name else "")])
        vehicles_json = json.dumps([{"license_plate": v.license_plate or ""} for v in vehicles])
        monthly_daily_routes_json = json.dumps([
            {
                "id": dr.id,
                "route_id": dr.route_id,
                "date": dr.date.strftime("%Y-%m-%d"),
                "distance_km": dr.distance_km or 0,
                "driver_name": dr.driver_name or "",
                "license_plate": dr.license_plate or "",
                "status": dr.status or "Online",
                "notes": dr.notes or ""
            } for dr in monthly_daily_routes
        ])
        
        return templates.TemplateResponse("daily_new.html", {
            "request": request,
            "current_user": current_user,
            "routes": [],  # Không dùng cho mode by-route
            "all_routes": all_routes,  # Tất cả tuyến cho mode by-route
            "employees": employees,
            "vehicles": vehicles,
            "daily_routes": [],  # Không dùng cho mode by-route
            "monthly_daily_routes": monthly_daily_routes,  # Chuyến trong tháng
            "selected_date": today.strftime('%Y-%m-%d'),
            "selected_date_display": today.strftime('%d/%m/%Y'),
            "selected_month": selected_month,
            "selected_month_display": selected_month_display,
            "current_month": current_month,
            "current_month_display": current_month_display,
            "selected_route_id": selected_route_id,
            "selected_route_name": selected_route_name,
            "deleted_all": deleted_all,
            "previous_assignments": {},
            "routes_json": routes_json,
            "employees_json": employees_json,
            "vehicles_json": vehicles_json,
            "monthly_daily_routes_json": monthly_daily_routes_json
        })
    else:
        # Chế độ chấm công theo ngày (mặc định)
        # Xử lý ngày được chọn
        if selected_date:
            try:
                filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            except ValueError:
                filter_date = today
        else:
            filter_date = today
        
        # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
        def sort_routes_with_tang_cuong_at_bottom(routes):
            # Lọc ra routes không phải "Tăng Cường"
            normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
            
            # Lọc ra routes "Tăng Cường"
            tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
            
            # Sắp xếp routes bình thường theo A-Z
            normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower())
            
            # Ghép lại: routes bình thường + routes "Tăng Cường"
            return normal_routes_sorted + tang_cuong_routes
        
        routes = sort_routes_with_tang_cuong_at_bottom(routes)
        
        # Sắp xếp employees theo tên (A-Z) để dễ tìm kiếm trong dropdown
        employees = sorted(employees, key=lambda emp: emp.name.lower() if emp.name else "")
        
        # Lọc chuyến đã ghi nhận theo ngày được chọn
        daily_routes = db.query(DailyRoute).filter(DailyRoute.date == filter_date).all()
        
        # Sắp xếp daily_routes: Mã tuyến A-Z, tuyến "Tăng Cường" luôn ở cuối
        def sort_daily_routes_by_route_code(daily_routes):
            # Tách ra các chuyến không phải "Tăng Cường" và các chuyến "Tăng Cường"
            normal_daily_routes = []
            tang_cuong_daily_routes = []
            
            for dr in daily_routes:
                if dr.route and dr.route.route_code and dr.route.route_code.strip() == "Tăng Cường":
                    tang_cuong_daily_routes.append(dr)
                else:
                    normal_daily_routes.append(dr)
            
            # Sắp xếp các chuyến bình thường theo mã tuyến A-Z
            normal_daily_routes_sorted = sorted(
                normal_daily_routes, 
                key=lambda dr: (dr.route.route_code.lower() if dr.route and dr.route.route_code else "", dr.created_at)
            )
            
            # Sắp xếp các chuyến "Tăng Cường" theo thời gian tạo
            tang_cuong_daily_routes_sorted = sorted(
                tang_cuong_daily_routes,
                key=lambda dr: dr.created_at
            )
            
            # Ghép lại: chuyến bình thường (A-Z) + chuyến "Tăng Cường"
            return normal_daily_routes_sorted + tang_cuong_daily_routes_sorted
        
        daily_routes = sort_daily_routes_by_route_code(daily_routes)
        
        # Lấy danh sách route_id đã được chấm công trong ngày này
        completed_route_ids = {daily_route.route_id for daily_route in daily_routes}
        
        # Lọc ra các tuyến chưa được chấm công (ẩn các tuyến đã chấm công)
        # Ngoại trừ tuyến "Tăng Cường" - luôn hiển thị để có thể thêm nhiều chuyến
        available_routes = []
        for route in routes:
            # Tuyến "Tăng Cường" luôn hiển thị
            if route.route_code and route.route_code.strip() == "Tăng Cường":
                available_routes.append(route)
            # Các tuyến khác chỉ hiển thị nếu chưa được chấm công
            elif route.id not in completed_route_ids:
                available_routes.append(route)
        
        # Lấy dữ liệu chấm công trước đó để tự động điền
        previous_assignments = {}
        for route in available_routes:
            # Tìm chuyến gần nhất của tuyến này (trước ngày hiện tại)
            previous_route = db.query(DailyRoute).filter(
                DailyRoute.route_id == route.id,
                DailyRoute.date < filter_date,
                DailyRoute.driver_name.isnot(None),
                DailyRoute.driver_name != "",
                DailyRoute.license_plate.isnot(None),
                DailyRoute.license_plate != ""
            ).order_by(DailyRoute.date.desc()).first()
            
            if previous_route:
                previous_assignments[route.id] = {
                    'driver_name': previous_route.driver_name,
                    'license_plate': previous_route.license_plate
                }
        
        return templates.TemplateResponse("daily_new.html", {
            "request": request,
            "current_user": current_user,
            "routes": available_routes,  # Chỉ hiển thị tuyến chưa chấm công
            "all_routes": [],  # Không dùng cho mode by-date
            "employees": employees,
            "vehicles": vehicles,
            "daily_routes": daily_routes,
            "monthly_daily_routes": [],  # Không dùng cho mode by-date
            "selected_date": filter_date.strftime('%Y-%m-%d'),
            "selected_date_display": filter_date.strftime('%d/%m/%Y'),
            "selected_month": None,
            "selected_month_display": None,
            "current_month": None,
            "current_month_display": None,
            "deleted_all": deleted_all,
            "previous_assignments": previous_assignments,  # Dữ liệu để tự động điền
            "routes_json": "[]",
            "employees_json": "[]",
            "vehicles_json": "[]",
            "monthly_daily_routes_json": "[]"
        })

@app.post("/daily-new/add")
async def add_daily_new_route(request: Request, db: Session = Depends(get_db)):
    form_data = await request.form()
    
    # Lấy ngày được chọn từ form
    selected_date_str = form_data.get("date")
    if not selected_date_str:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = date.today()
    
    # Lấy tất cả routes và sắp xếp theo mã tuyến (A-Z)
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        # Lọc ra routes không phải "Tăng Cường"
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        
        # Lọc ra routes "Tăng Cường"
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        
        # Sắp xếp routes bình thường theo A-Z
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower())
        
        # Ghép lại: routes bình thường + routes "Tăng Cường"
        return normal_routes_sorted + tang_cuong_routes
    
    routes = sort_routes_with_tang_cuong_at_bottom(routes)
    
    # Xử lý từng route
    for route in routes:
        route_id = route.id
        
        # Lấy dữ liệu từ form cho route này
        distance_km = form_data.get(f"distance_km_{route_id}")
        driver_name = form_data.get(f"driver_name_{route_id}")
        license_plate = form_data.get(f"license_plate_{route_id}")
        status = form_data.get(f"status_{route_id}")
        notes = form_data.get(f"notes_{route_id}")
        
        # Chỉ tạo record nếu có ít nhất một trường được điền
        if distance_km or driver_name or license_plate or notes:
            daily_route = DailyRoute(
                route_id=route_id,
                date=selected_date,
                distance_km=float(distance_km) if distance_km else 0,
                cargo_weight=0,  # Set default value
                driver_name=driver_name or "",
                license_plate=license_plate or "",
                employee_name="",  # Empty since we removed this field
                status=status or "Online",  # Mặc định là Online
                notes=notes or ""
            )
            db.add(daily_route)
    
    db.commit()
    # Redirect về trang daily-new với ngày đã chọn
    return RedirectResponse(url=f"/daily-new?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)

@app.get("/daily-new/edit/{daily_route_id}", response_class=HTMLResponse)
async def edit_daily_new_route_page(request: Request, daily_route_id: int, db: Session = Depends(get_db)):
    """Trang sửa chuyến"""
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if not daily_route:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    # Lấy danh sách để hiển thị trong dropdown
    employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    # Sắp xếp employees theo tên (A-Z) để dễ tìm kiếm trong dropdown
    employees = sorted(employees, key=lambda emp: emp.name.lower() if emp.name else "")
    
    return templates.TemplateResponse("edit_daily_route.html", {
        "request": request,
        "daily_route": daily_route,
        "employees": employees,
        "vehicles": vehicles
    })

@app.post("/daily-new/edit/{daily_route_id}")
async def edit_daily_new_route(
    daily_route_id: int,
    distance_km: float = Form(0),
    driver_name: str = Form(""),
    license_plate: str = Form(""),
    status: str = Form("Online"),
    notes: str = Form(""),
    db: Session = Depends(get_db)
):
    """Cập nhật chuyến"""
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if not daily_route:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    # Cập nhật thông tin
    daily_route.distance_km = distance_km
    daily_route.driver_name = driver_name
    daily_route.license_plate = license_plate
    daily_route.status = status
    daily_route.notes = notes
    
    db.commit()
    
    # Redirect về trang daily-new với ngày của chuyến
    return RedirectResponse(url=f"/daily-new?selected_date={daily_route.date.strftime('%Y-%m-%d')}", status_code=303)

@app.post("/daily-new/delete/{daily_route_id}")
async def delete_daily_new_route(daily_route_id: int, db: Session = Depends(get_db)):
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if daily_route:
        # Lưu ngày của chuyến bị xóa để redirect về đúng ngày
        deleted_date = daily_route.date
        db.delete(daily_route)
        db.commit()
        return RedirectResponse(url=f"/daily-new?selected_date={deleted_date.strftime('%Y-%m-%d')}", status_code=303)
    return RedirectResponse(url="/daily-new", status_code=303)

@app.post("/daily-new/delete-all")
async def delete_all_daily_routes(request: Request, db: Session = Depends(get_db)):
    """Xóa tất cả chuyến đã ghi nhận trong một ngày"""
    form_data = await request.form()
    selected_date_str = form_data.get("date")
    
    if not selected_date_str:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    # Tìm và xóa tất cả chuyến trong ngày được chọn
    daily_routes = db.query(DailyRoute).filter(DailyRoute.date == selected_date).all()
    
    if daily_routes:
        for daily_route in daily_routes:
            db.delete(daily_route)
        db.commit()
    
    # Redirect về trang daily-new với ngày đã chọn và thông báo thành công
    return RedirectResponse(url=f"/daily-new?selected_date={selected_date.strftime('%Y-%m-%d')}&deleted_all=true", status_code=303)

@app.post("/daily-new/add-by-route")
async def add_daily_new_route_by_route(request: Request, db: Session = Depends(get_db)):
    """Lưu chấm công theo tuyến (theo tháng)"""
    form_data = await request.form()
    
    # Lấy tháng và tuyến được chọn từ form
    selected_month_str = form_data.get("selected_month")
    selected_route_id_str = form_data.get("selected_route_id")
    
    if not selected_month_str:
        return RedirectResponse(url="/daily-new?mode=by-route", status_code=303)
    
    try:
        year, month = map(int, selected_month_str.split('-'))
    except ValueError:
        return RedirectResponse(url="/daily-new?mode=by-route", status_code=303)
    
    # Lấy tất cả các trường từ form
    # Form có format: route_id_1, date_1, distance_km_1, driver_name_1, license_plate_1, status_1, notes_1
    # Tìm tất cả các key bắt đầu bằng route_id_ hoặc date_
    date_keys = [key for key in form_data.keys() if key.startswith("date_")]
    
    for date_key in date_keys:
        # Lấy index từ key (ví dụ: date_1 -> 1)
        index = date_key.split("_")[-1]
        
        # Lấy các giá trị tương ứng
        route_id = form_data.get(f"route_id_{index}")
        date_str = form_data.get(f"date_{index}")
        distance_km = form_data.get(f"distance_km_{index}")
        driver_name = form_data.get(f"driver_name_{index}")
        license_plate = form_data.get(f"license_plate_{index}")
        status = form_data.get(f"status_{index}")
        notes = form_data.get(f"notes_{index}")
        
        # Bỏ qua nếu không có route_id hoặc date
        if not route_id or not date_str:
            continue
        
        try:
            route_id_int = int(route_id)
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        
        # QUAN TRỌNG: Kiểm tra xem đã có record cho route_id và date này chưa (tránh trùng lặp)
        existing_record = db.query(DailyRoute).filter(
            DailyRoute.route_id == route_id_int,
            DailyRoute.date == selected_date
        ).first()
        
        # Chỉ tạo/cập nhật record nếu có ít nhất một trường được điền
        if distance_km or driver_name or license_plate or notes or status:
            if existing_record:
                # Cập nhật record hiện có
                existing_record.distance_km = float(distance_km) if distance_km else 0
                existing_record.driver_name = driver_name or ""
                existing_record.license_plate = license_plate or ""
                existing_record.status = status or "Online"
                existing_record.notes = notes or ""
            else:
                # Tạo record mới (đã kiểm tra không trùng ở trên)
                daily_route = DailyRoute(
                    route_id=route_id_int,
                    date=selected_date,
                    distance_km=float(distance_km) if distance_km else 0,
                    cargo_weight=0,
                    driver_name=driver_name or "",
                    license_plate=license_plate or "",
                    employee_name="",
                    status=status or "Online",
                    notes=notes or ""
                )
                db.add(daily_route)
        elif existing_record:
            # Nếu không có dữ liệu nào được điền và có record cũ, xóa record
            db.delete(existing_record)
    
    db.commit()
    
    # Redirect về trang daily-new với mode by-route, tháng và tuyến đã chọn
    redirect_url = f"/daily-new?mode=by-route&selected_month={selected_month_str}"
    if selected_route_id_str:
        redirect_url += f"&selected_route_id={selected_route_id_str}"
    return RedirectResponse(url=redirect_url, status_code=303)

@app.get("/salary/driver-details/{driver_name}")
async def get_driver_details(
    driver_name: str,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Lấy chi tiết chuyến của một lái xe cụ thể"""
    # Xử lý khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = db.query(DailyRoute).filter(
                DailyRoute.driver_name == driver_name,
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
        except ValueError:
            return {"error": "Invalid date format"}
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        daily_routes_query = db.query(DailyRoute).filter(
            DailyRoute.driver_name == driver_name,
            DailyRoute.date >= date(today.year, today.month, 1),
            DailyRoute.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
    
    # Lấy dữ liệu và join với Route để có thông tin tuyến
    daily_routes = daily_routes_query.join(Route).order_by(DailyRoute.date.desc()).all()
    
    # Format dữ liệu
    trip_details = []
    for trip in daily_routes:
        trip_details.append({
            'date': trip.date.strftime('%d/%m/%Y'),
            'route_code': trip.route.route_code,
            'route_name': trip.route.route_name,
            'license_plate': trip.license_plate,
            'distance_km': trip.distance_km,
            'cargo_weight': trip.cargo_weight,
            'notes': trip.notes or ''
        })
    
    return {"trip_details": trip_details}

@app.get("/salary/driver-details-page/{driver_name}", response_class=HTMLResponse)
async def driver_details_page(
    request: Request,
    driver_name: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Trang hiển thị chi tiết chuyến của một lái xe cụ thể"""
    # Nếu chưa đăng nhập, redirect về trang login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    # Xử lý khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = db.query(DailyRoute).filter(
                DailyRoute.driver_name == driver_name,
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
            period_text = f"từ {from_date_obj.strftime('%d/%m/%Y')} đến {to_date_obj.strftime('%d/%m/%Y')}"
        except ValueError:
            return RedirectResponse(url="/salary", status_code=303)
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        daily_routes_query = db.query(DailyRoute).filter(
            DailyRoute.driver_name == driver_name,
            DailyRoute.date >= date(today.year, today.month, 1),
            DailyRoute.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
        period_text = f"tháng {today.month}/{today.year}"
    
    # Lấy dữ liệu và join với Route để có thông tin tuyến
    daily_routes = daily_routes_query.join(Route).order_by(DailyRoute.date.desc()).all()
    
    # Tính thống kê
    total_trips = len(daily_routes)
    total_distance = sum(trip.distance_km for trip in daily_routes)
    total_cargo = sum(trip.cargo_weight for trip in daily_routes)
    routes_used = list(set(trip.route.route_code for trip in daily_routes))
    
    return templates.TemplateResponse("driver_details.html", {
        "request": request,
        "current_user": current_user,
        "driver_name": driver_name,
        "period_text": period_text,
        "daily_routes": daily_routes,
        "total_trips": total_trips,
        "total_distance": total_distance,
        "total_cargo": total_cargo,
        "routes_used": routes_used,
        "from_date": from_date,
        "to_date": to_date
    })



@app.get("/salary-simple", response_class=HTMLResponse)
async def salary_simple_page(request: Request):
    """Redirect đến trang báo cáo tổng hợp"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/report", status_code=302)

@app.get("/general-report", response_class=HTMLResponse)
async def general_report_page(
    request: Request, 
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    route_code: Optional[str] = None
):
    """Trang thống kê tổng hợp - báo cáo chi tiết hoạt động vận chuyển"""
    # Nếu chưa đăng nhập, redirect về trang login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Khởi tạo query cơ bản
    daily_routes_query = db.query(DailyRoute)
    
    # Áp dụng bộ lọc thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = daily_routes_query.filter(
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
        except ValueError:
            pass
    
    # Áp dụng các bộ lọc khác
    if driver_name:
        daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name.ilike(f"%{driver_name}%"))
    if license_plate:
        daily_routes_query = daily_routes_query.filter(DailyRoute.license_plate.ilike(f"%{license_plate}%"))
    if route_code:
        daily_routes_query = daily_routes_query.join(Route).filter(Route.route_code.ilike(f"%{route_code}%"))
    
    daily_routes = daily_routes_query.all()
    
    # Tính thống kê theo lái xe
    driver_stats = {}
    for daily_route in daily_routes:
        driver_name_key = daily_route.driver_name
        license_plate_key = daily_route.license_plate
        if driver_name_key and driver_name_key not in driver_stats:
            driver_stats[driver_name_key] = {
                'driver_name': driver_name_key,
                'license_plate': license_plate_key or 'N/A',
                'trip_count': 0,
                'total_distance': 0,
                'total_cargo': 0,
                'routes': set()
            }
        
        if driver_name_key:
            driver_stats[driver_name_key]['trip_count'] += 1
            driver_stats[driver_name_key]['total_distance'] += daily_route.distance_km
            driver_stats[driver_name_key]['total_cargo'] += daily_route.cargo_weight
            driver_stats[driver_name_key]['routes'].add(daily_route.route.route_code)
            if license_plate_key:
                driver_stats[driver_name_key]['license_plate'] = license_plate_key
    
    # Convert to list
    salary_data = []
    for driver_name_key, stats in driver_stats.items():
        salary_data.append({
            'driver_name': driver_name_key,
            'license_plate': stats['license_plate'],
            'trip_count': stats['trip_count'],
            'total_distance': stats['total_distance'],
            'total_cargo': stats['total_cargo'],
            'routes': list(stats['routes'])
        })
    
    salary_data.sort(key=lambda x: x['trip_count'], reverse=True)
    
    # Tạo dữ liệu chi tiết từng chuyến
    trip_details = []
    for daily_route in daily_routes:
        if daily_route.driver_name:
            trip_details.append({
                'driver_name': daily_route.driver_name,
                'license_plate': daily_route.license_plate or 'N/A',
                'date': daily_route.date,
                'route_code': daily_route.route.route_code,
                'route_name': daily_route.route.route_name,
                'distance_km': daily_route.distance_km,
                'cargo_weight': daily_route.cargo_weight,
                'notes': daily_route.notes or ''
            })
    
    trip_details.sort(key=lambda x: (x['driver_name'], x['date']))
    
    # Lấy danh sách cho dropdown
    routes = db.query(Route).all()
    employees = db.query(Employee).all()
    vehicles = db.query(Vehicle).all()
    
    # Template data - CHỈ TRUYỀN KHI CÓ GIÁ TRỊ
    template_data = {
        "request": request,
        "current_user": current_user,
        "salary_data": salary_data,
        "trip_details": trip_details,
        "employees": employees,
        "vehicles": vehicles,
        "routes": routes,
        "total_routes": len(daily_routes),
        "total_distance": sum(dr.distance_km for dr in daily_routes),
        "total_cargo": sum(dr.cargo_weight for dr in daily_routes)
    }
    
    # Chỉ thêm khi có giá trị
    if from_date:
        template_data["from_date"] = from_date
    if to_date:
        template_data["to_date"] = to_date
    if driver_name:
        template_data["driver_name"] = driver_name
    if license_plate:
        template_data["license_plate"] = license_plate
    if route_code:
        template_data["route_code"] = route_code
    
    return templates.TemplateResponse("salary_simple.html", template_data)

@app.get("/salary-simple/export-excel")
async def export_salary_simple_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    route_code: Optional[str] = None
):
    """Redirect đến general-report export"""
    from fastapi.responses import RedirectResponse
    params = []
    if from_date:
        params.append(f"from_date={from_date}")
    if to_date:
        params.append(f"to_date={to_date}")
    if driver_name:
        params.append(f"driver_name={driver_name}")
    if license_plate:
        params.append(f"license_plate={license_plate}")
    if route_code:
        params.append(f"route_code={route_code}")
    
    url = "/general-report/export-excel"
    if params:
        url += "?" + "&".join(params)
    
    return RedirectResponse(url=url, status_code=302)

@app.get("/general-report/export-excel")
async def export_general_report_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    route_code: Optional[str] = None
):
    """Xuất Excel danh sách chi tiết từng chuyến cho general-report"""
    # Sử dụng lại logic lọc từ salary_simple_page
    daily_routes_query = db.query(DailyRoute)
    
    # Áp dụng bộ lọc thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = daily_routes_query.filter(
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
        except ValueError:
            pass
    
    # Áp dụng các bộ lọc khác
    if driver_name:
        daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name.ilike(f"%{driver_name}%"))
    if license_plate:
        daily_routes_query = daily_routes_query.filter(DailyRoute.license_plate.ilike(f"%{license_plate}%"))
    if route_code:
        daily_routes_query = daily_routes_query.join(Route).filter(Route.route_code.ilike(f"%{route_code}%"))
    
    daily_routes = daily_routes_query.all()
    
    # Tạo dữ liệu chi tiết từng chuyến
    trip_details = []
    for daily_route in daily_routes:
        if daily_route.driver_name:
            trip_details.append({
                'stt': len(trip_details) + 1,
                'ngay_chay': daily_route.date.strftime('%d/%m/%Y'),
                'ten_lai_xe': daily_route.driver_name,
                'bien_so_xe': daily_route.license_plate or 'N/A',
                'ma_tuyen': daily_route.route.route_code,
                'ten_tuyen': daily_route.route.route_name,
                'km': daily_route.distance_km,
                'tai_trong': daily_route.cargo_weight,
                'ghi_chu': daily_route.notes or ''
            })
    
    # Tạo CSV content với UTF-8 BOM để Excel hiển thị đúng tiếng Việt
    csv_content = "\ufeff"  # UTF-8 BOM
    csv_content += "STT,Ngày chạy,Tên lái xe,Biển số xe,Mã tuyến,Tên tuyến,Km,Tải trọng,Ghi chú\n"
    
    for trip in trip_details:
        # Escape các ký tự đặc biệt trong CSV
        def escape_csv_field(field):
            if field is None:
                return ""
            field_str = str(field)
            # Nếu chứa dấu phẩy, dấu ngoặc kép hoặc xuống dòng thì bọc trong dấu ngoặc kép
            if ',' in field_str or '"' in field_str or '\n' in field_str:
                field_str = field_str.replace('"', '""')  # Escape dấu ngoặc kép
                field_str = f'"{field_str}"'
            return field_str
        
        csv_content += f"{trip['stt']},{escape_csv_field(trip['ngay_chay'])},{escape_csv_field(trip['ten_lai_xe'])},{escape_csv_field(trip['bien_so_xe'])},{escape_csv_field(trip['ma_tuyen'])},{escape_csv_field(trip['ten_tuyen'])},{trip['km']},{trip['tai_trong']},{escape_csv_field(trip['ghi_chu'])}\n"
    
    # Tạo tên file
    if from_date and to_date:
        filename = f"chi_tiet_chuyen_{from_date}_den_{to_date}.csv"
    else:
        today = date.today()
        filename = f"chi_tiet_chuyen_{today.month}_{today.year}.csv"
    
    # Trả về file CSV với encoding UTF-8
    return Response(
        content=csv_content.encode('utf-8-sig'),  # UTF-8 with BOM
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "Content-Type": "text/csv; charset=utf-8"
        }
    )



# ===== FUEL MANAGEMENT ROUTES =====

@app.get("/fuel", response_class=HTMLResponse)
async def fuel_page(request: Request):
    """Redirect đến trang báo cáo tổng hợp"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/report", status_code=302)

@app.get("/fuel-report", response_class=HTMLResponse)
async def fuel_report_page(
    request: Request, 
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """Trang tổng hợp đổ dầu - báo cáo chi tiết"""
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    # Xử lý khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            fuel_records_query = db.query(FuelRecord).filter(
                FuelRecord.date >= from_date_obj,
                FuelRecord.date <= to_date_obj
            )
        except ValueError:
            fuel_records_query = db.query(FuelRecord)
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        fuel_records_query = db.query(FuelRecord).filter(
            FuelRecord.date >= date(today.year, today.month, 1),
            FuelRecord.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
    
    fuel_records = fuel_records_query.order_by(FuelRecord.date.desc(), FuelRecord.license_plate).all()
    
    # Tính tổng số lít dầu đã đổ
    total_liters_pumped = sum(record.liters_pumped for record in fuel_records)
    
    # Lấy danh sách xe để hiển thị trong dropdown
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    # Tạo template data
    template_data = {
        "request": request,
        "current_user": current_user,
        "fuel_records": fuel_records,
        "vehicles": vehicles,
        "total_liters_pumped": total_liters_pumped,
        "total_records": len(fuel_records)
    }
    
    if from_date:
        template_data["from_date"] = from_date
    if to_date:
        template_data["to_date"] = to_date
    
    return templates.TemplateResponse("fuel.html", template_data)

@app.post("/fuel/add")
async def add_fuel_record(
    request: Request,
    db: Session = Depends(get_db)
):
    """Thêm bản ghi đổ dầu mới"""
    form_data = await request.form()
    
    # Lấy dữ liệu từ form
    date_str = form_data.get("date")
    fuel_type = form_data.get("fuel_type", "Dầu DO 0,05S-II")
    license_plate = form_data.get("license_plate")
    fuel_price_per_liter = float(form_data.get("fuel_price_per_liter", 0))
    liters_pumped = float(form_data.get("liters_pumped", 0))
    notes = form_data.get("notes", "")
    
    if not date_str or not license_plate:
        return RedirectResponse(url="/fuel-report", status_code=303)
    
    try:
        fuel_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        fuel_date = date.today()
    
    # Tính toán số tiền dầu đã đổ = Đơn giá dầu × Số lít dầu đã đổ (làm tròn đến đồng)
    cost_pumped = round(fuel_price_per_liter * liters_pumped)
    
    # Tạo bản ghi mới
    fuel_record = FuelRecord(
        date=fuel_date,
        fuel_type=fuel_type,
        license_plate=license_plate,
        fuel_price_per_liter=fuel_price_per_liter,
        liters_pumped=liters_pumped,
        cost_pumped=cost_pumped,
        notes=notes
    )
    
    db.add(fuel_record)
    db.commit()
    
    # Redirect với tham số thời gian nếu có
    redirect_url = "/fuel-report"
    from_date = form_data.get("from_date")
    to_date = form_data.get("to_date")
    if from_date and to_date:
        redirect_url += f"?from_date={from_date}&to_date={to_date}"
    
    return RedirectResponse(url=redirect_url, status_code=303)

@app.post("/fuel/delete/{fuel_record_id}")
async def delete_fuel_record(
    fuel_record_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Xóa bản ghi đổ dầu"""
    fuel_record = db.query(FuelRecord).filter(FuelRecord.id == fuel_record_id).first()
    if fuel_record:
        db.delete(fuel_record)
        db.commit()
    
    # Redirect về trang fuel
    return RedirectResponse(url="/fuel-report", status_code=303)

@app.get("/fuel/edit/{fuel_record_id}", response_class=HTMLResponse)
async def edit_fuel_record_page(
    request: Request,
    fuel_record_id: int,
    db: Session = Depends(get_db)
):
    """Trang sửa bản ghi đổ dầu"""
    fuel_record = db.query(FuelRecord).filter(FuelRecord.id == fuel_record_id).first()
    if not fuel_record:
        return RedirectResponse(url="/fuel-report", status_code=303)
    
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    return templates.TemplateResponse("edit_fuel.html", {
        "request": request,
        "fuel_record": fuel_record,
        "vehicles": vehicles
    })

@app.post("/fuel/edit/{fuel_record_id}")
async def edit_fuel_record(
    fuel_record_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Cập nhật bản ghi đổ dầu"""
    fuel_record = db.query(FuelRecord).filter(FuelRecord.id == fuel_record_id).first()
    if not fuel_record:
        return RedirectResponse(url="/fuel-report", status_code=303)
    
    form_data = await request.form()
    
    # Cập nhật dữ liệu
    date_str = form_data.get("date")
    if date_str:
        try:
            fuel_record.date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    fuel_record.fuel_type = form_data.get("fuel_type", "Dầu DO 0,05S-II")
    fuel_record.license_plate = form_data.get("license_plate")
    fuel_record.fuel_price_per_liter = float(form_data.get("fuel_price_per_liter", 0))
    fuel_record.liters_pumped = float(form_data.get("liters_pumped", 0))
    fuel_record.notes = form_data.get("notes", "")
    
    # Tính toán lại số tiền dầu đã đổ = Đơn giá dầu × Số lít dầu đã đổ (làm tròn đến đồng)
    fuel_record.cost_pumped = round(fuel_record.fuel_price_per_liter * fuel_record.liters_pumped)
    
    db.commit()
    return RedirectResponse(url="/fuel-report", status_code=303)

@app.get("/fuel/download-template")
async def download_fuel_template(db: Session = Depends(get_db)):
    """Tải mẫu Excel để import dữ liệu đổ dầu"""
    # Lấy danh sách xe để hiển thị trong mẫu
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    vehicle_list = [v.license_plate for v in vehicles]
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Mẫu Import Đổ Dầu"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề
    ws.merge_cells('A1:F1')
    ws['A1'] = "MẪU IMPORT DỮ LIỆU ĐỔ DẦU"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Hướng dẫn
    ws.merge_cells('A2:F2')
    ws['A2'] = "Vui lòng điền dữ liệu theo đúng định dạng bên dưới"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True)
    
    # Header bảng
    headers = [
        "STT", "Ngày đổ dầu (dd/mm/yyyy)", "Biển số xe", 
        "Số lượng dầu đổ (lít)", "Đơn giá (đồng/lít)", "Thành tiền (đồng)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu mẫu
    sample_data = [
        [1, "01/01/2025", "51A-12345", 50.000, 19020, 951000],
        [2, "02/01/2025", "51B-67890", 45.500, 19020, 865410],
        [3, "03/01/2025", "51C-11111", 60.000, 19020, 1141200]
    ]
    
    for row, data in enumerate(sample_data, 5):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Định dạng số
    for row in range(5, 8):
        # Số lượng dầu - 3 chữ số thập phân
        ws.cell(row=row, column=4).number_format = '#,##0.000'
        # Đơn giá - 2 chữ số thập phân
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        # Thành tiền - không có chữ số thập phân
        ws.cell(row=row, column=6).number_format = '#,##0'
    
    # Thêm sheet hướng dẫn
    ws2 = wb.create_sheet("Hướng dẫn")
    ws2['A1'] = "HƯỚNG DẪN SỬ DỤNG"
    ws2['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "1. Định dạng cột:",
        "   - STT: Số thứ tự (tự động)",
        "   - Ngày đổ dầu: Định dạng dd/mm/yyyy (ví dụ: 01/01/2025)",
        "   - Biển số xe: Phải khớp với danh sách xe trong hệ thống",
        "   - Số lượng dầu đổ: Cho phép 3 chữ số thập phân (ví dụ: 50.000)",
        "   - Đơn giá: Số chính xác (ví dụ: 19020)",
        "   - Thành tiền: Có thể để trống, hệ thống sẽ tự tính",
        "",
        "2. Danh sách biển số xe hợp lệ:",
    ]
    
    for i, instruction in enumerate(instructions, 2):
        ws2.cell(row=i, column=1, value=instruction)
    
    # Thêm danh sách xe
    for i, vehicle in enumerate(vehicle_list, len(instructions) + 2):
        ws2.cell(row=i, column=1, value=f"   - {vehicle}")
    
    # Điều chỉnh độ rộng cột
    column_widths = [8, 20, 15, 20, 20, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws2.column_dimensions['A'].width = 50
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    today = date.today()
    filename = f"Mau_Import_DoDau_{today.strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.post("/fuel/import-excel")
async def import_fuel_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Import dữ liệu đổ dầu từ file Excel"""
    try:
        # Kiểm tra định dạng file
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return JSONResponse(
                status_code=400,
                content={
                    "success": False, 
                    "error": "Định dạng file không hợp lệ",
                    "error_type": "file_format",
                    "details": "Chỉ chấp nhận file Excel (.xlsx hoặc .xls)",
                    "suggestion": "Vui lòng chọn file Excel có định dạng .xlsx hoặc .xls"
                }
            )
        
        # Đọc file Excel
        content = await file.read()
        if len(content) == 0:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "error": "File rỗng",
                    "error_type": "empty_file",
                    "details": "File Excel không chứa dữ liệu",
                    "suggestion": "Vui lòng kiểm tra lại file Excel có chứa dữ liệu"
                }
            )
        
        try:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
        except Exception as e:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "error": "Không thể đọc file Excel",
                    "error_type": "file_corrupted",
                    "details": f"Lỗi kỹ thuật: {str(e)}",
                    "suggestion": "Vui lòng kiểm tra file Excel không bị hỏng và có định dạng đúng"
                }
            )
        
        # Lấy danh sách xe hợp lệ
        vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
        valid_license_plates = {v.license_plate for v in vehicles}
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        # Bỏ qua header (dòng 1-4)
        for row_num in range(5, ws.max_row + 1):
            try:
                # Đọc dữ liệu từ Excel
                stt = ws.cell(row=row_num, column=1).value
                date_str = ws.cell(row=row_num, column=2).value
                license_plate = ws.cell(row=row_num, column=3).value
                liters_pumped = ws.cell(row=row_num, column=4).value
                fuel_price_per_liter = ws.cell(row=row_num, column=5).value
                cost_pumped = ws.cell(row=row_num, column=6).value
                
                # Bỏ qua dòng trống
                if not date_str or not license_plate:
                    continue
                
                # Validation dữ liệu với thông báo chi tiết
                validation_errors = []
                
                # Kiểm tra ngày (cột B)
                if isinstance(date_str, str):
                    try:
                        fuel_date = datetime.strptime(date_str, "%d/%m/%Y").date()
                    except ValueError:
                        validation_errors.append({
                            "column": "B (Ngày đổ)",
                            "error": "Định dạng ngày không đúng",
                            "value": str(date_str),
                            "suggestion": "Định dạng đúng: dd/mm/yyyy (ví dụ: 25/09/2025)"
                        })
                elif isinstance(date_str, datetime):
                    fuel_date = date_str.date()
                else:
                    validation_errors.append({
                        "column": "B (Ngày đổ)",
                        "error": "Ngày không hợp lệ",
                        "value": str(date_str),
                        "suggestion": "Vui lòng nhập ngày theo định dạng dd/mm/yyyy"
                    })
                
                # Kiểm tra biển số xe (cột C)
                if not license_plate:
                    validation_errors.append({
                        "column": "C (Biển số xe)",
                        "error": "Biển số xe không được để trống",
                        "value": "",
                        "suggestion": "Vui lòng nhập biển số xe"
                    })
                elif str(license_plate).strip() not in valid_license_plates:
                    validation_errors.append({
                        "column": "C (Biển số xe)",
                        "error": "Biển số xe không tồn tại trong hệ thống",
                        "value": str(license_plate),
                        "suggestion": f"Biển số xe hợp lệ: {', '.join(list(valid_license_plates)[:5])}{'...' if len(valid_license_plates) > 5 else ''}"
                    })
                
                # Kiểm tra số lít dầu (cột D)
                try:
                    liters_pumped = float(liters_pumped) if liters_pumped is not None else 0
                    if liters_pumped <= 0:
                        validation_errors.append({
                            "column": "D (Số lít đã đổ)",
                            "error": "Số lít dầu phải lớn hơn 0",
                            "value": str(liters_pumped),
                            "suggestion": "Vui lòng nhập số lít dầu lớn hơn 0 (ví dụ: 50.5)"
                        })
                except (ValueError, TypeError):
                    validation_errors.append({
                        "column": "D (Số lít đã đổ)",
                        "error": "Số lít dầu không hợp lệ",
                        "value": str(liters_pumped),
                        "suggestion": "Vui lòng nhập số lít dầu là số (ví dụ: 50.5, 100)"
                    })
                
                # Kiểm tra đơn giá (cột E)
                try:
                    fuel_price_per_liter = float(fuel_price_per_liter) if fuel_price_per_liter is not None else 0
                    if fuel_price_per_liter <= 0:
                        validation_errors.append({
                            "column": "E (Giá xăng dầu)",
                            "error": "Đơn giá phải lớn hơn 0",
                            "value": str(fuel_price_per_liter),
                            "suggestion": "Vui lòng nhập đơn giá lớn hơn 0 (ví dụ: 25000)"
                        })
                except (ValueError, TypeError):
                    validation_errors.append({
                        "column": "E (Giá xăng dầu)",
                        "error": "Đơn giá không hợp lệ",
                        "value": str(fuel_price_per_liter),
                        "suggestion": "Vui lòng nhập đơn giá là số (ví dụ: 25000, 25000.5)"
                    })
                
                # Tính thành tiền nếu không có
                if cost_pumped is None or cost_pumped == "":
                    cost_pumped = round(fuel_price_per_liter * liters_pumped)
                else:
                    try:
                        cost_pumped = float(cost_pumped)
                    except (ValueError, TypeError):
                        cost_pumped = round(fuel_price_per_liter * liters_pumped)
                
                # Nếu có lỗi validation, bỏ qua dòng này
                if validation_errors:
                    errors.append({
                        "row": row_num,
                        "errors": validation_errors
                    })
                    skipped_count += 1
                    continue
                
                # Kiểm tra trùng lặp (cùng ngày, cùng xe)
                existing_record = db.query(FuelRecord).filter(
                    FuelRecord.date == fuel_date,
                    FuelRecord.license_plate == str(license_plate).strip()
                ).first()
                
                if existing_record:
                    errors.append({
                        "row": row_num,
                        "errors": [{
                            "column": "Tổng hợp",
                            "error": "Bản ghi trùng lặp",
                            "value": f"Xe {license_plate} - Ngày {fuel_date.strftime('%d/%m/%Y')}",
                            "suggestion": "Đã tồn tại bản ghi đổ dầu cho xe này vào ngày này. Vui lòng kiểm tra lại dữ liệu."
                        }]
                    })
                    skipped_count += 1
                    continue
                
                # Tạo bản ghi mới
                fuel_record = FuelRecord(
                    date=fuel_date,
                    fuel_type="Dầu DO 0,05S-II",  # Mặc định
                    license_plate=str(license_plate).strip(),
                    fuel_price_per_liter=fuel_price_per_liter,
                    liters_pumped=liters_pumped,
                    cost_pumped=cost_pumped,
                    notes=f"Import từ Excel - dòng {row_num}"
                )
                
                db.add(fuel_record)
                imported_count += 1
                
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "errors": [{
                        "column": "Tổng hợp",
                        "error": "Lỗi xử lý dữ liệu",
                        "value": f"Lỗi kỹ thuật: {str(e)}",
                        "suggestion": "Vui lòng kiểm tra định dạng dữ liệu trong dòng này"
                    }]
                })
                skipped_count += 1
                continue
        
        # Commit tất cả thay đổi
        db.commit()
        
        # Tạo response chi tiết
        response_data = {
            "success": True,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "total_errors": len(errors),
            "summary": {
                "total_rows_processed": ws.max_row - 4,  # Trừ header
                "successful_imports": imported_count,
                "failed_imports": skipped_count,
                "success_rate": f"{(imported_count / max(1, ws.max_row - 4)) * 100:.1f}%" if ws.max_row > 4 else "0%"
            }
        }
        
        if errors:
            response_data["errors"] = errors[:20]  # Hiển thị 20 lỗi đầu tiên
            if len(errors) > 20:
                response_data["has_more_errors"] = True
                response_data["remaining_errors"] = len(errors) - 20
            response_data["error_summary"] = {
                "validation_errors": len([e for e in errors if any(err.get("column") != "Tổng hợp" for err in e.get("errors", []))]),
                "duplicate_errors": len([e for e in errors if any("trùng lặp" in err.get("error", "") for err in e.get("errors", []))]),
                "technical_errors": len([e for e in errors if any("Lỗi xử lý" in err.get("error", "") for err in e.get("errors", []))])
            }
        
        return JSONResponse(content=response_data)
        
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={
                "success": False, 
                "error": "Lỗi hệ thống",
                "error_type": "system_error",
                "details": f"Lỗi kỹ thuật: {str(e)}",
                "suggestion": "Vui lòng thử lại hoặc liên hệ quản trị viên nếu lỗi vẫn tiếp tục"
            }
        )

@app.get("/fuel/export-excel")
async def export_fuel_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Redirect đến fuel-report export"""
    from fastapi.responses import RedirectResponse
    params = []
    if from_date:
        params.append(f"from_date={from_date}")
    if to_date:
        params.append(f"to_date={to_date}")
    
    url = "/fuel-report/export-excel"
    if params:
        url += "?" + "&".join(params)
    
    return RedirectResponse(url=url, status_code=302)

@app.get("/fuel-report/export-excel")
async def export_fuel_report_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Xuất Excel báo cáo đổ dầu"""
    # Xử lý khoảng thời gian (sử dụng logic giống như fuel_page)
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            fuel_records_query = db.query(FuelRecord).filter(
                FuelRecord.date >= from_date_obj,
                FuelRecord.date <= to_date_obj
            )
        except ValueError:
            fuel_records_query = db.query(FuelRecord)
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        fuel_records_query = db.query(FuelRecord).filter(
            FuelRecord.date >= date(today.year, today.month, 1),
            FuelRecord.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
    
    fuel_records = fuel_records_query.order_by(FuelRecord.date.desc(), FuelRecord.license_plate).all()

@app.get("/theo-doi-dau-v2", response_class=HTMLResponse)
async def theo_doi_dau_v2_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """Trang Theo dõi dầu V2 - Hiển thị bản ghi đổ dầu với bộ lọc theo tháng"""
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Lấy danh sách xe từ vehicles, chỉ lấy xe đang active
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    # Sắp xếp: Xe Nhà trước, Xe Đối tác sau
    xe_nha = [v for v in vehicles if v.vehicle_type == "Xe Nhà"]
    xe_doi_tac = [v for v in vehicles if v.vehicle_type == "Xe Đối tác" or (v.vehicle_type and v.vehicle_type != "Xe Nhà")]
    
    # Sắp xếp từng nhóm theo biển số xe
    xe_nha.sort(key=lambda x: x.license_plate)
    xe_doi_tac.sort(key=lambda x: x.license_plate)
    
    # Ghép lại: Xe Nhà trước, Xe Đối tác sau
    sorted_vehicles = xe_nha + xe_doi_tac
    
    # Xác định tháng/năm để lọc (mặc định là tháng hiện tại)
    today = date.today()
    selected_month = month if month is not None else today.month
    selected_year = year if year is not None else today.year
    
    # Validate tháng/năm
    if selected_month < 1 or selected_month > 12:
        selected_month = today.month
    if selected_year < 2000 or selected_year > 2100:
        selected_year = today.year
    
    # Lọc bản ghi đổ dầu theo tháng/năm
    # Tính ngày đầu và cuối tháng
    from calendar import monthrange
    days_in_month = monthrange(selected_year, selected_month)[1]
    start_date = date(selected_year, selected_month, 1)
    end_date = date(selected_year, selected_month, days_in_month)
    
    # Lọc theo khoảng ngày (tương thích với SQLite)
    fuel_records_query = db.query(FuelRecord).filter(
        and_(
            FuelRecord.date >= start_date,
            FuelRecord.date <= end_date
        )
    )
    
    # Sắp xếp theo ngày giảm dần
    fuel_records = fuel_records_query.order_by(FuelRecord.date.desc(), FuelRecord.id.desc()).all()
    
    # Tính tổng số lít và tổng tiền theo tháng
    total_liters = sum(record.liters_pumped for record in fuel_records)
    total_cost = sum(record.cost_pumped for record in fuel_records)
    
    # Lấy lịch sử giá dầu, sắp xếp theo ngày áp dụng giảm dần
    diesel_prices = db.query(DieselPriceHistory).order_by(DieselPriceHistory.application_date.desc()).all()
    
    return templates.TemplateResponse("theo_doi_dau_v2.html", {
        "request": request,
        "current_user": current_user,
        "fuel_records": fuel_records,
        "vehicles": sorted_vehicles,
        "diesel_prices": diesel_prices,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "total_liters": total_liters,
        "total_cost": total_cost
    })

@app.get("/api/do-dau/detail/{license_plate}")
async def get_fuel_detail(
    license_plate: str,
    request: Request,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """API lấy chi tiết đổ dầu của một xe - hỗ trợ lọc theo khoảng ngày hoặc tháng"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    # Ưu tiên sử dụng from_date và to_date nếu có
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
        except ValueError:
            # Nếu định dạng không hợp lệ, fallback về tháng hiện tại
            today = date.today()
            from_date_obj = date(today.year, today.month, 1)
            if today.month == 12:
                to_date_obj = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                to_date_obj = date(today.year, today.month + 1, 1) - timedelta(days=1)
    else:
        # Nếu không có from_date/to_date, sử dụng selected_month (backward compatibility)
        selected_month = request.query_params.get("selected_month")
        
        if selected_month:
            try:
                year, month = selected_month.split('-')
                year, month = int(year), int(month)
            except ValueError:
                today = date.today()
                year, month = today.year, today.month
        else:
            today = date.today()
            year, month = today.year, today.month
        
        # Tính ngày đầu và cuối tháng
        from_date_obj = date(year, month, 1)
        if month == 12:
            to_date_obj = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            to_date_obj = date(year, month + 1, 1) - timedelta(days=1)
    
    # Lấy các bản ghi đổ dầu của xe trong khoảng thời gian
    fuel_records = db.query(FuelRecord).filter(
        FuelRecord.license_plate == license_plate,
        FuelRecord.date >= from_date_obj,
        FuelRecord.date <= to_date_obj
    ).order_by(FuelRecord.date.desc()).all()
    
    records_data = []
    for record in fuel_records:
        records_data.append({
            'id': record.id,
            'date': record.date.strftime('%Y-%m-%d'),
            'unit_price': record.fuel_price_per_liter or 0.0,
            'liters': record.liters_pumped or 0.0,
            'total_amount': record.cost_pumped or 0.0,
            'notes': record.notes or ''
        })
    
    return JSONResponse({
        "license_plate": license_plate,
        "from_date": from_date_obj.strftime('%Y-%m-%d'),
        "to_date": to_date_obj.strftime('%Y-%m-%d'),
        "records": records_data
    })

@app.post("/api/do-dau/add")
async def add_fuel_record_api(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API thêm bản ghi đổ dầu mới"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        data = await request.json()
        date_str = data.get("date")
        license_plate = data.get("license_plate")
        unit_price = float(data.get("unit_price", 0))
        liters = float(data.get("liters", 0))
        notes = data.get("notes", "")
        
        if not date_str or not license_plate:
            return JSONResponse({"error": "Thiếu thông tin bắt buộc"}, status_code=400)
        
        fuel_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        # Tính thành tiền = Số lít × Đơn giá (làm tròn đến đồng)
        total_amount = round(unit_price * liters)
        
        # Tạo bản ghi mới
        fuel_record = FuelRecord(
            date=fuel_date,
            fuel_type="Dầu DO 0,05S-II",
            license_plate=license_plate,
            fuel_price_per_liter=unit_price,
            liters_pumped=liters,
            cost_pumped=total_amount,
            notes=notes
        )
        
        db.add(fuel_record)
        db.commit()
        db.refresh(fuel_record)
        
        return JSONResponse({
            "success": True,
            "id": fuel_record.id,
            "message": "Thêm bản ghi thành công"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/api/do-dau/add-bulk")
async def add_fuel_records_bulk_api(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API thêm nhiều bản ghi đổ dầu cùng lúc"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        data = await request.json()
        records = data.get("records", [])
        
        if not records or len(records) == 0:
            return JSONResponse({"error": "Không có dữ liệu để thêm"}, status_code=400)
        
        added_records = []
        errors = []
        
        for idx, record_data in enumerate(records):
            try:
                date_str = record_data.get("date")
                license_plate = record_data.get("license_plate")
                unit_price = float(record_data.get("unit_price", 0))
                liters = float(record_data.get("liters", 0))
                notes = record_data.get("notes", "")
                
                if not date_str or not license_plate:
                    errors.append(f"Dòng {idx + 1}: Thiếu thông tin bắt buộc")
                    continue
                
                if unit_price <= 0 or liters <= 0:
                    errors.append(f"Dòng {idx + 1}: Đơn giá và số lít phải lớn hơn 0")
                    continue
                
                fuel_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                
                # Tính thành tiền = Số lít × Đơn giá (làm tròn đến đồng)
                total_amount = round(unit_price * liters)
                
                # Tạo bản ghi mới
                fuel_record = FuelRecord(
                    date=fuel_date,
                    fuel_type="Dầu DO 0,05S-II",
                    license_plate=license_plate,
                    fuel_price_per_liter=unit_price,
                    liters_pumped=liters,
                    cost_pumped=total_amount,
                    notes=notes
                )
                
                db.add(fuel_record)
                added_records.append(fuel_record)
            except Exception as e:
                errors.append(f"Dòng {idx + 1}: {str(e)}")
                continue
        
        if added_records:
            db.commit()
            for record in added_records:
                db.refresh(record)
        
        return JSONResponse({
            "success": True,
            "added_count": len(added_records),
            "total_count": len(records),
            "errors": errors,
            "message": f"Đã thêm thành công {len(added_records)}/{len(records)} bản ghi"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.put("/api/do-dau/edit/{record_id}")
async def edit_fuel_record_api(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API sửa bản ghi đổ dầu"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        fuel_record = db.query(FuelRecord).filter(FuelRecord.id == record_id).first()
        if not fuel_record:
            return JSONResponse({"error": "Không tìm thấy bản ghi"}, status_code=404)
        
        data = await request.json()
        date_str = data.get("date")
        license_plate = data.get("license_plate")
        unit_price = float(data.get("unit_price", 0))
        liters = float(data.get("liters", 0))
        notes = data.get("notes", "")
        
        if date_str:
            fuel_record.date = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        if license_plate:
            fuel_record.license_plate = license_plate
        
        fuel_record.fuel_price_per_liter = unit_price
        fuel_record.liters_pumped = liters
        fuel_record.notes = notes
        
        # Tính lại thành tiền
        fuel_record.cost_pumped = round(unit_price * liters)
        
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Cập nhật bản ghi thành công"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.delete("/api/do-dau/delete/{record_id}")
async def delete_fuel_record_api(
    record_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API xóa bản ghi đổ dầu"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        fuel_record = db.query(FuelRecord).filter(FuelRecord.id == record_id).first()
        if not fuel_record:
            return JSONResponse({"error": "Không tìm thấy bản ghi"}, status_code=404)
        
        db.delete(fuel_record)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Xóa bản ghi thành công"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/do-dau/all")
async def get_all_fuel_records(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API lấy tất cả bản ghi đổ dầu cho sheet Chi tiết"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        # Lấy tất cả bản ghi đổ dầu, sắp xếp theo ngày giảm dần
        fuel_records = db.query(FuelRecord).order_by(FuelRecord.date.desc(), FuelRecord.id.desc()).all()
        
        records_data = []
        for record in fuel_records:
            records_data.append({
                'id': record.id,
                'license_plate': record.license_plate,
                'date': record.date.strftime('%Y-%m-%d'),
                'unit_price': record.fuel_price_per_liter or 0.0,
                'liters': record.liters_pumped or 0.0,
                'total_amount': record.cost_pumped or 0.0,
                'person': record.notes or ''  # Sử dụng notes cho "Người đổ"
            })
        
        return JSONResponse({
            "success": True,
            "records": records_data
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/do-dau/filter-by-month")
async def get_fuel_records_by_month(
    request: Request,
    db: Session = Depends(get_db),
    month_year: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """API lấy bản ghi đổ dầu theo tháng/năm (format: YYYY-MM)"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        # Xác định tháng/năm để lọc (mặc định là tháng hiện tại)
        today = date.today()
        
        if month_year:
            try:
                # Parse YYYY-MM format
                year, month = map(int, month_year.split('-'))
            except (ValueError, AttributeError):
                year, month = today.year, today.month
        else:
            year, month = today.year, today.month
        
        # Validate tháng/năm
        if month < 1 or month > 12:
            month = today.month
        if year < 2000 or year > 2100:
            year = today.year
        
        # Tính ngày đầu và cuối tháng
        from calendar import monthrange
        days_in_month = monthrange(year, month)[1]
        start_date = date(year, month, 1)
        end_date = date(year, month, days_in_month)
        
        # Lọc theo khoảng ngày
        fuel_records_query = db.query(FuelRecord).filter(
            and_(
                FuelRecord.date >= start_date,
                FuelRecord.date <= end_date
            )
        )
        
        # Sắp xếp theo ngày giảm dần
        fuel_records = fuel_records_query.order_by(FuelRecord.date.desc(), FuelRecord.id.desc()).all()
        
        # Tính tổng số lít và tổng tiền theo tháng
        total_liters = sum(record.liters_pumped for record in fuel_records)
        total_cost = sum(record.cost_pumped for record in fuel_records)
        
        # Chuyển đổi sang JSON
        records_data = []
        for record in fuel_records:
            records_data.append({
                'id': record.id,
                'date': record.date.strftime('%Y-%m-%d'),
                'license_plate': record.license_plate,
                'unit_price': record.fuel_price_per_liter or 0.0,
                'liters': record.liters_pumped or 0.0,
                'total_amount': record.cost_pumped or 0.0,
                'notes': record.notes or ''
            })
        
        return JSONResponse({
            "success": True,
            "records": records_data,
            "selected_month": month,
            "selected_year": year,
            "total_liters": total_liters,
            "total_cost": total_cost
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/do-dau/totals")
async def get_fuel_totals(
    request: Request,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    license_plate: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """API lấy tổng hợp đổ dầu cho tất cả các xe (sheet Total) với bộ lọc ngày và biển số xe"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        # Bắt đầu query
        query = db.query(FuelRecord)
        
        # Áp dụng bộ lọc ngày nếu có
        if from_date:
            try:
                from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
                query = query.filter(FuelRecord.date >= from_date_obj)
            except ValueError:
                pass  # Bỏ qua nếu định dạng ngày không hợp lệ
        
        if to_date:
            try:
                to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
                query = query.filter(FuelRecord.date <= to_date_obj)
            except ValueError:
                pass  # Bỏ qua nếu định dạng ngày không hợp lệ
        
        # Áp dụng bộ lọc biển số xe nếu có (bỏ qua nếu là "Tất cả" hoặc rỗng)
        if license_plate and license_plate.strip() and license_plate.strip() != "Tất cả":
            query = query.filter(FuelRecord.license_plate == license_plate.strip())
        
        # Lấy các bản ghi đổ dầu theo bộ lọc
        fuel_records = query.all()
        
        # Tính tổng theo từng biển số xe
        totals_by_vehicle = {}
        for record in fuel_records:
            lp = record.license_plate
            if lp not in totals_by_vehicle:
                totals_by_vehicle[lp] = {
                    'total_liters': 0.0,
                    'total_cost': 0.0
                }
            totals_by_vehicle[lp]['total_liters'] += record.liters_pumped or 0.0
            totals_by_vehicle[lp]['total_cost'] += record.cost_pumped or 0.0
        
        # Chuyển đổi thành danh sách để dễ sắp xếp
        totals_list = [
            {
                'license_plate': lp,
                'total_liters': data['total_liters'],
                'total_cost': data['total_cost']
            }
            for lp, data in totals_by_vehicle.items()
        ]
        # Sắp xếp theo biển số xe
        totals_list.sort(key=lambda x: x['license_plate'])
        
        return JSONResponse({
            "success": True,
            "totals": totals_list
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/fuel-quota/compare")
async def compare_fuel_quota_with_actual(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    license_plate: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    def _normalize_text_no_accents(text: str) -> str:
        """Normalize Vietnamese text for stable sorting/comparison (remove accents, uppercase, trim)."""
        if text is None:
            return ""
        s = unicodedata.normalize("NFKD", str(text))
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return " ".join(s.strip().upper().split())

    def _is_tang_cuong_route(route_code: str) -> bool:
        """
        Identify 'Tăng cường' routes.
        We treat any route that contains 'TANG CUONG' (accent-insensitive) as reinforcement.
        """
        norm = _normalize_text_no_accents(route_code)
        return "TANG CUONG" in norm

    def _sort_fuel_quota_trips(trips: list[dict]) -> list[dict]:
        """
        Sorting rules (for UI + Excel consistency):
        - Group by route_code
        - Within each route_code: date ascending
        - 'Tăng cường' group always at the end (still date ascending inside)
        """
        def _to_date_obj(v):
            if v is None:
                return date.min
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            # Expect ISO yyyy-mm-dd from API payload
            try:
                return datetime.strptime(str(v), "%Y-%m-%d").date()
            except Exception:
                return date.min

        def _key(x: dict):
            route = (x.get("route_code") or "").strip()
            is_tc = _is_tang_cuong_route(route)
            # Force all reinforcement routes to the same group key, so they cluster at the end.
            group = "ZZZ_TANG_CUONG" if is_tc else _normalize_text_no_accents(route)
            d = _to_date_obj(x.get("date"))
            return (is_tc, group, d, _normalize_text_no_accents(route))

        return sorted(trips, key=_key)

    """
    So sánh dầu khoán (từ timekeeping) với dầu thực tế đã đổ trong khoảng thời gian.
    - Chỉ áp dụng cho xe nhà.
    - Chỉ tính các chuyến có Km > 0 và có đơn giá dầu hợp lệ tại ngày chạy.
    """
    if current_user is None:
        return JSONResponse({"success": False, "message": "Bạn cần đăng nhập"}, status_code=401)
    
    # Validate input
    if not from_date or not to_date or not license_plate:
        return JSONResponse({"success": False, "message": "Thiếu tham số từ ngày, đến ngày hoặc biển số xe"}, status_code=400)
    
    try:
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        return JSONResponse({"success": False, "message": "Định dạng ngày không hợp lệ (yyyy-mm-dd)"}, status_code=400)
    
    if from_date_obj > to_date_obj:
        return JSONResponse({"success": False, "message": "Từ ngày phải nhỏ hơn hoặc bằng Đến ngày"}, status_code=400)
    
    # Xác thực xe
    vehicle = db.query(Vehicle).filter(
        Vehicle.license_plate == license_plate.strip(),
        Vehicle.status == 1
    ).first()
    
    if not vehicle:
        return JSONResponse({"success": False, "message": "Không tìm thấy xe"}, status_code=404)
    
    if vehicle.vehicle_type != "Xe Nhà":
        return JSONResponse({"success": False, "message": "Chỉ áp dụng cho xe nhà"}, status_code=400)
    
    if vehicle.fuel_consumption is None or vehicle.fuel_consumption <= 0:
        return JSONResponse({"success": False, "message": "Xe chưa có định mức nhiên liệu, vui lòng cập nhật trước khi tính khoán dầu"}, status_code=400)
    
    fuel_consumption = vehicle.fuel_consumption
    
    # Lấy dữ liệu chấm công theo khoảng ngày và biển số
    # CHỈ LẤY CÁC BẢN GHI CÓ STATUS = ON/ONLINE/Onl (BẮT BUỘC)
    details = db.query(TimekeepingDetail).filter(
        TimekeepingDetail.license_plate == license_plate.strip(),
        TimekeepingDetail.date >= from_date_obj,
        TimekeepingDetail.date <= to_date_obj,
        or_(
            TimekeepingDetail.status == "Onl",
            TimekeepingDetail.status == "ONLINE",
            TimekeepingDetail.status == "ON"
        )
    ).all()
    
    trips_data = []
    total_quota_liters = 0.0
    total_quota_cost = 0
    skipped_no_distance = 0
    skipped_no_price = 0
    skipped_off_status = 0
    
    skipped_route_off = 0
    
    for detail in details:
        # Kiểm tra an toàn: bỏ qua nếu status là OFF (case-insensitive)
        if detail.status and detail.status.strip().upper() == "OFF":
            skipped_off_status += 1
            continue
        
        # 🔍 KIỂM TRA ROUTE STATUS: Nếu route bị OFF trong ngày đó → KHÔNG tính dầu
        route_code_to_check = detail.route_code or detail.route_name or ""
        if route_code_to_check:
            if is_route_off_on_date(db, route_code_to_check, detail.date, license_plate.strip()):
                skipped_route_off += 1
                continue
        
        distance_km = detail.distance_km or 0
        if distance_km <= 0:
            skipped_no_distance += 1
            continue
        
        fuel_price_record = get_fuel_price_by_date(db, detail.date)
        if fuel_price_record is None or fuel_price_record.unit_price is None:
            skipped_no_price += 1
            continue
        
        dk_liters = round((distance_km * fuel_consumption) / 100.0, 2)
        fuel_cost = int(round(dk_liters * fuel_price_record.unit_price))
        
        trips_data.append({
            "date": detail.date.isoformat() if detail.date else "",
            "license_plate": detail.license_plate or license_plate.strip(),
            "route_code": detail.route_code or detail.route_name or "",
            "distance_km": round(distance_km, 2),
            "dk_liters": dk_liters,
            "fuel_price": fuel_price_record.unit_price,
            "fuel_cost": fuel_cost,
            "status": detail.status or "Onl",
            "driver_name": detail.driver_name or ""
        })
        
        total_quota_liters += dk_liters
        total_quota_cost += fuel_cost
    
    # Sắp xếp theo quy tắc UI: nhóm theo Mã tuyến, ngày tăng dần trong nhóm; 'Tăng cường' luôn ở cuối
    trips_data = _sort_fuel_quota_trips(trips_data)
    
    # Tổng dầu thực tế đã đổ
    fuel_records = db.query(FuelRecord).filter(
        FuelRecord.license_plate == license_plate.strip(),
        FuelRecord.date >= from_date_obj,
        FuelRecord.date <= to_date_obj
    ).all()
    
    actual_liters = sum(record.liters_pumped or 0 for record in fuel_records)
    actual_cost = sum(record.cost_pumped or 0 for record in fuel_records)
    
    diff_liters = round(total_quota_liters - actual_liters, 2)
    diff_cost = int(round(total_quota_cost - actual_cost))
    
    return JSONResponse({
        "success": True,
        "trips": trips_data,
        "totals": {
            "quota_liters": round(total_quota_liters, 2),
            "quota_cost": int(total_quota_cost)
        },
        "actual": {
            "liters": round(actual_liters, 2),
            "cost": int(round(actual_cost))
        },
        "difference": {
            "liters": diff_liters,
            "cost": diff_cost
        },
        "meta": {
            "skipped_no_distance": skipped_no_distance,
            "skipped_no_price": skipped_no_price,
            "skipped_off_status": skipped_off_status,
            "skipped_route_off": skipped_route_off,
            "license_plate": license_plate.strip(),
            "from_date": from_date_obj.isoformat(),
            "to_date": to_date_obj.isoformat()
        }
    })

@app.get("/api/fuel-quota/export-excel")
async def export_fuel_quota_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    license_plate: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    def _normalize_text_no_accents(text: str) -> str:
        """Normalize Vietnamese text for stable sorting/comparison (remove accents, uppercase, trim)."""
        if text is None:
            return ""
        s = unicodedata.normalize("NFKD", str(text))
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return " ".join(s.strip().upper().split())

    def _is_tang_cuong_route(route_code: str) -> bool:
        norm = _normalize_text_no_accents(route_code)
        return "TANG CUONG" in norm

    def _sort_fuel_quota_trips(trips: list[dict]) -> list[dict]:
        def _to_date_obj(v):
            if v is None:
                return date.min
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            try:
                return datetime.strptime(str(v), "%Y-%m-%d").date()
            except Exception:
                return date.min

        def _key(x: dict):
            route = (x.get("route_code") or "").strip()
            is_tc = _is_tang_cuong_route(route)
            group = "ZZZ_TANG_CUONG" if is_tc else _normalize_text_no_accents(route)
            d = _to_date_obj(x.get("date"))
            return (is_tc, group, d, _normalize_text_no_accents(route))

        return sorted(trips, key=_key)

    """Xuất Excel bảng khoán dầu - So sánh dầu khoán với dầu thực tế"""
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Validate input
    if not from_date or not to_date or not license_plate:
        return JSONResponse({"success": False, "message": "Thiếu tham số từ ngày, đến ngày hoặc biển số xe"}, status_code=400)
    
    try:
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        return JSONResponse({"success": False, "message": "Định dạng ngày không hợp lệ (yyyy-mm-dd)"}, status_code=400)
    
    if from_date_obj > to_date_obj:
        return JSONResponse({"success": False, "message": "Từ ngày phải nhỏ hơn hoặc bằng Đến ngày"}, status_code=400)
    
    # Xác thực xe
    vehicle = db.query(Vehicle).filter(
        Vehicle.license_plate == license_plate.strip(),
        Vehicle.status == 1
    ).first()
    
    if not vehicle:
        return JSONResponse({"success": False, "message": "Không tìm thấy xe"}, status_code=404)
    
    if vehicle.vehicle_type != "Xe Nhà":
        return JSONResponse({"success": False, "message": "Chỉ áp dụng cho xe nhà"}, status_code=400)
    
    if vehicle.fuel_consumption is None or vehicle.fuel_consumption <= 0:
        return JSONResponse({"success": False, "message": "Xe chưa có định mức nhiên liệu"}, status_code=400)
    
    fuel_consumption = vehicle.fuel_consumption
    
    # Lấy dữ liệu chấm công theo khoảng ngày và biển số
    # CHỈ LẤY CÁC BẢN GHI CÓ STATUS = ON/ONLINE/Onl (BẮT BUỘC)
    details = db.query(TimekeepingDetail).filter(
        TimekeepingDetail.license_plate == license_plate.strip(),
        TimekeepingDetail.date >= from_date_obj,
        TimekeepingDetail.date <= to_date_obj,
        or_(
            TimekeepingDetail.status == "Onl",
            TimekeepingDetail.status == "ONLINE",
            TimekeepingDetail.status == "ON"
        )
    ).all()
    
    trips_data = []
    total_quota_liters = 0.0
    total_quota_cost = 0
    
    for detail in details:
        # Kiểm tra an toàn: bỏ qua nếu status là OFF (case-insensitive)
        if detail.status and detail.status.strip().upper() == "OFF":
            continue
        
        # 🔍 KIỂM TRA ROUTE STATUS: Nếu route bị OFF trong ngày đó → KHÔNG tính dầu
        route_code_to_check = detail.route_code or detail.route_name or ""
        if route_code_to_check:
            if is_route_off_on_date(db, route_code_to_check, detail.date, license_plate.strip()):
                continue
        
        distance_km = detail.distance_km or 0
        if distance_km <= 0:
            continue
        
        fuel_price_record = get_fuel_price_by_date(db, detail.date)
        if fuel_price_record is None or fuel_price_record.unit_price is None:
            continue
        
        dk_liters = round((distance_km * fuel_consumption) / 100.0, 2)
        fuel_cost = int(round(dk_liters * fuel_price_record.unit_price))
        
        trips_data.append({
            "date": detail.date,
            "license_plate": detail.license_plate or license_plate.strip(),
            "route_code": detail.route_code or detail.route_name or "",
            "distance_km": round(distance_km, 2),
            "dk_liters": dk_liters,
            "fuel_price": fuel_price_record.unit_price,
            "fuel_cost": fuel_cost,
            "status": detail.status or "Onl",
            "driver_name": detail.driver_name or ""
        })
        
        total_quota_liters += dk_liters
        total_quota_cost += fuel_cost
    
    # Sắp xếp theo quy tắc UI: nhóm theo Mã tuyến, ngày tăng dần trong nhóm; 'Tăng cường' luôn ở cuối
    trips_data = _sort_fuel_quota_trips(trips_data)
    
    # Tổng dầu thực tế đã đổ
    fuel_records = db.query(FuelRecord).filter(
        FuelRecord.license_plate == license_plate.strip(),
        FuelRecord.date >= from_date_obj,
        FuelRecord.date <= to_date_obj
    ).all()
    
    actual_liters = sum(record.liters_pumped or 0 for record in fuel_records)
    actual_cost = sum(record.cost_pumped or 0 for record in fuel_records)
    
    diff_liters = round(total_quota_liters - actual_liters, 2)
    diff_cost = int(round(total_quota_cost - actual_cost))
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Khoán dầu"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề báo cáo
    ws.merge_cells('A1:H1')
    ws['A1'] = "BẢNG KHOÁN DẦU"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thông tin khoảng thời gian và biển số xe
    ws.merge_cells('A2:H2')
    date_text = f"Biển số xe: {license_plate.strip()} - Từ ngày: {from_date_obj.strftime('%d/%m/%Y')} - Đến ngày: {to_date_obj.strftime('%d/%m/%Y')}"
    ws['A2'] = date_text
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True)
    
    # Header bảng
    headers = ["Ngày", "Biển số xe", "Mã tuyến", "Km chuyến", "DK (lít)", "Tiền dầu", "Trạng thái", "Lái xe"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Dữ liệu chi tiết
    row_num = 4
    for trip in trips_data:
        ws.cell(row=row_num, column=1).value = trip["date"].strftime('%d/%m/%Y') if trip["date"] else ""
        ws.cell(row=row_num, column=2).value = trip["license_plate"]
        ws.cell(row=row_num, column=3).value = trip["route_code"]
        ws.cell(row=row_num, column=4).value = trip["distance_km"]
        ws.cell(row=row_num, column=4).number_format = '0.00'
        ws.cell(row=row_num, column=5).value = trip["dk_liters"]
        ws.cell(row=row_num, column=5).number_format = '0.00'
        ws.cell(row=row_num, column=6).value = trip["fuel_cost"]
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        status_label = "OFF" if (trip["status"] or "").lower().startswith("off") else "ON"
        ws.cell(row=row_num, column=7).value = status_label
        ws.cell(row=row_num, column=8).value = trip["driver_name"]
        
        # Border cho các ô
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row_num += 1
    
    # Dòng tổng hợp
    summary_font = Font(bold=True)
    summary_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Tổng khoán
    ws.cell(row=row_num, column=1).value = "Tổng khoán"
    ws.merge_cells(f'A{row_num}:D{row_num}')
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    ws.cell(row=row_num, column=5).value = round(total_quota_liters, 2)
    ws.cell(row=row_num, column=5).number_format = '0.00'
    ws.cell(row=row_num, column=6).value = total_quota_cost
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    
    row_num += 1
    
    # Dầu thực tế
    ws.cell(row=row_num, column=1).value = "Dầu thực tế"
    ws.merge_cells(f'A{row_num}:D{row_num}')
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    ws.cell(row=row_num, column=5).value = round(actual_liters, 2)
    ws.cell(row=row_num, column=5).number_format = '0.00'
    ws.cell(row=row_num, column=6).value = actual_cost
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    
    row_num += 1
    
    # Chênh lệch
    ws.cell(row=row_num, column=1).value = "Chênh lệch (Khoán - Thực tế)"
    ws.merge_cells(f'A{row_num}:D{row_num}')
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    ws.cell(row=row_num, column=5).value = diff_liters
    ws.cell(row=row_num, column=5).number_format = '0.00'
    if diff_liters < 0:
        ws.cell(row=row_num, column=5).font = Font(bold=True, color="E74C3C")
    elif diff_liters > 0:
        ws.cell(row=row_num, column=5).font = Font(bold=True, color="27AE60")
    
    ws.cell(row=row_num, column=6).value = diff_cost
    ws.cell(row=row_num, column=6).number_format = '#,##0'
    if diff_cost < 0:
        ws.cell(row=row_num, column=6).font = Font(bold=True, color="E74C3C")
    elif diff_cost > 0:
        ws.cell(row=row_num, column=6).font = Font(bold=True, color="27AE60")
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    
    # Tạo file Excel trong memory
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Tên file
    from_date_str = from_date_obj.strftime('%d-%m-%Y')
    to_date_str = to_date_obj.strftime('%d-%m-%Y')
    filename = f"Khoan_dau_{license_plate.strip().replace('-', '_')}_Tu_{from_date_str}_Den_{to_date_str}.xlsx"
    
    return Response(
        content=excel_file.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ===== API ENDPOINTS CHO QUẢN LÝ GIÁ DẦU =====

@app.get("/api/diesel-price/all")
async def get_all_diesel_prices(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API lấy tất cả lịch sử giá dầu"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        prices = db.query(DieselPriceHistory).order_by(DieselPriceHistory.application_date.desc()).all()
        prices_list = [
            {
                "id": p.id,
                "application_date": p.application_date.strftime("%Y-%m-%d"),
                "unit_price": p.unit_price,
                "created_at": p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else "",
                "updated_at": p.updated_at.strftime("%Y-%m-%d %H:%M:%S") if p.updated_at else ""
            }
            for p in prices
        ]
        return JSONResponse({
            "success": True,
            "prices": prices_list
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/api/diesel-price/add")
async def add_diesel_price(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API thêm giá dầu mới"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        data = await request.json()
        application_date_str = data.get("application_date")
        unit_price = data.get("unit_price")
        
        if not application_date_str or unit_price is None:
            return JSONResponse({"error": "Thiếu thông tin bắt buộc"}, status_code=400)
        
        # Chuyển đổi ngày
        try:
            application_date = datetime.strptime(application_date_str, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({"error": "Định dạng ngày không hợp lệ"}, status_code=400)
        
        # Kiểm tra giá phải là số nguyên
        try:
            unit_price_int = int(unit_price)
            if unit_price_int <= 0:
                return JSONResponse({"error": "Đơn giá phải lớn hơn 0"}, status_code=400)
        except (ValueError, TypeError):
            return JSONResponse({"error": "Đơn giá phải là số nguyên"}, status_code=400)
        
        # Kiểm tra xem đã có giá cho ngày này chưa
        existing_price = db.query(DieselPriceHistory).filter(
            DieselPriceHistory.application_date == application_date
        ).first()
        
        if existing_price:
            return JSONResponse({
                "error": "Ngày này đã có giá dầu",
                "existing_id": existing_price.id,
                "existing_price": existing_price.unit_price
            }, status_code=400)
        
        # Tạo bản ghi mới
        diesel_price = DieselPriceHistory(
            application_date=application_date,
            unit_price=unit_price_int
        )
        
        db.add(diesel_price)
        db.commit()
        db.refresh(diesel_price)
        
        return JSONResponse({
            "success": True,
            "id": diesel_price.id,
            "message": "Thêm giá dầu thành công"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.put("/api/diesel-price/edit/{price_id}")
async def edit_diesel_price(
    price_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API sửa giá dầu"""
    if current_user is None:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        diesel_price = db.query(DieselPriceHistory).filter(DieselPriceHistory.id == price_id).first()
        if not diesel_price:
            return JSONResponse({"error": "Không tìm thấy bản ghi giá dầu"}, status_code=404)
        
        data = await request.json()
        application_date_str = data.get("application_date")
        unit_price = data.get("unit_price")
        
        # Cập nhật ngày áp dụng nếu có
        if application_date_str:
            try:
                new_application_date = datetime.strptime(application_date_str, "%Y-%m-%d").date()
                # Kiểm tra xem ngày mới có trùng với bản ghi khác không
                if new_application_date != diesel_price.application_date:
                    existing_price = db.query(DieselPriceHistory).filter(
                        DieselPriceHistory.application_date == new_application_date,
                        DieselPriceHistory.id != price_id
                    ).first()
                    if existing_price:
                        return JSONResponse({
                            "error": "Ngày này đã có giá dầu",
                            "existing_id": existing_price.id
                        }, status_code=400)
                diesel_price.application_date = new_application_date
            except ValueError:
                return JSONResponse({"error": "Định dạng ngày không hợp lệ"}, status_code=400)
        
        # Cập nhật đơn giá nếu có
        if unit_price is not None:
            try:
                unit_price_int = int(unit_price)
                if unit_price_int <= 0:
                    return JSONResponse({"error": "Đơn giá phải lớn hơn 0"}, status_code=400)
                diesel_price.unit_price = unit_price_int
            except (ValueError, TypeError):
                return JSONResponse({"error": "Đơn giá phải là số nguyên"}, status_code=400)
        
        diesel_price.updated_at = datetime.utcnow()
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Cập nhật giá dầu thành công"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo đổ dầu"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề báo cáo
    ws.merge_cells('A1:H1')
    ws['A1'] = "BÁO CÁO ĐỔ DẦU"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thông tin thời gian
    period_text = ""
    if from_date and to_date:
        period_text = f"Từ ngày: {from_date} đến ngày: {to_date}"
    else:
        today = date.today()
        period_text = f"Tháng: {today.month}/{today.year}"
    
    ws.merge_cells('A2:H2')
    ws['A2'] = period_text
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Header bảng
    headers = [
        "STT", "Ngày đổ", "Loại dầu", "Biển số xe", 
        "Giá xăng dầu (đồng/lít)", "Số lít đã đổ", "Số tiền đã đổ (VNĐ)", "Ghi chú"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu
    for row, record in enumerate(fuel_records, 5):
        ws.cell(row=row, column=1, value=row-4)  # STT
        ws.cell(row=row, column=2, value=record.date.strftime('%d/%m/%Y'))  # Ngày đổ
        ws.cell(row=row, column=3, value=record.fuel_type)  # Loại dầu
        ws.cell(row=row, column=4, value=record.license_plate)  # Biển số xe
        ws.cell(row=row, column=5, value=record.fuel_price_per_liter)  # Giá xăng dầu
        ws.cell(row=row, column=6, value=record.liters_pumped)  # Số lít đã đổ
        ws.cell(row=row, column=7, value=record.cost_pumped)  # Số tiền đã đổ
        ws.cell(row=row, column=8, value=record.notes or '')  # Ghi chú
    
    # Định dạng số
    for row in range(5, 5 + len(fuel_records)):
        # Giá xăng dầu - 2 chữ số thập phân
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        # Số lít - 3 chữ số thập phân
        ws.cell(row=row, column=6).number_format = '#,##0.000'
        # Số tiền - không có chữ số thập phân
        ws.cell(row=row, column=7).number_format = '#,##0'
    
    # Dòng tổng cộng
    if fuel_records:
        total_row = 5 + len(fuel_records)
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=sum(r.liters_pumped for r in fuel_records)).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=sum(r.cost_pumped for r in fuel_records)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value="").font = Font(bold=True)
        
        # Định dạng số cho dòng tổng cộng
        ws.cell(row=total_row, column=6).number_format = '#,##0.000'
        ws.cell(row=total_row, column=7).number_format = '#,##0'
    
    # Điều chỉnh độ rộng cột
    column_widths = [8, 12, 20, 15, 20, 15, 18, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    today = date.today()
    filename = f"BaoCao_DoDau_{today.strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

# ===== SALARY CALCULATION ROUTES =====

@app.get("/api/employees")
async def get_employees_api(db: Session = Depends(get_db)):
    """API để lấy danh sách nhân viên cho dropdown"""
    employees = db.query(Employee).filter(Employee.status == 1).all()
    return [
        {
            "id": emp.id,
            "name": emp.name
        }
        for emp in employees
    ]

@app.get("/salary-calculation", response_class=HTMLResponse)
async def salary_calculation_page(
    request: Request, 
    db: Session = Depends(get_db),
    selected_month: Optional[str] = None,
    selected_employee: Optional[str] = None,
    selected_route: Optional[str] = None,
    selected_vehicle: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Chỉ Admin mới được truy cập
    if current_user["role"] != "Admin":
        return RedirectResponse(url="/daily-new", status_code=303)
    """Trang bảng tính lương"""
    import calendar
    
    # Xử lý tháng được chọn
    if selected_month:
        try:
            # selected_month format: "2025-01"
            year, month = selected_month.split('-')
            year, month = int(year), int(month)
        except ValueError:
            # Nếu format không đúng, dùng tháng hiện tại
            today = date.today()
            year, month = today.year, today.month
    else:
        # Nếu không có tháng được chọn, dùng tháng hiện tại
        today = date.today()
        year, month = today.year, today.month
    
    # Tính số ngày trong tháng
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Lấy dữ liệu chuyến trong tháng được chọn
    from_date = date(year, month, 1)
    to_date = date(year, month, days_in_month)
    
    # Xây dựng query cơ bản
    daily_routes_query = db.query(DailyRoute).filter(
        DailyRoute.date >= from_date,
        DailyRoute.date <= to_date,
        DailyRoute.driver_name.isnot(None),
        DailyRoute.driver_name != ""
    )
    
    # Thêm filter theo nhân viên nếu được chọn
    if selected_employee and selected_employee != "all":
        # Tìm nhân viên theo ID hoặc tên
        try:
            employee_id = int(selected_employee)
            employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
            if employee:
                daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == employee.name)
        except ValueError:
            # Nếu không phải số, coi như tên nhân viên
            daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == selected_employee)
    
    # Join với Route để có thể filter theo route_code
    daily_routes_query = daily_routes_query.join(Route)
    
    # Thêm filter theo mã tuyến nếu được chọn
    if selected_route and selected_route != "all":
        daily_routes_query = daily_routes_query.filter(Route.route_code == selected_route)
    
    daily_routes = daily_routes_query.order_by(Route.route_code, DailyRoute.date).all()
    
    # Tính lương cho từng chuyến và lấy biển số xe
    salary_data = []
    for daily_route in daily_routes:
        # Tính lương theo công thức khác nhau tùy loại tuyến
        daily_salary = 0
        salary_type = "standard"  # Mặc định là tuyến chuẩn
        
        # Kiểm tra nếu là tuyến "Tăng Cường"
        if daily_route.route.route_code and daily_route.route.route_code.strip() == "Tăng Cường":
            salary_type = "tang_cuong"  # Luôn đánh dấu là tuyến Tăng Cường
            # Công thức cho tuyến "Tăng Cường":
            # - Nếu km < 25km: Áp dụng mức lương tuyến nội thành cố định 66.667 VNĐ
            # - Nếu km >= 25km: Số km thực tế × 1,100 đ
            if daily_route.distance_km and daily_route.distance_km > 0:
                if daily_route.distance_km < 25:
                    daily_salary = 66667  # Mức lương cố định cho tuyến ngắn (< 25km)
                else:
                    daily_salary = daily_route.distance_km * 1100
        else:
            # Công thức cho tuyến thường: Lương tuyến/tháng / 30
            if daily_route.route.monthly_salary and daily_route.route.monthly_salary > 0:
                daily_salary = daily_route.route.monthly_salary / 30
        
        # Lấy biển số xe từ daily-new với điều kiện lọc chính xác:
        # Tên nhân viên + Mã tuyến + Ngày chạy
        license_plate_display = "Chưa cập nhật"
        if daily_route.driver_name:
            # Tìm chuyến có cùng: tên lái xe + route_id + ngày chạy
            matching_routes = db.query(DailyRoute).filter(
                DailyRoute.driver_name == daily_route.driver_name,
                DailyRoute.route_id == daily_route.route_id,
                DailyRoute.date == daily_route.date,
                DailyRoute.license_plate.isnot(None),
                DailyRoute.license_plate != ""
            ).order_by(DailyRoute.created_at.desc()).all()  # Sắp xếp theo thời gian tạo mới nhất
            
            if matching_routes:
                # Lấy danh sách biển số xe duy nhất từ các chuyến khớp
                license_plates = list(set([route.license_plate for route in matching_routes if route.license_plate]))
                
                if license_plates:
                    if len(license_plates) == 1:
                        license_plate_display = license_plates[0]
                    else:
                        # Nếu có nhiều biển số, hiển thị phân tách bằng dấu phẩy
                        license_plate_display = ", ".join(license_plates)
        
        # Kiểm tra filter theo biển số xe
        should_include = True
        if selected_vehicle and selected_vehicle != "all":
            # Chỉ bao gồm nếu biển số xe khớp với filter
            if selected_vehicle not in license_plate_display:
                should_include = False
        
        if should_include:
            # Lấy thông tin loại xe dựa trên biển số xe
            vehicle_type = "Xe Nhà"  # Mặc định
            if license_plate_display and license_plate_display != "Chưa cập nhật":
                # Lấy biển số xe đầu tiên nếu có nhiều biển số
                first_license_plate = license_plate_display.split(", ")[0]
                vehicle = db.query(Vehicle).filter(Vehicle.license_plate == first_license_plate).first()
                if vehicle and vehicle.vehicle_type:
                    vehicle_type = vehicle.vehicle_type
            
            salary_data.append({
                'driver_name': daily_route.driver_name,
                'route_code': daily_route.route.route_code,
                'route_name': daily_route.route.route_name,
                'date': daily_route.date,
                'license_plate': license_plate_display,
                'vehicle_type': vehicle_type,  # Thêm thông tin loại xe
                'daily_salary': daily_salary,
                'monthly_salary': daily_route.route.monthly_salary or 0,
                'days_in_month': 30,  # Chuẩn hóa tháng 30 ngày
                'salary_type': salary_type,  # "standard" hoặc "tang_cuong"
                'distance_km': daily_route.distance_km or 0  # Số km thực tế cho tuyến Tăng Cường
            })
    
    # Lấy danh sách lái xe, tuyến và xe để hiển thị
    employees = db.query(Employee).filter(Employee.status == 1).all()
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        # Lọc ra routes không phải "Tăng Cường"
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        
        # Lọc ra routes "Tăng Cường"
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        
        # Sắp xếp routes bình thường theo A-Z
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower())
        
        # Ghép lại: routes bình thường + routes "Tăng Cường"
        return normal_routes_sorted + tang_cuong_routes
    
    routes = sort_routes_with_tang_cuong_at_bottom(routes)
    
    # Tính tổng lương theo loại tuyến
    total_standard_salary = sum(item['daily_salary'] for item in salary_data if item['salary_type'] == 'standard')
    total_tang_cuong_salary = sum(item['daily_salary'] for item in salary_data if item['salary_type'] == 'tang_cuong')
    total_salary = total_standard_salary + total_tang_cuong_salary
    
    # Tạo template data
    template_data = {
        "request": request,
        "current_user": current_user,
        "salary_data": salary_data,
        "employees": employees,
        "routes": routes,
        "vehicles": vehicles,
        "selected_month": f"{year}-{month:02d}",
        "selected_month_display": f"{month}/{year}",
        "selected_employee": selected_employee or "all",
        "selected_route": selected_route or "all",
        "selected_vehicle": selected_vehicle or "all",
        "days_in_month": days_in_month,
        "total_trips": len(salary_data),
        "total_salary": total_salary,
        "total_standard_salary": total_standard_salary,
        "total_tang_cuong_salary": total_tang_cuong_salary
    }
    
    return templates.TemplateResponse("salary_calculation.html", template_data)

def calculate_trip_salary(result: TimekeepingDetail, db: Session) -> float:
    """
    Tính lương chuyến (Lương chuyến) dựa trên các quy tắc:
    
    1. Tuyến Nội thành: 66.667 đ / chuyến (NA_005, NA_005-1, NA_013-02, NA_013-02-1, NA_013-03, NA_013-04, NA_014)
    2. Tính theo Km chuyến:
       - NA_004, V_HT_07: Km chuyến × 1.100
       - NA_002, V_HT_08: Km chuyến × 1.280
       - NA_010, NA_013, NA_013-01: Km chuyến × 1.500
       - NA_017: Km chuyến × 1.380
    3. Tuyến tính theo lương tháng: NA_012, V_HT_03 (Lương tuyến/tháng ÷ 30)
    4. Tuyến cố định:
       - V_HT_01: 66.667 đ / chuyến
       - NA_021, V_HT_09: 150.000 đ / chuyến
    5. Tuyến Tăng Cường (ưu tiên):
       - Tăng cường – Nội thành: 66.667 đ
       - Tăng cường – Nội tỉnh hoặc Liên tỉnh: Km chuyến × 1.100
    """
    # Nếu status là OFF, lương = 0
    if result.status and (result.status.strip().upper() == "OFF"):
        return 0.0
    
    route_code = (result.route_code or "").strip() if result.route_code else ""
    route_type = (result.route_type or "").strip() if result.route_type else ""
    distance_km = result.distance_km or 0
    
    # 5. Kiểm tra Tuyến Tăng Cường (ưu tiên cao nhất)
    is_tang_cuong = (
        route_code == "Tăng Cường" or
        (result.route_name and "Tăng Cường" in result.route_name)
    )
    
    if is_tang_cuong:
        # Kiểm tra route_type để xác định loại Tăng Cường
        route_type_lower = route_type.lower()
        if "nội thành" in route_type_lower:
            # Tăng cường – Nội thành: 66.667 đ
            return 66667.0
        elif "nội tỉnh" in route_type_lower or "liên tỉnh" in route_type_lower:
            # Tăng cường – Nội tỉnh hoặc Liên tỉnh: Km chuyến × 1.100
            return distance_km * 1100.0
        else:
            # Mặc định cho Tăng Cường không rõ loại: dùng công thức Nội tỉnh
            return distance_km * 1100.0
    
    # 1. Tuyến Nội thành (cố định 66.667 đ / chuyến)
    noi_thanh_routes = [
        "NA_005", "NA_005-1",
        "NA_013-02", "NA_013-02-1",
        "NA_013-03", "NA_013-04",
        "NA_014"
    ]
    if route_code in noi_thanh_routes:
        return 66667.0
    
    # 2. Tính theo Km chuyến
    if route_code in ["NA_004", "V_HT_07"]:
        return distance_km * 1100.0
    elif route_code in ["NA_002", "V_HT_08"]:
        return distance_km * 1280.0
    elif route_code in ["NA_010", "NA_013", "NA_013-01"]:
        return distance_km * 1500.0
    elif route_code == "NA_017":
        return distance_km * 1380.0
    
    # 3. Tuyến tính theo lương tháng
    if route_code in ["NA_012", "V_HT_03"]:
        # Lấy lương tuyến/tháng từ bảng Route
        route = db.query(Route).filter(Route.route_code == route_code).first()
        if route and route.monthly_salary and route.monthly_salary > 0:
            return route.monthly_salary / 30.0
        else:
            # Nếu không tìm thấy, trả về 0
            return 0.0
    
    # 4. Tuyến cố định theo chuyến
    if route_code == "V_HT_01":
        return 66667.0
    elif route_code in ["NA_021", "V_HT_09"]:
        return 150000.0
    
    # Mặc định: trả về 0 nếu không khớp với bất kỳ quy tắc nào
    return 0.0

# ==================== MONTHLY SALARY SUMMARY SERVICE ====================

def get_fuel_monthly_summary_by_driver(db: Session, driver_name: str, month: str) -> dict:
    """
    Lấy tổng hợp dầu theo lái xe và tháng, sử dụng CÙNG LOGIC với tab "Khoán dầu".
    Đây là NGUỒN DỮ LIỆU CHUẨN cho dầu trong bảng lương tổng.
    
    Args:
        db: Database session
        driver_name: Tên lái xe
        month: Tháng định dạng "YYYY-MM" (ví dụ: "2025-01")
    
    Returns:
        Dictionary với các key:
        - fuel_quota_liter: Tổng dầu khoán (lít) - từ các chuyến có Km > 0 và có giá dầu
        - fuel_used_liter: Tổng dầu đã đổ (lít) - từ fuel_records
        - fuel_money: Tổng tiền dầu khoán (VNĐ)
    """
    try:
        # Parse month
        year, month_num = map(int, month.split('-'))
        start_date = date(year, month_num, 1)
        # Tính ngày cuối cùng của tháng
        if month_num == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month_num + 1, 1) - timedelta(days=1)
        
        # Lấy tất cả chuyến của lái xe trong tháng (chỉ Xe Nhà)
        # Logic giống với compare_fuel_quota_with_actual
        # CHỈ LẤY CÁC BẢN GHI CÓ STATUS = ON/ONLINE/Onl (BẮT BUỘC)
        details = db.query(TimekeepingDetail).filter(
            TimekeepingDetail.driver_name == driver_name.strip(),
            TimekeepingDetail.date >= start_date,
            TimekeepingDetail.date <= end_date,
            or_(
                TimekeepingDetail.status == "Onl",
                TimekeepingDetail.status == "ONLINE",
                TimekeepingDetail.status == "ON"
            )
        ).all()
        
        total_quota_liters = 0.0
        total_quota_cost = 0
        
        # Lấy danh sách license_plate từ các chuyến để kiểm tra Xe Nhà
        license_plates_set = set()
        for detail in details:
            if detail.license_plate:
                license_plates_set.add(detail.license_plate.strip())
        
        # Lấy thông tin xe để kiểm tra vehicle_type
        vehicles_info = {}
        if license_plates_set:
            vehicles = db.query(Vehicle).filter(
                Vehicle.license_plate.in_(list(license_plates_set)),
                Vehicle.status == 1
            ).all()
            for vehicle in vehicles:
                vehicles_info[vehicle.license_plate] = {
                    'vehicle_type': vehicle.vehicle_type,
                    'fuel_consumption': vehicle.fuel_consumption
                }
        
        # Tính dầu khoán - CHỈ cho Xe Nhà, có Km > 0, và có giá dầu
        for detail in details:
            # Kiểm tra an toàn: bỏ qua nếu status là OFF (case-insensitive)
            if detail.status and detail.status.strip().upper() == "OFF":
                continue
            
            distance_km = detail.distance_km or 0
            if distance_km <= 0:
                continue
            
            license_plate = (detail.license_plate or "").strip()
            if not license_plate:
                continue
            
            # Chỉ tính cho Xe Nhà
            vehicle_info = vehicles_info.get(license_plate)
            if not vehicle_info or vehicle_info['vehicle_type'] != 'Xe Nhà':
                continue
            
            # 🔍 KIỂM TRA ROUTE STATUS: Nếu route bị OFF trong ngày đó → KHÔNG tính dầu
            route_code_to_check = detail.route_code or detail.route_name or ""
            if route_code_to_check:
                if is_route_off_on_date(db, route_code_to_check, detail.date, license_plate):
                    continue
            
            # Kiểm tra định mức nhiên liệu
            fuel_consumption = vehicle_info.get('fuel_consumption')
            if not fuel_consumption or fuel_consumption <= 0:
                continue
            
            # Lấy giá dầu theo ngày chuyến
            fuel_price_record = get_fuel_price_by_date(db, detail.date)
            if fuel_price_record is None or fuel_price_record.unit_price is None:
                continue
            
            # Tính dầu khoán - CÙNG LOGIC với tab Khoán dầu
            dk_liters = round((distance_km * fuel_consumption) / 100.0, 2)
            fuel_cost = int(round(dk_liters * fuel_price_record.unit_price))
            
            total_quota_liters += dk_liters
            total_quota_cost += fuel_cost
        
        # Tính dầu đã đổ từ fuel_records
        # Lấy danh sách license_plate từ các chuyến của lái xe (chỉ Xe Nhà)
        xe_nha_plates = []
        for license_plate in license_plates_set:
            vehicle_info = vehicles_info.get(license_plate)
            if vehicle_info and vehicle_info['vehicle_type'] == 'Xe Nhà':
                xe_nha_plates.append(license_plate)
        
        fuel_used = 0.0
        if xe_nha_plates:
            fuel_used_query = db.query(func.sum(FuelRecord.liters_pumped)).filter(
                FuelRecord.date >= start_date,
                FuelRecord.date <= end_date,
                FuelRecord.license_plate.in_(xe_nha_plates)
            )
            fuel_used = fuel_used_query.scalar() or 0.0
        
        return {
            "fuel_quota_liter": round(total_quota_liters, 2),
            "fuel_used_liter": round(fuel_used, 2),
            "fuel_money": int(total_quota_cost)
        }
    
    except Exception as e:
        print(f"Error getting fuel monthly summary for driver {driver_name}, month {month}: {e}")
        import traceback
        traceback.print_exc()
        return {
            "fuel_quota_liter": 0.0,
            "fuel_used_liter": 0.0,
            "fuel_money": 0
        }

def calculate_monthly_salary_summary(db: Session, month: str) -> list:
    """
    Tính bảng lương tổng theo tháng cho tất cả nhân viên.
    
    Args:
        db: Database session
        month: Tháng định dạng "YYYY-MM" (ví dụ: "2025-01")
    
    Returns:
        List of dictionaries với các key:
        - user_id: ID nhân viên
        - month: Tháng (YYYY-MM)
        - full_name: Tên đầy đủ
        - working_days: Số ngày công
        - total_trips: Tổng số chuyến
        - trip_salary: Tổng lương chuyến (VNĐ)
        - fuel_quota: Tổng dầu khoán (lít)
        - fuel_used: Tổng dầu đã đổ (lít)
        - fuel_money_diff: Số tiền dầu dư (VNĐ) - có thể âm hoặc dương
        - fuel_price: Giá dầu trung bình (VNĐ/lít)
    """
    try:
        # Parse month
        year, month_num = map(int, month.split('-'))
        start_date = date(year, month_num, 1)
        # Tính ngày cuối cùng của tháng
        if month_num == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month_num + 1, 1) - timedelta(days=1)
        
        # Lấy tất cả nhân viên đang làm việc (lái xe)
        employees = db.query(Employee).filter(
            Employee.status == 1,
            Employee.employee_status == "Đang làm việc"
        ).all()
        
        results = []
        
        for employee in employees:
            employee_name = employee.name.strip() if employee.name else ""
            if not employee_name:
                continue
            
            # 1. Tính NGÀY CÔNG: COUNT DISTINCT ngày có status = "Onl" hoặc "ONLINE" hoặc "ON"
            working_days_query = db.query(func.count(func.distinct(TimekeepingDetail.date))).filter(
                TimekeepingDetail.driver_name == employee_name,
                TimekeepingDetail.date >= start_date,
                TimekeepingDetail.date <= end_date,
                or_(
                    TimekeepingDetail.status == "Onl",
                    TimekeepingDetail.status == "ONLINE",
                    TimekeepingDetail.status == "ON"
                )
            )
            working_days = working_days_query.scalar() or 0
            
            # 2. Tính SỐ CHUYẾN: COUNT chuyến có status = "Onl" hoặc "ONLINE" hoặc "ON"
            total_trips_query = db.query(func.count(TimekeepingDetail.id)).filter(
                TimekeepingDetail.driver_name == employee_name,
                TimekeepingDetail.date >= start_date,
                TimekeepingDetail.date <= end_date,
                or_(
                    TimekeepingDetail.status == "Onl",
                    TimekeepingDetail.status == "ONLINE",
                    TimekeepingDetail.status == "ON"
                )
            )
            total_trips = total_trips_query.scalar() or 0
            
            # 3. Tính LƯƠNG CHUYẾN: SUM của calculate_trip_salary() hoặc total_amount
            # Lấy tất cả chuyến trong tháng
            trips = db.query(TimekeepingDetail).filter(
                TimekeepingDetail.driver_name == employee_name,
                TimekeepingDetail.date >= start_date,
                TimekeepingDetail.date <= end_date,
                or_(
                    TimekeepingDetail.status == "Onl",
                    TimekeepingDetail.status == "ONLINE",
                    TimekeepingDetail.status == "ON"
                )
            ).all()
            
            trip_salary = 0.0
            for trip in trips:
                # Tính lương chuyến
                salary = calculate_trip_salary(trip, db)
                trip_salary += salary
            
            # Làm tròn lương chuyến
            trip_salary = round(trip_salary, 0)
            
            # 4. LẤY DỮ LIỆU DẦU TỪ NGUỒN CHUẨN: Tab "Khoán dầu"
            # KHÔNG tính dầu từ trips nữa, chỉ lấy từ get_fuel_monthly_summary_by_driver()
            fuel_summary = get_fuel_monthly_summary_by_driver(db, employee_name, month)
            fuel_quota_total = fuel_summary["fuel_quota_liter"]
            fuel_used = fuel_summary["fuel_used_liter"]
            fuel_money_total = fuel_summary["fuel_money"]
            
            # 5. Tính SỐ TIỀN DẦU DƯ: fuel_money - (fuel_used × giá dầu trung bình)
            # Hoặc đơn giản hơn: (fuel_quota - fuel_used) × giá dầu trung bình
            # Nhưng để đảm bảo khớp với tab Khoán dầu, ta tính từ fuel_money đã có
            # fuel_money = tổng tiền dầu khoán từ các chuyến
            # Tính giá dầu trung bình từ fuel_money và fuel_quota
            avg_fuel_price = 0
            if fuel_quota_total > 0:
                avg_fuel_price = fuel_money_total / fuel_quota_total
            else:
                # Nếu không có dầu khoán, lấy giá dầu cuối cùng của tháng
                fuel_price_record = get_fuel_price_by_date(db, end_date)
                if fuel_price_record and fuel_price_record.unit_price:
                    avg_fuel_price = fuel_price_record.unit_price
            
            # Tính tiền dầu dư: (fuel_quota - fuel_used) × giá dầu trung bình
            fuel_money_diff = (fuel_quota_total - fuel_used) * avg_fuel_price
            fuel_money_diff = round(fuel_money_diff, 0)
            
            # Kiểm tra xem có dữ liệu đã lưu cho tháng này không
            saved_salary = db.query(SalaryMonthly).filter(
                SalaryMonthly.employee_id == employee.id,
                SalaryMonthly.month == month_num,
                SalaryMonthly.year == year
            ).first()
            
            # Nếu có dữ liệu đã lưu, dùng dữ liệu đó; nếu không, dùng 0
            bao_hiem_xh = saved_salary.bao_hiem_xh if saved_salary else 0
            rua_xe = saved_salary.rua_xe if saved_salary else 0
            tien_trach_nhiem = saved_salary.tien_trach_nhiem if saved_salary else 0
            ung_luong = saved_salary.ung_luong if saved_salary else 0
            sua_xe = saved_salary.sua_xe if saved_salary else 0
            
            results.append({
                "user_id": employee.id,
                "month": month,
                "full_name": employee_name,
                "working_days": working_days,
                "total_trips": total_trips,
                "trip_salary": int(trip_salary),
                "fuel_quota": round(fuel_quota_total, 2),
                "fuel_used": round(fuel_used, 2),
                "fuel_money_diff": int(fuel_money_diff),
                "fuel_price": int(avg_fuel_price) if avg_fuel_price > 0 else 0,
                # Các cột manual: lấy từ saved data nếu có, nếu không thì = 0
                "bao_hiem_xh": bao_hiem_xh,
                "rua_xe": rua_xe,
                "tien_trach_nhiem": tien_trach_nhiem,
                "ung_luong": ung_luong,
                "sua_xe": sua_xe
            })
        
        # Sắp xếp theo tên: những người có tên cụ thể sẽ hiển thị ở dòng dưới cùng
        # Danh sách người cần đẩy xuống cuối
        bottom_names = {
            "Mr Ba",
            "Lê Bá Thắng",
            "Nguyễn Công Hảo",
            "Nguyễn Trang Kiều",
            "Nguyễn Văn Luận"
        }
        
        def sort_key(item):
            full_name = item["full_name"]
            # Nếu tên trong danh sách bottom_names, trả về (1, name) để đẩy xuống cuối
            # Nếu không, trả về (0, name) để giữ ở trên
            if full_name in bottom_names:
                return (1, full_name)
            else:
                return (0, full_name)
        
        results.sort(key=sort_key)
        
        return results
    
    except Exception as e:
        print(f"Error calculating monthly salary summary: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_partner_vehicle_unit_price(license_plate: str, route_type: str, route_code: str, route_name: str) -> float:
    """
    Lấy đơn giá theo km cho xe đối tác:
    - Nội thành: 0 (vì tính theo chuyến cố định)
    - Xe 37H-076.36: 5,175 đ/km
    - Xe 37H-083.68: 4,801 đ/km
    """
    license_plate = (license_plate or "").strip() if license_plate else ""
    route_type = (route_type or "").strip() if route_type else ""
    route_code = (route_code or "").strip() if route_code else ""
    route_name = (route_name or "").strip() if route_name else ""
    
    # Kiểm tra nếu là tuyến Nội thành
    noi_thanh_lower = "nội thành"
    if (route_type.lower() == noi_thanh_lower or 
        route_code.lower() == noi_thanh_lower or 
        noi_thanh_lower in route_name.lower()):
        return 0.0  # Nội thành không tính theo km
    
    # Xe 37H-076.36: 5,175 đ/km
    if license_plate == "37H-076.36":
        return 5175.0
    
    # Xe 37H-083.68: 4,801 đ/km
    if license_plate == "37H-083.68":
        return 4801.0
    
    # Mặc định: 0
    return 0.0

def calculate_partner_vehicle_payment(result: TimekeepingDetail, db: Session) -> float:
    """
    Tính tiền cho xe đối tác dựa trên các quy tắc:
    
    1. Tuyến "Nội thành": 204.545 đ / chuyến (cố định, không cộng phí cầu đường)
    2. Tính theo Km chuyến (ngoài Nội thành):
       - Xe 37H-076.36: (Km chuyến × 5.175 đ) + Phí cầu đường
       - Xe 37H-083.68: (Km chuyến × 4.801 đ) + Phí cầu đường
    
    Công thức: Thành tiền = đơn giá × km chuyến + Phí cầu đường
    
    Ưu tiên: Nếu route_type = "Nội thành" → áp dụng giá cố định
    Nếu không → áp dụng đơn giá km theo từng xe đối tác + phí cầu đường
    """
    # Nếu status là OFF, tiền = 0
    if result.status and (result.status.strip().upper() == "OFF"):
        return 0.0
    
    route_type = (result.route_type or "").strip() if result.route_type else ""
    route_code = (result.route_code or "").strip() if result.route_code else ""
    route_name = (result.route_name or "").strip() if result.route_name else ""
    license_plate = (result.license_plate or "").strip() if result.license_plate else ""
    distance_km = result.distance_km or 0
    bridge_fee = result.bridge_fee or 0  # Phí cầu đường
    
    # Trường hợp 1: Tuyến "Nội thành" - giá cố định 204.545 đ / chuyến (không cộng phí cầu đường)
    # Kiểm tra route_type, route_code hoặc route_name (case-insensitive)
    noi_thanh_lower = "nội thành"
    if (route_type.lower() == noi_thanh_lower or 
        route_code.lower() == noi_thanh_lower or 
        noi_thanh_lower in route_name.lower()):
        return 204545.0
    
    # Trường hợp 2: Tính theo Km (ngoài Nội thành) + Phí cầu đường
    # Xe 37H-076.36: (Km chuyến × 5.175 đ) + Phí cầu đường
    if license_plate == "37H-076.36":
        return (distance_km * 5175.0) + bridge_fee
    
    # Xe 37H-083.68: (Km chuyến × 4.801 đ) + Phí cầu đường
    if license_plate == "37H-083.68":
        return (distance_km * 4801.0) + bridge_fee
    
    # Mặc định: nếu không khớp với bất kỳ quy tắc nào, trả về 0
    return 0.0

# ==================== MONTHLY SALARY SUMMARY API ====================

@app.get("/api/salary-summary")
async def get_salary_summary(
    month: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    API: Lấy bảng lương tổng theo tháng
    
    Args:
        month: Tháng định dạng "YYYY-MM" (ví dụ: "2025-01")
               Nếu không có, mặc định là tháng hiện tại
    
    Returns:
        JSON response với danh sách bảng lương tổng
    """
    try:
        # Nếu không có month, dùng tháng hiện tại
        if not month:
            today = date.today()
            month = f"{today.year}-{today.month:02d}"
        
        # Validate format
        try:
            year, month_num = map(int, month.split('-'))
            if month_num < 1 or month_num > 12:
                return JSONResponse({
                    "success": False,
                    "message": "Tháng không hợp lệ. Format: YYYY-MM"
                }, status_code=400)
        except ValueError:
            return JSONResponse({
                "success": False,
                "message": "Format tháng không đúng. Format: YYYY-MM"
            }, status_code=400)
        
        # Tính bảng lương tổng
        results = calculate_monthly_salary_summary(db, month)
        
        return JSONResponse({
            "success": True,
            "month": month,
            "data": results,
            "count": len(results)
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi tính bảng lương: {str(e)}"
        }, status_code=500)

@app.get("/salary-summary", response_class=HTMLResponse)
async def salary_summary_page(
    request: Request,
    month: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Trang hiển thị bảng lương tổng theo tháng
    """
    # Nếu không có month, dùng tháng hiện tại
    if not month:
        today = date.today()
        month = f"{today.year}-{today.month:02d}"
    
    # Validate format
    try:
        year, month_num = map(int, month.split('-'))
        if month_num < 1 or month_num > 12:
            month = f"{date.today().year}-{date.today().month:02d}"
    except ValueError:
        month = f"{date.today().year}-{date.today().month:02d}"
    
    # Tính bảng lương tổng
    salary_data = calculate_monthly_salary_summary(db, month)
    
    # Tính tổng các cột
    totals = {
        "working_days": sum(item["working_days"] for item in salary_data),
        "total_trips": sum(item["total_trips"] for item in salary_data),
        "trip_salary": sum(item["trip_salary"] for item in salary_data),
        # Giữ lại các trường fuel để tương thích (có thể dùng cho export)
        "fuel_quota": round(sum(item.get("fuel_quota", 0) for item in salary_data), 2),
        "fuel_used": round(sum(item.get("fuel_used", 0) for item in salary_data), 2),
        "fuel_money_diff": sum(item.get("fuel_money_diff", 0) for item in salary_data)
    }
    
    return templates.TemplateResponse("salary_summary.html", {
        "request": request,
        "current_user": current_user,
        "month": month,
        "salary_data": salary_data,
        "totals": totals
    })

@app.post("/api/salary-summary/save")
async def save_salary_summary(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    API: Lưu snapshot lương tháng cho các lái xe
    Body: {
        "month": "2025-01",
        "salary_data": [
            {
                "user_id": 1,
                "bao_hiem_xh": 100000,
                "rua_xe": 50000,
                "tien_trach_nhiem": 200000,
                "ung_luong": 500000,
                "sua_xe": 300000
            },
            ...
        ]
    }
    """
    try:
        # Lấy dữ liệu từ request body
        body = await request.json()
        month = body.get("month")
        salary_data = body.get("salary_data", [])
        
        if not month:
            return JSONResponse({
                "success": False,
                "message": "Thiếu thông tin tháng"
            }, status_code=400)
        
        # Validate format
        try:
            year, month_num = map(int, month.split('-'))
            if month_num < 1 or month_num > 12:
                return JSONResponse({
                    "success": False,
                    "message": "Tháng không hợp lệ"
                }, status_code=400)
        except ValueError:
            return JSONResponse({
                "success": False,
                "message": "Format tháng không đúng. Format: YYYY-MM"
            }, status_code=400)
        
        # Lưu từng bản ghi
        saved_count = 0
        for item in salary_data:
            user_id = item.get("user_id")
            if not user_id:
                continue
            
            # Kiểm tra xem employee có tồn tại không
            employee = db.query(Employee).filter(Employee.id == user_id).first()
            if not employee:
                continue
            
            # Lấy giá trị các trường manual (mặc định 0 nếu không có)
            bao_hiem_xh = int(item.get("bao_hiem_xh", 0) or 0)
            rua_xe = int(item.get("rua_xe", 0) or 0)
            tien_trach_nhiem = int(item.get("tien_trach_nhiem", 0) or 0)
            ung_luong = int(item.get("ung_luong", 0) or 0)
            sua_xe = int(item.get("sua_xe", 0) or 0)
            
            # Tìm bản ghi đã tồn tại
            existing = db.query(SalaryMonthly).filter(
                SalaryMonthly.employee_id == user_id,
                SalaryMonthly.month == month_num,
                SalaryMonthly.year == year
            ).first()
            
            if existing:
                # Cập nhật bản ghi đã tồn tại
                existing.bao_hiem_xh = bao_hiem_xh
                existing.rua_xe = rua_xe
                existing.tien_trach_nhiem = tien_trach_nhiem
                existing.ung_luong = ung_luong
                existing.sua_xe = sua_xe
                existing.updated_at = datetime.utcnow()
            else:
                # Tạo bản ghi mới
                new_record = SalaryMonthly(
                    employee_id=user_id,
                    month=month_num,
                    year=year,
                    bao_hiem_xh=bao_hiem_xh,
                    rua_xe=rua_xe,
                    tien_trach_nhiem=tien_trach_nhiem,
                    ung_luong=ung_luong,
                    sua_xe=sua_xe
                )
                db.add(new_record)
            
            saved_count += 1
        
        # Commit tất cả thay đổi
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": f"Đã lưu lương tháng {month} cho {saved_count} nhân viên",
            "saved_count": saved_count
        })
    
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi lưu dữ liệu: {str(e)}"
        }, status_code=500)

@app.get("/api/salary-summary/export-excel")
@app.post("/api/salary-summary/export-excel")
async def export_salary_summary_excel(
    request: Request,
    month: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Export bảng lương tổng ra file Excel
    Hỗ trợ cả GET (tương thích ngược) và POST (với dữ liệu từ input fields)
    """
    try:
        # Nếu là POST request, lấy dữ liệu từ body
        manual_salary_data = None
        if request.method == "POST":
            try:
                body = await request.json()
                month = body.get("month") or month
                manual_salary_data = body.get("salary_data")
            except:
                pass
        
        # Nếu không có month, dùng tháng hiện tại
        if not month:
            today = date.today()
            month = f"{today.year}-{today.month:02d}"
        
        # Validate format
        try:
            year, month_num = map(int, month.split('-'))
            if month_num < 1 or month_num > 12:
                month = f"{date.today().year}-{date.today().month:02d}"
        except ValueError:
            month = f"{date.today().year}-{date.today().month:02d}"
        
        # Nếu có dữ liệu từ POST (manual input), dùng dữ liệu đó
        # Nếu không, tính từ database
        if manual_salary_data:
            salary_data = manual_salary_data
        else:
            # Tính bảng lương tổng từ database
            salary_data_db = calculate_monthly_salary_summary(db, month)
            # Convert sang format giống với manual data
            salary_data = []
            for item in salary_data_db:
                salary_data.append({
                    "full_name": item["full_name"],
                    "working_days": item["working_days"],
                    "total_trips": item["total_trips"],
                    "trip_salary": item["trip_salary"],
                    "bao_hiem_xh": item.get("bao_hiem_xh", 0),
                    "rua_xe": item.get("rua_xe", 0),
                    "tien_trach_nhiem": item.get("tien_trach_nhiem", 0),
                    "ung_luong": item.get("ung_luong", 0),
                    "sua_xe": item.get("sua_xe", 0),
                    "con_lai": item["trip_salary"] - item.get("bao_hiem_xh", 0) - item.get("tien_trach_nhiem", 0) - item.get("ung_luong", 0) + item.get("rua_xe", 0) + item.get("sua_xe", 0)
                })
        
        if not salary_data:
            return JSONResponse({
                "success": False,
                "message": "Không có dữ liệu để xuất Excel"
            }, status_code=404)
        
        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Bang Luong {month}"
        
        # Header style
        header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:K1')
        ws['A1'] = f"BẢNG LƯƠNG TỔNG THEO THÁNG {month}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Headers - Cập nhật với các cột mới
        headers = [
            "STT", 
            "Họ tên", 
            "Ngày công", 
            "Số chuyến", 
            "Lương chuyến (VNĐ)",
            "Bảo hiểm XH (VNĐ)",
            "Rửa xe (VNĐ)",
            "Tiền trách nhiệm (VNĐ)",
            "Ứng lương (VNĐ)",
            "Sửa xe (VNĐ)",
            "Còn lại (VNĐ)"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        for row_idx, item in enumerate(salary_data, start=4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)  # STT
            ws.cell(row=row_idx, column=2, value=item.get("full_name", ""))
            ws.cell(row=row_idx, column=3, value=item.get("working_days", 0))
            ws.cell(row=row_idx, column=4, value=item.get("total_trips", 0))
            
            # Lương chuyến
            trip_salary = float(item.get("trip_salary", 0))
            ws.cell(row=row_idx, column=5, value=trip_salary).number_format = '#,##0'
            
            # Bảo hiểm XH
            bao_hiem_xh = float(item.get("bao_hiem_xh", 0))
            ws.cell(row=row_idx, column=6, value=bao_hiem_xh).number_format = '#,##0'
            
            # Rửa xe
            rua_xe = float(item.get("rua_xe", 0))
            ws.cell(row=row_idx, column=7, value=rua_xe).number_format = '#,##0'
            
            # Tiền trách nhiệm
            tien_trach_nhiem = float(item.get("tien_trach_nhiem", 0))
            ws.cell(row=row_idx, column=8, value=tien_trach_nhiem).number_format = '#,##0'
            
            # Ứng lương
            ung_luong = float(item.get("ung_luong", 0))
            ws.cell(row=row_idx, column=9, value=ung_luong).number_format = '#,##0'
            
            # Sửa xe
            sua_xe = float(item.get("sua_xe", 0))
            ws.cell(row=row_idx, column=10, value=sua_xe).number_format = '#,##0'
            
            # Còn lại (tính từ dữ liệu hoặc lấy trực tiếp)
            con_lai = float(item.get("con_lai", 0))
            if con_lai == 0:
                # Tính lại nếu chưa có
                con_lai = trip_salary - bao_hiem_xh - tien_trach_nhiem - ung_luong + rua_xe + sua_xe
            ws.cell(row=row_idx, column=11, value=con_lai).number_format = '#,##0'
            ws.cell(row=row_idx, column=11).font = Font(bold=True, color="1976d2")
        
        # Total row
        totals = {
            "working_days": sum(item.get("working_days", 0) for item in salary_data),
            "total_trips": sum(item.get("total_trips", 0) for item in salary_data),
            "trip_salary": sum(float(item.get("trip_salary", 0)) for item in salary_data),
            "bao_hiem_xh": sum(float(item.get("bao_hiem_xh", 0)) for item in salary_data),
            "rua_xe": sum(float(item.get("rua_xe", 0)) for item in salary_data),
            "tien_trach_nhiem": sum(float(item.get("tien_trach_nhiem", 0)) for item in salary_data),
            "ung_luong": sum(float(item.get("ung_luong", 0)) for item in salary_data),
            "sua_xe": sum(float(item.get("sua_xe", 0)) for item in salary_data),
            "con_lai": sum(float(item.get("con_lai", 0)) for item in salary_data)
        }
        
        total_row = len(salary_data) + 4
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=totals["working_days"]).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=totals["total_trips"]).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=totals["trip_salary"]).number_format = '#,##0'
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=totals["bao_hiem_xh"]).number_format = '#,##0'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=totals["rua_xe"]).number_format = '#,##0'
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=totals["tien_trach_nhiem"]).number_format = '#,##0'
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=totals["ung_luong"]).number_format = '#,##0'
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=totals["sua_xe"]).number_format = '#,##0'
        ws.cell(row=total_row, column=10).font = Font(bold=True)
        ws.cell(row=total_row, column=11, value=totals["con_lai"]).number_format = '#,##0'
        ws.cell(row=total_row, column=11).font = Font(bold=True, color="1976d2")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        filename = f"Bang_Luong_Tong_{month}.xlsx"
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename*=UTF-8\'\'{quote(filename)}'
            }
        )
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi xuất Excel: {str(e)}"
        }, status_code=500)

@app.get("/salary-calculation-v2", response_class=HTMLResponse)
async def salary_calculation_v2_page(
    request: Request,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    tab: Optional[str] = None,
    search: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """Trang Bảng Tính Lương Ver 2.0 - Hỗ trợ 2 tab: Tính lương lái xe và Tính tiền xe đối tác"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    redirect_response = check_and_redirect_access(current_user["role"], "/salary-calculation-v2", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    # Xác định tab hiện tại (mặc định là "driver" - tính lương lái xe)
    current_tab = tab if tab in ["driver", "partner", "summary"] else "driver"
    
    # Lấy danh sách lái xe từ TimekeepingDetail (chỉ cho tab driver)
    drivers_query = db.query(TimekeepingDetail.driver_name).distinct()
    drivers_list = [row[0] for row in drivers_query.filter(TimekeepingDetail.driver_name.isnot(None), TimekeepingDetail.driver_name != "").all()]
    drivers_list.sort()
    
    # Lấy danh sách xe đối tác (chỉ cho tab partner)
    partner_vehicles = db.query(Vehicle).filter(
        Vehicle.status == 1,
        Vehicle.vehicle_type == "Xe Đối tác"
    ).all()
    partner_vehicle_plates = [v.license_plate for v in partner_vehicles]
    
    # Tính giá trị mặc định: từ ngày đầu tháng đến ngày cuối tháng hiện tại
    today = date.today()
    first_day_of_month = date(today.year, today.month, 1)
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    last_day_of_month = date(today.year, today.month, days_in_month)
    
    # Nếu không có from_date hoặc to_date, sử dụng giá trị mặc định
    if not from_date:
        from_date = first_day_of_month.strftime("%Y-%m-%d")
    if not to_date:
        to_date = last_day_of_month.strftime("%Y-%m-%d")
    
    # Chỉ thực hiện tìm kiếm khi người dùng nhấn nút Tìm kiếm (search=1)
    has_search = search == "1"
    
    # Mặc định: chưa tìm kiếm thì không trả về kết quả để tránh gọi DB không cần thiết
    results = [] if has_search else None
    selected_driver = None
    selected_license_plate = None
    total_driver_trip_salary = 0  # Khởi tạo tổng lương chuyến
    
    # Thực hiện tìm kiếm với giá trị mặc định hoặc giá trị được cung cấp
    if has_search and from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            
            # Validate dates
            if from_date_obj > to_date_obj:
                # Nếu ngày bắt đầu > ngày kết thúc, không tìm kiếm
                pass
            else:
                # Query TimekeepingDetail
                query = db.query(TimekeepingDetail).filter(
                    TimekeepingDetail.date >= from_date_obj,
                    TimekeepingDetail.date <= to_date_obj
                )
                
                # Nếu là tab "partner" (xe đối tác), chỉ lấy các chuyến của xe đối tác
                if current_tab == "partner":
                    # Lọc chỉ các chuyến có biển số xe là xe đối tác
                    if partner_vehicle_plates:
                        query = query.filter(TimekeepingDetail.license_plate.in_(partner_vehicle_plates))
                    else:
                        # Nếu không có xe đối tác nào, trả về kết quả rỗng
                        query = query.filter(TimekeepingDetail.license_plate == None)
                    
                    # Filter theo biển số xe nếu có (chỉ cho tab partner)
                    if license_plate and license_plate.strip():
                        query = query.filter(TimekeepingDetail.license_plate == license_plate.strip())
                        selected_license_plate = license_plate.strip()
                else:
                    # Tab "driver": Filter theo lái xe nếu có
                    if driver_name and driver_name.strip():
                        query = query.filter(TimekeepingDetail.driver_name == driver_name.strip())
                        selected_driver = driver_name.strip()
                
                # Lấy tất cả kết quả trước khi sắp xếp
                all_results = query.all()
                
                # Nếu là tab partner, lọc thêm để đảm bảo chỉ lấy xe đối tác
                if current_tab == "partner":
                    filtered_results = []
                    for result in all_results:
                        if result.license_plate and result.license_plate in partner_vehicle_plates:
                            filtered_results.append(result)
                    all_results = filtered_results
                
                # Tách ra 2 nhóm: tuyến thường và tuyến "Tăng Cường" (chỉ cho tab driver)
                normal_results = []
                tang_cuong_results = []
                
                for result in all_results:
                    if current_tab == "driver":
                        # Kiểm tra xem có phải tuyến "Tăng Cường" không
                        is_tang_cuong = (
                            (result.route_code and result.route_code.strip() == "Tăng Cường") or
                            (result.route_name and "Tăng Cường" in result.route_name)
                        )
                        
                        if is_tang_cuong:
                            tang_cuong_results.append(result)
                        else:
                            normal_results.append(result)
                    else:
                        # Tab partner: không cần tách Tăng Cường
                        normal_results.append(result)
                
                # Sắp xếp mỗi nhóm:
                # 1. Theo mã tuyến (route_code) - ưu tiên cao nhất
                # 2. Sau đó theo ngày (date)
                def sort_key(result):
                    # route_code có thể None, nên xử lý an toàn
                    route_code = (result.route_code or "").strip() if result.route_code else ""
                    date_val = result.date or date.min
                    return (route_code, date_val)
                
                # Sắp xếp nhóm tuyến thường
                normal_results_sorted = sorted(normal_results, key=sort_key)
                
                # Sắp xếp nhóm tuyến "Tăng Cường" (chỉ cho tab driver)
                if current_tab == "driver":
                    tang_cuong_results_sorted = sorted(tang_cuong_results, key=sort_key)
                    # Ghép lại: tuyến thường trước, tuyến "Tăng Cường" sau
                    results = normal_results_sorted + tang_cuong_results_sorted
                else:
                    results = normal_results_sorted
                
                # Tính lương/tiền chuyến cho từng kết quả
                results_with_payment = []
                for result in results:
                    if current_tab == "partner":
                        # Tính tiền xe đối tác
                        payment = calculate_partner_vehicle_payment(result, db)
                        # Lấy đơn giá và phí cầu đường để hiển thị
                        unit_price = get_partner_vehicle_unit_price(
                            result.license_plate,
                            result.route_type,
                            result.route_code,
                            result.route_name
                        )
                        bridge_fee = result.bridge_fee or 0
                        # Tab partner không tính dầu khoán
                        fuel_data = {
                            "dk_liters": 0.0,
                            "fuel_cost": 0,
                            "fuel_price": None,
                            "fuel_consumption": None,
                            "warning": None
                        }
                    else:
                        # Tính lương lái xe
                        payment = calculate_trip_salary(result, db)
                        unit_price = 0
                        bridge_fee = 0
                        # Tính dầu khoán cho tab driver
                        fuel_data = calculate_fuel_quota(result, db)
                    
                    # Tạo dictionary với thông tin result và tiền/lương đã tính
                    result_dict = {
                        "result": result,
                        "trip_salary": payment,
                        "unit_price": unit_price,
                        "bridge_fee": bridge_fee,
                        "fuel_data": fuel_data
                    }
                    results_with_payment.append(result_dict)
                
                results = results_with_payment
                
                # Tính tổng lương chuyến cho tab driver (chỉ tính các chuyến không OFF)
                if current_tab == "driver" and results:
                    for item in results:
                        result = item.get("result")
                        trip_salary = item.get("trip_salary", 0)
                        # Chỉ tính các chuyến có status không phải OFF
                        if result and hasattr(result, 'status') and result.status not in ['OFF', 'Off']:
                            total_driver_trip_salary += trip_salary
        except ValueError:
            # Nếu format ngày không đúng, bỏ qua
            pass
    
    return templates.TemplateResponse("salary_calculation_v2.html", {
        "request": request,
        "current_user": current_user,
        "drivers": drivers_list,
        "from_date": from_date,
        "to_date": to_date,
        "selected_driver": selected_driver,
        "selected_license_plate": selected_license_plate,
        "results": results,
        "current_tab": current_tab,
        "partner_vehicles": partner_vehicle_plates,
        "has_search": has_search,
        "total_driver_trip_salary": total_driver_trip_salary
    })

@app.get("/salary-calculation-v2/export-excel")
async def export_salary_calculation_v2_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    tab: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """Xuất Excel bảng tính lương Ver 2.0 - Hỗ trợ cả tab driver và partner"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    redirect_response = check_and_redirect_access(current_user["role"], "/salary-calculation-v2", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    # Xác định tab hiện tại (mặc định là "driver")
    current_tab = tab if tab in ["driver", "partner"] else "driver"
    
    # Lấy danh sách xe đối tác (chỉ cho tab partner)
    partner_vehicles = db.query(Vehicle).filter(
        Vehicle.status == 1,
        Vehicle.vehicle_type == "Xe Đối tác"
    ).all()
    partner_vehicle_plates = [v.license_plate for v in partner_vehicles]
    
    results = []
    
    # Nếu có điều kiện tìm kiếm, thực hiện tìm kiếm (sử dụng logic giống như salary_calculation_v2_page)
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            
            # Validate dates
            if from_date_obj > to_date_obj:
                # Nếu ngày bắt đầu > ngày kết thúc, trả về file rỗng
                pass
            else:
                # Query TimekeepingDetail
                query = db.query(TimekeepingDetail).filter(
                    TimekeepingDetail.date >= from_date_obj,
                    TimekeepingDetail.date <= to_date_obj
                )
                
                # Nếu là tab "partner" (xe đối tác), chỉ lấy các chuyến của xe đối tác
                if current_tab == "partner":
                    # Lọc chỉ các chuyến có biển số xe là xe đối tác
                    if partner_vehicle_plates:
                        query = query.filter(TimekeepingDetail.license_plate.in_(partner_vehicle_plates))
                    else:
                        # Nếu không có xe đối tác nào, trả về kết quả rỗng
                        query = query.filter(TimekeepingDetail.license_plate == None)
                    
                    # Filter theo biển số xe nếu có (chỉ cho tab partner)
                    if license_plate and license_plate.strip():
                        query = query.filter(TimekeepingDetail.license_plate == license_plate.strip())
                else:
                    # Tab "driver": Filter theo lái xe nếu có
                    if driver_name and driver_name.strip():
                        query = query.filter(TimekeepingDetail.driver_name == driver_name.strip())
                
                # Lấy tất cả kết quả trước khi sắp xếp
                all_results = query.all()
                
                # Nếu là tab partner, lọc thêm để đảm bảo chỉ lấy xe đối tác
                if current_tab == "partner":
                    filtered_results = []
                    for result in all_results:
                        if result.license_plate and result.license_plate in partner_vehicle_plates:
                            filtered_results.append(result)
                    all_results = filtered_results
                
                # Tách ra 2 nhóm: tuyến thường và tuyến "Tăng Cường" (chỉ cho tab driver)
                normal_results = []
                tang_cuong_results = []
                
                for result in all_results:
                    if current_tab == "driver":
                        # Kiểm tra xem có phải tuyến "Tăng Cường" không
                        is_tang_cuong = (
                            (result.route_code and result.route_code.strip() == "Tăng Cường") or
                            (result.route_name and "Tăng Cường" in result.route_name)
                        )
                        
                        if is_tang_cuong:
                            tang_cuong_results.append(result)
                        else:
                            normal_results.append(result)
                    else:
                        # Tab partner: không cần tách Tăng Cường
                        normal_results.append(result)
                
                # Sắp xếp mỗi nhóm:
                # 1. Theo mã tuyến (route_code) - ưu tiên cao nhất
                # 2. Sau đó theo ngày (date)
                def sort_key(result):
                    # route_code có thể None, nên xử lý an toàn
                    route_code = (result.route_code or "").strip() if result.route_code else ""
                    date_val = result.date or date.min
                    return (route_code, date_val)
                
                # Sắp xếp nhóm tuyến thường
                normal_results_sorted = sorted(normal_results, key=sort_key)
                
                # Sắp xếp nhóm tuyến "Tăng Cường" (chỉ cho tab driver)
                if current_tab == "driver":
                    tang_cuong_results_sorted = sorted(tang_cuong_results, key=sort_key)
                    # Ghép lại: tuyến thường trước, tuyến "Tăng Cường" sau
                    results = normal_results_sorted + tang_cuong_results_sorted
                else:
                    results = normal_results_sorted
                
                # Tính lương/tiền chuyến cho từng kết quả
                results_with_payment = []
                for result in results:
                    if current_tab == "partner":
                        # Tính tiền xe đối tác
                        payment = calculate_partner_vehicle_payment(result, db)
                        # Lấy đơn giá và phí cầu đường để hiển thị
                        unit_price = get_partner_vehicle_unit_price(
                            result.license_plate,
                            result.route_type,
                            result.route_code,
                            result.route_name
                        )
                        bridge_fee = result.bridge_fee or 0
                        # Tab partner không tính dầu khoán
                        fuel_data = {
                            "dk_liters": 0.0,
                            "fuel_cost": 0,
                            "fuel_price": None,
                            "fuel_consumption": None,
                            "warning": None
                        }
                    else:
                        # Tính lương lái xe
                        payment = calculate_trip_salary(result, db)
                        unit_price = 0
                        bridge_fee = 0
                        # Tính dầu khoán cho tab driver
                        fuel_data = calculate_fuel_quota(result, db)
                    
                    result_dict = {
                        "result": result,
                        "trip_salary": payment,
                        "unit_price": unit_price,
                        "bridge_fee": bridge_fee,
                        "fuel_data": fuel_data
                    }
                    results_with_payment.append(result_dict)
                
                results = results_with_payment
        except ValueError:
            # Nếu format ngày không đúng, trả về file rỗng
            pass
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng tính lương V2"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Xác định số cột cho merge cells
    merge_range = 'A1:M1' if current_tab == "partner" else 'A1:L1'
    
    # Tiêu đề báo cáo
    ws.merge_cells(merge_range)
    if current_tab == "partner":
        ws['A1'] = "BẢNG TÍNH TIỀN XE ĐỐI TÁC VER 2.0"
    else:
        ws['A1'] = "BẢNG TÍNH LƯƠNG VER 2.0"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thông tin khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            date_text = f"Từ ngày: {from_date_obj.strftime('%d/%m/%Y')} - Đến ngày: {to_date_obj.strftime('%d/%m/%Y')}"
        except:
            date_text = "Khoảng thời gian: Chưa xác định"
    else:
        date_text = "Khoảng thời gian: Chưa xác định"
    
    ws.merge_cells(merge_range.replace('1', '2'))
    ws['A2'] = date_text
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True)
    
    # Thông tin filter
    if current_tab == "partner":
        if license_plate and license_plate.strip():
            filter_text = f"Biển số xe: {license_plate.strip()}"
        else:
            filter_text = "Xe đối tác: Tất cả"
    else:
        if driver_name and driver_name.strip():
            filter_text = f"Lái xe: {driver_name.strip()}"
        else:
            filter_text = "Lái xe: Tất cả"
    
    ws.merge_cells(merge_range.replace('1', '3'))
    ws['A3'] = filter_text
    ws['A3'].alignment = Alignment(horizontal="center")
    ws['A3'].font = Font(italic=True)
    
    # Header bảng
    payment_column_name = "Tiền chuyến" if current_tab == "partner" else "Lương chuyến"
    if current_tab == "partner":
        headers = [
            "STT", "Ngày", "Biển số xe", "Mã tuyến", "Lộ trình",
            "Km chuyến", "Đơn giá", "Phí cầu đường", "Trạng thái", "Lái xe", "Mã chuyến", payment_column_name, "Ghi chú"
        ]
    else:
        headers = [
            "STT", "Ngày", "Biển số xe", "Mã tuyến", "Lộ trình",
            "Km chuyến", "DK", "Tiền dầu", "Trạng thái", "Lái xe", payment_column_name, "Ghi chú"
        ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu
    for idx, item in enumerate(results, 6):
        # Lấy result và trip_salary từ item
        result = item.get("result") if isinstance(item, dict) else item
        trip_salary = item.get("trip_salary", 0) if isinstance(item, dict) else 0
        
        # STT
        ws.cell(row=idx, column=1, value=idx-5)
        
        # Ngày
        if result.date:
            ws.cell(row=idx, column=2, value=result.date.strftime('%d/%m/%Y'))
        else:
            ws.cell(row=idx, column=2, value='')
        
        # Biển số xe
        ws.cell(row=idx, column=3, value=result.license_plate or '')
        
        # Mã tuyến
        ws.cell(row=idx, column=4, value=result.route_code or '')
        
        # Lộ trình
        ws.cell(row=idx, column=5, value=result.itinerary or '')
        
        # Km chuyến
        if result.distance_km:
            ws.cell(row=idx, column=6, value=result.distance_km)
            ws.cell(row=idx, column=6).number_format = '#,##0.0'
        else:
            ws.cell(row=idx, column=6, value=0)
            ws.cell(row=idx, column=6).number_format = '#,##0.0'
        
        # Đơn giá và Phí cầu đường (chỉ cho tab partner)
        if current_tab == "partner":
            # Đơn giá
            unit_price = item.get("unit_price", 0) if isinstance(item, dict) else 0
            ws.cell(row=idx, column=7, value=unit_price)
            ws.cell(row=idx, column=7).number_format = '#,##0'
            
            # Phí cầu đường
            bridge_fee = item.get("bridge_fee", 0) if isinstance(item, dict) else 0
            ws.cell(row=idx, column=8, value=bridge_fee)
            ws.cell(row=idx, column=8).number_format = '#,##0'
            
            # Trạng thái (cột 9)
            status_value = result.status or 'ON'
            if status_value == 'OFF' or status_value == 'Off':
                ws.cell(row=idx, column=9, value='OFF')
            else:
                ws.cell(row=idx, column=9, value='ON')
            
            # Lái xe (cột 10)
            ws.cell(row=idx, column=10, value=result.driver_name or '')
            
            # Mã chuyến (cột 11)
            ws.cell(row=idx, column=11, value=result.trip_code or '')
            
            # Tiền chuyến (cột 12)
            if result.status == 'OFF' or result.status == 'Off':
                ws.cell(row=idx, column=12, value=0)
            else:
                ws.cell(row=idx, column=12, value=trip_salary)
            ws.cell(row=idx, column=12).number_format = '#,##0'
            
            # Ghi chú (cột 13)
            ws.cell(row=idx, column=13, value=result.notes or '')
        else:
            # DK (cột 7)
            fuel_data = item.get("fuel_data", {}) if isinstance(item, dict) else {}
            if fuel_data.get("warning"):
                ws.cell(row=idx, column=7, value=fuel_data.get("warning", ""))
            elif fuel_data.get("dk_liters") is not None and fuel_data.get("dk_liters", 0) > 0:
                ws.cell(row=idx, column=7, value=fuel_data.get("dk_liters", 0))
                ws.cell(row=idx, column=7).number_format = '#,##0.00'
            else:
                ws.cell(row=idx, column=7, value='')
            
            # Tiền dầu (cột 8)
            # Chỉ hiển thị tiền dầu nếu đúng khoán và có giá trị > 0
            assignment_status = fuel_data.get("assignment_status")
            if fuel_data.get("warning"):
                ws.cell(row=idx, column=8, value='')
            elif assignment_status == "valid" and fuel_data.get("fuel_cost") is not None and fuel_data.get("fuel_cost", 0) > 0:
                ws.cell(row=idx, column=8, value=fuel_data.get("fuel_cost", 0))
                ws.cell(row=idx, column=8).number_format = '#,##0'
            elif assignment_status == "invalid" or assignment_status == "no_assignment":
                # Không tính tiền dầu - hiển thị 0 hoặc -- cho xe đối tác
                if fuel_data.get("assignment_reason") == "Xe đối tác":
                    ws.cell(row=idx, column=8, value='--')
                else:
                    ws.cell(row=idx, column=8, value=0)
                    ws.cell(row=idx, column=8).number_format = '#,##0'
            elif fuel_data.get("fuel_cost") is not None and fuel_data.get("fuel_cost", 0) > 0:
                ws.cell(row=idx, column=8, value=fuel_data.get("fuel_cost", 0))
                ws.cell(row=idx, column=8).number_format = '#,##0'
            else:
                ws.cell(row=idx, column=8, value=0)
                ws.cell(row=idx, column=8).number_format = '#,##0'
            
            # Trạng thái (cột 9)
            status_value = result.status or 'ON'
            if status_value == 'OFF' or status_value == 'Off':
                ws.cell(row=idx, column=9, value='OFF')
            else:
                ws.cell(row=idx, column=9, value='ON')
            
            # Lái xe (cột 10)
            ws.cell(row=idx, column=10, value=result.driver_name or '')
            
            # Lương chuyến (cột 11)
            if result.status == 'OFF' or result.status == 'Off':
                ws.cell(row=idx, column=11, value=0)
            else:
                ws.cell(row=idx, column=11, value=trip_salary)
            ws.cell(row=idx, column=11).number_format = '#,##0'
            
            # Ghi chú (cột 12)
            ws.cell(row=idx, column=12, value=result.notes or '')
    
    # Định dạng số cho cột lương chuyến (nếu cần format lại)
    salary_column = 12 if current_tab == "partner" else 11
    for row in range(6, 6 + len(results)):
        cell = ws.cell(row=row, column=salary_column)
        if cell.value == 0 or cell.value == '':
            pass
        else:
            cell.number_format = '#,##0'
    
    # Dòng tổng cộng
    if results:
        total_row = 5 + len(results) + 1
        # Tính tổng lương chuyến
        total_salary = sum(item.get("trip_salary", 0) if isinstance(item, dict) else 0 for item in results)
        
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value="").font = Font(bold=True)
        
        # Tổng km
        total_km = sum(
            (item.get("result") if isinstance(item, dict) else item).distance_km or 0 
            for item in results
        )
        ws.cell(row=total_row, column=6, value=total_km).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '#,##0.0'
        
        if current_tab == "partner":
            ws.cell(row=total_row, column=7, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=8, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=9, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=10, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=11, value="").font = Font(bold=True)
            # Tổng tiền chuyến (cột 12)
            ws.cell(row=total_row, column=12, value=total_salary).font = Font(bold=True)
            ws.cell(row=total_row, column=12).number_format = '#,##0'
            ws.cell(row=total_row, column=13, value="").font = Font(bold=True)
        else:
            # Tổng DK (cột 7)
            total_dk = sum(
                item.get("fuel_data", {}).get("dk_liters", 0) if isinstance(item, dict) else 0
                for item in results
            )
            ws.cell(row=total_row, column=7, value=total_dk).font = Font(bold=True)
            ws.cell(row=total_row, column=7).number_format = '#,##0.00'
            
            # Tổng tiền dầu (cột 8)
            total_fuel_cost = sum(
                item.get("fuel_data", {}).get("fuel_cost", 0) if isinstance(item, dict) else 0
                for item in results
            )
            ws.cell(row=total_row, column=8, value=total_fuel_cost).font = Font(bold=True)
            ws.cell(row=total_row, column=8).number_format = '#,##0'
            
            ws.cell(row=total_row, column=9, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=10, value="").font = Font(bold=True)
            # Tổng lương chuyến (cột 11)
            ws.cell(row=total_row, column=11, value=total_salary).font = Font(bold=True)
            ws.cell(row=total_row, column=11).number_format = '#,##0'
            ws.cell(row=total_row, column=12, value="").font = Font(bold=True)
    
    # Điều chỉnh độ rộng cột
    if current_tab == "partner":
        column_widths = [8, 12, 15, 15, 20, 12, 12, 15, 12, 20, 15, 18, 30]
    else:
        column_widths = [8, 12, 15, 15, 20, 12, 12, 15, 12, 20, 18, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            filename = f"BangTinhLuong_V2_{from_date_obj.strftime('%Y%m%d')}_{to_date_obj.strftime('%Y%m%d')}.xlsx"
        except:
            today = date.today()
            filename = f"BangTinhLuong_V2_{today.strftime('%Y%m%d')}.xlsx"
    else:
        today = date.today()
        filename = f"BangTinhLuong_V2_{today.strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/salary-calculation/export-excel")
async def export_salary_calculation_excel(
    db: Session = Depends(get_db),
    selected_month: Optional[str] = None,
    selected_employee: Optional[str] = None,
    selected_route: Optional[str] = None,
    selected_vehicle: Optional[str] = None
):
    """Xuất Excel bảng tính lương"""
    import calendar
    
    # Xử lý tháng được chọn (sử dụng logic giống như salary_calculation_page)
    if selected_month:
        try:
            year, month = selected_month.split('-')
            year, month = int(year), int(month)
        except ValueError:
            today = date.today()
            year, month = today.year, today.month
    else:
        today = date.today()
        year, month = today.year, today.month
    
    # Tính số ngày trong tháng
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Lấy dữ liệu chuyến trong tháng được chọn
    from_date = date(year, month, 1)
    to_date = date(year, month, days_in_month)
    
    # Xây dựng query cơ bản (sử dụng logic giống như salary_calculation_page)
    daily_routes_query = db.query(DailyRoute).filter(
        DailyRoute.date >= from_date,
        DailyRoute.date <= to_date,
        DailyRoute.driver_name.isnot(None),
        DailyRoute.driver_name != ""
    )
    
    # Thêm filter theo nhân viên nếu được chọn
    if selected_employee and selected_employee != "all":
        try:
            employee_id = int(selected_employee)
            employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
            if employee:
                daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == employee.name)
        except ValueError:
            daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == selected_employee)
    
    # Join với Route để có thể filter theo route_code
    daily_routes_query = daily_routes_query.join(Route)
    
    # Thêm filter theo mã tuyến nếu được chọn
    if selected_route and selected_route != "all":
        daily_routes_query = daily_routes_query.filter(Route.route_code == selected_route)
    
    daily_routes = daily_routes_query.order_by(Route.route_code, DailyRoute.date).all()
    
    # Tính lương cho từng chuyến và lấy biển số xe (sử dụng logic giống như salary_calculation_page)
    salary_data = []
    for daily_route in daily_routes:
        # Tính lương theo công thức khác nhau tùy loại tuyến
        daily_salary = 0
        salary_type = "standard"  # Mặc định là tuyến chuẩn
        
        # Kiểm tra nếu là tuyến "Tăng Cường"
        if daily_route.route.route_code and daily_route.route.route_code.strip() == "Tăng Cường":
            salary_type = "tang_cuong"  # Luôn đánh dấu là tuyến Tăng Cường
            # Công thức cho tuyến "Tăng Cường":
            # - Nếu km < 25km: Áp dụng mức lương tuyến nội thành cố định 66.667 VNĐ
            # - Nếu km >= 25km: Số km thực tế × 1,100 đ
            if daily_route.distance_km and daily_route.distance_km > 0:
                if daily_route.distance_km < 25:
                    daily_salary = 66667  # Mức lương cố định cho tuyến ngắn (< 25km)
                else:
                    daily_salary = daily_route.distance_km * 1100
        else:
            # Công thức cho tuyến thường: Lương tuyến/tháng / 30
            if daily_route.route.monthly_salary and daily_route.route.monthly_salary > 0:
                daily_salary = daily_route.route.monthly_salary / 30
        
        # Lấy biển số xe từ daily-new với điều kiện lọc chính xác
        license_plate_display = "Chưa cập nhật"
        if daily_route.driver_name:
            matching_routes = db.query(DailyRoute).filter(
                DailyRoute.driver_name == daily_route.driver_name,
                DailyRoute.route_id == daily_route.route_id,
                DailyRoute.date == daily_route.date,
                DailyRoute.license_plate.isnot(None),
                DailyRoute.license_plate != ""
            ).order_by(DailyRoute.created_at.desc()).all()
            
            if matching_routes:
                license_plates = list(set([route.license_plate for route in matching_routes if route.license_plate]))
                
                if license_plates:
                    if len(license_plates) == 1:
                        license_plate_display = license_plates[0]
                    else:
                        license_plate_display = ", ".join(license_plates)
        
        # Kiểm tra filter theo biển số xe
        should_include = True
        if selected_vehicle and selected_vehicle != "all":
            # Chỉ bao gồm nếu biển số xe khớp với filter
            if selected_vehicle not in license_plate_display:
                should_include = False
        
        if should_include:
            # Lấy thông tin loại xe dựa trên biển số xe
            vehicle_type = "Xe Nhà"  # Mặc định
            if license_plate_display and license_plate_display != "Chưa cập nhật":
                # Lấy biển số xe đầu tiên nếu có nhiều biển số
                first_license_plate = license_plate_display.split(", ")[0]
                vehicle = db.query(Vehicle).filter(Vehicle.license_plate == first_license_plate).first()
                if vehicle and vehicle.vehicle_type:
                    vehicle_type = vehicle.vehicle_type
            
            salary_data.append({
                'driver_name': daily_route.driver_name,
                'route_code': daily_route.route.route_code,
                'route_name': daily_route.route.route_name,
                'date': daily_route.date,
                'license_plate': license_plate_display,
                'vehicle_type': vehicle_type,  # Thêm thông tin loại xe
                'daily_salary': daily_salary,
                'salary_type': salary_type,  # "standard" hoặc "tang_cuong"
                'distance_km': daily_route.distance_km or 0  # Số km thực tế cho tuyến Tăng Cường
            })
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng tính lương"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề báo cáo
    ws.merge_cells('A1:H1')
    ws['A1'] = "BẢNG TÍNH LƯƠNG"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thông tin tháng
    month_text = f"Tháng: {month}/{year}"
    ws.merge_cells('A2:H2')
    ws['A2'] = month_text
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True)
    
    # Header bảng
    headers = [
        "STT", "Họ và tên lái xe", "Mã tuyến", 
        "Ngày chạy", "Biển số xe", "Số km", "Lương chuyến (XN)", "Lương chuyến (XĐT)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu
    for row, item in enumerate(salary_data, 5):
        ws.cell(row=row, column=1, value=row-4)  # STT
        ws.cell(row=row, column=2, value=item['driver_name'])  # Họ và tên lái xe
        ws.cell(row=row, column=3, value=item['route_code'])  # Mã tuyến
        ws.cell(row=row, column=4, value=item['date'].strftime('%d/%m/%Y'))  # Ngày chạy
        ws.cell(row=row, column=5, value=item['license_plate'])  # Biển số xe
        
        # Số km - chỉ hiển thị cho tuyến Tăng Cường
        if item['salary_type'] == 'tang_cuong' and item['distance_km'] > 0:
            ws.cell(row=row, column=6, value=item['distance_km'])
        else:
            ws.cell(row=row, column=6, value='')
        
        # Lương chuyến theo loại xe
        if item.get('vehicle_type') == 'Xe Đối tác':
            ws.cell(row=row, column=7, value='')  # Lương chuyến (XN) - trống
            ws.cell(row=row, column=8, value=item['daily_salary'])  # Lương chuyến (XĐT)
        else:
            ws.cell(row=row, column=7, value=item['daily_salary'])  # Lương chuyến (XN)
            ws.cell(row=row, column=8, value='')  # Lương chuyến (XĐT) - trống
    
    # Định dạng số cho các cột
    for row in range(5, 5 + len(salary_data)):
        ws.cell(row=row, column=6).number_format = '#,##0.0'  # Số km - 1 chữ số thập phân
        ws.cell(row=row, column=7).number_format = '#,##0'  # Lương chuyến (XN)
        ws.cell(row=row, column=8).number_format = '#,##0'  # Lương chuyến (XĐT)
    
    # Dòng tổng cộng
    if salary_data:
        total_row = 5 + len(salary_data)
        total_xn_salary = sum(item['daily_salary'] for item in salary_data if item.get('vehicle_type') != 'Xe Đối tác')
        total_xdt_salary = sum(item['daily_salary'] for item in salary_data if item.get('vehicle_type') == 'Xe Đối tác')
        
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_xn_salary).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_xdt_salary).font = Font(bold=True)
        
        # Định dạng số cho dòng tổng cộng
        ws.cell(row=total_row, column=7).number_format = '#,##0'
        ws.cell(row=total_row, column=8).number_format = '#,##0'
    
    # Điều chỉnh độ rộng cột
    column_widths = [8, 25, 15, 15, 20, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    filename = f"BangTinhLuong_{month:02d}_{year}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

async def create_daily_revenue_finance_record(selected_date: date, db: Session):
    """Tự động tạo/cập nhật bản ghi thu nhập trong finance-report từ doanh thu hàng ngày"""
    try:
        # Lấy tổng doanh thu của ngày
        revenue_records = db.query(RevenueRecord).filter(RevenueRecord.date == selected_date).all()
        
        print(f"Processing date {selected_date}: Found {len(revenue_records)} revenue records")
        
        # Kiểm tra xem đã có bản ghi finance cho ngày này chưa
        # Tìm bản ghi doanh thu tự động: transaction_type = "Thu" và category = "Doanh thu vận chuyển"
        existing_finance_record = db.query(FinanceTransaction).filter(
            FinanceTransaction.date == selected_date,
            FinanceTransaction.transaction_type == "Thu",
            FinanceTransaction.category == "Doanh thu vận chuyển"
        ).first()
        
        # Nếu không có revenue records, xóa finance record nếu có
        if not revenue_records:
            if existing_finance_record:
                db.delete(existing_finance_record)
                db.commit()
                print(f"Deleted finance record for date {selected_date} (no revenue records)")
            return
        
        # Tính tổng doanh thu - chỉ tính cho các chuyến có trạng thái ON (Online)
        # Bỏ qua các chuyến có status OFF
        total_revenue = 0
        online_count = 0
        offline_count = 0
        for record in revenue_records:
            # Chỉ tính doanh thu cho các chuyến có status = "Online" hoặc "ON"
            if record.status and record.status.upper() in ["ONLINE", "ON"]:
                online_count += 1
                if record.manual_total > 0:
                    total_revenue += record.manual_total
                else:
                    total_revenue += record.total_amount
            else:
                offline_count += 1
        
        print(f"Date {selected_date}: Online={online_count}, Offline={offline_count}, Total revenue={total_revenue}")
        
        if existing_finance_record:
            # Cập nhật bản ghi hiện có - luôn cập nhật kể cả khi total_revenue = 0
            existing_finance_record.amount = total_revenue
            existing_finance_record.total = total_revenue
            existing_finance_record.note = f"Tự động cập nhật từ {len(revenue_records)} tuyến doanh thu (Online: {online_count}, Offline: {offline_count})"
            existing_finance_record.updated_at = datetime.utcnow()
            db.commit()
            print(f"Updated finance record for date {selected_date} with total: {total_revenue}")
        else:
            # Tạo bản ghi mới - luôn tạo nếu có revenue records, kể cả khi total_revenue = 0
            finance_record = FinanceTransaction(
                transaction_type="Thu",
                category="Doanh thu vận chuyển",
                date=selected_date,
                description=f"Doanh thu hàng ngày {selected_date.strftime('%d/%m/%Y')}",
                route_code="Tổng hợp",
                amount=total_revenue,
                vat=0,
                discount1=0,
                discount2=0,
                total=total_revenue,
                note=f"Tự động tạo từ {len(revenue_records)} tuyến doanh thu (Online: {online_count}, Offline: {offline_count})"
            )
            
            db.add(finance_record)
            db.commit()
            print(f"Created finance record for date {selected_date} with total: {total_revenue} (from {len(revenue_records)} revenue records)")
            
    except Exception as e:
        print(f"Error creating/updating daily revenue finance record for {selected_date}: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()

@app.get("/finance-report", response_class=HTMLResponse)
async def finance_report_page(
    request: Request, 
    db: Session = Depends(get_db),
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user = Depends(get_current_user)
):
    # Mặc định là tháng hiện tại nếu không có tham số
    if not month or not year:
        current_date = datetime.now()
        month = month or current_date.month
        year = year or current_date.year
    
    # Tự động tạo bản ghi tài chính cho tất cả các ngày trong tháng có doanh thu
    from calendar import monthrange
    days_in_month = monthrange(year, month)[1]
    
    # Lấy tất cả các ngày có revenue records trong tháng
    start_date = date(year, month, 1)
    end_date = date(year, month, days_in_month)
    
    # Query tất cả revenue records trong tháng để đếm số ngày
    all_revenue_records = db.query(RevenueRecord).filter(
        and_(
            RevenueRecord.date >= start_date,
            RevenueRecord.date <= end_date
        )
    ).all()
    
    # Lấy danh sách các ngày duy nhất có revenue records
    revenue_dates_set = {record.date for record in all_revenue_records}
    print(f"[Finance Report] Processing {month}/{year}: Found {len(revenue_dates_set)} unique days with revenue records")
    print(f"[Finance Report] Revenue dates: {sorted(revenue_dates_set)}")
    
    # Lấy danh sách các ngày đã có finance records
    existing_finance_records = db.query(FinanceTransaction).filter(
        and_(
            extract('month', FinanceTransaction.date) == month,
            extract('year', FinanceTransaction.date) == year,
            FinanceTransaction.transaction_type == "Thu",
            FinanceTransaction.category == "Doanh thu vận chuyển"
        )
    ).all()
    existing_dates_set = {record.date for record in existing_finance_records}
    print(f"[Finance Report] Already have {len(existing_dates_set)} finance records for {month}/{year}")
    print(f"[Finance Report] Existing finance dates: {sorted(existing_dates_set)}")
    
    # Tạo finance records cho các ngày có revenue nhưng chưa có finance record
    created_count = 0
    updated_count = 0
    for revenue_date in sorted(revenue_dates_set):
        try:
            # Kiểm tra xem đã có finance record cho ngày này chưa
            existing_finance = db.query(FinanceTransaction).filter(
                FinanceTransaction.date == revenue_date,
                FinanceTransaction.transaction_type == "Thu",
                FinanceTransaction.category == "Doanh thu vận chuyển"
            ).first()
            
            # Luôn gọi hàm để đảm bảo cập nhật đúng (nó sẽ tự kiểm tra và tạo/cập nhật)
            await create_daily_revenue_finance_record(revenue_date, db)
            
            if not existing_finance:
                created_count += 1
                print(f"[Finance Report] ✓ Created new finance record for {revenue_date}")
            else:
                updated_count += 1
                print(f"[Finance Report] ✓ Updated existing finance record for {revenue_date}")
        except Exception as e:
            print(f"[Finance Report] ✗ Error processing date {revenue_date}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"[Finance Report] Summary: Created {created_count} new, Updated {updated_count} existing finance records for {month}/{year}")
    
    # Lấy dữ liệu tài chính từ bảng FinanceTransaction riêng biệt
    finance_data = db.query(FinanceTransaction).filter(
        and_(
            extract('month', FinanceTransaction.date) == month,
            extract('year', FinanceTransaction.date) == year
        )
    ).order_by(FinanceTransaction.date.desc()).all()
    
    # Tính tổng từ bảng mới
    total_income = sum(item.total for item in finance_data if item.transaction_type == "Thu")
    total_expense = sum(item.total for item in finance_data if item.transaction_type == "Chi")
    total_balance = total_income - total_expense
    
    return templates.TemplateResponse("finance_report.html", {
        "request": request,
        "current_user": current_user,
        "finance_data": finance_data,
        "total_income": total_income,
        "total_expense": total_expense,
        "total_balance": total_balance,
        "selected_month": month,
        "selected_year": year
    })

@app.get("/finance-report/export")
async def export_finance_report_excel(
    db: Session = Depends(get_db),
    month: Optional[int] = None,
    year: Optional[int] = None
):
    # Mặc định là tháng hiện tại nếu không có tham số
    if not month or not year:
        current_date = datetime.now()
        month = month or current_date.month
        year = year or current_date.year
    
    # Tự động tạo bản ghi tài chính cho tất cả các ngày trong tháng có doanh thu
    from calendar import monthrange
    days_in_month = monthrange(year, month)[1]
    
    # Lấy tất cả các ngày có revenue records trong tháng
    start_date = date(year, month, 1)
    end_date = date(year, month, days_in_month)
    
    revenue_dates = db.query(RevenueRecord.date).filter(
        and_(
            RevenueRecord.date >= start_date,
            RevenueRecord.date <= end_date
        )
    ).distinct().all()
    
    # Tạo finance records cho các ngày có revenue nhưng chưa có finance record
    created_count = 0
    for (revenue_date,) in revenue_dates:
        try:
            # Kiểm tra xem đã có finance record cho ngày này chưa
            existing_finance = db.query(FinanceTransaction).filter(
                FinanceTransaction.date == revenue_date,
                FinanceTransaction.transaction_type == "Thu",
                FinanceTransaction.category == "Doanh thu vận chuyển"
            ).first()
            
            # Nếu chưa có, tạo mới
            if not existing_finance:
                await create_daily_revenue_finance_record(revenue_date, db)
                created_count += 1
        except Exception as e:
            print(f"Error creating finance record for date {revenue_date}: {e}")
            continue
    
    # Lấy dữ liệu tài chính từ bảng FinanceTransaction
    finance_data = db.query(FinanceTransaction).filter(
        and_(
            extract('month', FinanceTransaction.date) == month,
            extract('year', FinanceTransaction.date) == year
        )
    ).order_by(FinanceTransaction.date).all()
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"BaoCaoTaiChinh_{month:02d}_{year}"
    
    # Tiêu đề
    ws.cell(row=1, column=1, value=f"BÁO CÁO TÀI CHÍNH THÁNG {month}/{year}").font = Font(bold=True, size=16)
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Thông tin thời gian
    ws.merge_cells('A2:K2')
    ws.cell(row=2, column=1, value=f"Xuất báo cáo ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1).font = Font(italic=True)
    
    # Header bảng
    headers = [
        "Ngày", "Danh mục", "Diễn giải", "Mã tuyến", 
        "Số tiền (chưa VAT)", "VAT (%)", "CK1 (%)", "CK2 (%)", 
        "Thành tiền", "Ghi chú"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dữ liệu
    for row, item in enumerate(finance_data, 5):
        ws.cell(row=row, column=1, value=item.date.strftime('%d/%m/%Y') if item.date else '')
        ws.cell(row=row, column=2, value=item.transaction_type or '')
        ws.cell(row=row, column=3, value=item.description or '')
        ws.cell(row=row, column=4, value=item.route_code or '')
        ws.cell(row=row, column=5, value=item.amount or 0)
        ws.cell(row=row, column=6, value=item.vat or 0)
        ws.cell(row=row, column=7, value=item.discount1 or 0)
        ws.cell(row=row, column=8, value=item.discount2 or 0)
        ws.cell(row=row, column=9, value=item.total or 0)
        ws.cell(row=row, column=10, value=item.note or '')
        
        # Định dạng số cho các cột tiền
        ws.cell(row=row, column=5).number_format = '#,##0'  # Số tiền chưa VAT
        ws.cell(row=row, column=9).number_format = '#,##0'  # Thành tiền
        
        # Định dạng phần trăm cho VAT và chiết khấu
        ws.cell(row=row, column=6).number_format = '0.0"%"'  # VAT
        ws.cell(row=row, column=7).number_format = '0.0"%"'  # CK1
        ws.cell(row=row, column=8).number_format = '0.0"%"'  # CK2
    
    # Dòng tổng cộng
    if finance_data:
        total_row = 5 + len(finance_data)
        total_amount = sum(item.amount or 0 for item in finance_data)
        total_final = sum(item.total or 0 for item in finance_data)
        
        # Tính tổng thu và chi
        total_income = sum(item.total or 0 for item in finance_data if item.transaction_type == 'Thu')
        total_expense = sum(item.total or 0 for item in finance_data if item.transaction_type == 'Chi')
        net_balance = total_income - total_expense
        
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_amount).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=total_final).font = Font(bold=True)
        ws.cell(row=total_row, column=10, value="").font = Font(bold=True)
        
        # Định dạng số cho dòng tổng
        ws.cell(row=total_row, column=5).number_format = '#,##0'
        ws.cell(row=total_row, column=9).number_format = '#,##0'
        
        # Thêm dòng tổng kết
        summary_row = total_row + 2
        ws.cell(row=summary_row, column=1, value="TỔNG KẾT:").font = Font(bold=True, size=12)
        ws.cell(row=summary_row + 1, column=1, value="Tổng thu:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=2, value=total_income).font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=2).number_format = '#,##0'
        ws.cell(row=summary_row + 2, column=1, value="Tổng chi:").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=2, value=total_expense).font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=2).number_format = '#,##0'
        ws.cell(row=summary_row + 3, column=1, value="Lợi nhuận:").font = Font(bold=True)
        ws.cell(row=summary_row + 3, column=2, value=net_balance).font = Font(bold=True)
        ws.cell(row=summary_row + 3, column=2).number_format = '#,##0'
        
        # Màu sắc cho lợi nhuận
        if net_balance > 0:
            ws.cell(row=summary_row + 3, column=2).font = Font(bold=True, color="00AA00")
        elif net_balance < 0:
            ws.cell(row=summary_row + 3, column=2).font = Font(bold=True, color="AA0000")
    
    # Điều chỉnh độ rộng cột
    column_widths = [12, 12, 30, 15, 18, 10, 10, 10, 18, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Thêm border cho toàn bộ bảng
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Áp dụng border cho tất cả cells có dữ liệu
    max_row = 5 + len(finance_data) + 5  # +5 cho tổng kết
    for row in range(1, max_row + 1):
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    filename = f"BaoCaoTaiChinh_{month:02d}_{year}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/finance-report/create-sample-data")
async def create_sample_finance_data(db: Session = Depends(get_db)):
    """Tạo dữ liệu mẫu cho báo cáo tài chính"""
    current_date = datetime.now()
    
    # Dữ liệu mẫu cho tháng hiện tại
    sample_data = [
        {
            "date": current_date.replace(day=1),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_002",
            "income": 5000000,
            "expense": 0,
            "balance": 5000000
        },
        {
            "date": current_date.replace(day=2),
            "category": "Chi",
            "description": "Chi phí đổ dầu xe 51A-12345",
            "income": 0,
            "expense": 2000000,
            "balance": -2000000
        },
        {
            "date": current_date.replace(day=3),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_004",
            "income": 4500000,
            "expense": 0,
            "balance": 4500000
        },
        {
            "date": current_date.replace(day=5),
            "category": "Chi",
            "description": "Chi phí sửa chữa xe 51A-67890",
            "income": 0,
            "expense": 1500000,
            "balance": -1500000
        },
        {
            "date": current_date.replace(day=10),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_002",
            "income": 4800000,
            "expense": 0,
            "balance": 4800000
        },
        {
            "date": current_date.replace(day=12),
            "category": "Chi",
            "description": "Chi phí đổ dầu xe 51A-12345",
            "income": 0,
            "expense": 1800000,
            "balance": -1800000
        },
        {
            "date": current_date.replace(day=15),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_004",
            "income": 5200000,
            "expense": 0,
            "balance": 5200000
        },
        {
            "date": current_date.replace(day=18),
            "category": "Chi",
            "description": "Chi phí bảo hiểm xe",
            "income": 0,
            "expense": 3000000,
            "balance": -3000000
        },
        {
            "date": current_date.replace(day=20),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_002",
            "income": 4600000,
            "expense": 0,
            "balance": 4600000
        },
        {
            "date": current_date.replace(day=25),
            "category": "Chi",
            "description": "Chi phí đổ dầu xe 51A-67890",
            "income": 0,
            "expense": 2200000,
            "balance": -2200000
        },
        {
            "date": current_date.replace(day=28),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_004",
            "income": 5100000,
            "expense": 0,
            "balance": 5100000
        },
        {
            "date": current_date.replace(day=30),
            "category": "Chi",
            "description": "Chi phí lương lái xe",
            "income": 0,
            "expense": 8000000,
            "balance": -8000000
        }
    ]
    
    # Xóa dữ liệu cũ nếu có
    db.query(FinanceTransaction).delete()
    
    # Thêm dữ liệu mẫu vào bảng mới
    for data in sample_data:
        # Chuyển đổi dữ liệu từ format cũ sang format mới
        transaction = FinanceTransaction(
            transaction_type=data["category"],
            category=data["category"],
            date=data["date"],
            description=data["description"],
            route_code=data.get("route_code", ""),
            amount=data["amount_before_vat"],
            vat=data["vat_rate"],
            discount1=data["discount1_rate"],
            discount2=data["discount2_rate"],
            total=data["final_amount"],
            note=data.get("notes", ""),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        db.add(transaction)
    
    db.commit()
    
    return JSONResponse({
        "message": f"Đã tạo {len(sample_data)} bản ghi tài chính mẫu cho tháng {current_date.month}/{current_date.year}",
        "count": len(sample_data)
    })

@app.post("/finance-report/add")
async def add_finance_record(
    request: Request,
    db: Session = Depends(get_db)
):
    """Thêm bản ghi tài chính mới"""
    try:
        form_data = await request.form()
        
        # Lấy dữ liệu từ form
        date_str = form_data.get("date")
        category = form_data.get("category")
        description = form_data.get("description")
        route_code = form_data.get("route_code", "")
        
        # Xử lý các trường số, đảm bảo không bị lỗi khi chuỗi rỗng
        amount_before_vat_str = form_data.get("amount_before_vat", "0")
        vat_rate_str = form_data.get("vat_rate", "0")
        discount1_rate_str = form_data.get("discount1_rate", "0")
        discount2_rate_str = form_data.get("discount2_rate", "0")
        
        # Convert sang float, xử lý trường hợp chuỗi rỗng
        amount_before_vat = float(amount_before_vat_str) if amount_before_vat_str else 0.0
        vat_rate = float(vat_rate_str) if vat_rate_str else 0.0
        discount1_rate = float(discount1_rate_str) if discount1_rate_str else 0.0
        discount2_rate = float(discount2_rate_str) if discount2_rate_str else 0.0
        
        notes = form_data.get("notes", "")
        
        # Parse ngày
        from datetime import datetime
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        # Tính thành tiền theo công thức
        # Thành tiền = Số tiền + (Số tiền * VAT/100) - (Số tiền * CK1/100) - (Số tiền * CK2/100)
        vat_amount = amount_before_vat * (vat_rate / 100)
        discount1_amount = amount_before_vat * (discount1_rate / 100)
        discount2_amount = amount_before_vat * (discount2_rate / 100)
        final_amount = amount_before_vat + vat_amount - discount1_amount - discount2_amount
        
        # Tạo bản ghi mới trong bảng FinanceTransaction riêng biệt
        finance_transaction = FinanceTransaction(
            transaction_type=category,  # Thu/Chi
            category=category,  # Danh mục
            date=date_obj,  # Ngày thu/chi
            description=description,  # Diễn giải
            route_code=route_code,  # Mã tuyến (nếu có)
            amount=amount_before_vat,  # Số tiền chưa VAT
            vat=vat_rate,  # VAT (%)
            discount1=discount1_rate,  # Chiết khấu 1 (%)
            discount2=discount2_rate,  # Chiết khấu 2 (%)
            total=final_amount,  # Thành tiền
            note=notes,  # Ghi chú
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        
        db.add(finance_transaction)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã thêm bản ghi tài chính thành công",
            "record_id": finance_transaction.id
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi thêm bản ghi: {str(e)}"
        }, status_code=400)

@app.get("/finance-report/get/{record_id}")
async def get_finance_record(record_id: int, db: Session = Depends(get_db)):
    """Lấy thông tin bản ghi tài chính theo ID"""
    try:
        finance_record = db.query(FinanceTransaction).filter(FinanceTransaction.id == record_id).first()
        
        if not finance_record:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy bản ghi tài chính"
            }, status_code=404)
        
        return JSONResponse({
            "success": True,
            "data": {
                "id": finance_record.id,
                "transaction_type": finance_record.transaction_type,
                "date": finance_record.date.strftime("%Y-%m-%d") if finance_record.date else None,
                "description": finance_record.description,
                "route_code": finance_record.route_code,
                "amount": finance_record.amount,
                "vat": finance_record.vat,
                "discount1": finance_record.discount1,
                "discount2": finance_record.discount2,
                "total": finance_record.total,
                "note": finance_record.note
            }
        })
        
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi lấy thông tin bản ghi: {str(e)}"
        }, status_code=500)

@app.post("/finance-report/edit")
async def edit_finance_record(
    request: Request,
    db: Session = Depends(get_db)
):
    """Sửa bản ghi tài chính"""
    try:
        form_data = await request.form()
        
        # Lấy ID bản ghi cần sửa
        record_id = form_data.get("record_id")
        if not record_id:
            return JSONResponse({
                "success": False,
                "message": "Thiếu ID bản ghi"
            }, status_code=400)
        
        # Tìm bản ghi trong database
        finance_record = db.query(FinanceTransaction).filter(FinanceTransaction.id == record_id).first()
        if not finance_record:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy bản ghi tài chính"
            }, status_code=404)
        
        # Lấy dữ liệu từ form
        date_str = form_data.get("date")
        category = form_data.get("category")
        description = form_data.get("description")
        route_code = form_data.get("route_code", "")
        
        # Xử lý các trường số, đảm bảo không bị lỗi khi chuỗi rỗng
        amount_before_vat_str = form_data.get("amount_before_vat", "0")
        vat_rate_str = form_data.get("vat_rate", "0")
        discount1_rate_str = form_data.get("discount1_rate", "0")
        discount2_rate_str = form_data.get("discount2_rate", "0")
        
        # Convert sang float, xử lý trường hợp chuỗi rỗng
        amount_before_vat = float(amount_before_vat_str) if amount_before_vat_str else 0.0
        vat_rate = float(vat_rate_str) if vat_rate_str else 0.0
        discount1_rate = float(discount1_rate_str) if discount1_rate_str else 0.0
        discount2_rate = float(discount2_rate_str) if discount2_rate_str else 0.0
        
        notes = form_data.get("notes", "")
        
        # Parse ngày
        from datetime import datetime
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        # Tính thành tiền theo công thức
        vat_amount = amount_before_vat * (vat_rate / 100)
        discount1_amount = amount_before_vat * (discount1_rate / 100)
        discount2_amount = amount_before_vat * (discount2_rate / 100)
        final_amount = amount_before_vat + vat_amount - discount1_amount - discount2_amount
        
        # Cập nhật bản ghi
        finance_record.transaction_type = category
        finance_record.category = category
        finance_record.date = date_obj
        finance_record.description = description
        finance_record.route_code = route_code
        finance_record.amount = amount_before_vat
        finance_record.vat = vat_rate
        finance_record.discount1 = discount1_rate
        finance_record.discount2 = discount2_rate
        finance_record.total = final_amount
        finance_record.note = notes
        finance_record.updated_at = datetime.utcnow()
        
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã cập nhật bản ghi tài chính thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi cập nhật bản ghi: {str(e)}"
        }, status_code=400)

@app.delete("/finance-report/delete/{record_id}")
async def delete_finance_record(record_id: int, db: Session = Depends(get_db)):
    """Xóa bản ghi tài chính"""
    try:
        finance_record = db.query(FinanceTransaction).filter(FinanceTransaction.id == record_id).first()
        
        if not finance_record:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy bản ghi tài chính"
            }, status_code=404)
        
        db.delete(finance_record)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã xóa bản ghi tài chính thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi xóa bản ghi: {str(e)}"
        }, status_code=500)

# ===== FINANCIAL STATISTICS ROUTES =====

@app.get("/financial-statistics", response_class=HTMLResponse)
async def financial_statistics_page(
    request: Request,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """Trang thống kê tài chính với form tìm kiếm"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    redirect_response = check_and_redirect_access(current_user["role"], "/financial-statistics", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    # Nếu không có from_date hoặc to_date, set mặc định là đầu tháng và cuối tháng hiện tại
    if not from_date or not to_date:
        import calendar
        today = date.today()
        # Ngày đầu tháng: ngày 1 của tháng hiện tại
        first_day_of_month = date(today.year, today.month, 1)
        # Ngày cuối tháng: sử dụng calendar.monthrange để tính chính xác số ngày trong tháng
        days_in_month = calendar.monthrange(today.year, today.month)[1]
        last_day_of_month = date(today.year, today.month, days_in_month)
        
        # Set giá trị mặc định nếu chưa có
        if not from_date:
            from_date = first_day_of_month.strftime("%Y-%m-%d")
        if not to_date:
            to_date = last_day_of_month.strftime("%Y-%m-%d")
    
    # Khởi tạo kết quả tìm kiếm và các biến tính toán với giá trị mặc định
    search_results = []
    total_revenue = 0
    discount_ghn = 0
    remaining_after_ghn = 0
    discount_vo_gia = 0
    remaining_after_vo_gia = 0
    vat = 0
    final_total = 0
    
    # Nếu có tham số tìm kiếm, thực hiện tìm kiếm
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            
            # Query revenue records với các điều kiện
            query = db.query(RevenueRecord).filter(
                and_(
                    RevenueRecord.date >= from_date_obj,
                    RevenueRecord.date <= to_date_obj
                )
            )
            
            # Chỉ lấy các chuyến có status Online/ON
            query = query.filter(
                RevenueRecord.status.in_(["Online", "ON", "ONLINE"])
            )
            
            revenue_records = query.all()
            
            # Nhóm theo route_id và tính tổng doanh thu
            # Xử lý riêng cho tuyến "Tăng Cường" - tổng hợp tất cả các chuyến tăng cường
            route_revenue_dict = {}
            tang_cuong_revenue = 0
            tang_cuong_notes = []
            
            for record in revenue_records:
                route = db.query(Route).filter(Route.id == record.route_id).first()
                route_code = route.route_code if route else "N/A"
                
                # Tính doanh thu: ưu tiên manual_total, nếu không có thì dùng total_amount
                revenue_amount = record.manual_total if record.manual_total > 0 else record.total_amount
                
                # Xử lý riêng cho tuyến "Tăng Cường" (so sánh không phân biệt hoa thường)
                if route_code and route_code.strip().upper().replace(" ", "") == "TĂNGCƯỜNG":
                    tang_cuong_revenue += revenue_amount
                    if record.notes:
                        tang_cuong_notes.append(record.notes)
                else:
                    # Các tuyến khác: nhóm theo route_id
                    route_id = record.route_id
                    if route_id not in route_revenue_dict:
                        route_revenue_dict[route_id] = {
                            "route_code": route_code,
                            "revenue": 0,
                            "notes": []
                        }
                    
                    route_revenue_dict[route_id]["revenue"] += revenue_amount
                    
                    # Thêm ghi chú nếu có
                    if record.notes:
                        route_revenue_dict[route_id]["notes"].append(record.notes)
            
            # Chuyển đổi thành danh sách để hiển thị
            search_results = []
            
            # Thêm các tuyến thường (không phải Tăng Cường)
            for route_id, data in route_revenue_dict.items():
                search_results.append({
                    "route_code": data["route_code"],
                    "revenue": data["revenue"],
                    "notes": "<br>".join(set(data["notes"])) if data["notes"] else ""
                })
                total_revenue += data["revenue"]
            
            # Sắp xếp các tuyến thường theo mã tuyến
            search_results.sort(key=lambda x: x["route_code"])
            
            # Thêm tuyến "Tăng Cường" vào cuối nếu có doanh thu
            if tang_cuong_revenue > 0:
                search_results.append({
                    "route_code": "TĂNG CƯỜNG",
                    "revenue": tang_cuong_revenue,
                    "notes": "<br>".join(set(tang_cuong_notes)) if tang_cuong_notes else ""
                })
                total_revenue += tang_cuong_revenue
            
        except Exception as e:
            print(f"Error in financial statistics search: {e}")
            import traceback
            traceback.print_exc()
    
    # Tính toán các khoản chiết khấu và VAT
    discount_ghn = total_revenue * 0.05  # 5%
    remaining_after_ghn = total_revenue - discount_ghn
    discount_vo_gia = remaining_after_ghn * 0.05  # 5%
    remaining_after_vo_gia = remaining_after_ghn - discount_vo_gia
    vat = remaining_after_vo_gia * 0.08  # 8%
    final_total = remaining_after_vo_gia + vat
    
    return templates.TemplateResponse("financial_statistics.html", {
        "request": request,
        "current_user": current_user,
        "from_date": from_date or "",
        "to_date": to_date or "",
        "search_results": search_results,
        "total_revenue": total_revenue,
        "discount_ghn": discount_ghn,
        "discount_vo_gia": discount_vo_gia,
        "remaining_after_ghn": remaining_after_ghn,
        "remaining_after_vo_gia": remaining_after_vo_gia,
        "vat": vat,
        "final_total": final_total
    })

@app.get("/api/financial-statistics/details")
async def financial_statistics_details(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    route_code: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    filter_route_code: Optional[str] = None,
    driver_name: Optional[str] = None
):
    """API để lấy chi tiết các tuyến trong thống kê tài chính"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return JSONResponse(
            status_code=401,
            content={"success": False, "error": "Chưa đăng nhập"}
        )
    
    if not check_page_access(current_user["role"], "/financial-statistics"):
        return JSONResponse(
            status_code=403,
            content={"success": False, "error": "Không có quyền truy cập"}
        )
    
    if not from_date or not to_date or not route_code:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Thiếu tham số bắt buộc"}
        )
    
    try:
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
        
        # Query revenue records với các điều kiện
        query = db.query(RevenueRecord).filter(
            and_(
                RevenueRecord.date >= from_date_obj,
                RevenueRecord.date <= to_date_obj
            )
        )
        
        # Lọc theo route_code được chọn (từ button Chi tiết)
        # Xử lý đặc biệt cho tuyến "Tăng Cường"
        route_code_normalized = route_code.strip().upper().replace(" ", "")
        if route_code_normalized == "TĂNGCƯỜNG" or route_code_normalized == "TANGCUONG":
            # Lấy tất cả các record có route_code là "Tăng Cường"
            # Tìm route có route_code là "Tăng Cường" (không phân biệt hoa thường, bỏ dấu cách)
            all_routes = db.query(Route).all()
            tang_cuong_route = None
            for r in all_routes:
                if r.route_code:
                    r_code_normalized = r.route_code.strip().upper().replace(" ", "")
                    if r_code_normalized == "TĂNGCƯỜNG" or r_code_normalized == "TANGCUONG":
                        tang_cuong_route = r
                        break
            
            if tang_cuong_route:
                query = query.filter(RevenueRecord.route_id == tang_cuong_route.id)
            else:
                return JSONResponse(
                    status_code=200,
                    content={"success": True, "details": []}
                )
        else:
            # Lấy record theo route_code cụ thể
            route = db.query(Route).filter(Route.route_code == route_code.strip()).first()
            if route:
                query = query.filter(RevenueRecord.route_id == route.id)
            else:
                return JSONResponse(
                    status_code=200,
                    content={"success": True, "details": []}
                )
        
        # Chỉ lấy các chuyến có status Online/ON
        query = query.filter(
            RevenueRecord.status.in_(["Online", "ON", "ONLINE"])
        )
        
        # Join với Route để đảm bảo relationship được load
        query = query.join(Route, RevenueRecord.route_id == Route.id)
        
        # Sắp xếp theo ngày và route_code
        revenue_records = query.order_by(RevenueRecord.date, RevenueRecord.route_id).all()
        
        # Debug: In ra số lượng records tìm được
        print(f"DEBUG: Found {len(revenue_records)} revenue records for route_code={route_code}, from_date={from_date}, to_date={to_date}")
        
        # Chuyển đổi sang dictionary
        details = []
        for record in revenue_records:
            # Lấy route từ relationship (đã được join)
            route = record.route
            if not route:
                # Nếu không có route, thử query lại
                route = db.query(Route).filter(Route.id == record.route_id).first()
            
            route_code_val = route.route_code if route else "N/A"
            route_name_full = route.route_name if route else "N/A"
            
            # Lộ trình: ưu tiên record.route_name (cho tuyến tăng cường), nếu không có thì dùng route.route_name
            route_name = record.route_name if record.route_name else route_name_full
            
            # Lấy thành tiền từ RevenueRecord: ưu tiên manual_total, nếu không có thì dùng total_amount
            # Đảm bảo đồng bộ với logic trong financial_statistics_page
            total_amount = record.manual_total if record.manual_total > 0 else record.total_amount
            
            details.append({
                'date': record.date.strftime('%Y-%m-%d') if record.date else None,
                'route_code': route_code_val,
                'route_name': route_name,
                'route_name_full': route_name_full,
                'distance_km': float(record.distance_km or 0),
                'unit_price': int(record.unit_price or 0),
                'bridge_fee': int(record.bridge_fee or 0),
                'loading_fee': int(record.loading_fee or 0),
                'late_penalty': int(record.late_penalty or 0),
                'total_amount': float(total_amount or 0),  # Thành tiền từ RevenueRecord
                'driver_name': record.driver_name or '',
                'notes': record.notes or ''
            })
        
        # Debug: In ra số lượng details
        print(f"DEBUG: Returning {len(details)} details")
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "details": details
            }
        )
        
    except Exception as e:
        print(f"Error in financial statistics details API: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e), "message": f"Lỗi khi lấy dữ liệu chi tiết: {str(e)}"}
        )

# ==================== TIMEKEEPING V1 ROUTES ====================

@app.get("/timekeeping-v1", response_class=HTMLResponse)
async def timekeeping_v1_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Trang Bảng chấm công V1"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    redirect_response = check_and_redirect_access(current_user["role"], "/timekeeping-v1", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    # Lấy danh sách các bảng chấm công đã tạo
    timekeeping_tables = db.query(TimekeepingTable).order_by(TimekeepingTable.created_at.desc()).all()
    
    # Chuyển đổi sang dictionary để có thể serialize JSON
    timekeeping_tables_data = []
    for table in timekeeping_tables:
        timekeeping_tables_data.append({
            "id": table.id,
            "name": table.name,
            "from_date": table.from_date.isoformat() if table.from_date else None,
            "to_date": table.to_date.isoformat() if table.to_date else None,
            "created_at": table.created_at.isoformat() if table.created_at else None
        })
    
    return templates.TemplateResponse("timekeeping_v1.html", {
        "request": request,
        "current_user": current_user,
        "timekeeping_tables": timekeeping_tables,
        "timekeeping_tables_data": timekeeping_tables_data
    })

@app.post("/timekeeping-v1/create")
async def create_timekeeping_table(
    request: Request,
    db: Session = Depends(get_db),
    name: str = Form(...),
    from_date: str = Form(...),
    to_date: str = Form(...),
    current_user = Depends(get_current_user)
):
    """Tạo bảng chấm công mới"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return JSONResponse({
            "success": False,
            "message": "Bạn cần đăng nhập để thực hiện thao tác này"
        }, status_code=401)
    
    if not check_page_access(current_user["role"], "/timekeeping-v1"):
        return JSONResponse({
            "success": False,
            "message": "Bạn không có quyền truy cập"
        }, status_code=403)
    
    try:
        # Validate dates
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
        
        if from_date_obj > to_date_obj:
            return JSONResponse({
                "success": False,
                "message": "Ngày bắt đầu phải nhỏ hơn hoặc bằng ngày kết thúc"
            }, status_code=400)
        
        # Validate name
        if not name or not name.strip():
            return JSONResponse({
                "success": False,
                "message": "Tên bảng chấm công là bắt buộc"
            }, status_code=400)
        
        # Tạo bảng chấm công mới
        new_table = TimekeepingTable(
            name=name.strip(),
            from_date=from_date_obj,
            to_date=to_date_obj
        )
        
        db.add(new_table)
        db.commit()
        db.refresh(new_table)
        
        return JSONResponse({
            "success": True,
            "message": "Tạo bảng chấm công thành công",
            "data": {
                "id": new_table.id,
                "name": new_table.name,
                "from_date": new_table.from_date.strftime("%Y-%m-%d"),
                "to_date": new_table.to_date.strftime("%Y-%m-%d"),
                "created_at": new_table.created_at.strftime("%Y-%m-%d %H:%M:%S")
            }
        })
        
    except ValueError as e:
        return JSONResponse({
            "success": False,
            "message": f"Định dạng ngày không hợp lệ: {str(e)}"
        }, status_code=400)
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi tạo bảng chấm công: {str(e)}"
        }, status_code=500)

@app.get("/timekeeping-v1/detail/{table_id}", response_class=HTMLResponse)
async def timekeeping_v1_detail_page(
    request: Request,
    table_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Trang chi tiết bảng chấm công"""
    # Kiểm tra quyền truy cập
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    redirect_response = check_and_redirect_access(current_user["role"], "/timekeeping-v1", current_user["id"], db)
    if redirect_response:
        return redirect_response
    
    # Lấy thông tin bảng chấm công
    timekeeping_table = db.query(TimekeepingTable).filter(TimekeepingTable.id == table_id).first()
    
    if not timekeeping_table:
        return RedirectResponse(url="/timekeeping-v1", status_code=303)
    
    # Lấy dữ liệu từ các bảng hiện có
    # Lấy tất cả employees để có thể bao gồm các lái xe đã gán (cho dữ liệu lịch sử)
    all_employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    # Chỉ lấy routes có route_status = "ONL" khi tạo bảng chấm công mới
    routes = db.query(Route).filter(
        Route.is_active == 1, 
        Route.status == 1,
        Route.route_status == "ONL"
    ).all()

    # Tính dải ngày theo khoảng đã chọn
    date_range = []
    current_date = timekeeping_table.from_date
    while current_date <= timekeeping_table.to_date:
        date_range.append(current_date.isoformat())
        current_date += timedelta(days=1)

    # Ngày hiệu lực giá mới: 18/12/2025
    new_price_effective_date = date(2025, 12, 18)
    
    # Chuyển routes sang dict + thêm sheet Tăng Cường nếu chưa có
    routes_data = []
    route_prices_by_date = {}  # Dictionary để lưu giá theo route_id và ngày
    
    for route in routes:
        # Lấy giá từ RoutePrice theo ngày hiệu lực (từ 18/12/2025)
        # Lấy giá mới nhất có application_date >= 18/12/2025
        route_price = db.query(RoutePrice).filter(
            RoutePrice.route_id == route.id,
            RoutePrice.application_date >= new_price_effective_date
        ).order_by(RoutePrice.application_date.desc()).first()
        
        # Nếu có giá trong RoutePrice, sử dụng giá đó; nếu không, fallback về giá từ Route
        unit_price = route_price.unit_price if route_price else (route.unit_price or 0)
        
        # Lưu giá theo route_id để frontend có thể sử dụng
        route_prices_by_date[route.id] = unit_price
        
        routes_data.append({
            "route_code": route.route_code or "",
            "route_name": route.route_name or "",
            "route_type": route.route_type or "",
            "distance": route.distance or 0,
            "unit_price": unit_price,
            "bridge_fee": route.bridge_fee or 0,
            "loading_fee": route.loading_fee or 0,
            "route_id": route.id  # Thêm route_id để frontend có thể map
        })

    has_tang_cuong = any(
        (r.get("route_code") or "").strip().lower() == "tăng cường"
        or (r.get("route_name") or "").strip().lower() == "tăng cường"
        for r in routes_data
    )
    if not has_tang_cuong:
        routes_data.append({
            "route_code": "TĂNG CƯỜNG",
            "route_name": "TĂNG CƯỜNG",
            "route_type": "Tăng cường",
            "distance": 0,
            "unit_price": 0,
            "bridge_fee": 0,
            "loading_fee": 0
        })

    routes_data = sorted(
        routes_data,
        key=lambda r: (r.get("route_code") or r.get("route_name") or "").lower()
    )

    # Dữ liệu đã lưu - cần lấy trước để biết các lái xe đã gán
    saved_details = db.query(TimekeepingDetail).filter(TimekeepingDetail.table_id == table_id).all()
    
    # Lấy danh sách tên lái xe đã được gán trong dữ liệu đã lưu (để giữ lại trong dropdown)
    assigned_driver_names = set()
    for detail in saved_details:
        if detail.driver_name:
            assigned_driver_names.add(detail.driver_name.strip())
    
    # Dropdown data - chỉ lấy employees có trạng thái "Đang làm việc"
    # Nhưng cũng bao gồm các employees đã được gán (để giữ dữ liệu lịch sử)
    employees_data = []
    matched_driver_names = set()  # Track which assigned names have been matched to employees
    
    for emp in all_employees:
        if emp.status == 1:
            # Chỉ thêm vào dropdown nếu:
            # 1. employee_status == "Đang làm việc" (cho phép chọn mới)
            # 2. HOẶC tên đã được gán trong dữ liệu đã lưu (giữ dữ liệu lịch sử)
            emp_name = emp.name or ""
            is_active = (emp.employee_status or "Đang làm việc") == "Đang làm việc"
            is_assigned = emp_name.strip() in assigned_driver_names
            
            if is_active or is_assigned:
                if is_assigned:
                    matched_driver_names.add(emp_name.strip())
                employees_data.append({
                    "id": emp.id, 
                    "name": emp_name,
                    "employee_status": emp.employee_status or "Đang làm việc",
                    "is_active": is_active  # Flag để frontend biết có thể chọn hay không
                })
    
    # Thêm các tên lái xe đã gán nhưng không khớp với employee nào (fallback cho dữ liệu lịch sử)
    for driver_name in assigned_driver_names:
        if driver_name not in matched_driver_names:
            # Tên này không khớp với employee nào, thêm vào như một option disabled
            employees_data.append({
                "id": None,
                "name": driver_name,
                "employee_status": "Không còn trong hệ thống",
                "is_active": False  # Không cho phép chọn mới
            })
    vehicles_data = [{"id": veh.id, "license_plate": veh.license_plate or ""} for veh in vehicles if veh.status == 1]

    # Dữ liệu đã lưu (đã lấy ở trên)
    details_by_sheet = {}
    for detail in saved_details:
        sheet_key = detail.sheet_name or detail.route_code or detail.route_name or "TĂNG CƯỜNG"
        if sheet_key not in details_by_sheet:
            details_by_sheet[sheet_key] = []
        details_by_sheet[sheet_key].append({
            "id": detail.id,
            "sheet_name": sheet_key,
            "route_code": detail.route_code or "",
            "route_name": detail.route_name or "",
            "route_type": detail.route_type or "",
            "itinerary": detail.itinerary or "",
            "date": detail.date.isoformat() if detail.date else "",
            "license_plate": detail.license_plate or "",
            "driver_name": detail.driver_name or "",
            "trip_code": detail.trip_code or "",
            "notes": detail.notes or "",
            "status": detail.status or "Onl",
            "distance_km": detail.distance_km or 0,
            "unit_price": detail.unit_price or 0,
            "bridge_fee": detail.bridge_fee or 0,
            "loading_fee": detail.loading_fee or 0,
            "total_amount": detail.total_amount or 0
        })

    for sheet_key, items in details_by_sheet.items():
        details_by_sheet[sheet_key] = sorted(items, key=lambda x: x.get("date") or "")

    return templates.TemplateResponse("timekeeping_v1_detail.html", {
        "request": request,
        "current_user": current_user,
        "timekeeping_table": timekeeping_table,
        "employees": employees_data,
        "vehicles": vehicles_data,
        "routes": routes_data,
        "date_range": date_range,
        "timekeeping_details": details_by_sheet
    })


@app.post("/api/timekeeping-v1/{table_id}/save")
async def save_timekeeping_detail(
    table_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Lưu dữ liệu chấm công chi tiết"""
    if current_user is None:
        return JSONResponse({"success": False, "message": "Bạn cần đăng nhập"}, status_code=401)
    if not check_page_access(current_user["role"], "/timekeeping-v1", current_user["id"], db):
        return JSONResponse({"success": False, "message": "Không có quyền truy cập"}, status_code=403)

    table = db.query(TimekeepingTable).filter(TimekeepingTable.id == table_id).first()
    if not table:
        return JSONResponse({"success": False, "message": "Không tìm thấy bảng chấm công"}, status_code=404)

    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"success": False, "message": "Payload không hợp lệ"}, status_code=400)

    scope = payload.get("scope", "sheet")
    sheet_name = payload.get("sheet_name") or ""
    entries = payload.get("entries", [])

    def parse_date_safe(date_str: str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return None

    # Xóa dữ liệu cũ theo phạm vi
    try:
        if scope == "all":
            db.query(TimekeepingDetail).filter(TimekeepingDetail.table_id == table_id).delete()
        else:
            db.query(TimekeepingDetail).filter(
                TimekeepingDetail.table_id == table_id,
                TimekeepingDetail.sheet_name == sheet_name
            ).delete()
        db.commit()
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": f"Lỗi khi xóa dữ liệu cũ: {e}"}, status_code=500)

    # Lưu mới
    try:
        records_to_add = []
        for entry in entries:
            entry_date = parse_date_safe(entry.get("date"))
            if not entry_date:
                continue
            if entry_date < table.from_date or entry_date > table.to_date:
                continue

            # Lấy status, mặc định là "Onl"
            entry_status = entry.get("status") or "Onl"
            
            # Nếu status là OFF, đảm bảo total_amount = 0
            entry_total = float(entry.get("total_amount") or 0)
            if entry_status == "OFF":
                entry_total = 0

            detail = TimekeepingDetail(
                table_id=table_id,
                sheet_name=entry.get("sheet_name") or sheet_name or entry.get("route_code") or entry.get("route_name") or "",
                route_code=entry.get("route_code") or "",
                route_name=entry.get("route_name") or "",
                route_type=entry.get("route_type") or "",
                itinerary=entry.get("itinerary") or "",
                date=entry_date,
                license_plate=entry.get("license_plate") or "",
                driver_name=entry.get("driver_name") or "",
                trip_code=entry.get("trip_code") or "",
                notes=entry.get("notes") or "",
                status=entry_status,
                distance_km=float(entry.get("distance_km") or 0),
                unit_price=float(entry.get("unit_price") or 0),
                bridge_fee=float(entry.get("bridge_fee") or 0),
                loading_fee=float(entry.get("loading_fee") or 0),
                total_amount=entry_total
            )
            records_to_add.append(detail)

        if records_to_add:
            db.bulk_save_objects(records_to_add)
        db.commit()
        return JSONResponse({"success": True, "message": "Lưu dữ liệu thành công"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": f"Lỗi khi lưu dữ liệu: {e}"}, status_code=500)

@app.get("/api/timekeeping-v1/{table_id}/export-excel")
async def export_timekeeping_excel(
    table_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Xuất bảng chấm công ra file Excel"""
    if current_user is None:
        return JSONResponse({"success": False, "message": "Bạn cần đăng nhập"}, status_code=401)
    if not check_page_access(current_user["role"], "/timekeeping-v1", current_user["id"], db):
        return JSONResponse({"success": False, "message": "Không có quyền truy cập"}, status_code=403)
    
    # Lấy thông tin bảng chấm công
    table = db.query(TimekeepingTable).filter(TimekeepingTable.id == table_id).first()
    if not table:
        return JSONResponse({"success": False, "message": "Không tìm thấy bảng chấm công"}, status_code=404)
    
    # Lấy tất cả dữ liệu chi tiết, sắp xếp theo sheet_name và date
    details = db.query(TimekeepingDetail).filter(
        TimekeepingDetail.table_id == table_id
    ).order_by(TimekeepingDetail.sheet_name, TimekeepingDetail.date).all()
    
    # Hàm sanitize filename - di chuyển lên đây để dùng ở nhiều nơi
    def sanitize_filename(text):
        """Loại bỏ ký tự đặc biệt khỏi tên file, chỉ trả về ASCII"""
        if not text:
            return ""
        # Chuyển đổi ký tự có dấu thành không dấu
        text = unicodedata.normalize('NFKD', str(text))
        text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
        # Chỉ giữ lại chữ cái, số, khoảng trắng, dấu gạch - và đảm bảo là ASCII
        text = ''.join(c if (c.isalnum() or c in (' ', '-', '_')) and ord(c) < 128 else '_' for c in text)
        # Loại bỏ khoảng trắng ở đầu và cuối, thay bằng dấu gạch dưới
        text = text.strip().replace(' ', '_')
        return text[:30] if text else "file"  # Giới hạn độ dài
    
    # Tạo workbook Excel
    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định
    
    # Chuẩn bị thông tin ngày tháng
    from_date_str = table.from_date.strftime('%d/%m/%Y')
    to_date_str = table.to_date.strftime('%d/%m/%Y')
    
    # Nhóm dữ liệu theo sheet_name
    details_by_sheet = {}
    for detail in details:
        sheet_name = detail.sheet_name or "TĂNG CƯỜNG"
        if sheet_name not in details_by_sheet:
            details_by_sheet[sheet_name] = []
        details_by_sheet[sheet_name].append(detail)
    
    # Tạo sheet cho mỗi tuyến
    for sheet_name, sheet_details in details_by_sheet.items():
        # Sanitize tên sheet để tránh lỗi với ký tự đặc biệt
        safe_sheet_name = sheet_name[:31]  # Excel giới hạn 31 ký tự cho tên sheet
        # Loại bỏ ký tự không hợp lệ cho tên sheet Excel
        safe_sheet_name = ''.join(c for c in safe_sheet_name if c not in ['\\', '/', '?', '*', '[', ']', ':'])
        if not safe_sheet_name:
            safe_sheet_name = "Sheet"
        ws = wb.create_sheet(title=safe_sheet_name)
        
        # Định dạng header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Tiêu đề bảng chấm công
        ws.merge_cells('A1:O1')
        ws['A1'] = f"BẢNG CHẤM CÔNG - {table.name.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Thông tin thời gian
        ws.merge_cells('A2:O2')
        ws['A2'] = f"Từ ngày: {from_date_str} - Đến ngày: {to_date_str}"
        ws['A2'].font = Font(size=11)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Header row
        headers = [
            "STT", "Ngày", "Biển số", "Lái xe", "Mã chuyến", "Ghi chú", 
            "Trạng thái", "Km", "Đơn giá", "Phí cầu", "Phí bốc", "Tổng tiền"
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Dữ liệu
        row = 5
        for idx, detail in enumerate(sheet_details, 1):
            date_str = detail.date.strftime('%d/%m/%Y') if detail.date else ""
            
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=detail.license_plate or "")
            ws.cell(row=row, column=4, value=detail.driver_name or "")
            ws.cell(row=row, column=5, value=detail.trip_code or "")
            ws.cell(row=row, column=6, value=detail.notes or "")
            ws.cell(row=row, column=7, value=detail.status or "Onl").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=8, value=detail.distance_km or 0).number_format = '#,##0.000'
            ws.cell(row=row, column=9, value=detail.unit_price or 0).number_format = '#,##0'
            ws.cell(row=row, column=10, value=detail.bridge_fee or 0).number_format = '#,##0'
            ws.cell(row=row, column=11, value=detail.loading_fee or 0).number_format = '#,##0'
            ws.cell(row=row, column=12, value=detail.total_amount or 0).number_format = '#,##0'
            
            # Thêm border cho tất cả các ô
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = border_style
            
            row += 1
        
        # Dòng tổng cộng
        if sheet_details:
            total_row = row
            ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
            for col in range(2, 8):
                ws.cell(row=total_row, column=col, value="").font = Font(bold=True)
            
            total_distance = sum(d.distance_km or 0 for d in sheet_details)
            total_amount = sum(d.total_amount or 0 for d in sheet_details)
            
            ws.cell(row=total_row, column=8, value=total_distance).font = Font(bold=True)
            ws.cell(row=total_row, column=8).number_format = '#,##0.000'
            ws.cell(row=total_row, column=9, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=10, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=11, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=12, value=total_amount).font = Font(bold=True)
            ws.cell(row=total_row, column=12).number_format = '#,##0'
            
            # Thêm border cho dòng tổng cộng
            for col in range(1, 13):
                ws.cell(row=total_row, column=col).border = border_style
        
        # Điều chỉnh độ rộng cột
        column_widths = [6, 12, 12, 20, 12, 20, 10, 10, 12, 12, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Đặt chiều cao cho header
        ws.row_dimensions[4].height = 30
    
    # Nếu không có dữ liệu, tạo một sheet trống
    if not details_by_sheet:
        ws = wb.create_sheet(title="DuLieu")
        # Đảm bảo text trong cell không gây lỗi encoding
        safe_table_name = sanitize_filename(table.name) or "BANG CHAM CONG"
        ws['A1'] = f"BANG CHAM CONG - {safe_table_name.upper()}"
        ws['A2'] = f"Tu ngay: {from_date_str} - Den ngay: {to_date_str}"
        ws['A3'] = "Chua co du lieu"
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file - chỉ sử dụng ASCII để tránh lỗi encoding
    safe_name = sanitize_filename(table.name) or "BangChamCong"
    filename = f"BangChamCong_{safe_name}_{table.from_date.strftime('%Y%m%d')}.xlsx"
    # Đảm bảo filename chỉ chứa ASCII
    filename = filename.encode('ascii', 'ignore').decode('ascii')
    
    # Tạo header Content-Disposition - sử dụng quote để encode an toàn
    encoded_filename = quote(filename, safe='-_.')
    content_disposition = f"attachment; filename*=UTF-8''{encoded_filename}"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition}
    )

@app.delete("/api/timekeeping-v1/{table_id}/delete")
async def delete_timekeeping_table(
    table_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Xóa bảng chấm công và tất cả dữ liệu liên quan"""
    if current_user is None:
        return JSONResponse({"success": False, "message": "Bạn cần đăng nhập"}, status_code=401)
    if not check_page_access(current_user["role"], "/timekeeping-v1", current_user["id"], db):
        return JSONResponse({"success": False, "message": "Không có quyền truy cập"}, status_code=403)
    
    # Lấy thông tin bảng chấm công
    table = db.query(TimekeepingTable).filter(TimekeepingTable.id == table_id).first()
    if not table:
        return JSONResponse({"success": False, "message": "Không tìm thấy bảng chấm công"}, status_code=404)
    
    try:
        # Xóa tất cả dữ liệu chi tiết trước
        db.query(TimekeepingDetail).filter(TimekeepingDetail.table_id == table_id).delete()
        
        # Xóa bảng chấm công
        db.delete(table)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Xóa bảng chấm công thành công"
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi xóa bảng chấm công: {str(e)}"
        }, status_code=500)

@app.get("/api/timekeeping-v1/{table_id}/filter")
async def filter_timekeeping_data(
    table_id: int,
    db: Session = Depends(get_db),
    driver_name: Optional[str] = None,
    route_code: Optional[str] = None,
    license_plate: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """Lọc dữ liệu chấm công theo các điều kiện"""
    if current_user is None:
        return JSONResponse({"success": False, "message": "Bạn cần đăng nhập"}, status_code=401)
    if not check_page_access(current_user["role"], "/timekeeping-v1", current_user["id"], db):
        return JSONResponse({"success": False, "message": "Không có quyền truy cập"}, status_code=403)
    
    # Lấy thông tin bảng chấm công
    table = db.query(TimekeepingTable).filter(TimekeepingTable.id == table_id).first()
    if not table:
        return JSONResponse({"success": False, "message": "Không tìm thấy bảng chấm công"}, status_code=404)
    
    try:
        # Xây dựng query filter
        query = db.query(TimekeepingDetail).filter(TimekeepingDetail.table_id == table_id)
        
        if driver_name:
            query = query.filter(TimekeepingDetail.driver_name == driver_name)
        
        if route_code:
            query = query.filter(TimekeepingDetail.route_code == route_code)
        
        if license_plate:
            query = query.filter(TimekeepingDetail.license_plate == license_plate)
        
        # Lấy dữ liệu và sắp xếp theo route_code
        details = query.order_by(TimekeepingDetail.route_code, TimekeepingDetail.date).all()
        
        # Chuyển đổi sang dictionary
        result_data = []
        for detail in details:
            result_data.append({
                "sheet_name": detail.sheet_name or "",
                "route_code": detail.route_code or "",
                "route_name": detail.route_name or "",
                "route_type": detail.route_type or "",
                "itinerary": detail.itinerary or "",
                "date": detail.date.isoformat() if detail.date else "",
                "license_plate": detail.license_plate or "",
                "driver_name": detail.driver_name or "",
                "trip_code": detail.trip_code or "",
                "notes": detail.notes or "",
                "status": detail.status or "Onl",
                "distance_km": detail.distance_km or 0,
                "unit_price": detail.unit_price or 0,
                "bridge_fee": detail.bridge_fee or 0,
                "loading_fee": detail.loading_fee or 0,
                "total_amount": detail.total_amount or 0
            })
        
        return JSONResponse({
            "success": True,
            "data": result_data,
            "count": len(result_data)
        })
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi lọc dữ liệu: {str(e)}"
        }, status_code=500)

@app.get("/api/timekeeping-v1/{table_id}/export-filtered-excel")
async def export_filtered_timekeeping_excel(
    table_id: int,
    db: Session = Depends(get_db),
    driver_name: Optional[str] = None,
    route_code: Optional[str] = None,
    license_plate: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    """Xuất Excel dữ liệu chấm công đã được lọc"""
    if current_user is None:
        return JSONResponse({"success": False, "message": "Bạn cần đăng nhập"}, status_code=401)
    if not check_page_access(current_user["role"], "/timekeeping-v1", current_user["id"], db):
        return JSONResponse({"success": False, "message": "Không có quyền truy cập"}, status_code=403)
    
    # Lấy thông tin bảng chấm công
    table = db.query(TimekeepingTable).filter(TimekeepingTable.id == table_id).first()
    if not table:
        return JSONResponse({"success": False, "message": "Không tìm thấy bảng chấm công"}, status_code=404)
    
    try:
        # Xây dựng query filter (giống như endpoint filter)
        query = db.query(TimekeepingDetail).filter(TimekeepingDetail.table_id == table_id)
        
        if driver_name:
            query = query.filter(TimekeepingDetail.driver_name == driver_name)
        
        if route_code:
            query = query.filter(TimekeepingDetail.route_code == route_code)
        
        if license_plate:
            query = query.filter(TimekeepingDetail.license_plate == license_plate)
        
        # Lấy dữ liệu và sắp xếp theo route_code
        details = query.order_by(TimekeepingDetail.route_code, TimekeepingDetail.date).all()
        
        # Tạo workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Kết quả lọc"
        
        # Định dạng header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Tiêu đề bảng chấm công
        ws.merge_cells('A1:N1')
        ws['A1'] = f"BẢNG CHẤM CÔNG - {table.name.upper()} (Đã lọc)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Thông tin thời gian và điều kiện lọc
        from_date_str = table.from_date.strftime('%d/%m/%Y')
        to_date_str = table.to_date.strftime('%d/%m/%Y')
        filter_conditions = []
        if driver_name:
            filter_conditions.append(f"Lái xe: {driver_name}")
        if route_code:
            filter_conditions.append(f"Mã tuyến: {route_code}")
        if license_plate:
            filter_conditions.append(f"Biển số: {license_plate}")
        
        ws.merge_cells('A2:N2')
        filter_text = f"Từ ngày: {from_date_str} - Đến ngày: {to_date_str}"
        if filter_conditions:
            filter_text += f" | Điều kiện: {', '.join(filter_conditions)}"
        ws['A2'] = filter_text
        ws['A2'].font = Font(size=11)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Header row
        headers = [
            "STT", "Ngày", "Biển số xe", "Mã tuyến", "Status", "Lộ trình",
            "Km chuyến", "Đơn giá", "Phí cầu đường", "Phí chờ tải",
            "Thành tiền", "Lái xe", "Mã chuyến", "Ghi chú"
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Dữ liệu
        row = 5
        for idx, detail in enumerate(details, 1):
            date_str = detail.date.strftime('%d/%m/%Y') if detail.date else ""
            
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=detail.license_plate or "")
            ws.cell(row=row, column=4, value=detail.route_code or "")
            ws.cell(row=row, column=5, value=detail.status or "Onl").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=6, value=detail.itinerary or "")
            ws.cell(row=row, column=7, value=detail.distance_km or 0).number_format = '#,##0.000'
            ws.cell(row=row, column=8, value=detail.unit_price or 0).number_format = '#,##0'
            ws.cell(row=row, column=9, value=detail.bridge_fee or 0).number_format = '#,##0'
            ws.cell(row=row, column=10, value=detail.loading_fee or 0).number_format = '#,##0'
            ws.cell(row=row, column=11, value=detail.total_amount or 0).number_format = '#,##0'
            ws.cell(row=row, column=12, value=detail.driver_name or "")
            ws.cell(row=row, column=13, value=detail.trip_code or "")
            ws.cell(row=row, column=14, value=detail.notes or "")
            
            # Thêm border cho tất cả các ô
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = border_style
            
            row += 1
        
        # Dòng tổng cộng
        if details:
            total_row = row
            ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
            for col in range(2, 7):
                ws.cell(row=total_row, column=col, value="").font = Font(bold=True)
            
            total_distance = sum(d.distance_km or 0 for d in details)
            total_amount = sum(d.total_amount or 0 for d in details)
            
            ws.cell(row=total_row, column=7, value=total_distance).font = Font(bold=True)
            ws.cell(row=total_row, column=7).number_format = '#,##0.000'
            ws.cell(row=total_row, column=8, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=9, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=10, value="").font = Font(bold=True)
            ws.cell(row=total_row, column=11, value=total_amount).font = Font(bold=True)
            ws.cell(row=total_row, column=11).number_format = '#,##0'
            for col in range(12, 15):
                ws.cell(row=total_row, column=col, value="").font = Font(bold=True)
            
            # Thêm border cho dòng tổng cộng
            for col in range(1, 15):
                ws.cell(row=total_row, column=col).border = border_style
        
        # Điều chỉnh độ rộng cột
        column_widths = [6, 12, 12, 12, 10, 25, 12, 12, 12, 12, 15, 20, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Đặt chiều cao cho header
        ws.row_dimensions[4].height = 30
        
        # Lưu vào memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tạo tên file - chỉ sử dụng ASCII để tránh lỗi encoding
        def sanitize_filename(text):
            """Loại bỏ ký tự đặc biệt khỏi tên file, chỉ trả về ASCII"""
            if not text:
                return ""
            # Chuyển đổi ký tự có dấu thành không dấu
            text = unicodedata.normalize('NFKD', str(text))
            text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
            # Chỉ giữ lại chữ cái, số, khoảng trắng, dấu gạch - và đảm bảo là ASCII
            text = ''.join(c if (c.isalnum() or c in (' ', '-', '_')) and ord(c) < 128 else '_' for c in text)
            # Loại bỏ khoảng trắng ở đầu và cuối, thay bằng dấu gạch dưới
            text = text.strip().replace(' ', '_')
            return text[:30] if text else "file"  # Giới hạn độ dài
        
        safe_name = sanitize_filename(table.name) or "BangChamCong"
        filter_suffix = ""
        if driver_name:
            safe_driver = sanitize_filename(driver_name)
            if safe_driver:
                filter_suffix += f"_LaiXe_{safe_driver}"
        if route_code:
            safe_route = sanitize_filename(route_code)
            if safe_route:
                filter_suffix += f"_Tuyen_{safe_route}"
        if license_plate:
            safe_plate = sanitize_filename(license_plate)
            if safe_plate:
                filter_suffix += f"_BienSo_{safe_plate}"
        
        # Đảm bảo filename chỉ chứa ASCII
        filename = f"BangChamCong_Loc_{safe_name}{filter_suffix}_{table.from_date.strftime('%Y%m%d')}.xlsx"
        # Kiểm tra và đảm bảo filename chỉ chứa ASCII
        filename = filename.encode('ascii', 'ignore').decode('ascii')
        
        # Tạo header Content-Disposition - đơn giản như các endpoint khác
        # Sử dụng quote để encode filename an toàn
        encoded_filename = quote(filename, safe='-_.')
        content_disposition = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": content_disposition}
        )
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi xuất Excel: {str(e)}"
        }, status_code=500)

# ==================== ACCOUNT MANAGEMENT ====================

def validate_password(password: str) -> Tuple[bool, str]:
    """Kiểm tra mật khẩu có thỏa mãn password policy không"""
    if len(password) < 8:
        return False, "Mật khẩu phải có ít nhất 8 ký tự"
    
    if not re.search(r'[A-Z]', password):
        return False, "Mật khẩu phải có ít nhất 1 chữ in hoa (A-Z)"
    
    if not re.search(r'[a-z]', password):
        return False, "Mật khẩu phải có ít nhất 1 chữ thường (a-z)"
    
    if not re.search(r'[0-9]', password):
        return False, "Mật khẩu phải có ít nhất 1 chữ số (0-9)"
    
    return True, ""

@app.get("/statistics", response_class=HTMLResponse)
async def statistics_page(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Trang thống kê - đang xây dựng"""
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Lấy danh sách routes và vehicles để hiển thị trong dropdown của tab tài chính
    routes = db.query(Route).all()
    vehicles = db.query(Vehicle).all()
    
    return templates.TemplateResponse("statistics.html", {
        "request": request,
        "current_user": current_user,
        "routes": routes,
        "vehicles": vehicles
    })

@app.get("/statistics/finance", response_class=HTMLResponse)
async def statistics_finance_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    route_code: Optional[str] = None,
    license_plate: Optional[str] = None
):
    """Trang thống kê tài chính - tổng hợp doanh thu theo tuyến"""
    # Nếu chưa đăng nhập, redirect về login
    if current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    
    # Khởi tạo query cơ bản
    revenue_query = db.query(RevenueRecord).join(Route)
    
    # Áp dụng bộ lọc thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            revenue_query = revenue_query.filter(
                RevenueRecord.date >= from_date_obj,
                RevenueRecord.date <= to_date_obj
            )
        except ValueError:
            pass
    
    # Áp dụng các bộ lọc khác
    if route_code:
        revenue_query = revenue_query.filter(Route.route_code.ilike(f"%{route_code}%"))
    if license_plate:
        revenue_query = revenue_query.filter(RevenueRecord.license_plate.ilike(f"%{license_plate}%"))
    
    revenue_records = revenue_query.all()
    
    # Tính tổng hợp doanh thu theo mã tuyến
    revenue_by_route = {}
    for record in revenue_records:
        route_code_key = record.route.route_code if record.route else "N/A"
        if route_code_key not in revenue_by_route:
            revenue_by_route[route_code_key] = {
                'route_code': route_code_key,
                'total_revenue': 0
            }
        revenue_by_route[route_code_key]['total_revenue'] += record.total_amount or 0
    
    # Convert to list và sắp xếp
    revenue_summary = []
    for route_code_key, data in revenue_by_route.items():
        revenue_summary.append({
            'route_code': route_code_key,
            'total_revenue': data['total_revenue']
        })
    
    revenue_summary.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    # Lấy danh sách cho dropdown
    routes = db.query(Route).all()
    vehicles = db.query(Vehicle).all()
    
    # Template data
    template_data = {
        "request": request,
        "current_user": current_user,
        "revenue_summary": revenue_summary,
        "routes": routes,
        "vehicles": vehicles
    }
    
    # Chỉ thêm khi có giá trị và format ngày
    if from_date:
        template_data["from_date"] = from_date
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            template_data["from_date_formatted"] = from_date_obj.strftime('%d/%m/%Y')
        except:
            template_data["from_date_formatted"] = from_date
    if to_date:
        template_data["to_date"] = to_date
        try:
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            template_data["to_date_formatted"] = to_date_obj.strftime('%d/%m/%Y')
        except:
            template_data["to_date_formatted"] = to_date
    if route_code:
        template_data["route_code"] = route_code
    if license_plate:
        template_data["license_plate"] = license_plate
    
    return templates.TemplateResponse("statistics.html", template_data)

@app.get("/statistics/finance/details")
async def statistics_finance_details(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
    route_code: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    license_plate: Optional[str] = None
):
    """API trả về chi tiết doanh thu theo tuyến"""
    # Nếu chưa đăng nhập
    if current_user is None:
        return JSONResponse(
            status_code=401,
            content={"success": False, "message": "Chưa đăng nhập"}
        )
    
    # Khởi tạo query
    revenue_query = db.query(RevenueRecord).join(Route)
    
    # Áp dụng bộ lọc mã tuyến (bắt buộc)
    if route_code:
        revenue_query = revenue_query.filter(Route.route_code.ilike(f"%{route_code}%"))
    else:
        return JSONResponse(
            status_code=400,
            content={"success": False, "message": "Thiếu mã tuyến"}
        )
    
    # Áp dụng bộ lọc thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            revenue_query = revenue_query.filter(
                RevenueRecord.date >= from_date_obj,
                RevenueRecord.date <= to_date_obj
            )
        except ValueError:
            pass
    
    # Áp dụng bộ lọc biển số xe
    if license_plate:
        revenue_query = revenue_query.filter(RevenueRecord.license_plate.ilike(f"%{license_plate}%"))
    
    revenue_records = revenue_query.order_by(RevenueRecord.date.desc()).all()
    
    # Chuyển đổi sang dictionary
    details = []
    for record in revenue_records:
        # Tên tuyến: route.route_name
        route_name = record.route.route_name if record.route else 'N/A'
        # Mã tuyến: route.route_code
        route_code = record.route.route_code if record.route else 'N/A'
        # Lộ trình: record.route_name (cho tuyến tăng cường) hoặc route.route_name
        route_path = record.route_name or route_name
        details.append({
            'route_code': route_code,
            'route_name': route_name,
            'route_path': route_path,
            'distance_km': record.distance_km or 0,
            'unit_price': record.unit_price or 0,
            'bridge_fee': record.bridge_fee or 0,
            'loading_fee': record.loading_fee or 0,
            'late_penalty': record.late_penalty or 0,
            'total_amount': record.total_amount or 0,
            'license_plate': record.license_plate or 'N/A',
            'driver_name': record.driver_name or 'N/A',
            'date': record.date.strftime('%d/%m/%Y') if record.date else 'N/A'
        })
    
    return JSONResponse(content={
        "success": True,
        "data": details
    })

@app.get("/accounts", response_class=HTMLResponse)
async def accounts_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """(Removed) Trang Tài khoản đã được gỡ bỏ theo yêu cầu."""
    return templates.TemplateResponse("blank.html", {"request": request, "current_user": current_user})

# ==================== REMOVED: Account Management Routes ====================
# Các routes sau đã bị xóa để đơn giản hóa module tài khoản:
# - /accounts/add
# - /accounts/update/{account_id}
# - /accounts/reset-password/{account_id}
# - /accounts/lock/{account_id}
# - /accounts/unlock/{account_id}
# - /accounts/delete/{account_id}
# - /accounts/{account_id}/permissions
# - /user-management
# - /role-management
# - /permission-management
# - /api/users, /api/users/{user_id}/roles
# - /api/roles, /api/roles/{role_id}
# - /api/permissions, /api/roles/{role_id}/permissions
# 
# Chỉ giữ lại route /accounts để hiển thị danh sách tài khoản nội bộ (chỉ admin)

# ==================== ADMINISTRATIVE MODULE - DOCUMENTS ====================

@app.get("/administrative", response_class=HTMLResponse)
async def administrative_page(
    request: Request,
    db: Session = Depends(get_db),
    category: Optional[str] = None,
    status: Optional[str] = None,
    document_type: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    tax_period_month: Optional[int] = None,
    tax_period_year: Optional[int] = None,
    current_user = Depends(get_current_user)
):
    """Trang quản lý tài liệu hành chính"""
    # Check permission
    if not check_permission(db, current_user["id"], "/administrative", "view"):
        return RedirectResponse(url="/access-denied", status_code=303)
    
    # Build query
    query = db.query(Document)
    
    # Filter by category
    if category and category in ["legal", "administrative", "tax"]:
        query = query.filter(Document.category == category)
    
    # Filter by document type
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    # Filter by status
    if status and status in ["active", "expired", "archived"]:
        query = query.filter(Document.status == status)
    
    # Search by title (case-insensitive for SQLite)
    if search:
        query = query.filter(Document.title.like(f"%{search}%"))
    
    # Filter by date range (issued_date)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            query = query.filter(Document.issued_date >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            query = query.filter(Document.issued_date <= date_to_obj)
        except ValueError:
            pass
    
    # Filter by tax period (month/year) - for tax category documents
    if tax_period_month and tax_period_year:
        # Filter documents where issued_date matches the tax period
        # Create start and end dates for the month
        from calendar import monthrange
        start_date = date(tax_period_year, tax_period_month, 1)
        last_day = monthrange(tax_period_year, tax_period_month)[1]
        end_date = date(tax_period_year, tax_period_month, last_day)
        query = query.filter(
            Document.issued_date >= start_date,
            Document.issued_date <= end_date
        )
    
    # Order by created_at desc
    documents = query.order_by(Document.created_at.desc()).all()
    
    # Create a mapping of employee IDs to employee objects for quick lookup
    employee_ids = [doc.related_entity_id for doc in documents if doc.related_entity_type == "employee" and doc.related_entity_id]
    employees_dict = {}
    if employee_ids:
        employees_list = db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        employees_dict = {emp.id: emp for emp in employees_list}
    
    # Attach employee info to documents
    for doc in documents:
        if doc.related_entity_type == "employee" and doc.related_entity_id:
            doc.related_employee = employees_dict.get(doc.related_entity_id)
    
    # Get counts by category
    legal_count = db.query(Document).filter(Document.category == "legal").count()
    administrative_count = db.query(Document).filter(Document.category == "administrative").count()
    tax_count = db.query(Document).filter(Document.category == "tax").count()
    
    # Get unique document types for filter dropdown
    document_types = db.query(Document.document_type).distinct().order_by(Document.document_type).all()
    document_types_list = [dt[0] for dt in document_types if dt[0]]
    
    # Get employees for dropdown (for administrative documents)
    employees = db.query(Employee).filter(Employee.status == 1).order_by(Employee.name).all()
    
    return templates.TemplateResponse("administrative.html", {
        "request": request,
        "current_user": current_user,
        "documents": documents,
        "category": category,
        "status": status,
        "document_type": document_type,
        "search": search or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "tax_period_month": tax_period_month,
        "tax_period_year": tax_period_year,
        "legal_count": legal_count,
        "administrative_count": administrative_count,
        "tax_count": tax_count,
        "document_types": document_types_list,
        "employees": employees,
        "today": date.today(),
        "db": db  # Pass db for RBAC permission checks in templates
    })

@app.get("/api/documents", response_class=JSONResponse)
async def get_documents_api(
    request: Request,
    db: Session = Depends(get_db),
    category: Optional[str] = None,
    document_type: Optional[str] = None,
    status: Optional[str] = None,
    related_entity_type: Optional[str] = None,
    related_entity_id: Optional[int] = None,
    current_user = Depends(get_current_user)
):
    """API: Lấy danh sách documents"""
    
    try:
        query = db.query(Document)
        
        if category:
            query = query.filter(Document.category == category)
        if document_type:
            query = query.filter(Document.document_type == document_type)
        if status:
            query = query.filter(Document.status == status)
        if related_entity_type:
            query = query.filter(Document.related_entity_type == related_entity_type)
        if related_entity_id:
            query = query.filter(Document.related_entity_id == related_entity_id)
        
        documents = query.order_by(Document.created_at.desc()).all()
        
        result = []
        for doc in documents:
            result.append({
                "id": doc.id,
                "category": doc.category,
                "document_type": doc.document_type,
                "related_entity_type": doc.related_entity_type,
                "related_entity_id": doc.related_entity_id,
                "title": doc.title,
                "file_path": doc.file_path,
                "file_url": f"/{doc.file_path}",
                "issued_date": doc.issued_date.isoformat() if doc.issued_date else None,
                "expiry_date": doc.expiry_date.isoformat() if doc.expiry_date else None,
                "status": doc.status,
                "created_at": doc.created_at.isoformat() if doc.created_at else None
            })
        
        return JSONResponse({"success": True, "data": result})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# Removed all account management routes - see comment at line 10665
# Removed all RBAC routes - see comment at line 10665
# Removed all permission management routes - see comment at line 10665

# (cleanup) Removed leftover orphaned code block (was part of old /accounts/update route)

@app.post("/accounts/reset-password/{account_id}")
async def reset_password(*args, **kwargs):
    """(Removed)"""
    return JSONResponse({"success": False, "message": "Removed"}, status_code=404)

@app.post("/accounts/lock/{account_id}")
async def lock_account(*args, **kwargs):
    """(Removed)"""
    return JSONResponse({"success": False, "message": "Removed"}, status_code=404)

@app.post("/accounts/unlock/{account_id}")
async def unlock_account(*args, **kwargs):
    """(Removed)"""
    return JSONResponse({"success": False, "message": "Removed"}, status_code=404)

@app.post("/accounts/delete/{account_id}")
async def delete_account(*args, **kwargs):
    """(Removed)"""
    return JSONResponse({"success": False, "message": "Removed"}, status_code=404)

# ==================== PERMISSION MANAGEMENT ROUTES ====================

@app.get("/accounts/{account_id}/permissions")
async def get_user_permissions(
    account_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Lấy danh sách permissions của user - chỉ Admin"""
    if current_user is None or current_user["role"] != "Admin":
        return JSONResponse({
            "success": False,
            "message": "Không có quyền thực hiện thao tác này"
        }, status_code=403)
    
    try:
        account = db.query(Account).filter(Account.id == account_id).first()
        if not account:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy tài khoản"
            }, status_code=404)
        
        # Admin luôn có toàn quyền
        if account.role == "Admin":
            return JSONResponse({
                "success": True,
                "data": {
                    "is_admin": True,
                    "permissions": []
                }
            })
        
        # Lấy tất cả permissions
        all_permissions = db.query(Permission).order_by(Permission.page_path).all()
        
        # Lấy permissions của user
        user_permissions = db.query(UserPermission).filter(UserPermission.user_id == account_id).all()
        user_permission_ids = [up.permission_id for up in user_permissions]
        
        # Nếu user chưa có permissions nào, mặc định tất cả đều được phép
        # (has_permission = true cho tất cả)
        # Nếu user đã có permissions, chỉ những permission có trong user_permissions mới được phép
        is_new_user = len(user_permission_ids) == 0
        
        # Format dữ liệu
        permissions_data = []
        for perm in all_permissions:
            # Mặc định: nếu user mới (chưa có permissions), tất cả đều được phép
            # Nếu user đã có permissions, chỉ những permission có trong danh sách mới được phép
            has_permission = True if is_new_user else (perm.id in user_permission_ids)
            
            permissions_data.append({
                "id": perm.id,
                "name": perm.name,
                "description": perm.description,
                "page_path": perm.page_path,
                "action": perm.action,
                "has_permission": has_permission
            })
        
        return JSONResponse({
            "success": True,
            "data": {
                "is_admin": False,
                "permissions": permissions_data
            }
        })
        
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi lấy permissions: {str(e)}"
        }, status_code=500)

@app.post("/accounts/{account_id}/permissions")
async def update_user_permissions(*args, **kwargs):
    """(Removed)"""
    return JSONResponse({"success": False, "message": "Removed"}, status_code=404)

# ==================== RBAC ROUTES ====================

@app.get("/api/users", response_class=JSONResponse)
async def get_users_api(
    db: Session = Depends(get_db)
):
    """API: Lấy danh sách users với roles"""
    try:
        users = db.query(Account).all()
        result = []
        for user in users:
            user_roles = db.query(UserRole).filter(UserRole.user_id == user.id).all()
            roles = [{"id": ur.role_id, "name": ur.role.name} for ur in user_roles]
            result.append({
                "id": user.id,
                "username": user.username,
                "full_name": user.full_name,
                "email": user.email,
                "phone": user.phone,
                "status": user.status,
                "is_locked": user.is_locked,
                "roles": roles,
                "created_at": user.created_at.isoformat() if user.created_at else None
            })
        return JSONResponse({"success": True, "data": result})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/api/users/{user_id}/roles")
async def assign_user_roles(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """API: Gán roles cho user"""
    try:
        account = db.query(Account).filter(Account.id == user_id).first()
        if not account:
            return JSONResponse({"success": False, "message": "User not found"}, status_code=404)
        
        body = await request.json()
        role_ids = body.get("role_ids", [])
        
        if not isinstance(role_ids, list):
            return JSONResponse({"success": False, "message": "role_ids must be a list"}, status_code=400)
        
        # Delete existing roles
        db.query(UserRole).filter(UserRole.user_id == user_id).delete()
        
        # Add new roles
        for role_id in role_ids:
            role = db.query(Role).filter(Role.id == role_id).first()
            if role:
                user_role = UserRole(
                    user_id=user_id,
                    role_id=role_id,
                    assigned_by=0
                )
                db.add(user_role)
        
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=0,
            action="update",
            entity_type="user_roles",
            entity_id=user_id,
            new_values={"role_ids": role_ids},
            description=f"Updated roles for user: {account.username}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "message": "Roles assigned successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.get("/api/roles", response_class=JSONResponse)
async def get_roles_api(
    db: Session = Depends(get_db)
):
    """API: Lấy danh sách roles"""
    try:
        roles = db.query(Role).all()
        result = [{
            "id": r.id,
            "name": r.name,
            "description": r.description,
            "is_system_role": bool(r.is_system_role),
            "created_at": r.created_at.isoformat() if r.created_at else None
        } for r in roles]
        return JSONResponse({"success": True, "data": result})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/api/roles")
async def create_role(
    request: Request,
    db: Session = Depends(get_db)
):
    """API: Tạo role mới"""
    try:
        body = await request.json()
        name = body.get("name")
        description = body.get("description", "")
        
        if not name:
            return JSONResponse({"success": False, "message": "Role name is required"}, status_code=400)
        
        # Check if role exists
        existing = db.query(Role).filter(Role.name == name).first()
        if existing:
            return JSONResponse({"success": False, "message": "Role already exists"}, status_code=400)
        
        role = Role(name=name, description=description, is_system_role=0)
        db.add(role)
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=0,
            action="create",
            entity_type="role",
            entity_id=role.id,
            new_values={"name": name, "description": description},
            description=f"Created role: {name}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "data": {"id": role.id, "name": role.name}})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.put("/api/roles/{role_id}")
async def update_role(
    role_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """API: Cập nhật role"""
    try:
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            return JSONResponse({"success": False, "message": "Role not found"}, status_code=404)
        
        if role.is_system_role:
            return JSONResponse({"success": False, "message": "Cannot modify system role"}, status_code=400)
        
        body = await request.json()
        old_values = {"name": role.name, "description": role.description}
        
        if "name" in body:
            # Check if new name conflicts
            existing = db.query(Role).filter(Role.name == body["name"], Role.id != role_id).first()
            if existing:
                return JSONResponse({"success": False, "message": "Role name already exists"}, status_code=400)
            role.name = body["name"]
        
        if "description" in body:
            role.description = body["description"]
        
        role.updated_at = datetime.utcnow()
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=0,
            action="update",
            entity_type="role",
            entity_id=role_id,
            old_values=old_values,
            new_values={"name": role.name, "description": role.description},
            description=f"Updated role: {role.name}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "message": "Role updated successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.delete("/api/roles/{role_id}")
async def delete_role(
    role_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """API: Xóa role"""
    try:
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            return JSONResponse({"success": False, "message": "Role not found"}, status_code=404)
        
        if role.is_system_role:
            return JSONResponse({"success": False, "message": "Cannot delete system role"}, status_code=400)
        
        role_name = role.name
        db.delete(role)
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=0,
            action="delete",
            entity_type="role",
            entity_id=role_id,
            old_values={"name": role_name},
            description=f"Deleted role: {role_name}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "message": "Role deleted successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.get("/api/permissions", response_class=JSONResponse)
async def get_permissions_api(
    db: Session = Depends(get_db)
):
    """API: Lấy danh sách permissions, nhóm theo page"""
    try:
        permissions = db.query(Permission).order_by(Permission.page_path, Permission.action).all()
        
        # Group by page_path
        pages = {}
        for perm in permissions:
            if perm.page_path not in pages:
                pages[perm.page_path] = []
            pages[perm.page_path].append({
                "id": perm.id,
                "name": perm.name,
                "description": perm.description,
                "action": perm.action
            })
        
        result = [{"page_path": path, "permissions": perms} for path, perms in pages.items()]
        return JSONResponse({"success": True, "data": result})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.get("/api/roles/{role_id}/permissions", response_class=JSONResponse)
async def get_role_permissions(
    role_id: int,
    db: Session = Depends(get_db)
):
    """API: Lấy permissions của role"""
    try:
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            return JSONResponse({"success": False, "message": "Role not found"}, status_code=404)
        
        role_permissions = db.query(RolePermission).filter(RolePermission.role_id == role_id).all()
        permission_ids = [rp.permission_id for rp in role_permissions]
        
        return JSONResponse({"success": True, "data": {"permission_ids": permission_ids}})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/api/roles/{role_id}/permissions")
async def update_role_permissions(
    role_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """API: Cập nhật permissions của role"""
    try:
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            return JSONResponse({"success": False, "message": "Role not found"}, status_code=404)
        
        body = await request.json()
        permission_ids = body.get("permission_ids", [])
        
        if not isinstance(permission_ids, list):
            return JSONResponse({"success": False, "message": "permission_ids must be a list"}, status_code=400)
        
        # Delete existing permissions
        db.query(RolePermission).filter(RolePermission.role_id == role_id).delete()
        
        # Add new permissions
        for perm_id in permission_ids:
            permission = db.query(Permission).filter(Permission.id == perm_id).first()
            if permission:
                role_permission = RolePermission(role_id=role_id, permission_id=perm_id)
                db.add(role_permission)
        
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=0,
            action="update",
            entity_type="role_permissions",
            entity_id=role_id,
            new_values={"permission_ids": permission_ids},
            description=f"Updated permissions for role: {role.name}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "message": "Role permissions updated successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# ==================== ADMINISTRATIVE MODULE - DOCUMENTS ====================

@app.get("/administrative", response_class=HTMLResponse)
async def administrative_page(
    request: Request,
    db: Session = Depends(get_db),
    category: Optional[str] = None,
    status: Optional[str] = None,
    document_type: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    tax_period_month: Optional[int] = None,
    tax_period_year: Optional[int] = None,
    current_user = Depends(get_current_user)
):
    """Trang quản lý tài liệu hành chính"""
    
    # Build query
    query = db.query(Document)
    
    # Filter by category
    if category and category in ["legal", "administrative", "tax"]:
        query = query.filter(Document.category == category)
    
    # Filter by document type
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    # Filter by status
    if status and status in ["active", "expired", "archived"]:
        query = query.filter(Document.status == status)
    
    # Search by title (case-insensitive for SQLite)
    if search:
        query = query.filter(Document.title.like(f"%{search}%"))
    
    # Filter by date range (issued_date)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            query = query.filter(Document.issued_date >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            query = query.filter(Document.issued_date <= date_to_obj)
        except ValueError:
            pass
    
    # Filter by tax period (month/year) - for tax category documents
    if tax_period_month and tax_period_year:
        # Filter documents where issued_date matches the tax period
        # Create start and end dates for the month
        from calendar import monthrange
        start_date = date(tax_period_year, tax_period_month, 1)
        last_day = monthrange(tax_period_year, tax_period_month)[1]
        end_date = date(tax_period_year, tax_period_month, last_day)
        query = query.filter(
            Document.issued_date >= start_date,
            Document.issued_date <= end_date
        )
    
    # Order by created_at desc
    documents = query.order_by(Document.created_at.desc()).all()
    
    # Create a mapping of employee IDs to employee objects for quick lookup
    employee_ids = [doc.related_entity_id for doc in documents if doc.related_entity_type == "employee" and doc.related_entity_id]
    employees_dict = {}
    if employee_ids:
        employees_list = db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        employees_dict = {emp.id: emp for emp in employees_list}
    
    # Attach employee info to documents
    for doc in documents:
        if doc.related_entity_type == "employee" and doc.related_entity_id:
            doc.related_employee = employees_dict.get(doc.related_entity_id)
    
    # Get counts by category
    legal_count = db.query(Document).filter(Document.category == "legal").count()
    administrative_count = db.query(Document).filter(Document.category == "administrative").count()
    tax_count = db.query(Document).filter(Document.category == "tax").count()
    
    # Get unique document types for filter dropdown
    document_types = db.query(Document.document_type).distinct().order_by(Document.document_type).all()
    document_types_list = [dt[0] for dt in document_types if dt[0]]
    
    # Get employees for dropdown (for administrative documents)
    employees = db.query(Employee).filter(Employee.status == 1).order_by(Employee.name).all()
    
    return templates.TemplateResponse("administrative.html", {
        "request": request,
        "current_user": current_user,
        "documents": documents,
        "category": category,
        "status": status,
        "document_type": document_type,
        "search": search or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "tax_period_month": tax_period_month,
        "tax_period_year": tax_period_year,
        "legal_count": legal_count,
        "administrative_count": administrative_count,
        "tax_count": tax_count,
        "document_types": document_types_list,
        "employees": employees,
        "today": date.today(),
        "db": db  # Pass db for RBAC permission checks in templates
    })

@app.get("/api/documents", response_class=JSONResponse)
async def get_documents_api(
    request: Request,
    db: Session = Depends(get_db),
    category: Optional[str] = None,
    document_type: Optional[str] = None,
    status: Optional[str] = None,
    related_entity_type: Optional[str] = None,
    related_entity_id: Optional[int] = None,
    current_user = Depends(get_current_user)
):
    """API: Lấy danh sách documents"""
    
    try:
        query = db.query(Document)
        
        if category:
            query = query.filter(Document.category == category)
        if document_type:
            query = query.filter(Document.document_type == document_type)
        if status:
            query = query.filter(Document.status == status)
        if related_entity_type:
            query = query.filter(Document.related_entity_type == related_entity_type)
        if related_entity_id:
            query = query.filter(Document.related_entity_id == related_entity_id)
        
        documents = query.order_by(Document.created_at.desc()).all()
        
        result = []
        for doc in documents:
            result.append({
                "id": doc.id,
                "category": doc.category,
                "document_type": doc.document_type,
                "related_entity_type": doc.related_entity_type,
                "related_entity_id": doc.related_entity_id,
                "title": doc.title,
                "file_path": doc.file_path,
                "file_url": f"/{doc.file_path}",
                "issued_date": doc.issued_date.isoformat() if doc.issued_date else None,
                "expiry_date": doc.expiry_date.isoformat() if doc.expiry_date else None,
                "status": doc.status,
                "description": doc.description,
                "notes": doc.notes,
                "created_at": doc.created_at.isoformat() if doc.created_at else None,
                "created_by": doc.created_by,
                "creator_name": doc.creator.username if doc.creator else None
            })
        
        return JSONResponse({"success": True, "data": result})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.get("/api/documents/{document_id}", response_class=JSONResponse)
async def get_document_api(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API: Lấy thông tin một document"""
    if not check_permission(db, current_user["id"], "/administrative", "view"):
        return JSONResponse({"success": False, "message": "Insufficient permissions"}, status_code=403)
    
    try:
        doc = db.query(Document).filter(Document.id == document_id).first()
        if not doc:
            return JSONResponse({"success": False, "message": "Document not found"}, status_code=404)
        
        return JSONResponse({
            "success": True,
            "data": {
                "id": doc.id,
                "category": doc.category,
                "document_type": doc.document_type,
                "related_entity_type": doc.related_entity_type,
                "related_entity_id": doc.related_entity_id,
                "title": doc.title,
                "file_path": doc.file_path,
                "file_url": f"/{doc.file_path}",
                "issued_date": doc.issued_date.isoformat() if doc.issued_date else None,
                "expiry_date": doc.expiry_date.isoformat() if doc.expiry_date else None,
                "status": doc.status,
                "description": doc.description,
                "notes": doc.notes,
                "created_at": doc.created_at.isoformat() if doc.created_at else None,
                "created_by": doc.created_by,
                "creator_name": doc.creator.username if doc.creator else None
            }
        })
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/api/documents")
async def create_document(
    request: Request,
    db: Session = Depends(get_db),
    category: str = Form(...),
    document_type: str = Form(...),
    title: str = Form(...),
    file: UploadFile = File(...),
    related_entity_type: Optional[str] = Form(None),
    related_entity_id: Optional[int] = Form(None),
    issued_date: Optional[str] = Form(None),
    expiry_date: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    current_user = Depends(get_current_user)
):
    """API: Tạo document mới"""
    
    try:
        # Validate category
        if category not in ["legal", "administrative", "tax"]:
            return JSONResponse({"success": False, "message": "Invalid category"}, status_code=400)
        
        # Validate file type
        is_valid, error_msg = validate_document_file(file.filename)
        if not is_valid:
            return JSONResponse({"success": False, "message": error_msg}, status_code=400)
        
        # Validate file size (10MB limit)
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        file_content = await file.read()
        file_size = len(file_content)
        if file_size > MAX_FILE_SIZE:
            return JSONResponse({"success": False, "message": f"File size exceeds 10MB limit. File size: {(file_size / 1024 / 1024):.2f}MB"}, status_code=400)
        
        # Reset file pointer for saving
        await file.seek(0)
        
        # Ensure directories exist
        ensure_document_dirs()
        
        # Determine folder based on category and document type
        folder = get_document_category_folder(category, document_type)
        category_dir = os.path.join(DOCUMENTS_UPLOAD_DIR, folder)
        ensure_directory_exists(category_dir)
        
        # Generate unique filename while preserving original name
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_ext = os.path.splitext(file.filename)[1]
        # Sanitize original filename
        original_name = os.path.splitext(file.filename)[0]
        # Remove/replace unsafe characters
        safe_original_name = re.sub(r'[^\w\s-]', '', original_name).strip()
        safe_original_name = re.sub(r'[-\s]+', '-', safe_original_name)
        safe_filename = f"{safe_original_name}_{timestamp}{file_ext}"
        file_path = os.path.join(category_dir, safe_filename)
        
        # Save file (reuse content already read for size validation)
        with open(file_path, "wb") as buffer:
            buffer.write(file_content)
        
        # Relative path for database
        relative_path = file_path.replace("\\", "/")
        
        # Parse dates
        issued_date_obj = None
        if issued_date:
            try:
                issued_date_obj = datetime.strptime(issued_date, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        expiry_date_obj = None
        if expiry_date:
            try:
                expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        # Determine status based on expiry date
        status = "active"
        if expiry_date_obj and expiry_date_obj < date.today():
            status = "expired"
        
        # Create document
        document = Document(
            category=category,
            document_type=document_type,
            related_entity_type=related_entity_type if related_entity_type else None,
            related_entity_id=related_entity_id if related_entity_id else None,
            title=title,
            file_path=relative_path,
            issued_date=issued_date_obj,
            expiry_date=expiry_date_obj,
            status=status,
            description=description,
            notes=notes,
            created_by=current_user["id"],
            updated_by=current_user["id"]
        )
        
        db.add(document)
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=current_user["id"],
            action="create",
            entity_type="document",
            entity_id=document.id,
            new_values={"title": title, "category": category, "document_type": document_type},
            description=f"Created document: {title}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({
            "success": True,
            "message": "Document created successfully",
            "data": {"id": document.id}
        })
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.put("/api/documents/{document_id}")
async def update_document(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    category: Optional[str] = Form(None),
    document_type: Optional[str] = Form(None),
    title: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    related_entity_type: Optional[str] = Form(None),
    related_entity_id: Optional[int] = Form(None),
    issued_date: Optional[str] = Form(None),
    expiry_date: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    current_user = Depends(get_current_user)
):
    """API: Cập nhật document"""
    
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            return JSONResponse({"success": False, "message": "Document not found"}, status_code=404)
        
        old_values = {
            "category": document.category,
            "document_type": document.document_type,
            "title": document.title,
            "status": document.status
        }
        
        # Update fields
        if category and category in ["legal", "administrative", "tax"]:
            document.category = category
        if document_type:
            document.document_type = document_type
        if title:
            document.title = title
        if related_entity_type is not None:
            document.related_entity_type = related_entity_type if related_entity_type else None
        if related_entity_id is not None:
            document.related_entity_id = related_entity_id if related_entity_id else None
        if description is not None:
            document.description = description
        if notes is not None:
            document.notes = notes
        if status and status in ["active", "expired", "archived"]:
            document.status = status
        
        # Parse dates
        if issued_date:
            try:
                document.issued_date = datetime.strptime(issued_date, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        if expiry_date:
            try:
                document.expiry_date = datetime.strptime(expiry_date, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        # Update file if provided (check both file object and filename to ensure a real file was uploaded)
        if file and file.filename:
            # Validate file type
            is_valid, error_msg = validate_document_file(file.filename)
            if not is_valid:
                return JSONResponse({"success": False, "message": error_msg}, status_code=400)
            
            # Validate file size (10MB limit)
            MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
            file_content = await file.read()
            file_size = len(file_content)
            if file_size > MAX_FILE_SIZE:
                return JSONResponse({"success": False, "message": f"File size exceeds 10MB limit. File size: {(file_size / 1024 / 1024):.2f}MB"}, status_code=400)
            
            # Delete old file
            if document.file_path and os.path.exists(document.file_path):
                try:
                    os.remove(document.file_path)
                except Exception:
                    pass
            
            # Save new file
            folder = get_document_category_folder(document.category, document.document_type)
            category_dir = os.path.join(DOCUMENTS_UPLOAD_DIR, folder)
            ensure_directory_exists(category_dir)
            
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            file_ext = os.path.splitext(file.filename)[1]
            # Preserve original filename
            original_name = os.path.splitext(file.filename)[0]
            safe_original_name = re.sub(r'[^\w\s-]', '', original_name).strip()
            safe_original_name = re.sub(r'[-\s]+', '-', safe_original_name)
            safe_filename = f"{safe_original_name}_{timestamp}{file_ext}"
            file_path = os.path.join(category_dir, safe_filename)
            
            with open(file_path, "wb") as buffer:
                buffer.write(file_content)
            
            document.file_path = file_path.replace("\\", "/")
        
        # Update status based on expiry date
        if document.expiry_date and document.expiry_date < date.today():
            document.status = "expired"
        
        document.updated_by = current_user["id"]
        document.updated_at = datetime.utcnow()
        
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=current_user["id"],
            action="update",
            entity_type="document",
            entity_id=document_id,
            old_values=old_values,
            new_values={
                "category": document.category,
                "document_type": document.document_type,
                "title": document.title,
                "status": document.status
            },
            description=f"Updated document: {document.title}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "message": "Document updated successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.delete("/api/documents/{document_id}")
async def delete_document(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """API: Xóa document"""
    
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            return JSONResponse({"success": False, "message": "Document not found"}, status_code=404)
        
        # Delete file
        if document.file_path and os.path.exists(document.file_path):
            try:
                os.remove(document.file_path)
            except Exception as e:
                print(f"Error deleting file: {e}")
        
        title = document.title
        db.delete(document)
        db.commit()
        
        # Audit log
        create_audit_log(
            db=db,
            user_id=current_user["id"],
            action="delete",
            entity_type="document",
            entity_id=document_id,
            old_values={"title": title},
            description=f"Deleted document: {title}",
            ip_address=get_client_ip(request)
        )
        
        return JSONResponse({"success": True, "message": "Document deleted successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# Document view/download/print routes
@app.get("/documents/view/{document_id}", response_class=HTMLResponse)
async def view_document(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """View document in browser"""
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            return HTMLResponse("<h1>Document not found</h1>", status_code=404)
        
        if not os.path.exists(document.file_path):
            return HTMLResponse("<h1>File not found on server</h1>", status_code=404)
        
        file_ext = os.path.splitext(document.file_path)[1].lower()
        
        # For images, display inline with HTML page
        if file_ext in [".jpg", ".jpeg", ".png"]:
            # Get the file URL - need to serve it via a route
            file_url = f"/documents/file/{document_id}"
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{document.title}</title>
                <meta charset="UTF-8">
                <style>
                    * {{
                        margin: 0;
                        padding: 0;
                        box-sizing: border-box;
                    }}
                    body {{
                        font-family: Arial, sans-serif;
                        background: #f5f5f5;
                        padding: 20px;
                    }}
                    .container {{
                        max-width: 1200px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 8px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                        overflow: hidden;
                    }}
                    .header {{
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        color: white;
                        padding: 20px;
                        display: flex;
                        justify-content: space-between;
                        align-items: center;
                    }}
                    .header h1 {{
                        font-size: 24px;
                        font-weight: 600;
                    }}
                    .controls {{
                        display: flex;
                        gap: 10px;
                    }}
                    .btn {{
                        padding: 10px 20px;
                        background: rgba(255,255,255,0.2);
                        color: white;
                        border: 1px solid rgba(255,255,255,0.3);
                        border-radius: 5px;
                        cursor: pointer;
                        font-size: 14px;
                        text-decoration: none;
                        display: inline-flex;
                        align-items: center;
                        gap: 8px;
                        transition: background 0.2s;
                    }}
                    .btn:hover {{
                        background: rgba(255,255,255,0.3);
                    }}
                    .image-container {{
                        padding: 20px;
                        text-align: center;
                        background: #fafafa;
                        min-height: 400px;
                        display: flex;
                        align-items: center;
                        justify-content: center;
                    }}
                    .image-container img {{
                        max-width: 100%;
                        max-height: calc(100vh - 200px);
                        height: auto;
                        border-radius: 4px;
                        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
                        cursor: zoom-in;
                    }}
                    .image-container img.zoomed {{
                        max-width: none;
                        max-height: none;
                        width: auto;
                        height: auto;
                        cursor: zoom-out;
                    }}
                    .info {{
                        padding: 15px 20px;
                        border-top: 1px solid #e1e8ed;
                        background: #f8f9fa;
                        font-size: 14px;
                        color: #666;
                    }}
                    @media print {{
                        .header, .controls, .info {{
                            display: none;
                        }}
                        .image-container {{
                            padding: 0;
                            background: white;
                        }}
                        .image-container img {{
                            max-width: 100%;
                            max-height: 100vh;
                            box-shadow: none;
                        }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>📄 {document.title}</h1>
                        <div class="controls">
                            <a href="/documents/download/{document_id}" class="btn">
                                <span>⬇️</span> Download
                            </a>
                            <a href="/documents/print/{document_id}" target="_blank" class="btn">
                                <span>🖨️</span> Print
                            </a>
                            <button onclick="window.close()" class="btn">
                                <span>✕</span> Close
                            </button>
                        </div>
                    </div>
                    <div class="image-container">
                        <img src="{file_url}" alt="{document.title}" id="documentImage" onclick="toggleZoom(this)">
                    </div>
                    <div class="info">
                        <strong>Category:</strong> {document.category.title()} | 
                        <strong>Type:</strong> {document.document_type} | 
                        <strong>Uploaded:</strong> {document.created_at.strftime('%Y-%m-%d %H:%M') if document.created_at else 'N/A'}
                    </div>
                </div>
                <script>
                    function toggleZoom(img) {{
                        img.classList.toggle('zoomed');
                    }}
                </script>
            </body>
            </html>
            """
            return HTMLResponse(html_content)
        # For PDF, serve inline
        elif file_ext == ".pdf":
            from fastapi.responses import FileResponse
            return FileResponse(
                document.file_path,
                media_type="application/pdf",
                headers={"Content-Disposition": "inline"}
            )
        else:
            # For DOC/DOCX, redirect to download
            return RedirectResponse(url=f"/documents/download/{document_id}", status_code=303)
    
    except Exception as e:
        return HTMLResponse(f"<h1>Error: {str(e)}</h1>", status_code=500)

@app.get("/documents/file/{document_id}")
async def get_document_file(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Serve document file directly (for images)"""
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if not os.path.exists(document.file_path):
            raise HTTPException(status_code=404, detail="File not found on server")
        
        from fastapi.responses import FileResponse
        
        file_ext = os.path.splitext(document.file_path)[1].lower()
        media_types = {
            ".pdf": "application/pdf",
            ".doc": "application/msword",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png"
        }
        media_type = media_types.get(file_ext, "application/octet-stream")
        
        return FileResponse(
            document.file_path,
            media_type=media_type,
            headers={"Content-Disposition": "inline"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/documents/download/{document_id}")
async def download_document(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Download document file"""
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if not os.path.exists(document.file_path):
            raise HTTPException(status_code=404, detail="File not found on server")
        
        from fastapi.responses import FileResponse
        
        file_ext = os.path.splitext(document.file_path)[1].lower()
        media_types = {
            ".pdf": "application/pdf",
            ".doc": "application/msword",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png"
        }
        media_type = media_types.get(file_ext, "application/octet-stream")
        
        # Get original filename from file_path or use title
        filename = os.path.basename(document.file_path)
        # Try to extract original name (before timestamp)
        if "_" in filename:
            parts = filename.rsplit("_", 1)
            if len(parts) == 2 and parts[1].replace(file_ext, "").isdigit():
                filename = parts[0] + file_ext
        
        # Use RFC 5987 encoding for filenames with non-ASCII characters
        # Format: filename*=UTF-8''<url-encoded-filename>
        encoded_filename = quote(filename, safe='')
        content_disposition = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        return FileResponse(
            document.file_path,
            media_type=media_type,
            headers={"Content-Disposition": content_disposition}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/documents/print/{document_id}", response_class=HTMLResponse)
async def print_document(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Print document page (opens document in new window with print dialog)"""
    try:
        document = db.query(Document).filter(Document.id == document_id).first()
        if not document:
            return HTMLResponse("<h1>Document not found</h1>", status_code=404)
        
        file_ext = os.path.splitext(document.file_path)[1].lower()
        
        # For images, display with print support
        if file_ext in [".jpg", ".jpeg", ".png"]:
            file_url = f"/documents/file/{document_id}"
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Print {document.title}</title>
                <meta charset="UTF-8">
                <style>
                    * {{
                        margin: 0;
                        padding: 0;
                        box-sizing: border-box;
                    }}
                    body {{
                        font-family: Arial, sans-serif;
                        background: #f5f5f5;
                        padding: 20px;
                    }}
                    .print-controls {{
                        position: fixed;
                        top: 10px;
                        right: 10px;
                        z-index: 1000;
                        background: white;
                        padding: 10px;
                        border-radius: 5px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.2);
                    }}
                    button {{
                        padding: 10px 20px;
                        margin: 5px;
                        background: #667eea;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        cursor: pointer;
                        font-size: 14px;
                    }}
                    button:hover {{
                        background: #5568d3;
                    }}
                    .image-container {{
                        text-align: center;
                        padding: 20px;
                        background: white;
                        border-radius: 8px;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    }}
                    .image-container img {{
                        max-width: 100%;
                        height: auto;
                    }}
                    @media print {{
                        .print-controls {{
                            display: none;
                        }}
                        body {{
                            padding: 0;
                            background: white;
                        }}
                        .image-container {{
                            box-shadow: none;
                            padding: 0;
                        }}
                        .image-container img {{
                            max-width: 100%;
                            max-height: 100vh;
                        }}
                    }}
                </style>
            </head>
            <body>
                <div class="print-controls">
                    <button onclick="window.print()">🖨️ Print</button>
                    <button onclick="window.close()">✕ Close</button>
                </div>
                <div class="image-container">
                    <img src="{file_url}" alt="{document.title}">
                </div>
            </body>
            </html>
            """
            return HTMLResponse(html_content)
        # For PDF, embed with print button
        elif file_ext == ".pdf":
            file_url = f"/documents/view/{document_id}"
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Print {document.title}</title>
                <style>
                    body {{ margin: 0; padding: 20px; }}
                    .print-controls {{
                        position: fixed;
                        top: 10px;
                        right: 10px;
                        z-index: 1000;
                        background: white;
                        padding: 10px;
                        border-radius: 5px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.2);
                    }}
                    button {{
                        padding: 10px 20px;
                        margin: 5px;
                        background: #667eea;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        cursor: pointer;
                        font-size: 14px;
                    }}
                    button:hover {{ background: #5568d3; }}
                    iframe {{
                        width: 100%;
                        height: calc(100vh - 40px);
                        border: none;
                    }}
                </style>
            </head>
            <body>
                <div class="print-controls">
                    <button onclick="window.print()">🖨️ Print</button>
                    <button onclick="window.close()">✕ Close</button>
                </div>
                <iframe src="{file_url}"></iframe>
            </body>
            </html>
            """
            return HTMLResponse(html_content)
        else:
            # For DOC/DOCX, show download message
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Print {document.title}</title>
                <style>
                    body {{
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        height: 100vh;
                        margin: 0;
                        font-family: Arial, sans-serif;
                    }}
                    .message {{
                        text-align: center;
                        padding: 40px;
                        background: #f5f5f5;
                        border-radius: 10px;
                        max-width: 500px;
                    }}
                    button {{
                        padding: 12px 24px;
                        margin: 10px;
                        background: #667eea;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        cursor: pointer;
                        font-size: 16px;
                    }}
                    button:hover {{ background: #5568d3; }}
                </style>
            </head>
            <body>
                <div class="message">
                    <h2>📄 {document.title}</h2>
                    <p>This document type cannot be printed directly in the browser.</p>
                    <p>Please download the file and print it using your document viewer.</p>
                    <button onclick="window.location.href='/documents/download/{document_id}'">Download File</button>
                    <button onclick="window.close()">Close</button>
                </div>
            </body>
            </html>
            """
            return HTMLResponse(html_content)
    
    except Exception as e:
        return HTMLResponse(f"<h1>Error: {str(e)}</h1>", status_code=500)

# Khởi tạo permissions sẽ được gọi sau khi initialize_permissions được định nghĩa
def init_permissions_on_startup():
    """Khởi tạo permissions khi khởi động ứng dụng - chỉ chạy nếu migration thành công"""
    # Kiểm tra migration đã thành công chưa
    global rbac_migration_success
    
    # Nếu migration thất bại, không chạy init permissions
    if not rbac_migration_success:
        print("Skipping permissions initialization: RBAC migration failed or not completed")
        return
    
    db = SessionLocal()
    try:
        initialize_permissions(db)
    except Exception as e:
        print(f"Error initializing permissions: {e}")
    finally:
        db.close()

# Chạy khởi tạo permissions khi khởi động ứng dụng (chỉ nếu migration thành công)
init_permissions_on_startup()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
